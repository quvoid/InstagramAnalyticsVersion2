"""
Scan GRT Jewellers Facebook Page (Posts, Videos, Collabs, Boosted Partnership posts)
"""

import sys, re, json, html
from datetime import datetime, timezone
from curl_cffi import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

session = requests.Session(impersonate="chrome120")
headers = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    "accept-language": "en-US,en;q=0.9",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
}

print(f"\n{'='*75}")
print("  Scanning Facebook Page: https://www.facebook.com/grtjewellers/")
print(f"{'='*75}\n")

target_urls = [
    ("Main Page", "https://www.facebook.com/grtjewellers/"),
    ("Videos Tab", "https://www.facebook.com/grtjewellers/videos/"),
    ("Posts Tab", "https://www.facebook.com/grtjewellers/posts/"),
    ("Mobile Videos", "https://m.facebook.com/grtjewellers/videos/"),
]

all_extracted_posts = []
seen_ids = set()

def scan_json_tree(obj, source_tag=""):
    if isinstance(obj, dict):
        # Look for story / video / node representations
        is_node = False
        post_id = obj.get("post_id") or obj.get("id") or obj.get("video_id")
        ts = obj.get("creation_time") or obj.get("publish_time") or 0
        msg_obj = obj.get("message") or obj.get("savable_description") or obj.get("name")
        feedback = obj.get("feedback") or {}
        
        # Check comet section nesting
        if not ts and "comet_sections" in obj:
            try:
                ts = obj["comet_sections"]["context_layout"]["story"]["comet_sections"]["metadata"][0]["story"]["creation_time"]
            except Exception:
                pass
                
        if not msg_obj and "comet_sections" in obj:
            try:
                msg_obj = obj["comet_sections"]["content"]["story"]["message"]
            except Exception:
                pass
        
        msg = ""
        if isinstance(msg_obj, dict):
            msg = msg_obj.get("text", "")
        elif isinstance(msg_obj, str):
            msg = msg_obj
            
        reacts = 0
        comments = 0
        shares = 0
        if isinstance(feedback, dict):
            r_c = feedback.get("reaction_count")
            reacts = r_c.get("count", 0) if isinstance(r_c, dict) else (r_c or 0)
            
            c_c = feedback.get("comments_count") or feedback.get("comment_count")
            comments = c_c.get("total_count", 0) if isinstance(c_c, dict) else (c_c or 0)
            
            s_c = feedback.get("share_count")
            shares = s_c.get("count", 0) if isinstance(s_c, dict) else (s_c or 0)
            
        views = obj.get("video_view_count") or obj.get("play_count") or 0
        url = obj.get("url") or obj.get("permalink_url") or ""
        
        # Actors & Sponsors
        actors = obj.get("actors", [])
        actor_names = [a.get("name") for a in actors if isinstance(a, dict) and a.get("name")]
        
        sponsors = obj.get("sponsor_tags") or obj.get("branded_content_sponsor_relationship") or []
        sponsor_names = [s.get("sponsor", {}).get("name") for s in sponsors if isinstance(s, dict)]
        
        # If there's valid post content
        if (msg or reacts or views or ts) and (post_id or url):
            key = str(post_id) if post_id else url
            if key and key not in seen_ids and len(key) > 5:
                seen_ids.add(key)
                
                dt = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d") if ts else "N/A"
                if not url and post_id:
                    url = f"https://www.facebook.com/grtjewellers/posts/{post_id}"
                
                # Check for partnership / influencer keywords in caption or coauthors
                msg_l = msg.lower()
                partner_matched = []
                known_partners = [
                    ("R. Ashwin / Prithi Ashwin", ["ashwin", "prithi"]),
                    ("Athulya Ravi", ["athulya"]),
                    ("Siri Hanmanth", ["siri hanmanth", "sirihanmanth"]),
                    ("Roshni Haripriyan", ["roshni"]),
                    ("Megha Shetty", ["megha shetty"]),
                    ("Faria Abdullah", ["faria"]),
                    ("Niharika Konidela", ["niharika"]),
                    ("Shankar Mahadevan", ["shankar mahadevan"]),
                    ("Ritu Varma", ["ritu varma"]),
                    ("Divya Uruduga", ["divya uruduga"]),
                    ("Teju Ashwini", ["teju ashwini"]),
                    ("Bhumika Basavaraj", ["bhumika"]),
                    ("Platinum Days of Love / Men of Platinum", ["platinum", "menofplatinum", "evara"]),
                    ("Event Art / Marvellous Margazhi", ["margazhi", "eventart", "rithvik", "shweta"]),
                ]
                
                for p_name, keywords in known_partners:
                    if any(k in msg_l for k in keywords):
                        partner_matched.append(p_name)
                
                if sponsor_names:
                    partner_matched.extend(sponsor_names)
                if len(actor_names) > 1:
                    partner_matched.extend(actor_names[1:])
                
                is_paid_collab = bool(partner_matched) or any(t in msg_l for t in ["paid partnership", "collab", "collaboration", "ambassador", "#ad", "sponsored"])
                partner_label = ", ".join(list(dict.fromkeys(partner_matched))) if partner_matched else ("Organic Brand Post" if not is_paid_collab else "Influencer / Brand Campaign")
                
                all_extracted_posts.append({
                    "post_id": str(post_id),
                    "date": dt,
                    "url": url,
                    "partner": partner_label,
                    "is_paid_collab": is_paid_collab,
                    "reactions": reacts,
                    "comments": comments,
                    "shares": shares,
                    "video_views": views,
                    "caption": msg,
                    "source": source_tag
                })
        
        for k, v in obj.items():
            scan_json_tree(v, source_tag)
    elif isinstance(obj, list):
        for item in obj:
            scan_json_tree(item, source_tag)

for tag, page_url in target_urls:
    print(f"Fetching {tag} ({page_url}) ...", end=" ", flush=True)
    try:
        r = session.get(page_url, headers=headers, timeout=20)
        if r.status_code == 200:
            scripts = re.findall(r'<script type="application/json"[^>]*>(.*?)</script>', r.text, re.DOTALL)
            print(f"✓ HTTP 200 ({len(r.text):,} bytes, {len(scripts)} JSON nodes)")
            for s in scripts:
                try:
                    data = json.loads(s)
                    scan_json_tree(data, tag)
                except Exception:
                    pass
        else:
            print(f"⚠ HTTP {r.status_code}")
    except Exception as e:
        print(f"⚠ Error: {e}")

print(f"\n{'='*75}")
print(f"Total Unique Facebook Posts Extracted: {len(all_extracted_posts)}")
partner_posts = [p for p in all_extracted_posts if p["is_paid_collab"]]
print(f"Total Paid / Creator Partnership Posts Found on Facebook: {len(partner_posts)}")
print(f"{'='*75}\n")

# Save detailed JSON
with open("grt_facebook_all_posts.json", "w", encoding="utf-8") as f:
    json.dump(all_extracted_posts, f, ensure_ascii=False, indent=2)

# Build Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "GRT Facebook Partnerships"
ws.sheet_view.showGridLines = True

thin = Side(style="thin", color="CCCCCC")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="1877F2")
hdr_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
bf = Font(name="Calibri", bold=True, size=10, color="000000")
nf = Font(name="Calibri", bold=False, size=10, color="000000")
lnk = Font(name="Calibri", bold=False, size=10, color="0563C1", underline="single")

ws.merge_cells("A1:J1")
ws["A1"] = f"GRT Jewellers Facebook Page — Detected Partnership & Creator Campaign Posts ({len(all_extracted_posts)} Posts Total)"
ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="0E3E85")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 34

headers_list = [
    ("#", 5),
    ("Date", 13),
    ("Partner / Influencer", 30),
    ("Classification", 22),
    ("Post URL", 48),
    ("Reactions", 12),
    ("Comments", 12),
    ("Shares", 12),
    ("Video Views", 14),
    ("Caption / Copy", 65),
]

for col_idx, (h_text, width) in enumerate(headers_list, 1):
    c = ws.cell(row=2, column=col_idx, value=h_text)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bdr
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[2].height = 24
ws.freeze_panes = "A3"

# Sort posts: Paid Collabs first, then by date
all_extracted_posts.sort(key=lambda x: (not x["is_paid_collab"], x["date"]), reverse=False)

for idx, p in enumerate(all_extracted_posts, 1):
    row = idx + 2
    c_type = "Paid Creator / Campaign" if p["is_paid_collab"] else "Organic Product Post"
    
    vals = [
        idx,
        p["date"],
        p["partner"],
        c_type,
        p["url"],
        p["reactions"],
        p["comments"],
        p["shares"],
        p["video_views"],
        p["caption"],
    ]
    
    for col_idx, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.border = bdr
        if col_idx == 1:
            c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 3:
            c.font = bf; c.alignment = Alignment(horizontal="left", vertical="center")
            if p["is_paid_collab"]:
                c.fill = PatternFill("solid", fgColor="E7F3FF")
        elif col_idx == 4:
            c.font = bf; c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 5:
            c.font = lnk; c.alignment = Alignment(horizontal="left", vertical="center")
            if val: c.hyperlink = val
        elif col_idx in (6, 7, 8, 9):
            c.font = nf; c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "#,##0"
        elif col_idx == 10:
            c.font = nf; c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24

excel_filename = "grt_facebook_all_partnerships.xlsx"
wb.save(excel_filename)
print(f"✓ Saved Excel report to {excel_filename}")
