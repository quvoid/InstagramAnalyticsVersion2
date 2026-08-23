"""
Rebuild Master Excel and CSV files with the 100% verified API ground truth (70 Toggle ON posts)
"""

import sys, json, csv, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

SRC_EXPORT = "jewellery_brands_full_export.xlsx"
EXCEL_OUTPUT = "jewellery_brands_master_analysis.xlsx"
CSV_COLLABS = "All_Brands_Paid_Collabs.csv"
CSV_SUMMARY = "Brand_Creator_Summary.csv"

# Load API ground truth results
with open("api_toggle_ground_truth.json", encoding="utf-8") as f:
    api_results = json.load(f)

# Load raw source workbook
wb_src = openpyxl.load_workbook(SRC_EXPORT, data_only=True)
all_sheets = wb_src.sheetnames

overview_sheets = {"State-Brand Overview", "Creators", "All Brands - Paid Collabs", "Brand-Creator Summary"}
brand_sheets = [s for s in all_sheets if s not in overview_sheets]

NON_CREATOR_HANDLES = {
    "@grt.diamonds", "@grt.silverjewellery", "@grt.silverarticles", "@grtoriana",
    "@kalyan_jewellers_uk", "@kalyanjewellersusa", "@laksyah_", "@bluestone_india",
    "@tanishqjewellery", "@mia_by_tanishq", "@joyalukkasindia", "@joyalukkas_uk",
    "@sencogoldanddiamonds", "@malabargoldanddiamonds",
    "@elleindia", "@chennaitimestoi", "@eventart_india", "@coresocial05",
    "@platinumevara", "@platinumdaysoflove", "@menofplatinum"
}

all_paid_posts = []
brand_stats = {}

for s in brand_sheets:
    ws = wb_src[s]
    brand_posts = []
    brand_creators = set()
    
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 11)]
        h = str(vals[0] or "").strip()
        fol = vals[1]
        l = vals[2]
        c = vals[3]
        er = vals[4]
        date = vals[5]
        url = str(vals[6] or "").strip()
        via = str(vals[7] or "").strip()
        cap = str(vals[8] or "").strip()
        stat = vals[9]
        
        if not h or not url or url == "None":
            continue
            
        if "owned by partner" in via.lower():
            if h.lower() in {x.lower() for x in NON_CREATOR_HANDLES}:
                continue
                
            brand_creators.add(h.lower())
            
            # Check 100% verified API Ground Truth
            api_info = api_results.get(url, {})
            is_on = api_info.get("is_paid_partnership", False)
            toggle_str = "🟢 ON (Formal Label)" if is_on else "⚪ OFF (Collab Only)"
            
            # Clean followers
            fol_num = 0
            try:
                if isinstance(fol, (int, float)):
                    fol_num = int(fol)
                elif isinstance(fol, str):
                    fol_num = int(fol.replace(",", "").replace("%", "").strip())
            except Exception:
                fol_num = 0
                
            # Clean likes
            likes_num = 0
            try:
                if isinstance(l, (int, float)):
                    likes_num = int(l)
                elif isinstance(l, str):
                    likes_num = int(l.replace(",", "").strip())
            except Exception:
                likes_num = 0
                
            # Clean comments
            com_num = 0
            try:
                if isinstance(c, (int, float)):
                    com_num = int(c)
                elif isinstance(c, str):
                    com_num = int(c.replace(",", "").strip())
            except Exception:
                com_num = 0
                
            # Clean ER
            er_num = 0.0
            try:
                if isinstance(er, (int, float)):
                    er_num = float(er)
                    if er_num > 1: er_num = er_num / 100
                elif isinstance(er, str):
                    er_num = float(er.replace("%", "").strip()) / 100
            except Exception:
                er_num = 0.0
                
            date_str = str(date)[:10] if date else "N/A"
            if len(date_str) > 10:
                date_str = date_str[:10]
                
            post_obj = {
                "brand": s,
                "handle": h,
                "toggle": toggle_str,
                "is_on": is_on,
                "followers": fol_num,
                "avg_likes": likes_num,
                "avg_comments": com_num,
                "avg_er": er_num,
                "date": date_str,
                "url": url,
                "via": via,
                "caption": cap[:200].replace("\n", " ").replace("\r", " ")
            }
            brand_posts.append(post_obj)
            all_paid_posts.append(post_obj)
            
    # Brand stats
    fols = [p["followers"] for p in brand_posts if p["followers"] > 0]
    ers = [p["avg_er"] for p in brand_posts if p["avg_er"] > 0]
    avg_f = int(sum(fols)/len(fols)) if fols else 0
    avg_e = round(sum(ers)/len(ers), 4) if ers else 0.0
    latest_dt = max([p["date"] for p in brand_posts if p["date"] != "N/A"], default="N/A")
    top_c = list(dict.fromkeys([p["handle"] for p in brand_posts]))[:4]
    
    on_cnt = sum(1 for p in brand_posts if p["is_on"])
    off_cnt = len(brand_posts) - on_cnt
    
    brand_stats[s] = {
        "brand": s,
        "posts": len(brand_posts),
        "creators": len(brand_creators),
        "toggle_on": on_cnt,
        "toggle_off": off_cnt,
        "toggle_pct": on_cnt / len(brand_posts) if brand_posts else 0.0,
        "avg_followers": avg_f,
        "avg_er": avg_e,
        "latest_date": latest_dt,
        "top_creators": ", ".join(top_c) if top_c else "None"
    }

total_on = sum(1 for p in all_paid_posts if p["is_on"])
total_off = len(all_paid_posts) - total_on
unique_creators_total = len(set(p["handle"].lower() for p in all_paid_posts))

print(f"Total Posts: {len(all_paid_posts)}")
print(f"Total Toggle ON: {total_on} ({total_on/len(all_paid_posts)*100:.1f}%)")
print(f"Total Toggle OFF: {total_off} ({total_off/len(all_paid_posts)*100:.1f}%)")
print(f"Total Unique Creators: {unique_creators_total}")

# Sort posts by Date descending
all_paid_posts.sort(key=lambda x: x["date"], reverse=True)


# ── 1. Write CSV 1: All_Brands_Paid_Collabs.csv ───────────────
with open(CSV_COLLABS, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow([
        "#", "Brand Name", "Creator Handle", "Paid Partnership Toggle",
        "Followers", "Avg Likes/Post", "Avg Comments/Post", "Avg ER%",
        "Post Date", "Post URL", "Detection Method", "Caption Preview"
    ])
    for idx, p in enumerate(all_paid_posts, 1):
        writer.writerow([
            idx,
            p["brand"],
            p["handle"],
            p["toggle"],
            p["followers"],
            p["avg_likes"],
            p["avg_comments"],
            f"{p['avg_er']*100:.2f}%",
            p["date"],
            p["url"],
            p["via"],
            p["caption"]
        ])
print(f"✓ Created {CSV_COLLABS}")


# ── 2. Write CSV 2: Brand_Creator_Summary.csv ─────────────────
sorted_brands = sorted(brand_stats.values(), key=lambda x: x["posts"], reverse=True)
with open(CSV_SUMMARY, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow([
        "#", "Brand Name", "Paid Collab Posts", "Unique Creators",
        "Toggle ON", "Toggle OFF", "Toggle ON %",
        "Avg Creator Followers", "Avg Creator ER%", "Latest Collab Date", "Top Creator / Partner Samples"
    ])
    for idx, b in enumerate(sorted_brands, 1):
        writer.writerow([
            idx,
            b["brand"],
            b["posts"],
            b["creators"],
            b["toggle_on"],
            b["toggle_off"],
            f"{b['toggle_pct']*100:.1f}%",
            b["avg_followers"],
            f"{b['avg_er']*100:.2f}%",
            b["latest_date"],
            b["top_creators"]
        ])
print(f"✓ Created {CSV_SUMMARY}")


# ── 3. Build Formatted Excel Workbook ─────────────────────────
wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

thin = Side(style="thin", color="D0D3D4")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

f_title = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
f_hdr = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
f_bold = Font(name="Calibri", bold=True, size=10, color="000000")
f_norm = Font(name="Calibri", bold=False, size=10, color="000000")
f_link = Font(name="Calibri", bold=False, size=10, color="0563C1", underline="single")

# Soft Mint Green styling for Toggle ON
fill_green_row = PatternFill("solid", fgColor="D4EFDF") # Mint Green
f_green_bold = Font(name="Calibri", bold=True, size=10, color="145A32")
f_green_norm = Font(name="Calibri", bold=False, size=10, color="145A32")
f_green_link = Font(name="Calibri", bold=True, size=10, color="0B5345", underline="single")

fill_off_pill = PatternFill("solid", fgColor="F2F4F4")
f_off_pill = Font(name="Calibri", bold=False, size=10, color="5D6D7E")


# Sheet 1: Brand-Creator Summary
ws_s1 = wb_out.create_sheet("Brand-Creator Summary")
ws_s1.sheet_view.showGridLines = True

ws_s1.merge_cells("A1:K1")
ws_s1["A1"] = f"Executive Summary — Paid Creator Partnerships by Jewellery Brand ({len(all_paid_posts)} Total Posts · {total_on} Toggle ON · {total_off} Toggle OFF)"
ws_s1["A1"].font = f_title
ws_s1["A1"].fill = PatternFill("solid", fgColor="0B2240")
ws_s1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_s1.row_dimensions[1].height = 32

sum_cols = [
    ("#", 5), ("Brand Name", 28), ("Paid Collab Posts", 17), ("Unique Creators", 16),
    ("Toggle ON", 13), ("Toggle OFF", 13), ("Toggle ON %", 14),
    ("Avg Creator Followers", 21), ("Avg Creator ER%", 15), ("Latest Collab Date", 17),
    ("Top Creator / Partner Samples", 45)
]

for col_idx, (h_text, width) in enumerate(sum_cols, 1):
    c = ws_s1.cell(row=2, column=col_idx, value=h_text)
    c.font = f_hdr
    c.fill = PatternFill("solid", fgColor="1F2D3D")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bdr
    ws_s1.column_dimensions[get_column_letter(col_idx)].width = width
ws_s1.row_dimensions[2].height = 24
ws_s1.freeze_panes = "A3"

for idx, b in enumerate(sorted_brands, 1):
    r = idx + 2
    vals = [
        idx, b["brand"], b["posts"], b["creators"],
        b["toggle_on"], b["toggle_off"], b["toggle_pct"],
        b["avg_followers"], b["avg_er"], b["latest_date"], b["top_creators"]
    ]
    for c_idx, val in enumerate(vals, 1):
        cell = ws_s1.cell(row=r, column=c_idx, value=val)
        cell.border = bdr
        if c_idx == 1:
            cell.font = f_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2:
            cell.font = f_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (3, 4):
            cell.font = f_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"
            if val > 0: cell.fill = PatternFill("solid", fgColor="EBF5FB")
        elif c_idx == 5:
            cell.font = f_green_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"
            if val > 0: cell.fill = fill_green_row
        elif c_idx == 6:
            cell.font = f_off_pill; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"
            if val > 0: cell.fill = fill_off_pill
        elif c_idx == 7:
            cell.font = f_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.0%"
        elif c_idx == 8:
            cell.font = f_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx == 9:
            cell.font = f_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
        elif c_idx == 10:
            cell.font = f_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 11:
            cell.font = f_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_s1.row_dimensions[r].height = 22


# Sheet 2: All Brands - Paid Collabs
ws_s2 = wb_out.create_sheet("All Brands - Paid Collabs")
ws_s2.sheet_view.showGridLines = True

ws_s2.merge_cells("A1:L1")
ws_s2["A1"] = f"Consolidated Creator-Owned Paid Partnerships ('Post owned by partner') — {len(all_paid_posts)} Total Posts (Toggle ON: {total_on} · Toggle OFF: {total_off})"
ws_s2["A1"].font = f_title
ws_s2["A1"].fill = PatternFill("solid", fgColor="1E5631")
ws_s2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_s2.row_dimensions[1].height = 32

collab_cols = [
    ("#", 5), ("Brand Name", 28), ("Creator Handle", 25), ("Paid Partnership Toggle", 25),
    ("Followers", 14), ("Avg Likes/Post", 14), ("Avg Comments/Post", 14), ("Avg ER%", 12),
    ("Post Date", 13), ("Post URL", 48), ("Detection Method", 28), ("Caption Preview", 65)
]

for col_idx, (h_text, width) in enumerate(collab_cols, 1):
    c = ws_s2.cell(row=2, column=col_idx, value=h_text)
    c.font = f_hdr
    c.fill = PatternFill("solid", fgColor="1E5631")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bdr
    ws_s2.column_dimensions[get_column_letter(col_idx)].width = width
ws_s2.row_dimensions[2].height = 24
ws_s2.freeze_panes = "A3"

for idx, p in enumerate(all_paid_posts, 1):
    r = idx + 2
    is_on = p["is_on"]
    vals = [
        idx, p["brand"], p["handle"], p["toggle"],
        p["followers"], p["avg_likes"], p["avg_comments"], p["avg_er"],
        p["date"], p["url"], p["via"], p["caption"]
    ]
    for c_idx, val in enumerate(vals, 1):
        cell = ws_s2.cell(row=r, column=c_idx, value=val)
        cell.border = bdr
        
        # If Toggle ON: highlight entire row in mint green
        if is_on:
            cell.fill = fill_green_row
            if c_idx in (2, 3, 4):
                cell.font = f_green_bold
            elif c_idx == 10:
                cell.font = f_green_link
                if val: cell.hyperlink = val
            else:
                cell.font = f_green_norm
        else:
            if c_idx == 1:
                cell.font = f_norm
            elif c_idx == 2:
                cell.font = f_bold
            elif c_idx == 3:
                cell.font = f_bold
                cell.fill = PatternFill("solid", fgColor="FDF2E9")
            elif c_idx == 4:
                cell.font = f_off_pill
                cell.fill = fill_off_pill
            elif c_idx == 10:
                cell.font = f_link
                if val: cell.hyperlink = val
            else:
                cell.font = f_norm
                
        # Alignment & Number Formats
        if c_idx == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (2, 3):
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 4:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (5, 6, 7):
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "#,##0"
        elif c_idx == 8:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.00%"
        elif c_idx == 9:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 10:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 11:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 12:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
    ws_s2.row_dimensions[r].height = 22

# Also copy all 34 brand sheets intact
for s in brand_sheets:
    ws_src_brand = wb_src[s]
    ws_target = wb_out.create_sheet(s)
    ws_target.sheet_view.showGridLines = True
    for r in range(1, ws_src_brand.max_row + 1):
        for c in range(1, ws_src_brand.max_column + 1):
            val = ws_src_brand.cell(row=r, column=c).value
            ws_target.cell(row=r, column=c, value=val)
    for col_l in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws_target.column_dimensions[col_l].width = ws_src_brand.column_dimensions[col_l].width or 20

wb_out.save(EXCEL_OUTPUT)
print(f"\n✅ Completed Master Build!")
print(f"   • Excel: {EXCEL_OUTPUT}")
print(f"   • CSV:   {CSV_COLLABS}")
print(f"   • CSV:   {CSV_SUMMARY}")
