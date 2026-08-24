"""
Deep 1-Year Scraper & 4-Tier Analysis for Footwear / Sneaker Brands:
1. Skechers India (@skechersindia)
2. Gully Labs (@gullylabs)
3. Comet (@thecometuniverse)

Scrapes all collaboration posts from the last 1 year (Aug 2025 - Aug 2026),
extracts exact post dates, direct URLs, metrics, and classifies every post into Tiers 1-4.
"""

import sys, os, json, time, re, csv
from datetime import datetime, timezone, timedelta
from collections import defaultdict, Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scrape_bulk import make_session, extract_shortcode, shortcode_to_id, COOKIES

sys.stdout.reconfigure(encoding="utf-8")

session = make_session()

mob_hdrs = {
    "User-Agent": "Instagram 269.0.0.18.75 Android (26/8.0.0; 480dpi; 1080x1920; OnePlus; ONEPLUS A3003; OnePlus3; qcom; en_US; 314665256)",
    "x-ig-app-id": "936619743392459",
}

# 1-Year Cutoff: 365 days ago
NOW_DT = datetime.now(timezone.utc)
CUTOFF_DT = NOW_DT - timedelta(days=365)
CUTOFF_TIMESTAMP = int(CUTOFF_DT.timestamp())

print(f"[*] Scraping posts from {CUTOFF_DT.strftime('%Y-%m-%d')} to {NOW_DT.strftime('%Y-%m-%d')} (Past 1 Year)\n")

BRANDS = [
    {"name": "Skechers India", "username": "skechersindia", "state": "Pan-India / Maharashtra (HQ: Mumbai)"},
    {"name": "Gully Labs", "username": "gullylabs", "state": "Pan-India / Delhi NCR (HQ: New Delhi, D2C)"},
    {"name": "Comet", "username": "thecometuniverse", "state": "Pan-India / Karnataka (HQ: Bengaluru, D2C)"},
]

INTERNAL_ACCOUNTS = {
    "skechersindia", "skechers", "skechersperformance", "skechers_south_asia",
    "gullylabs", "gully_labs",
    "thecometuniverse", "comet_universe", "wearcomet",
}

def resolve_user_id(username):
    url = f"https://i.instagram.com/api/v1/users/search/?q={username}"
    try:
        r = session.get(url, headers=mob_hdrs, cookies=COOKIES, timeout=10)
        if r.status_code == 200:
            for u in r.json().get("users", []):
                if u.get("username", "").lower() == username.lower():
                    return {
                        "pk": u.get("pk"),
                        "username": u.get("username"),
                        "full_name": u.get("full_name"),
                        "followers": u.get("follower_count", 0),
                    }
    except Exception as e:
        print(f"Error resolving @{username}: {e}", file=sys.stderr)
    return {"pk": None, "username": username, "full_name": username, "followers": 0}

def fetch_feed_1year(user_id, max_pages=35):
    items = []
    max_id = ""
    for p in range(1, max_pages + 1):
        url = f"https://i.instagram.com/api/v1/feed/user/{user_id}/"
        if max_id:
            url += f"?max_id={max_id}"
        try:
            r = session.get(url, headers=mob_hdrs, cookies=COOKIES, timeout=12)
            if r.status_code == 200:
                data = r.json()
                f_items = data.get("items", [])
                items.extend(f_items)
                max_id = data.get("next_max_id")
                
                # Check oldest item timestamp
                oldest_ts = min([it.get("taken_at", 0) for it in f_items if it.get("taken_at")], default=0)
                oldest_date = datetime.fromtimestamp(oldest_ts, tz=timezone.utc).strftime("%Y-%m-%d") if oldest_ts else "N/A"
                print(f"    [Feed] Page {p:>2}: {len(f_items)} items (Oldest in page: {oldest_date})", flush=True)
                
                if oldest_ts and oldest_ts < CUTOFF_TIMESTAMP:
                    print(f"    -> Reached 1-year cutoff ({oldest_date}) in feed!", flush=True)
                    break
                if not max_id or len(f_items) == 0:
                    break
                time.sleep(0.35)
            else:
                break
        except Exception:
            break
    return items

def fetch_clips_1year(user_id, max_pages=35):
    items = []
    max_id = ""
    for p in range(1, max_pages + 1):
        url = "https://i.instagram.com/api/v1/clips/user/"
        payload = {"target_user_id": str(user_id), "page_size": 30}
        if max_id:
            payload["max_id"] = str(max_id)
        try:
            r = session.post(url, headers=mob_hdrs, data=payload, cookies=COOKIES, timeout=12)
            if r.status_code == 200:
                data = r.json()
                clips = [it.get("media") for it in data.get("items", []) if it.get("media")]
                items.extend(clips)
                paging = data.get("paging_info", {})
                max_id = paging.get("max_id")
                
                oldest_ts = min([it.get("taken_at", 0) for it in clips if it.get("taken_at")], default=0)
                oldest_date = datetime.fromtimestamp(oldest_ts, tz=timezone.utc).strftime("%Y-%m-%d") if oldest_ts else "N/A"
                print(f"    [Clips] Page {p:>2}: {len(clips)} reels (Oldest in page: {oldest_date})", flush=True)
                
                if oldest_ts and oldest_ts < CUTOFF_TIMESTAMP:
                    print(f"    -> Reached 1-year cutoff ({oldest_date}) in clips!", flush=True)
                    break
                if not paging.get("more_available") or not max_id:
                    break
                time.sleep(0.35)
            else:
                break
        except Exception:
            break
    return items

def get_creator_followers(handle):
    try:
        sess = make_session()
        r = sess.get(f"https://i.instagram.com/api/v1/users/search/?q={handle}", headers=mob_hdrs, cookies=COOKIES, timeout=8)
        if r.status_code == 200:
            for u in r.json().get("users", []):
                if u.get("username", "").lower() == handle.lower():
                    pk = u.get("pk")
                    r2 = sess.get(f"https://i.instagram.com/api/v1/users/{pk}/info/", headers=mob_hdrs, cookies=COOKIES, timeout=8)
                    if r2.status_code == 200:
                        ud = r2.json().get("user", {})
                        return ud.get("follower_count") or u.get("follower_count") or 0
                    return u.get("follower_count") or 0
    except Exception:
        pass
    return 0

print("="*75, flush=True)
print("SCRAPING 1-YEAR CREATOR COLLABORATIONS FOR SKECHERS, GULLY LABS & COMET", flush=True)
print("="*75, flush=True)

all_collab_posts = []

for b in BRANDS:
    b_name = b["name"]
    b_user = b["username"]
    b_state = b["state"]
    
    print(f"\n[+] Resolving user profile for {b_name} (@{b_user})...", flush=True)
    u_info = resolve_user_id(b_user)
    b_pk = u_info["pk"]
    
    if not b_pk:
        print(f"[-] Could not resolve user ID for {b_user}", flush=True)
        continue
        
    print(f"    User ID: {b_pk} | Followers: {u_info['followers']:,} | Full Name: {u_info['full_name']}", flush=True)
    print("    Fetching 1-year Feed & Clips...", flush=True)
    
    f_items = fetch_feed_1year(b_pk, max_pages=35)
    c_items = fetch_clips_1year(b_pk, max_pages=35)
    
    seen_ids = set()
    unique_items = []
    for it in f_items + c_items:
        pk = str(it.get("pk") or it.get("id"))
        if pk and pk not in seen_ids:
            seen_ids.add(pk)
            unique_items.append(it)
            
    print(f"    -> Total unique posts fetched: {len(unique_items)}", flush=True)
    
    brand_posts = []
    for it in unique_items:
        taken_at = it.get("taken_at")
        if not taken_at:
            continue
            
        # Enforce 1-year cutoff
        if taken_at < CUTOFF_TIMESTAMP:
            continue
            
        date_str = datetime.fromtimestamp(taken_at, tz=timezone.utc).strftime("%Y-%m-%d")
        owner = it.get("user", {}).get("username", "").lower()
        coauthors = [c.get("username", "").lower() for c in it.get("coauthor_producers", [])]
        is_paid = bool(it.get("is_paid_partnership", False))
        code = it.get("code") or ""
        post_url = f"https://www.instagram.com/p/{code}/" if code else ""
        
        cap_obj = it.get("caption") or {}
        cap_text = cap_obj.get("text", "") if isinstance(cap_obj, dict) else str(cap_obj)
        
        play_count = it.get("play_count") or it.get("view_count") or 0
        like_count = it.get("like_count") or 0
        comment_count = it.get("comment_count") or 0
        
        if not play_count and like_count:
            play_count = int(like_count * 18.5)
            
        creator_handle = ""
        # Check partner ownership or coauthor
        if owner != b_user.lower():
            if owner not in INTERNAL_ACCOUNTS:
                creator_handle = f"@{owner}"
        elif coauthors:
            ext = [c for c in coauthors if c not in INTERNAL_ACCOUNTS and c != b_user.lower()]
            if ext:
                creator_handle = f"@{ext[0]}"
                
        if not creator_handle:
            continue
            
        brand_posts.append({
            "brand": b_name,
            "state": b_state,
            "handle": creator_handle,
            "raw_handle": creator_handle.replace("@", ""),
            "url": post_url,
            "shortcode": code,
            "media_id": str(it.get("pk") or ""),
            "date": date_str,
            "taken_at": taken_at,
            "is_paid_partnership": is_paid,
            "views": play_count,
            "likes": like_count,
            "comments": comment_count,
            "via": "Post owned by partner (collab)",
            "caption": cap_text[:250].replace("\n", " ").replace("\r", " ")
        })
        
    print(f"    -> Extracted {len(brand_posts)} creator collaborations within past 1 year for {b_name}!", flush=True)
    all_collab_posts.extend(brand_posts)

# Resolve creator followers concurrently
unique_handles = list(set(p["raw_handle"] for p in all_collab_posts))
print(f"\n[+] Resolving follower counts for {len(unique_handles)} unique creators...", flush=True)

followers_cache = {}
with ThreadPoolExecutor(max_workers=15) as executor:
    fut_to_h = {executor.submit(get_creator_followers, h): h for h in unique_handles}
    for fut in as_completed(fut_to_h):
        h = fut_to_h[fut]
        followers_cache[h] = fut.result()

# Evaluate Boost and 4-Tier Assignment
for p in all_collab_posts:
    raw_h = p["raw_handle"]
    fols = followers_cache.get(raw_h, 0)
    p["followers"] = fols
    
    views = p["views"]
    likes = p["likes"]
    comments = p["comments"]
    cap_lower = p["caption"].lower()
    
    # Check disclosure tags in caption
    has_disclosure = any(t in cap_lower for t in ["#ad", "#paidpartnership", "#sponsored", "#collab", "#brandpartner", "paid partnership"])
    is_formal_paid = p["is_paid_partnership"] or has_disclosure
    
    like_rate = round((likes / views) * 100, 2) if views > 0 else 0.0
    view_mult = round(views / fols, 2) if fols > 0 else 0.0
    er = round(((likes + comments) / fols) * 100, 2) if fols > 0 else 0.0
    
    p["like_rate_pct"] = like_rate
    p["view_multiplier"] = view_mult
    p["er_pct"] = er
    
    # Boost Detection Logic
    is_boosted = False
    if views >= 500000 and like_rate < 0.35:
        is_boosted = True
        boost_status = "🚀 Heavily Boosted (Paid Ad Spend)"
        reason = f"High view count ({views:,}) with sub-0.35% like rate ({like_rate}%) indicates paid video ads campaign"
    elif view_mult >= 5.0 and like_rate < 0.70:
        is_boosted = True
        boost_status = "🚀 Boosted (Paid Ad Spend)"
        reason = f"High view multiplier ({view_mult}x followers) combined with low engagement rate ({like_rate}%)"
    elif view_mult >= 3.0 and like_rate < 1.00 and views >= 80000:
        is_boosted = True
        boost_status = "🔍 Likely Boosted (Targeted Ad)"
        reason = f"Disproportionate views ({views:,}) relative to likes ({likes:,})"
    elif er >= 4.0 and like_rate >= 2.00:
        is_boosted = False
        boost_status = "📈 Viral Organic Reach"
        reason = f"High organic views ({views:,}) with strong like rate ({like_rate}%)"
    else:
        is_boosted = False
        boost_status = "⚪ Standard Organic"
        reason = "Baseline organic collab reach"
        
    # Tier Assignment
    if is_formal_paid and is_boosted:
        tier = 1
        tier_name = "Tier 1: Toggle ON + Boosted"
    elif is_formal_paid and not is_boosted:
        tier = 2
        tier_name = "Tier 2: Toggle ON + Organic"
    elif not is_formal_paid and is_boosted:
        tier = 3
        tier_name = "Tier 3: Toggle OFF + Boosted"
    else:
        tier = 4
        tier_name = "Tier 4: Toggle OFF + Organic (Noise)"
        
    p["tier"] = tier
    p["tier_name"] = tier_name
    p["is_boosted"] = is_boosted
    p["boost_status"] = boost_status
    p["boost_reason"] = reason

# Sort strictly by Tier ascending, then Views descending
all_collab_posts.sort(key=lambda x: (x["tier"], -x["views"]))

with open("footwear_1year_4tier_dataset.json", "w", encoding="utf-8") as f:
    json.dump(all_collab_posts, f, indent=2)

print(f"\n{'='*75}", flush=True)
print(f"1-Year Scrape Complete! Total Posts: {len(all_collab_posts)}", flush=True)
print(f"{'='*75}\n", flush=True)

# ─────────────────────────────────────────────────────────────
# GENERATE MASTER EXCEL & CSVS
# ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Styles
font_title = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_norm = Font(name="Calibri", size=10, bold=False, color="000000")
font_mute = Font(name="Calibri", size=9, bold=False, color="5D6D7E")
font_link = Font(name="Calibri", size=10, bold=False, color="0563C1", underline="single")

fill_t1_banner = PatternFill("solid", fgColor="145A32") # Dark Emerald
fill_t1_row = PatternFill("solid", fgColor="D4EFDF")    # Mint Green
font_t1_bold = Font(name="Calibri", size=10, bold=True, color="0E6251")
font_t1_link = Font(name="Calibri", size=10, bold=True, color="0B5345", underline="single")

fill_t2_banner = PatternFill("solid", fgColor="1E8449") # Forest Green
fill_t2_row = PatternFill("solid", fgColor="EAFAF1")    # Sage Green
font_t2_bold = Font(name="Calibri", size=10, bold=True, color="196F3D")
font_t2_link = Font(name="Calibri", size=10, bold=False, color="145A32", underline="single")

fill_t3_banner = PatternFill("solid", fgColor="B7950B") # Dark Gold
fill_t3_row = PatternFill("solid", fgColor="FEF9E7")    # Warm Soft Gold
font_t3_bold = Font(name="Calibri", size=10, bold=True, color="7D6608")
font_t3_link = Font(name="Calibri", size=10, bold=False, color="9A7D0A", underline="single")

fill_t4_banner = PatternFill("solid", fgColor="566573") # Slate Gray
fill_t4_row = PatternFill("solid", fgColor="FFFFFF")    # Clean White
font_t4_norm = Font(name="Calibri", size=10, bold=False, color="2C3E50")
font_t4_link = Font(name="Calibri", size=10, bold=False, color="2980B9", underline="single")

thin_line = Side(style="thin", color="D5D8DC")
border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

# Group by brand
brand_records = defaultdict(list)
for r in all_collab_posts:
    brand_records[r["brand"]].append(r)

# Sheet 1: Executive Summary
ws_sum = wb.create_sheet("Executive Summary")
ws_sum.sheet_view.showGridLines = True

ws_sum.merge_cells("A1:O1")
ws_sum["A1"] = f"Executive Summary — 1-Year Paid Creator Collab Hierarchy (Skechers India, Gully Labs, Comet) [{CUTOFF_DT.strftime('%b %Y')} – {NOW_DT.strftime('%b %Y')}]"
ws_sum["A1"].font = font_title
ws_sum["A1"].fill = PatternFill("solid", fgColor="0B2240")
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 32

sum_headers = [
    ("#", 5),
    ("Brand Name", 24),
    ("State / Origin (HQ)", 32),
    ("Total Collab Posts (1-Yr)", 18),
    ("Total Unique Creators", 18),
    ("🟢 Tier 1: Toggle ON + Boosted\n(Posts / Creators)", 24),
    ("🟢 Tier 2: Toggle ON + Organic\n(Posts / Creators)", 24),
    ("🚀 Tier 3: Toggle OFF + Boosted\n(Posts / Creators)", 24),
    ("⚪ Tier 4: Toggle OFF + Organic (Noise)\n(Posts / Creators)", 28),
    ("💎 Total High-Intent Paid\n(Tiers 1+2+3 Posts)", 22),
    ("High-Intent Paid %\n(Tiers 1+2+3)", 18),
    ("Avg Views / Post", 16),
    ("Avg Creator Followers", 20),
    ("Avg Creator ER%", 15),
    ("Top Creator / Ambassador Samples", 48)
]

for col_idx, (h_text, w) in enumerate(sum_headers, 1):
    c = ws_sum.cell(row=2, column=col_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="1B2631")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_cell
    ws_sum.column_dimensions[get_column_letter(col_idx)].width = w
ws_sum.row_dimensions[2].height = 36
ws_sum.freeze_panes = "A3"

brand_summary_data = []
for b in BRANDS:
    b_name = b["name"]
    posts = brand_records.get(b_name, [])
    tot_p = len(posts)
    tot_c = len(set(p["handle"].lower() for p in posts))
    
    t1_posts = [p for p in posts if p["tier"] == 1]
    t1_creators = len(set(p["handle"].lower() for p in t1_posts))
    
    t2_posts = [p for p in posts if p["tier"] == 2]
    t2_creators = len(set(p["handle"].lower() for p in t2_posts))
    
    t3_posts = [p for p in posts if p["tier"] == 3]
    t3_creators = len(set(p["handle"].lower() for p in t3_posts))
    
    t4_posts = [p for p in posts if p["tier"] == 4]
    t4_creators = len(set(p["handle"].lower() for p in t4_posts))
    
    high_intent = len(t1_posts) + len(t2_posts) + len(t3_posts)
    high_intent_pct = high_intent / tot_p if tot_p > 0 else 0.0
    
    views_list = [p["views"] for p in posts if p["views"] > 0]
    avg_v = int(sum(views_list) / len(views_list)) if views_list else 0
    
    fols_list = [p["followers"] for p in posts if p["followers"] > 0]
    avg_f = int(sum(fols_list) / len(fols_list)) if fols_list else 0
    
    ers_list = [p["er_pct"] for p in posts if p["er_pct"] > 0]
    avg_e = round(sum(ers_list) / len(ers_list), 2) if ers_list else 0.0
    
    top_c = list(dict.fromkeys([p["handle"] for p in posts if p["tier"] in (1, 2, 3)]))[:4]
    if not top_c:
        top_c = list(dict.fromkeys([p["handle"] for p in posts]))[:4]
        
    brand_summary_data.append({
        "brand": b_name,
        "state": b["state"],
        "posts": tot_p,
        "creators": tot_c,
        "t1_p": len(t1_posts), "t1_c": t1_creators,
        "t2_p": len(t2_posts), "t2_c": t2_creators,
        "t3_p": len(t3_posts), "t3_c": t3_creators,
        "t4_p": len(t4_posts), "t4_c": t4_creators,
        "high_intent_p": high_intent,
        "high_intent_pct": high_intent_pct,
        "avg_views": avg_v,
        "avg_followers": avg_f,
        "avg_er": avg_e,
        "top_creators": ", ".join(top_c)
    })

brand_summary_data.sort(key=lambda x: (x["high_intent_p"], x["posts"]), reverse=True)

for idx, b in enumerate(brand_summary_data, 1):
    row_num = idx + 2
    row_vals = [
        idx,
        b["brand"],
        b["state"],
        b["posts"],
        b["creators"],
        f"{b['t1_p']} posts ({b['t1_c']} creators)" if b["t1_p"] > 0 else "—",
        f"{b['t2_p']} posts ({b['t2_c']} creators)" if b["t2_p"] > 0 else "—",
        f"{b['t3_p']} posts ({b['t3_c']} creators)" if b["t3_p"] > 0 else "—",
        f"{b['t4_p']} posts ({b['t4_c']} creators)" if b["t4_p"] > 0 else "—",
        b["high_intent_p"],
        b["high_intent_pct"],
        b["avg_views"],
        b["avg_followers"],
        b["avg_er"] / 100 if b["avg_er"] else 0.0,
        b["top_creators"]
    ]
    
    for col_idx, val in enumerate(row_vals, 1):
        cell = ws_sum.cell(row=row_num, column=col_idx, value=val)
        cell.border = border_cell
        
        if col_idx == 1:
            cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 2:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif col_idx == 3:
            cell.font = Font(name="Calibri", size=9, bold=True, color="2C3E50"); cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = PatternFill("solid", fgColor="F4F6F6")
        elif col_idx in (4, 5):
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"
            cell.fill = PatternFill("solid", fgColor="EBF5FB")
        elif col_idx == 6:
            cell.font = font_t1_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = fill_t1_row
        elif col_idx == 7:
            cell.font = font_t2_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = fill_t2_row
        elif col_idx == 8:
            cell.font = font_t3_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = fill_t3_row
        elif col_idx == 9:
            cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = PatternFill("solid", fgColor="F8F9F9")
        elif col_idx == 10:
            cell.font = Font(name="Calibri", size=10, bold=True, color="1B4F72"); cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"
            if val > 0: cell.fill = PatternFill("solid", fgColor="D6EAF8")
        elif col_idx == 11:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.0%"
        elif col_idx in (12, 13):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif col_idx == 14:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
        elif col_idx == 15:
            cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
            
    ws_sum.row_dimensions[row_num].height = 22

# Helper for master and brand tabs
table_cols = [
    ("#", 5),
    ("Hierarchy Tier", 30),
    ("Brand Name", 22),
    ("Creator Handle", 24),
    ("Followers", 14),
    ("Views / Plays", 16),
    ("Likes", 14),
    ("Comments", 12),
    ("Like-to-View %", 15),
    ("Creator ER%", 13),
    ("Post Date", 13),
    ("Direct Instagram URL", 48),
    ("Boost Classification & Reason", 38),
    ("Caption Preview", 65)
]

def render_footwear_sheet(ws, records, sheet_title, state_str):
    ws.sheet_view.showGridLines = True
    
    tot_p = len(records)
    tot_c = len(set(p["handle"].lower() for p in records))
    t1_p = sum(1 for p in records if p["tier"] == 1)
    t2_p = sum(1 for p in records if p["tier"] == 2)
    t3_p = sum(1 for p in records if p["tier"] == 3)
    t4_p = sum(1 for p in records if p["tier"] == 4)
    high_intent = t1_p + t2_p + t3_p
    
    # Header Banner
    ws.merge_cells("A1:N1")
    ws["A1"] = f"👟 {sheet_title.upper()}  |  📍 STATE / REGION: {state_str}"
    ws["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0B2240")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    
    # Metadata Overview
    ws.merge_cells("A2:N2")
    ws["A2"] = f"1-Year Collab Posts: {tot_p}  •  Unique Creators: {tot_c}  •  💎 High-Intent Paid: {high_intent} (T1: {t1_p} | T2: {t2_p} | T3: {t3_p})  •  ⚪ Noise/Unboosted (T4): {t4_p}"
    ws["A2"].font = Font(name="Calibri", size=10, bold=True, color="1B4F72")
    ws["A2"].fill = PatternFill("solid", fgColor="EBF5FB")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22
    
    # Table Header Row
    for col_idx, (h_text, w) in enumerate(table_cols, 1):
        c = ws.cell(row=3, column=col_idx, value=h_text)
        c.font = font_hdr
        c.fill = PatternFill("solid", fgColor="1F2D3D")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_cell
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[3].height = 25
    ws.freeze_panes = "A4"
    
    records_by_tier = {1: [], 2: [], 3: [], 4: []}
    for r in records:
        records_by_tier[r["tier"]].append(r)
        
    current_row = 4
    global_index = 1
    
    tier_meta = [
        (1, "🟢 TIER 1: TOGGLE ON + 🚀 BOOSTED (Formal Paid Partnership Label + Paid Ad Spend)", fill_t1_banner, fill_t1_row, font_t1_bold, font_t1_link),
        (2, "🟢 TIER 2: TOGGLE ON + ⚪ ORGANIC (Formal Paid Partnership Label + Organic Reach Only)", fill_t2_banner, fill_t2_row, font_t2_bold, font_t2_link),
        (3, "🚀 TIER 3: TOGGLE OFF + 🚀 BOOSTED (Co-Author Collab + Heavy Paid Ad Spend Detected)", fill_t3_banner, fill_t3_row, font_t3_bold, font_t3_link),
        (4, "⚪ TIER 4: TOGGLE OFF + ⚪ ORGANIC (Standard Collab / Low Organic Reach / Noise)", fill_t4_banner, fill_t4_row, font_t4_norm, font_t4_link)
    ]
    
    for t_id, banner_text, fill_banner, fill_row, font_b, font_l in tier_meta:
        t_records = records_by_tier[t_id]
        if not t_records:
            continue
            
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        banner_cell = ws.cell(row=current_row, column=1, value=f"{banner_text} — {len(t_records)} Posts ({len(set(p['handle'].lower() for p in t_records))} Unique Creators)")
        banner_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        banner_cell.fill = fill_banner
        banner_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[current_row].height = 23
        current_row += 1
        
        for p in t_records:
            vals = [
                global_index,
                p["tier_name"],
                p["brand"],
                p["handle"],
                p["followers"],
                p["views"],
                p["likes"],
                p["comments"],
                p["like_rate_pct"] / 100 if p["like_rate_pct"] else 0.0,
                p["er_pct"] / 100 if p["er_pct"] else 0.0,
                p["date"],
                p["url"],
                f"{p['boost_status']}: {p['boost_reason']}" if p.get('boost_reason') else p['boost_status'],
                p["caption"]
            ]
            
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=current_row, column=c_idx, value=val)
                cell.border = border_cell
                cell.fill = fill_row
                
                if c_idx == 1:
                    cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 2:
                    cell.font = font_b; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx in (3, 4):
                    cell.font = font_b; cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx in (5, 6, 7, 8):
                    cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
                elif c_idx in (9, 10):
                    cell.font = font_b; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
                elif c_idx == 11:
                    cell.font = font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 12:
                    cell.font = font_l; cell.alignment = Alignment(horizontal="left", vertical="center")
                    if val: cell.hyperlink = val
                elif c_idx == 13:
                    cell.font = font_b; cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx == 14:
                    cell.font = font_mute; cell.alignment = Alignment(horizontal="left", vertical="center")
                    
            ws.row_dimensions[current_row].height = 21
            current_row += 1
            global_index += 1

# Sheet 2: All Brands Master
ws_all = wb.create_sheet("All Brands - Master Hierarchy")
render_footwear_sheet(
    ws_all,
    all_collab_posts,
    "CONSOLIDATED 3 SNEAKER & FOOTWEAR BRANDS",
    "PAST 1 YEAR (AUG 2025 – AUG 2026)"
)

# Sheets 3 to 5: Brand sheets
for b in BRANDS:
    b_name = b["name"]
    posts = brand_records.get(b_name, [])
    ws_brand = wb.create_sheet(b_name[:31])
    render_footwear_sheet(
        ws_brand,
        posts,
        b_name,
        b["state"]
    )

wb.save("footwear_sneaker_brands_master_analysis.xlsx")
print("\n✅ Master Excel Generated: footwear_sneaker_brands_master_analysis.xlsx", flush=True)

# Write CSV Exports
with open("Footwear_All_Brands_4Tier_Master.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow([
        "#", "Hierarchy Tier", "Brand Name", "State / Origin (HQ)", "Creator Handle", "Followers",
        "Views / Plays", "Likes", "Comments", "Like-to-View %", "Creator ER%",
        "Post Date", "Direct Instagram URL", "Boost Status & Reason", "Caption Preview"
    ])
    for idx, p in enumerate(all_collab_posts, 1):
        w.writerow([
            idx,
            p["tier_name"],
            p["brand"],
            p["state"],
            p["handle"],
            p["followers"],
            p["views"],
            p["likes"],
            p["comments"],
            f"{p['like_rate_pct']:.2f}%",
            f"{p['er_pct']:.2f}%",
            p["date"],
            p["url"],
            f"{p['boost_status']}: {p['boost_reason']}" if p.get('boost_reason') else p['boost_status'],
            p["caption"]
        ])
print("✓ Saved Footwear_All_Brands_4Tier_Master.csv", flush=True)

with open("Footwear_Brand_Summary.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow([
        "#", "Brand Name", "State / Origin (HQ)", "Total Collab Posts (1-Yr)", "Total Unique Creators",
        "Tier 1 (Toggle ON + Boosted)", "Tier 2 (Toggle ON + Organic)",
        "Tier 3 (Toggle OFF + Boosted)", "Tier 4 (Noise / Unboosted)",
        "Total High-Intent Paid Posts", "High-Intent Paid %",
        "Avg Estimated Views", "Avg Creator Followers", "Avg Creator ER%", "Top Creator Samples"
    ])
    for idx, b in enumerate(brand_summary_data, 1):
        w.writerow([
            idx,
            b["brand"],
            b["state"],
            b["posts"],
            b["creators"],
            f"{b['t1_p']} posts ({b['t1_c']} creators)" if b["t1_p"] > 0 else "—",
            f"{b['t2_p']} posts ({b['t2_c']} creators)" if b["t2_p"] > 0 else "—",
            f"{b['t3_p']} posts ({b['t3_c']} creators)" if b["t3_p"] > 0 else "—",
            f"{b['t4_p']} posts ({b['t4_c']} creators)" if b["t4_p"] > 0 else "—",
            b["high_intent_p"],
            f"{b['high_intent_pct']*100:.1f}%",
            b["avg_views"],
            b["avg_followers"],
            f"{b['avg_er']:.2f}%",
            b["top_creators"]
        ])
print("✓ Saved Footwear_Brand_Summary.csv\n", flush=True)
