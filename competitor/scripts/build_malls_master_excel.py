"""
Build Master Excel for 12 Malls (Pune & Hyderabad) including Lake Shore properties:
pune_hyderabad_malls_master_analysis.xlsx
"""

import sys, json, os, openpyxl
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

with open("pune_hyderabad_malls_1year_dataset.json", encoding="utf-8") as f:
    raw_data = json.load(f)

malls_results = raw_data["malls_results"]
creators_roster = raw_data["creators_roster"]

# Flatten all collabs
all_collabs = []
for mr in malls_results:
    for c in mr["collabs"]:
        all_collabs.append(c)

# Map creator profile info for fast lookup
creator_map = {c["raw_handle"]: c for c in creators_roster}

for c in all_collabs:
    prof = creator_map.get(c["raw_handle"], {})
    c["full_name"] = prof.get("full_name", c["raw_handle"])
    c["followers"] = prof.get("followers", 0)
    c["creator_tier"] = prof.get("tier", "🌱 Nano (<10K)")

# Sort all collabs by Tier ascending (Tier 1 -> Tier 2 -> Tier 3 -> Tier 4), then Views descending
tier_rank = {
    "Tier 1: Paid Partnership Toggle Active (Dark Ads)": 1,
    "Tier 2: Direct Creator Collab (Organic Grid + Potential Dark Boosting)": 2,
    "Tier 3: Boosted Organic Grid Post (High View Spike)": 3,
    "Tier 4: Pure Organic Barter Collab": 4
}
all_collabs.sort(key=lambda x: (tier_rank.get(x.get("paid_classification"), 4), -x.get("views", 0)))

# Workbook Setup
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Styles
font_title = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
font_sub = Font(name="Calibri", size=9, italic=True, color="E0E0E0")
font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_norm = Font(name="Calibri", size=10, bold=False, color="000000")
font_mute = Font(name="Calibri", size=9, bold=False, color="5D6D7E")
font_link = Font(name="Calibri", size=10, bold=False, color="0563C1", underline="single")

thin_line = Side(style="thin", color="D5D8DC")
border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

fill_client = PatternFill("solid", fgColor="D4EFDF") # Mint Green highlight for Lake Shore
fill_client_hdr = PatternFill("solid", fgColor="145A32") # Emerald for Lake Shore
fill_navy = PatternFill("solid", fgColor="0B2240")
fill_dark = PatternFill("solid", fgColor="1B2631")


# ==============================================================================
# TAB 1: EXECUTIVE SUMMARY
# ==============================================================================
ws_sum = wb.create_sheet("Executive Summary")
ws_sum.sheet_view.showGridLines = True
ws_sum.merge_cells("A1:N1")
ws_sum["A1"] = "Executive Summary — 1-Year Creator Collaborations & 4-Tier Hierarchy (Pune & Hyderabad Malls + Lake Shore)"
ws_sum["A1"].font = font_title
ws_sum["A1"].fill = fill_navy
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 30

sum_headers = [
    ("#", 5), ("Mall Name", 36), ("City", 14), ("Total Collab Posts", 16),
    ("Unique Creators", 16), ("🟢 Tier 1: Toggle ON + Boosted", 22),
    ("🟢 Tier 2: Toggle ON + Organic", 22), ("🚀 Tier 3: Toggle OFF + Boosted", 22),
    ("⚪ Tier 4: Toggle OFF + Organic", 22), ("💎 High-Intent Paid (T1+T2+T3)", 22),
    ("Paid Collabs %", 15), ("Total Video Views", 18), ("Avg Views / Post", 16),
    ("Top Creator Partnerships (Sample)", 40)
]

for c_idx, (h_text, w) in enumerate(sum_headers, 1):
    c = ws_sum.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_cell
    ws_sum.column_dimensions[get_column_letter(c_idx)].width = w
ws_sum.row_dimensions[2].height = 28
ws_sum.freeze_panes = "A3"

# Mall stats
row_idx = 3
pune_totals = defaultdict(int)
hyd_totals = defaultdict(int)

for idx, mr in enumerate(malls_results, 1):
    m_name = mr["mall_name"]
    city = mr["city"]
    collabs = mr["collabs"]
    is_client = mr.get("is_client", False)
    
    c_set = set(c["raw_handle"] for c in collabs)
    tot_c = len(collabs)
    uniq_creators = len(c_set)
    
    t1 = sum(1 for c in collabs if "Tier 1" in c.get("paid_classification", ""))
    t2 = sum(1 for c in collabs if "Tier 2" in c.get("paid_classification", ""))
    t3 = sum(1 for c in collabs if "Tier 3" in c.get("paid_classification", ""))
    t4 = sum(1 for c in collabs if "Tier 4" in c.get("paid_classification", ""))
    
    high_paid = t1 + t2 + t3
    paid_pct = (high_paid / tot_c * 100.0) if tot_c else 0.0
    tot_views = sum(c.get("views", 0) for c in collabs)
    avg_views = int(tot_views / tot_c) if tot_c else 0
    
    top_creators = [c["creator_handle"] for c in sorted(collabs, key=lambda x: x.get("views", 0), reverse=True)[:3]]
    top_str = ", ".join(dict.fromkeys(top_creators)) if top_creators else "—"

    if city == "Pune":
        pune_totals["collabs"] += tot_c; pune_totals["creators"] += uniq_creators
        pune_totals["t1"] += t1; pune_totals["t2"] += t2; pune_totals["t3"] += t3; pune_totals["t4"] += t4
        pune_totals["views"] += tot_views
    else:
        hyd_totals["collabs"] += tot_c; hyd_totals["creators"] += uniq_creators
        hyd_totals["t1"] += t1; hyd_totals["t2"] += t2; hyd_totals["t3"] += t3; hyd_totals["t4"] += t4
        hyd_totals["views"] += tot_views

    vals = [
        idx, f"★ {m_name} (CLIENT)" if is_client else m_name, city, tot_c, uniq_creators,
        t1, t2, t3, t4, high_paid, f"{paid_pct:.1f}%", tot_views, avg_views, top_str
    ]
    
    for c_idx, v in enumerate(vals, 1):
        cell = ws_sum.cell(row=row_idx, column=c_idx, value=v)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2: cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (3, 11): cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in range(4, 11): cell.font = font_bold if c_idx in (4, 5, 10) else font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx in (12, 13): cell.font = font_bold; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx == 14: cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        
        if is_client:
            cell.fill = fill_client

    ws_sum.row_dimensions[row_idx].height = 20
    row_idx += 1

# Grand Total Row
tot_collabs = len(all_collabs)
tot_creators = len(creators_roster)
g_t1 = pune_totals["t1"] + hyd_totals["t1"]
g_t2 = pune_totals["t2"] + hyd_totals["t2"]
g_t3 = pune_totals["t3"] + hyd_totals["t3"]
g_t4 = pune_totals["t4"] + hyd_totals["t4"]
g_paid = g_t1 + g_t2 + g_t3
g_paid_pct = (g_paid / tot_collabs * 100.0) if tot_collabs else 0.0
g_views = pune_totals["views"] + hyd_totals["views"]
g_avg = int(g_views / tot_collabs) if tot_collabs else 0

ws_sum.merge_cells(f"A{row_idx}:C{row_idx}")
ws_sum.cell(row=row_idx, column=1, value="GRAND TOTAL: ALL 12 MALLS").font = font_title
ws_sum.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")

g_vals = [tot_collabs, tot_creators, g_t1, g_t2, g_t3, g_t4, g_paid, f"{g_paid_pct:.1f}%", g_views, g_avg, "—"]
for c_idx, v in enumerate(g_vals, 4):
    cell = ws_sum.cell(row=row_idx, column=c_idx, value=v)
    cell.font = font_title; cell.alignment = Alignment(horizontal="right", vertical="center")
    if isinstance(v, int): cell.number_format = "#,##0"

for col in range(1, 15):
    ws_sum.cell(row=row_idx, column=col).border = border_cell
    ws_sum.cell(row=row_idx, column=col).fill = fill_dark
ws_sum.row_dimensions[row_idx].height = 24


# ==============================================================================
# TAB 2: MASTER HIERARCHY (ALL COLLABS)
# ==============================================================================
ws_mast = wb.create_sheet("Master Hierarchy (All Collabs)")
ws_mast.sheet_view.showGridLines = True
ws_mast.merge_cells("A1:O1")
ws_mast["A1"] = f"Master Collaborations Roster — All 12 Pune & Hyderabad Malls ({len(all_collabs)} Collab Posts)"
ws_mast["A1"].font = font_title
ws_mast["A1"].fill = fill_navy
ws_mast["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_mast.row_dimensions[1].height = 30

mast_headers = [
    ("#", 5), ("City", 12), ("Mall Name", 32), ("Creator Handle", 24),
    ("Creator Full Name", 26), ("Audience Followers", 18), ("Creator Scale Tier", 24),
    ("Date", 12), ("Video Views", 15), ("Likes", 14), ("Comments", 12),
    ("Like-to-View %", 14), ("Paid Partnership Toggle", 22), ("4-Tier Classification", 30), ("Instagram Post URL", 42)
]

for c_idx, (h_text, w) in enumerate(mast_headers, 1):
    c = ws_mast.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_mast.column_dimensions[get_column_letter(c_idx)].width = w
ws_mast.row_dimensions[2].height = 26
ws_mast.freeze_panes = "A3"

for idx, c in enumerate(all_collabs, 1):
    r_num = idx + 2
    is_client = c.get("is_client", False)
    r_vals = [
        idx, c.get("city"), f"★ {c.get('mall_name')}" if is_client else c.get("mall_name"),
        c.get("creator_handle"), c.get("full_name"), c.get("followers", 0),
        c.get("creator_tier"), c.get("date"), c.get("views", 0), c.get("likes", 0),
        c.get("comments", 0), c.get("like_to_view_ratio", "N/A"),
        "ON (Paid Toggle)" if "Tier 1" in c.get("paid_classification", "") else "OFF",
        c.get("paid_classification"), c.get("post_url")
    ]
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws_mast.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (2, 8, 13): cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (3, 4, 5): cell.font = font_bold if c_idx == 4 else font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (6, 9, 10, 11): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx == 12: cell.alignment = Alignment(horizontal="right", vertical="center")
        elif c_idx in (7, 14): cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 15: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None
        
        if is_client:
            cell.fill = fill_client
    ws_mast.row_dimensions[r_num].height = 20


# ==============================================================================
# TAB 3: UNIQUE CREATOR ROSTER (586 CREATORS)
# ==============================================================================
ws_roster = wb.create_sheet("Unique Creator Roster")
ws_roster.sheet_view.showGridLines = True
ws_roster.merge_cells("A1:J1")
ws_roster["A1"] = f"Master Creator Directory — All {len(creators_roster)} Unique Creators across Pune & Hyderabad"
ws_roster["A1"].font = font_title
ws_roster["A1"].fill = fill_navy
ws_roster["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_roster.row_dimensions[1].height = 30

roster_headers = [
    ("#", 5), ("Creator Handle", 24), ("Full Name", 26), ("Follower Count", 16),
    ("Audience Scale Tier", 24), ("Primary Mall Collaborated", 32),
    ("All Malls Collaborated", 42), ("Total Collab Posts Done", 20),
    ("Total Video Views Generated", 22), ("Instagram Profile URL", 40)
]
for c_idx, (h_text, w) in enumerate(roster_headers, 1):
    c = ws_roster.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_roster.column_dimensions[get_column_letter(c_idx)].width = w
ws_roster.row_dimensions[2].height = 26
ws_roster.freeze_panes = "A3"

for idx, cr in enumerate(creators_roster, 1):
    r_num = idx + 2
    r_vals = [
        idx, cr["handle"], cr["full_name"], cr["followers"], cr["tier"],
        cr["primary_mall"], cr["all_malls_collaborated"], cr["total_collabs_done"],
        cr["total_views_generated"], cr["profile_url"]
    ]
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws_roster.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (2, 3, 6, 7): cell.font = font_bold if c_idx == 2 else font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (4, 8, 9): cell.font = font_bold if c_idx == 4 else font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx == 5: cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 10: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None
    ws_roster.row_dimensions[r_num].height = 20


# ==============================================================================
# TABS 4 TO 15: 12 INDIVIDUAL MALL TABS
# ==============================================================================
for mr in malls_results:
    m_name = mr["mall_name"]
    clean_tab_name = m_name.replace(" (Lake Shore)", "").replace(" (Hyderabad)", "").replace("Mall", "").strip()[:28]
    ws_mall = wb.create_sheet(clean_tab_name)
    ws_mall.sheet_view.showGridLines = True
    
    ws_mall.merge_cells("A1:K1")
    tot_p_cnt = mr.get("total_posts_1year") or len(mr.get("all_posts", []))
    ws_mall["A1"] = f"{m_name} — 1-Year Collaborations Breakdown ({len(mr['collabs'])} Collabs | {tot_p_cnt} Total Posts)"
    ws_mall["A1"].font = font_title
    ws_mall["A1"].fill = fill_client_hdr if mr.get("is_client") else fill_navy
    ws_mall["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_mall.row_dimensions[1].height = 30
    
    m_headers = [
        ("#", 5), ("Creator Handle", 24), ("Full Name", 26), ("Followers", 16),
        ("Date", 12), ("Video Views", 15), ("Likes", 14), ("Comments", 12),
        ("4-Tier Classification", 30), ("Audio Track / Music", 35), ("Instagram URL", 40)
    ]
    for c_idx, (h_text, w) in enumerate(m_headers, 1):
        c = ws_mall.cell(row=2, column=c_idx, value=h_text)
        c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_mall.column_dimensions[get_column_letter(c_idx)].width = w
    ws_mall.row_dimensions[2].height = 26
    ws_mall.freeze_panes = "A3"
    
    for idx, c in enumerate(mr["collabs"], 1):
        r_num = idx + 2
        r_vals = [
            idx, c["creator_handle"], c.get("full_name", c["raw_handle"]), c.get("followers", 0),
            c["date"], c["views"], c["likes"], c["comments"],
            c.get("paid_classification", "Tier 4: Pure Organic Barter Collab"),
            c.get("audio_track", "Original Audio"), c["post_url"]
        ]
        for c_idx, val in enumerate(r_vals, 1):
            cell = ws_mall.cell(row=r_num, column=c_idx, value=val)
            cell.border = border_cell
            if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in (2, 3, 10): cell.font = font_bold if c_idx == 2 else font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx in (4, 6, 7, 8): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
            elif c_idx in (5, 9): cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx == 11: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None
        ws_mall.row_dimensions[r_num].height = 20


# ==============================================================================
# TAB 16: ALL POSTS DEEP-DIVE (CONTENT & CAPTIONS FOR 4,000+ POSTS)
# ==============================================================================
ws_deep = wb.create_sheet("All Posts Deep-Dive (Content)")
ws_deep.sheet_view.showGridLines = True
ws_deep.merge_cells("A1:K1")

all_scanned_posts = []
for mr in malls_results:
    for p in mr["all_posts"]:
        all_scanned_posts.append(p)

ws_deep["A1"] = f"Complete Content Deep-Dive — All 12 Malls ({len(all_scanned_posts)} Posts with Captions & Audio)"
ws_deep["A1"].font = font_title
ws_deep["A1"].fill = fill_navy
ws_deep["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_deep.row_dimensions[1].height = 30

deep_headers = [
    ("#", 5), ("City", 12), ("Mall Name", 32), ("Date", 12), ("Views", 14),
    ("Likes", 12), ("Comments", 12), ("Audio Track", 32), ("Collaborator(s)", 28),
    ("Full Caption / Description", 65), ("Instagram Link", 40)
]
for c_idx, (h_text, w) in enumerate(deep_headers, 1):
    c = ws_deep.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_deep.column_dimensions[get_column_letter(c_idx)].width = w
ws_deep.row_dimensions[2].height = 26
ws_deep.freeze_panes = "A3"

for idx, p in enumerate(all_scanned_posts, 1):
    r_num = idx + 2
    is_client = p.get("is_client", False)
    r_vals = [
        idx, p.get("city", "—"), f"★ {p.get('mall_name')}" if is_client else p.get("mall_name", "—"), p.get("date", "—"),
        p.get("views", 0), p.get("likes", 0), p.get("comments", 0), p.get("audio_track", "Original Audio"),
        p.get("all_creators_tagged") or p.get("creator_handle", "—"), p.get("caption", ""), p.get("post_url", "—")
    ]
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws_deep.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (2, 4): cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (3, 8, 9, 10): cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (5, 6, 7): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx == 11: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None
        
        if is_client:
            cell.fill = fill_client
    ws_deep.row_dimensions[r_num].height = 20

output_file = "pune_hyderabad_malls_master_analysis.xlsx"
wb.save(output_file)
print(f"\n✓ Master Workbook saved successfully with {len(wb.sheetnames)} tabs: {output_file}")
