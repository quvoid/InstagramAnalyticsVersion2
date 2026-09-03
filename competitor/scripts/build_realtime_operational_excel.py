"""
Build lakeshore_realtime_operational_intelligence.xlsx
Tabs:
1. 🕐 Google Popular Times & Hourly Ad Flighting Heatmap (24-hour curves & trigger rules)
2. 🏢 Tenant Brand Directory & Exclusivity Matrix (Rosters, anchor tenants, KOPA luxury USPs)
3. ❓ Google Community Q&A & Search Ad Engine (Pre-visit customer questions & Google Ads)
"""

import sys, json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

OUTPUT_PATH = "lakeshore_realtime_operational_intelligence.xlsx"

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Styles
font_title = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_norm = Font(name="Calibri", size=10, bold=False, color="111111")
font_code = Font(name="Consolas", size=9, bold=False, color="1B4F72")

thin_line = Side(style="thin", color="D5D8DC")
border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

fill_navy = PatternFill("solid", fgColor="0B2240")
fill_dark = PatternFill("solid", fgColor="212F3D")
fill_client = PatternFill("solid", fgColor="D4EFDF") # Soft green
fill_highlight = PatternFill("solid", fgColor="FCF3CF") # Soft yellow
fill_neg = PatternFill("solid", fgColor="FADBD8") # Soft red
fill_pos = PatternFill("solid", fgColor="EAFAF1")
fill_accent = PatternFill("solid", fgColor="EBF5FB")

def set_row(ws, r_num, vals, font=font_norm, fill=None, align_center_cols=None, align_right_cols=None, code_cols=None, height=22, wrap=True):
    align_center_cols = align_center_cols or []
    align_right_cols = align_right_cols or []
    code_cols = code_cols or []
    for c_idx, val in enumerate(vals, 1):
        c = ws.cell(row=r_num, column=c_idx, value=val)
        c.font = font_code if c_idx in code_cols else font
        c.border = border_cell
        if fill: c.fill = fill
        if c_idx in align_center_cols:
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        elif c_idx in align_right_cols:
            c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=wrap)
        else:
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    ws.row_dimensions[r_num].height = height


# ──────────────────────────────────────────────────────────────────────────────
# TAB 1: GOOGLE POPULAR TIMES & HOURLY AD FLIGHTING HEATMAP
# ──────────────────────────────────────────────────────────────────────────────
print("Building Tab 1: Google Popular Times & Hourly Busyness...")
ws1 = wb.create_sheet("1. Popular Times & Ad Triggers")
ws1.sheet_view.showGridLines = True
ws1.merge_cells("A1:H1")
ws1["A1"] = "GOOGLE MAPS POPULAR TIMES: HOURLY BUSYNESS PROFILES & REAL-TIME AD CONQUESTING TRIGGERS"
ws1["A1"].font = font_title; ws1["A1"].fill = fill_navy; ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 32

headers1 = [
    ("#", 5), ("Mall Name", 30), ("City", 12), ("Typical Dwell Time", 22),
    ("Peak Parking / Wait Time", 28), ("Sat Peak Capacity % (8 PM)", 24),
    ("Sun Peak Capacity % (7 PM)", 24), ("Real-Time Media Conquesting Rule", 55)
]
for c_idx, (h_text, w) in enumerate(headers1, 1):
    c = ws1.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.column_dimensions[get_column_letter(c_idx)].width = w
ws1.row_dimensions[2].height = 28
ws1.freeze_panes = "A3"

with open("google_popular_times_busyness_dataset.json", encoding="utf-8") as f:
    pop_data = json.load(f)["malls"]

for idx, m in enumerate(pop_data, 1):
    r = idx + 2
    is_client = m.get("is_client", False)
    sat_peak = m["weekly_busyness"]["Saturday"][19] # 7-8 PM
    sun_peak = m["weekly_busyness"]["Sunday"][18]  # 6-7 PM
    
    set_row(ws1, r, [
        idx, m.get("mall_name", ""), m.get("city", ""), m.get("typical_time_spent", ""),
        m.get("peak_wait_time", ""), f"{sat_peak}% Capacity", f"{sun_peak}% Capacity",
        m.get("conquest_trigger_rule", "")
    ], font=font_norm, fill=fill_client if is_client else (fill_neg if sat_peak >= 95 else (fill_highlight if sat_peak >= 85 else None)),
    align_center_cols=[1, 3, 4, 6, 7], height=34)


# ──────────────────────────────────────────────────────────────────────────────
# TAB 2: TENANT BRAND DIRECTORIES & EXCLUSIVITY MATRIX
# ──────────────────────────────────────────────────────────────────────────────
print("Building Tab 2: Tenant Brand Directories & Exclusivity...")
ws2 = wb.create_sheet("2. Tenant Brands & Exclusivity")
ws2.sheet_view.showGridLines = True
ws2.merge_cells("A1:H1")
ws2["A1"] = "MALL TENANT BRAND DIRECTORIES & EXCLUSIVITY AUDIT: 800+ STORES ACROSS 12 MALLS"
ws2["A1"].font = font_title; ws2["A1"].fill = fill_navy; ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

headers2 = [
    ("#", 5), ("Mall Name", 28), ("City", 12), ("Store Count", 14),
    ("Positioning Tier", 28), ("Anchor Tenant Brands", 42),
    ("Exclusive KOPA Luxury USPs", 45), ("Winning Ad Hook / Positioning Angle", 48)
]
for c_idx, (h_text, w) in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.column_dimensions[get_column_letter(c_idx)].width = w
ws2.row_dimensions[2].height = 28
ws2.freeze_panes = "A3"

with open("mall_tenant_brand_directories_dataset.json", encoding="utf-8") as f:
    tenant_data = json.load(f)["malls"]

for idx, m in enumerate(tenant_data, 1):
    r = idx + 2
    is_client = m.get("is_client", False)
    anchors = ", ".join(m.get("anchor_tenants", [])[:4])
    excl = "; ".join(m.get("exclusive_kopa_tenants", [])) if m.get("exclusive_kopa_tenants") else "Standard regional brand mix."
    
    set_row(ws2, r, [
        idx, m.get("mall_name", ""), m.get("city", ""), m.get("total_store_count", 0),
        m.get("positioning_tier", ""), anchors, excl, m.get("ad_hook_usp", "")
    ], font=font_norm, fill=fill_client if is_client else None,
    align_center_cols=[1, 3, 4], height=38)


# ──────────────────────────────────────────────────────────────────────────────
# TAB 3: GOOGLE COMMUNITY Q&A & SEARCH AD ENGINE
# ──────────────────────────────────────────────────────────────────────────────
print("Building Tab 3: Google Community Q&A & Search Ads...")
ws3 = wb.create_sheet("3. Community Q&A & Search Ads")
ws3.sheet_view.showGridLines = True
ws3.merge_cells("A1:G1")
ws3["A1"] = "GOOGLE MAPS COMMUNITY Q&A AUDIT & READY-TO-RUN GOOGLE SEARCH ADS"
ws3["A1"].font = font_title; ws3["A1"].fill = fill_navy; ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 32

headers3 = [
    ("#", 5), ("Mall Name", 26), ("Topic / Category", 24),
    ("Real Community Question Asked", 45), ("Community Verified Answer", 48),
    ("KOPA Competitive Advantage", 40), ("Ready-to-Run Google Search Ad Copy", 55)
]
for c_idx, (h_text, w) in enumerate(headers3, 1):
    c = ws3.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws3.column_dimensions[get_column_letter(c_idx)].width = w
ws3.row_dimensions[2].height = 28
ws3.freeze_panes = "A3"

with open("google_maps_community_qna_dataset.json", encoding="utf-8") as f:
    qna_data = json.load(f)["qna_data"]

row_idx = 3
counter = 1
for mall_obj in qna_data:
    m_name = mall_obj.get("mall_name", "")
    is_client = mall_obj.get("is_client", False)
    for qa in mall_obj.get("qa_pairs", []):
        set_row(ws3, row_idx, [
            counter, m_name, qa.get("topic", ""), qa.get("question", ""),
            qa.get("community_answer", ""), qa.get("kopa_competitive_advantage", ""),
            qa.get("google_search_ad_copy", "")
        ], font=font_norm, fill=fill_client if is_client else None,
        align_center_cols=[1], code_cols=[7], height=44)
        row_idx += 1
        counter += 1

wb.save(OUTPUT_PATH)
print(f"✓ Saved Real-Time Operational Intelligence Suite: {OUTPUT_PATH}")
