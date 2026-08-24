"""
Apply exact follower counts for the 20 handles and enforce pure Creator Tiers (Mega, Macro, Mid, Micro, Nano)
without any 'Brand / Media Partner' labels.
"""

import sys, json, openpyxl, csv
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

with open("og_20_results.json", encoding="utf-8") as f:
    og_20 = json.load(f)

with open("footwear_creators_profile_metrics.json", encoding="utf-8") as f:
    profiles = json.load(f)

with open("footwear_1year_4tier_dataset.json", encoding="utf-8") as f:
    posts = json.load(f)

def get_pure_tier(followers):
    if followers >= 1000000:
        return "🌟 Mega Creator / Celebrity (1M+)"
    elif followers >= 100000:
        return "🚀 Macro Creator (100K - 1M)"
    elif followers >= 50000:
        return "✨ Mid-Tier Creator (50K - 100K)"
    elif followers >= 10000:
        return "🎯 Micro Creator (10K - 50K)"
    else:
        return "🌱 Nano Creator (<10K)"

# Update 20 handles in profiles
for p in profiles:
    rh = p["raw_handle"].lower()
    if rh in og_20 and og_20[rh]["followers"] > 0:
        p["followers"] = og_20[rh]["followers"]
        if og_20[rh].get("full_name"):
            p["full_name"] = og_20[rh]["full_name"]
            
    # Apply pure tier
    p["creator_tier"] = get_pure_tier(p["followers"])
    
    # Recalculate ER%
    if p["followers"] > 0 and p.get("avg_likes", 0) > 0:
        p["avg_er"] = round(((p["avg_likes"] + p.get("avg_comments", 0)) / p["followers"]) * 100, 2)
    else:
        p["avg_er"] = 0.0

# Sort profiles by followers descending
profiles.sort(key=lambda x: x["followers"], reverse=True)

with open("footwear_creators_profile_metrics.json", "w", encoding="utf-8") as f:
    json.dump(profiles, f, indent=2)

print("✓ Updated profiles JSON with exact followers and pure tier labels")

# Update posts with new followers
prof_map = {p["raw_handle"].lower(): p for p in profiles}

for p in posts:
    rh = p["raw_handle"].lower()
    if rh in prof_map:
        fols = prof_map[rh]["followers"]
        p["followers"] = fols
        views = p["views"]
        likes = p["likes"]
        comments = p["comments"]
        
        like_rate = round((likes / views) * 100, 2) if views > 0 else 0.0
        view_mult = round(views / fols, 2) if fols > 0 else 0.0
        er = round(((likes + comments) / fols) * 100, 2) if fols > 0 else 0.0
        
        p["like_rate_pct"] = like_rate
        p["view_multiplier"] = view_mult
        p["er_pct"] = er

posts.sort(key=lambda x: (x["tier"], -x["views"]))
with open("footwear_1year_4tier_dataset.json", "w", encoding="utf-8") as f:
    json.dump(posts, f, indent=2)

print("✓ Updated posts JSON with exact followers and ER%")

# ─────────────────────────────────────────────────────────────
# REBUILD EXCEL WORKBOOK
# ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

font_title = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_norm = Font(name="Calibri", size=10, bold=False, color="000000")
font_mute = Font(name="Calibri", size=9, bold=False, color="5D6D7E")
font_link = Font(name="Calibri", size=10, bold=False, color="0563C1", underline="single")

fill_t1_banner = PatternFill("solid", fgColor="145A32") # Dark Emerald
fill_t1_row = PatternFill("solid", fgColor="D4EFDF")    # Mint Green
font_t1_bold = Font(name="Calibri", size=10, bold=True, color="0E6251")
font_t1_link = Font(name="Calibri", size=10, bold=True, color="0B5345", underline="single")

fill_t2_banner = PatternFill("solid", fgColor="1E8449") # Forest Green
fill_t2_row = PatternFill("solid", fgColor="EAFAF1")    # Sage Green
font_t2_bold = Font(name="Calibri", size=10, bold=True, color="196F3D")
font_t2_link = Font(name="Calibri", size=10, bold=False, color="145A32", underline="single")

fill_t3_banner = PatternFill("solid", fgColor="B7950B") # Dark Gold
fill_t3_row = PatternFill("solid", fgColor="FEF9E7")    # Warm Soft Gold
font_t3_bold = Font(name="Calibri", size=10, bold=True, color="7D6608")
font_t3_link = Font(name="Calibri", size=10, bold=False, color="9A7D0A", underline="single")

fill_t4_banner = PatternFill("solid", fgColor="566573") # Slate Gray
fill_t4_row = PatternFill("solid", fgColor="FFFFFF")    # Clean White
font_t4_norm = Font(name="Calibri", size=10, bold=False, color="2C3E50")
font_t4_link = Font(name="Calibri", size=10, bold=False, color="2980B9", underline="single")

thin_line = Side(style="thin", color="D5D8DC")
border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

BRANDS = [
    {"name": "Skechers India", "username": "skechersindia", "state": "Pan-India / Maharashtra (HQ: Mumbai)"},
    {"name": "Gully Labs", "username": "gullylabs", "state": "Pan-India / Delhi NCR (HQ: New Delhi, D2C)"},
    {"name": "Comet", "username": "thecometuniverse", "state": "Pan-India / Karnataka (HQ: Bengaluru, D2C)"},
]

brand_records = defaultdict(list)
for r in posts:
    brand_records[r["brand"]].append(r)

# 1. TAB 1: EXECUTIVE SUMMARY
ws_sum = wb.create_sheet("Executive Summary")
ws_sum.sheet_view.showGridLines = True

ws_sum.merge_cells("A1:O1")
ws_sum["A1"] = "Executive Summary — 1-Year Paid Creator Collab Hierarchy (Skechers India, Gully Labs, Comet) [Aug 2025 – Aug 2026]"
ws_sum["A1"].font = font_title
ws_sum["A1"].fill = PatternFill("solid", fgColor="0B2240")
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 32

sum_headers = [
    ("#", 5),
    ("Brand Name", 24),
    ("State / Origin (HQ)", 32),
    ("Total Collab Posts (1-Yr)", 18),
    ("Total Unique Creators", 18),
    ("🟢 Tier 1: Toggle ON + Boosted\n(Posts / Creators)", 24),
    ("🟢 Tier 2: Toggle ON + Organic\n(Posts / Creators)", 24),
    ("🚀 Tier 3: Toggle OFF + Boosted\n(Posts / Creators)", 24),
    ("⚪ Tier 4: Toggle OFF + Organic (Noise)\n(Posts / Creators)", 28),
    ("💎 Total High-Intent Paid\n(Tiers 1+2+3 Posts)", 22),
    ("High-Intent Paid %\n(Tiers 1+2+3)", 18),
    ("Avg Views / Post", 16),
    ("Avg Creator Followers", 20),
    ("Avg Creator ER%", 15),
    ("Top Creator / Ambassador Samples", 48)
]

for col_idx, (h_text, w) in enumerate(sum_headers, 1):
    c = ws_sum.cell(row=2, column=col_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="1B2631")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_cell
    ws_sum.column_dimensions[get_column_letter(col_idx)].width = w
ws_sum.row_dimensions[2].height = 36
ws_sum.freeze_panes = "A3"

brand_summary_data = []
for b in BRANDS:
    b_name = b["name"]
    b_posts = brand_records.get(b_name, [])
    tot_p = len(b_posts)
    tot_c = len(set(p["handle"].lower() for p in b_posts))
    
    t1_posts = [p for p in b_posts if p["tier"] == 1]
    t1_creators = len(set(p["handle"].lower() for p in t1_posts))
    
    t2_posts = [p for p in b_posts if p["tier"] == 2]
    t2_creators = len(set(p["handle"].lower() for p in t2_posts))
    
    t3_posts = [p for p in b_posts if p["tier"] == 3]
    t3_creators = len(set(p["handle"].lower() for p in t3_posts))
    
    t4_posts = [p for p in b_posts if p["tier"] == 4]
    t4_creators = len(set(p["handle"].lower() for p in t4_posts))
    
    high_intent = len(t1_posts) + len(t2_posts) + len(t3_posts)
    high_intent_pct = high_intent / tot_p if tot_p > 0 else 0.0
    
    views_list = [p["views"] for p in b_posts if p["views"] > 0]
    avg_v = int(sum(views_list) / len(views_list)) if views_list else 0
    
    fols_list = [p["followers"] for p in b_posts if p["followers"] > 0]
    avg_f = int(sum(fols_list) / len(fols_list)) if fols_list else 0
    
    ers_list = [p["er_pct"] for p in b_posts if p["er_pct"] > 0]
    avg_e = round(sum(ers_list) / len(ers_list), 2) if ers_list else 0.0
    
    top_c = list(dict.fromkeys([p["handle"] for p in b_posts if p["tier"] in (1, 2, 3)]))[:4]
    if not top_c:
        top_c = list(dict.fromkeys([p["handle"] for p in b_posts]))[:4]
        
    brand_summary_data.append({
        "brand": b_name,
        "state": b["state"],
        "posts": tot_p,
        "creators": tot_c,
        "t1_p": len(t1_posts), "t1_c": t1_creators,
        "t2_p": len(t2_posts), "t2_c": t2_creators,
        "t3_p": len(t3_posts), "t3_c": t3_creators,
        "t4_p": len(t4_posts), "t4_c": t4_creators,
        "high_intent_p": high_intent,
        "high_intent_pct": high_intent_pct,
        "avg_views": avg_v,
        "avg_followers": avg_f,
        "avg_er": avg_e,
        "top_creators": ", ".join(top_c)
    })

brand_summary_data.sort(key=lambda x: (x["high_intent_p"], x["posts"]), reverse=True)

for idx, b in enumerate(brand_summary_data, 1):
    row_num = idx + 2
    row_vals = [
        idx,
        b["brand"],
        b["state"],
        b["posts"],
        b["creators"],
        f"{b['t1_p']} posts ({b['t1_c']} creators)" if b["t1_p"] > 0 else "—",
        f"{b['t2_p']} posts ({b['t2_c']} creators)" if b["t2_p"] > 0 else "—",
        f"{b['t3_p']} posts ({b['t3_c']} creators)" if b["t3_p"] > 0 else "—",
        f"{b['t4_p']} posts ({b['t4_c']} creators)" if b["t4_p"] > 0 else "—",
        b["high_intent_p"],
        b["high_intent_pct"],
        b["avg_views"],
        b["avg_followers"],
        b["avg_er"] / 100 if b["avg_er"] else 0.0,
        b["top_creators"]
    ]
    
    for col_idx, val in enumerate(row_vals, 1):
        cell = ws_sum.cell(row=row_num, column=col_idx, value=val)
        cell.border = border_cell
        
        if col_idx == 1:
            cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 2:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif col_idx == 3:
            cell.font = Font(name="Calibri", size=9, bold=True, color="2C3E50"); cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = PatternFill("solid", fgColor="F4F6F6")
        elif col_idx in (4, 5):
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"
            cell.fill = PatternFill("solid", fgColor="EBF5FB")
        elif col_idx == 6:
            cell.font = font_t1_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = fill_t1_row
        elif col_idx == 7:
            cell.font = font_t2_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = fill_t2_row
        elif col_idx == 8:
            cell.font = font_t3_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = fill_t3_row
        elif col_idx == 9:
            cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = PatternFill("solid", fgColor="F8F9F9")
        elif col_idx == 10:
            cell.font = Font(name="Calibri", size=10, bold=True, color="1B4F72"); cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"
            if val > 0: cell.fill = PatternFill("solid", fgColor="D6EAF8")
        elif col_idx == 11:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.0%"
        elif col_idx in (12, 13):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif col_idx == 14:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
        elif col_idx == 15:
            cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
            
    ws_sum.row_dimensions[row_num].height = 22

# 2. TAB 2: CREATORS PROFILE METRICS
ws_prof = wb.create_sheet("Creators Profile Metrics", index=1)
ws_prof.sheet_view.showGridLines = True

ws_prof.merge_cells("A1:M1")
ws_prof["A1"] = f"Deduped Creator Profiles & Tier Classification ({len(profiles)} Creators across Skechers India, Gully Labs, Comet)"
ws_prof["A1"].font = font_title
ws_prof["A1"].fill = PatternFill("solid", fgColor="1B4F72")
ws_prof["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_prof.row_dimensions[1].height = 30

prof_headers = [
    ("#", 5),
    ("Creator Handle", 22),
    ("Creator Tier / Size", 30),
    ("Brands Collaborated With", 28),
    ("Full Name", 26),
    ("Verified", 10),
    ("Business / Pro", 14),
    ("Total Followers", 16),
    ("Following", 12),
    ("Total Posts", 14),
    ("Avg Likes / Post", 16),
    ("Avg Comments / Post", 18),
    ("Avg Profile ER%", 14),
]

for col_idx, (h_text, w) in enumerate(prof_headers, 1):
    c = ws_prof.cell(row=2, column=col_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="283747")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_cell
    ws_prof.column_dimensions[get_column_letter(col_idx)].width = w
ws_prof.row_dimensions[2].height = 25
ws_prof.freeze_panes = "A3"

tier_fills = {
    "🌟 Mega Creator / Celebrity (1M+)": PatternFill("solid", fgColor="E8F8F5"),
    "🚀 Macro Creator (100K - 1M)": PatternFill("solid", fgColor="FEF9E7"),
    "✨ Mid-Tier Creator (50K - 100K)": PatternFill("solid", fgColor="EBF5FB"),
    "🎯 Micro Creator (10K - 50K)": PatternFill("solid", fgColor="F4F6F7"),
    "🌱 Nano Creator (<10K)": PatternFill("solid", fgColor="FFFFFF"),
}

for idx, p in enumerate(profiles, 1):
    r_num = idx + 2
    r_vals = [
        idx,
        p["handle"],
        p["creator_tier"],
        p["brands"],
        p["full_name"],
        "Yes" if p["verified"] else "No",
        "Yes" if p["is_business"] else "No",
        p["followers"],
        p["following"],
        p["total_posts"],
        p["avg_likes"],
        p["avg_comments"],
        p["avg_er"] / 100 if p["avg_er"] else 0.0
    ]
    
    tier_fill = tier_fills.get(p["creator_tier"], PatternFill("solid", fgColor="FFFFFF"))
    
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws_prof.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        
        if c_idx == 1:
            cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.hyperlink = p["profile_url"]
        elif c_idx == 3:
            cell.font = Font(name="Calibri", size=10, bold=True, color="1B4F72")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = tier_fill
        elif c_idx in (4, 5):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (6, 7):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
            if val == "Yes": cell.fill = PatternFill("solid", fgColor="EAFAF1")
        elif c_idx in (8, 9, 10, 11, 12):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx == 13:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
            
    ws_prof.row_dimensions[r_num].height = 21

# 3. HELPER FUNCTION FOR MASTER & BRAND TABS
table_cols = [
    ("#", 5),
    ("Hierarchy Tier", 30),
    ("Brand Name", 22),
    ("Creator Handle", 24),
    ("Followers", 14),
    ("Views / Plays", 16),
    ("Likes", 14),
    ("Comments", 12),
    ("Like-to-View %", 15),
    ("Creator ER%", 13),
    ("Post Date", 13),
    ("Direct Instagram URL", 48),
    ("Boost Classification & Reason", 38),
    ("Caption Preview", 65)
]

def render_footwear_sheet(ws, records, sheet_title, state_str):
    ws.sheet_view.showGridLines = True
    
    tot_p = len(records)
    tot_c = len(set(p["handle"].lower() for p in records))
    t1_p = sum(1 for p in records if p["tier"] == 1)
    t2_p = sum(1 for p in records if p["tier"] == 2)
    t3_p = sum(1 for p in records if p["tier"] == 3)
    t4_p = sum(1 for p in records if p["tier"] == 4)
    high_intent = t1_p + t2_p + t3_p
    
    # Header Banner
    ws.merge_cells("A1:N1")
    ws["A1"] = f"👟 {sheet_title.upper()}  |  📍 STATE / REGION: {state_str}"
    ws["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0B2240")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    
    # Metadata Overview
    ws.merge_cells("A2:N2")
    ws["A2"] = f"1-Year Collab Posts: {tot_p}  •  Unique Creators: {tot_c}  •  💎 High-Intent Paid: {high_intent} (T1: {t1_p} | T2: {t2_p} | T3: {t3_p})  •  ⚪ Noise/Unboosted (T4): {t4_p}"
    ws["A2"].font = Font(name="Calibri", size=10, bold=True, color="1B4F72")
    ws["A2"].fill = PatternFill("solid", fgColor="EBF5FB")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22
    
    # Table Header Row
    for col_idx, (h_text, w) in enumerate(table_cols, 1):
        c = ws.cell(row=3, column=col_idx, value=h_text)
        c.font = font_hdr
        c.fill = PatternFill("solid", fgColor="1F2D3D")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_cell
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[3].height = 25
    ws.freeze_panes = "A4"
    
    records_by_tier = {1: [], 2: [], 3: [], 4: []}
    for r in records:
        records_by_tier[r["tier"]].append(r)
        
    current_row = 4
    global_index = 1
    
    tier_meta = [
        (1, "🟢 TIER 1: TOGGLE ON + 🚀 BOOSTED (Formal Paid Partnership Label + Paid Ad Spend)", fill_t1_banner, fill_t1_row, font_t1_bold, font_t1_link),
        (2, "🟢 TIER 2: TOGGLE ON + ⚪ ORGANIC (Formal Paid Partnership Label + Organic Reach Only)", fill_t2_banner, fill_t2_row, font_t2_bold, font_t2_link),
        (3, "🚀 TIER 3: TOGGLE OFF + 🚀 BOOSTED (Co-Author Collab + Heavy Paid Ad Spend Detected)", fill_t3_banner, fill_t3_row, font_t3_bold, font_t3_link),
        (4, "⚪ TIER 4: TOGGLE OFF + ⚪ ORGANIC (Standard Collab / Low Organic Reach / Noise)", fill_t4_banner, fill_t4_row, font_t4_norm, font_t4_link)
    ]
    
    for t_id, banner_text, fill_banner, fill_row, font_b, font_l in tier_meta:
        t_records = records_by_tier[t_id]
        if not t_records:
            continue
            
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        banner_cell = ws.cell(row=current_row, column=1, value=f"{banner_text} — {len(t_records)} Posts ({len(set(p['handle'].lower() for p in t_records))} Unique Creators)")
        banner_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        banner_cell.fill = fill_banner
        banner_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[current_row].height = 23
        current_row += 1
        
        for p in t_records:
            vals = [
                global_index,
                p["tier_name"],
                p["brand"],
                p["handle"],
                p["followers"],
                p["views"],
                p["likes"],
                p["comments"],
                p["like_rate_pct"] / 100 if p["like_rate_pct"] else 0.0,
                p["er_pct"] / 100 if p["er_pct"] else 0.0,
                p["date"],
                p["url"],
                f"{p['boost_status']}: {p['boost_reason']}" if p.get('boost_reason') else p['boost_status'],
                p["caption"]
            ]
            
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=current_row, column=c_idx, value=val)
                cell.border = border_cell
                cell.fill = fill_row
                
                if c_idx == 1:
                    cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 2:
                    cell.font = font_b; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx in (3, 4):
                    cell.font = font_b; cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx in (5, 6, 7, 8):
                    cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
                elif c_idx in (9, 10):
                    cell.font = font_b; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
                elif c_idx == 11:
                    cell.font = font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 12:
                    cell.font = font_l; cell.alignment = Alignment(horizontal="left", vertical="center")
                    if val: cell.hyperlink = val
                elif c_idx == 13:
                    cell.font = font_b; cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx == 14:
                    cell.font = font_mute; cell.alignment = Alignment(horizontal="left", vertical="center")
                    
            ws.row_dimensions[current_row].height = 21
            current_row += 1
            global_index += 1

# Tab 3: Master Hierarchy
ws_all = wb.create_sheet("All Brands - Master Hierarchy")
render_footwear_sheet(
    ws_all,
    posts,
    "CONSOLIDATED 3 SNEAKER & FOOTWEAR BRANDS",
    "PAST 1 YEAR (AUG 2025 – AUG 2026)"
)

# Tabs 4 to 6: Individual Brand sheets
for b in BRANDS:
    b_name = b["name"]
    b_posts = brand_records.get(b_name, [])
    ws_brand = wb.create_sheet(b_name[:31])
    render_footwear_sheet(
        ws_brand,
        b_posts,
        b_name,
        b["state"]
    )

wb.save("footwear_sneaker_brands_master_analysis.xlsx")
print("✅ Saved updated footwear_sneaker_brands_master_analysis.xlsx")

# ─────────────────────────────────────────────────────────────
# EXPORT CSVS
# ─────────────────────────────────────────────────────────────
with open("Footwear_Creator_Profile_Metrics.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow([
        "#", "Creator Handle", "Creator Tier / Size", "Brands Collaborated With", "Full Name",
        "Verified", "Business / Professional", "Total Followers", "Following", "Total Posts",
        "Avg Likes / Post", "Avg Comments / Post", "Avg Profile ER%", "Instagram Profile URL"
    ])
    for idx, p in enumerate(profiles, 1):
        w.writerow([
            idx,
            p["handle"],
            p["creator_tier"],
            p["brands"],
            p["full_name"],
            "Yes" if p["verified"] else "No",
            "Yes" if p["is_business"] else "No",
            p["followers"],
            p["following"],
            p["total_posts"],
            p["avg_likes"],
            p["avg_comments"],
            f"{p['avg_er']:.2f}%",
            p["profile_url"]
        ])
print("✓ Saved updated Footwear_Creator_Profile_Metrics.csv")

with open("Footwear_All_Brands_4Tier_Master.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow([
        "#", "Hierarchy Tier", "Brand Name", "State / Origin (HQ)", "Creator Handle", "Followers",
        "Views / Plays", "Likes", "Comments", "Like-to-View %", "Creator ER%",
        "Post Date", "Direct Instagram URL", "Boost Status & Reason", "Caption Preview"
    ])
    for idx, p in enumerate(posts, 1):
        w.writerow([
            idx,
            p["tier_name"],
            p["brand"],
            p["state"],
            p["handle"],
            p["followers"],
            p["views"],
            p["likes"],
            p["comments"],
            f"{p['like_rate_pct']:.2f}%",
            f"{p['er_pct']:.2f}%",
            p["date"],
            p["url"],
            f"{p['boost_status']}: {p['boost_reason']}" if p.get('boost_reason') else p['boost_status'],
            p["caption"]
        ])
print("✓ Saved updated Footwear_All_Brands_4Tier_Master.csv")

with open("Footwear_Brand_Summary.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow([
        "#", "Brand Name", "State / Origin (HQ)", "Total Collab Posts (1-Yr)", "Total Unique Creators",
        "Tier 1 (Toggle ON + Boosted)", "Tier 2 (Toggle ON + Organic)",
        "Tier 3 (Toggle OFF + Boosted)", "Tier 4 (Noise / Unboosted)",
        "Total High-Intent Paid Posts", "High-Intent Paid %",
        "Avg Estimated Views", "Avg Creator Followers", "Avg Creator ER%", "Top Creator Samples"
    ])
    for idx, b in enumerate(brand_summary_data, 1):
        w.writerow([
            idx,
            b["brand"],
            b["state"],
            b["posts"],
            b["creators"],
            f"{b['t1_p']} posts ({b['t1_c']} creators)" if b["t1_p"] > 0 else "—",
            f"{b['t2_p']} posts ({b['t2_c']} creators)" if b["t2_p"] > 0 else "—",
            f"{b['t3_p']} posts ({b['t3_c']} creators)" if b["t3_p"] > 0 else "—",
            f"{b['t4_p']} posts ({b['t4_c']} creators)" if b["t4_p"] > 0 else "—",
            b["high_intent_p"],
            f"{b['high_intent_pct']*100:.1f}%",
            b["avg_views"],
            b["avg_followers"],
            f"{b['avg_er']:.2f}%",
            b["top_creators"]
        ])
print("✓ Saved updated Footwear_Brand_Summary.csv")
