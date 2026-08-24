"""
Add Creator Tier Column (Mega, Macro, Mid-Tier, Micro, Nano, Brand)
to Creators Profile Metrics sheet in Excel and CSV
"""

import sys, json, openpyxl, csv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

with open("footwear_creators_profile_metrics.json", encoding="utf-8") as f:
    profiles = json.load(f)

BRAND_HANDLES = {
    "royalenfield", "sharktank.india", "myntra", "amazonfashionin", "bombaysweetshop",
    "casabacardiin", "cmf.tech", "districtupdates", "indiansneakerfestival", "kommunedelhincr",
    "medusaindia", "kauraverse", "leada.in", "niviasports", "rahasyafragrances", "parvaazmusic",
    "skecherscricket", "thethirdspacedelhi", "yuzenmatcha", "farakwear", "unoindiaofficial", "capsulindia", "delhiartweekend"
}

def classify_creator_tier(handle, followers):
    clean_h = handle.replace("@", "").lower()
    if clean_h in BRAND_HANDLES:
        return "🏢 Brand / Media Partner"
    
    if followers >= 1000000:
        return "🌟 Mega Creator / Celebrity (1M+)"
    elif followers >= 100000:
        return "🚀 Macro Creator (100K - 1M)"
    elif followers >= 50000:
        return "✨ Mid-Tier Creator (50K - 100K)"
    elif followers >= 10000:
        return "🎯 Micro Creator (10K - 50K)"
    else:
        return "🌱 Nano Creator (<10K)"

# Enrich profiles
for p in profiles:
    p["creator_tier"] = classify_creator_tier(p["raw_handle"], p["followers"])

with open("footwear_creators_profile_metrics.json", "w", encoding="utf-8") as f:
    json.dump(profiles, f, indent=2)

print("✓ Updated footwear_creators_profile_metrics.json with Creator Tier")

# Update Excel
wb = openpyxl.load_workbook("footwear_sneaker_brands_master_analysis.xlsx")

if "Creators Profile Metrics" in wb.sheetnames:
    del wb["Creators Profile Metrics"]

ws_prof = wb.create_sheet("Creators Profile Metrics", index=1)
ws_prof.sheet_view.showGridLines = True

font_title = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_norm = Font(name="Calibri", size=10, bold=False, color="000000")
font_mute = Font(name="Calibri", size=9, bold=False, color="5D6D7E")

thin_line = Side(style="thin", color="D5D8DC")
border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

ws_prof.merge_cells("A1:M1")
ws_prof["A1"] = f"Deduped Creator Profiles & Tier Classification ({len(profiles)} Creators across Skechers India, Gully Labs, Comet)"
ws_prof["A1"].font = font_title
ws_prof["A1"].fill = PatternFill("solid", fgColor="1B4F72")
ws_prof["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_prof.row_dimensions[1].height = 30

prof_headers = [
    ("#", 5),
    ("Creator Handle", 22),
    ("Creator Tier / Size", 30),
    ("Brands Collaborated With", 28),
    ("Full Name", 26),
    ("Verified", 10),
    ("Business / Pro", 14),
    ("Total Followers", 16),
    ("Following", 12),
    ("Total Posts", 14),
    ("Avg Likes / Post", 16),
    ("Avg Comments / Post", 18),
    ("Avg Profile ER%", 14),
]

for col_idx, (h_text, w) in enumerate(prof_headers, 1):
    c = ws_prof.cell(row=2, column=col_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="283747")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_cell
    ws_prof.column_dimensions[get_column_letter(col_idx)].width = w
ws_prof.row_dimensions[2].height = 25
ws_prof.freeze_panes = "A3"

# Tier coloring
tier_fills = {
    "🌟 Mega Creator / Celebrity (1M+)": PatternFill("solid", fgColor="E8F8F5"),
    "🚀 Macro Creator (100K - 1M)": PatternFill("solid", fgColor="FEF9E7"),
    "✨ Mid-Tier Creator (50K - 100K)": PatternFill("solid", fgColor="EBF5FB"),
    "🎯 Micro Creator (10K - 50K)": PatternFill("solid", fgColor="F4F6F7"),
    "🌱 Nano Creator (<10K)": PatternFill("solid", fgColor="FFFFFF"),
    "🏢 Brand / Media Partner": PatternFill("solid", fgColor="FDEDEC"),
}

for idx, p in enumerate(profiles, 1):
    r_num = idx + 2
    r_vals = [
        idx,
        p["handle"],
        p["creator_tier"],
        p["brands"],
        p["full_name"],
        "Yes" if p["verified"] else "No",
        "Yes" if p["is_business"] else "No",
        p["followers"],
        p["following"],
        p["total_posts"],
        p["avg_likes"],
        p["avg_comments"],
        p["avg_er"] / 100 if p["avg_er"] else 0.0
    ]
    
    tier_fill = tier_fills.get(p["creator_tier"], PatternFill("solid", fgColor="FFFFFF"))
    
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws_prof.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        
        if c_idx == 1:
            cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.hyperlink = p["profile_url"]
        elif c_idx == 3:
            cell.font = Font(name="Calibri", size=10, bold=True, color="1B4F72")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = tier_fill
        elif c_idx in (4, 5):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (6, 7):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
            if val == "Yes": cell.fill = PatternFill("solid", fgColor="EAFAF1")
        elif c_idx in (8, 9, 10, 11, 12):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx == 13:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
            
    ws_prof.row_dimensions[r_num].height = 21

wb.save("footwear_sneaker_brands_master_analysis.xlsx")
print("✓ Saved updated footwear_sneaker_brands_master_analysis.xlsx with Creator Tier column")

# Export CSV
with open("Footwear_Creator_Profile_Metrics.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow([
        "#", "Creator Handle", "Creator Tier / Size", "Brands Collaborated With", "Full Name",
        "Verified", "Business / Professional", "Total Followers", "Following", "Total Posts",
        "Avg Likes / Post", "Avg Comments / Post", "Avg Profile ER%", "Instagram Profile URL"
    ])
    for idx, p in enumerate(profiles, 1):
        w.writerow([
            idx,
            p["handle"],
            p["creator_tier"],
            p["brands"],
            p["full_name"],
            "Yes" if p["verified"] else "No",
            "Yes" if p["is_business"] else "No",
            p["followers"],
            p["following"],
            p["total_posts"],
            p["avg_likes"],
            p["avg_comments"],
            f"{p['avg_er']:.2f}%",
            p["profile_url"]
        ])
print("✓ Saved updated Footwear_Creator_Profile_Metrics.csv")
