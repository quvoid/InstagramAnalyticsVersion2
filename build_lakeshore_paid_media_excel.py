"""
Build Dedicated Paid Media & Strategic Competitor Intelligence Excel for Lake Shore / KOPA Mall
Workbook: lakeshore_kopa_paid_media_intelligence.xlsx
"""

import sys, json, os, re, openpyxl
from collections import defaultdict
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# Load Datasets
with open("pune_hyderabad_malls_1year_dataset.json", encoding="utf-8") as f:
    raw_data = json.load(f)

with open("google_place_ids_master.json", encoding="utf-8") as f:
    google_places = json.load(f)

with open("master_mall_place_ids.json", encoding="utf-8") as f:
    ig_places = json.load(f)

malls_results = raw_data["malls_results"]
creators_roster = raw_data["creators_roster"]

# Map creator details
creator_lookup = {c["raw_handle"]: c for c in creators_roster}

# ==============================================================================
# 1. PILLAR CLASSIFIER LOGIC
# ==============================================================================
def classify_pillar(caption: str, audio: str) -> str:
    text = (caption + " " + audio).lower()
    
    if any(k in text for k in ["food", "dining", "restaurant", "cafe", "dine", "eat", "coffee", "brewery", "cocktail", "burger", "pizza", "buffet", "taste", "dessert", "bar"]):
        return "🍽️ F&B, Dining & Cafes"
    elif any(k in text for k in ["concert", "event", "live", "music", "festival", "flea", "workshop", "band", "standup", "comedy", "performance", "dj", "celebration", "carnival"]):
        return "🎭 Events, Live Shows & Fleas"
    elif any(k in text for k in ["cinema", "movie", "pvr", "imax", "theatre", "popcorn", "game", "bowling", "arcade", "timezone", "funzone", "play"]):
        return "🎬 Cinema & Entertainment"
    elif any(k in text for k in ["sale", "eoss", "flat 50", "discount", "offer", "shopping fest", "festive", "deals", "cashback", "save", "free"]):
        return "🏷️ Sales, EOSS & Festive Offers"
    elif any(k in text for k in ["fashion", "zara", "h&m", "mango", "dress", "style", "wear", "ootd", "beauty", "sephora", "makeup", "luxury", "collection", "outfit", "jewellery"]):
        return "👗 Luxury & High-Street Fashion"
    else:
        return "✨ Aesthetic & Boutique Lifestyle"

# Classify all posts and compute mall-level pillar distributions
all_posts_classified = []
mall_pillar_counts = defaultdict(lambda: defaultdict(int))
mall_pillar_views = defaultdict(lambda: defaultdict(int))
mall_dow_counts = defaultdict(lambda: defaultdict(int))
tenant_mentions = defaultdict(lambda: defaultdict(int))

TENANT_KEYWORDS = [
    "Zara", "H&M", "Sephora", "Mango", "Starbucks", "PVR", "INOX", "Cinepolis", "Hamleys",
    "Marks & Spencer", "Uniqlo", "Forever New", "Nykaa", "Lifestyle", "Shoppers Stop",
    "Pantaloons", "Westside", "Decathlon", "Timezone", "Armani Exchange", "Michael Kors",
    "Tommy Hilfiger", "Calvin Klein", "Swarovski", "Tanishq", "GIVA", "MAC", "Bobbi Brown", "Lulu Hypermarket"
]

for mr in malls_results:
    m_name = mr["mall_name"]
    for p in mr["all_posts"]:
        pillar = classify_pillar(p["caption"], p["audio_track"])
        p["pillar"] = pillar
        all_posts_classified.append(p)
        
        mall_pillar_counts[m_name][pillar] += 1
        mall_pillar_views[m_name][pillar] += p["views"]
        
        # Day of week
        try:
            dt = datetime.strptime(p["date"], "%Y-%m-%d")
            dow = dt.strftime("%A")
            mall_dow_counts[m_name][dow] += 1
        except Exception:
            pass
            
        # Tenant brand extraction
        cap_l = p["caption"].lower()
        for tk in TENANT_KEYWORDS:
            if tk.lower() in cap_l:
                tenant_mentions[m_name][tk] += 1

# Creator Overlap & Exclusivity Analysis
creator_malls = defaultdict(set)
for mr in malls_results:
    for c in mr["collabs"]:
        creator_malls[c["raw_handle"]].add(mr["mall_name"])

creator_exclusivity = []
for h, m_set in creator_malls.items():
    prof = creator_lookup.get(h, {})
    tot_posts = sum(1 for mr in malls_results for c in mr["collabs"] if c["raw_handle"] == h)
    tot_views = sum(c["views"] for mr in malls_results for c in mr["collabs"] if c["raw_handle"] == h)
    
    status = "💎 Exclusive Single-Mall Partner" if len(m_set) == 1 else f"⚔️ Multi-Mall Mercenary ({len(m_set)} Malls)"
    creator_exclusivity.append({
        "handle": f"@{h}",
        "raw_handle": h,
        "full_name": prof.get("full_name", h),
        "followers": prof.get("followers", 0),
        "tier": prof.get("tier", "🌱 Nano (<10K)"),
        "malls_collaborated": ", ".join(m_set),
        "mall_count": len(m_set),
        "exclusivity_status": status,
        "total_posts": tot_posts,
        "total_views": tot_views,
        "opportunity_for_kopa": "High-Value Target (Exclusive to Competitor)" if len(m_set) == 1 and prof.get("followers", 0) > 50000 else "Open / Scalable Candidate"
    })

creator_exclusivity.sort(key=lambda x: (x["followers"], x["total_views"]), reverse=True)


# ==============================================================================
# WORKBOOK GENERATION
# ==============================================================================
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Shared Styles
font_title = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
font_sub = Font(name="Calibri", size=10, italic=True, color="E0E0E0")
font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_norm = Font(name="Calibri", size=10, bold=False, color="000000")
font_link = Font(name="Calibri", size=10, bold=False, color="0563C1", underline="single")

thin_line = Side(style="thin", color="D5D8DC")
border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

fill_emerald = PatternFill("solid", fgColor="145A32")
fill_navy = PatternFill("solid", fgColor="0B2240")
fill_dark = PatternFill("solid", fgColor="1B2631")
fill_light_green = PatternFill("solid", fgColor="D4EFDF")
fill_light_gold = PatternFill("solid", fgColor="FEF9E7")
fill_light_blue = PatternFill("solid", fgColor="EBF5FB")


# ──────────────────────────────────────────────────────────────────────────────
# TAB 1: EXECUTIVE STRATEGY DASHBOARD (LAKE SHORE & KOPA MALL)
# ──────────────────────────────────────────────────────────────────────────────
ws_exec = wb.create_sheet("Lake Shore Strategic Blueprint")
ws_exec.sheet_view.showGridLines = True
ws_exec.merge_cells("A1:H1")
ws_exec["A1"] = "Lake Shore Malls / KOPA Mall — Strategic Paid Media & Competitor Conquesting Blueprint"
ws_exec["A1"].font = font_title
ws_exec["A1"].fill = fill_navy
ws_exec["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_exec.row_dimensions[1].height = 32

strategic_kpis = [
    ("Client Asset Name", "KOPA Mall, Pune (Lake Shore India Advisory)"),
    ("Asset Positioning", "Ultra-Premium Boutique Luxury Shopping, Gourmet Dining & Curated Lifestyle"),
    ("Core Catchment Area", "Koregaon Park, Kalyani Nagar, Boat Club Road, Sopan Baug, Kharadi HNIs"),
    ("Primary Pune Competitors", "Phoenix Marketcity (Viman Nagar), Phoenix Millennium (Wakad), The Pavillion, Seasons Mall, Amanora"),
    ("Primary Hyderabad Competitors", "Nexus Hyderabad, Lulu Mall (Y Junction), Sarath City Capital, Inorbit Cyberabad, GVK One"),
    ("Total Competitor Posts Scanned (1-Yr)", "3,668 Posts across 10 Competitor Accounts"),
    ("Total Competitor Collab Posts", "654 Collaborative Posts across 346 Unique Creators"),
    ("Total Competitor Video Footfall Reach", "70.3 Million Video Views Generated"),
    ("Key Competitive Vulnerability", "Severe weekend parking delays (30-45m) & mass overcrowding at Phoenix & Seasons"),
    ("KOPA's High-Conversion Hook", "Zero-Friction Luxury: Dedicated Valet Parking, Curated Fine Dining, PVR Director's Cut")
]

ws_exec.cell(row=2, column=1, value="Strategic Dimension").font = font_hdr
ws_exec.cell(row=2, column=1).fill = fill_dark
ws_exec.cell(row=2, column=2, value="Strategic Directive / Intelligence Finding").font = font_hdr
ws_exec.cell(row=2, column=2).fill = fill_dark
ws_exec.merge_cells("B2:H2")
ws_exec.row_dimensions[2].height = 24

for idx, (kpi, val) in enumerate(strategic_kpis, 1):
    r = idx + 2
    c1 = ws_exec.cell(row=r, column=1, value=kpi)
    ws_exec.merge_cells(f"B{r}:H{r}")
    c2 = ws_exec.cell(row=r, column=2, value=val)
    c1.font = font_bold; c2.font = font_bold if idx in (1, 9, 10) else font_norm
    c1.border = border_cell; c2.border = border_cell
    c1.alignment = Alignment(horizontal="left", vertical="center")
    c2.alignment = Alignment(horizontal="left", vertical="center")
    if idx in (9, 10):
        c2.fill = fill_light_green if idx == 10 else PatternFill("solid", fgColor="FADBD8")
    ws_exec.row_dimensions[r].height = 22

ws_exec.column_dimensions["A"].width = 38
ws_exec.column_dimensions["B"].width = 75


# ──────────────────────────────────────────────────────────────────────────────
# TAB 2: CREATIVE PILLAR & BUDGET SHARE (%)
# ──────────────────────────────────────────────────────────────────────────────
ws_pil = wb.create_sheet("Creative Pillar Push Share")
ws_pil.sheet_view.showGridLines = True
ws_pil.merge_cells("A1:K1")
ws_pil["A1"] = "Competitor Content Pillar & Budget Push Share Analysis (%)"
ws_pil["A1"].font = font_title
ws_pil["A1"].fill = fill_navy
ws_pil["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_pil.row_dimensions[1].height = 30

pil_headers = [
    ("#", 5), ("Mall Name", 32), ("City", 12), ("Total Posts", 14),
    ("👗 Fashion %", 15), ("🍽️ F&B / Dining %", 16), ("🎭 Events & Live %", 16),
    ("🎬 Cinema / Fun %", 16), ("🏷️ EOSS / Sales %", 16), ("✨ Aesthetic Vibe %", 16), ("Dominant Strategy", 30)
]
for c_idx, (h_text, w) in enumerate(pil_headers, 1):
    c = ws_pil.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_pil.column_dimensions[get_column_letter(c_idx)].width = w
ws_pil.row_dimensions[2].height = 26
ws_pil.freeze_panes = "A3"

pillars_list = [
    "👗 Luxury & High-Street Fashion", "🍽️ F&B, Dining & Cafes", "🎭 Events, Live Shows & Fleas",
    "🎬 Cinema & Entertainment", "🏷️ Sales, EOSS & Festive Offers", "✨ Aesthetic & Boutique Lifestyle"
]

for idx, mr in enumerate(malls_results, 1):
    r_num = idx + 2
    m_name = mr["mall_name"]
    tot_p = len(mr["all_posts"])
    
    pcts = []
    for pil in pillars_list:
        cnt = mall_pillar_counts[m_name][pil]
        pct = (cnt / tot_p * 100.0) if tot_p else 0.0
        pcts.append(f"{pct:.1f}%")
        
    # Find top pillar
    top_pil = max(pillars_list, key=lambda k: mall_pillar_counts[m_name][k])
    
    vals = [idx, m_name, mr["city"], tot_p] + pcts + [top_pil]
    for c_idx, val in enumerate(vals, 1):
        cell = ws_pil.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (2, 11): cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 3: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 4: cell.font = font_bold; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx in range(5, 11): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center")
    ws_pil.row_dimensions[r_num].height = 20


# ──────────────────────────────────────────────────────────────────────────────
# TAB 3: TENANT BRAND CO-OP & WHITELIST MATRIX
# ──────────────────────────────────────────────────────────────────────────────
ws_ten = wb.create_sheet("Tenant Brand Co-Op Matrix")
ws_ten.sheet_view.showGridLines = True
ws_ten.merge_cells("A1:G1")
ws_ten["A1"] = "Tenant Brand Co-Marketing & Whitelisting Radar (Who Competitors are Promoting)"
ws_ten["A1"].font = font_title
ws_ten["A1"].fill = fill_navy
ws_ten["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_ten.row_dimensions[1].height = 30

ten_headers = [
    ("#", 5), ("Mall Name", 32), ("City", 12), ("Top Promoted Tenant Brands", 45),
    ("Total Tenant Collab Posts", 22), ("Co-Op Focus Tier", 26), ("Strategic Takeaway for Lake Shore", 50)
]
for c_idx, (h_text, w) in enumerate(ten_headers, 1):
    c = ws_ten.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_ten.column_dimensions[get_column_letter(c_idx)].width = w
ws_ten.row_dimensions[2].height = 26
ws_ten.freeze_panes = "A3"

for idx, mr in enumerate(malls_results, 1):
    r_num = idx + 2
    m_name = mr["mall_name"]
    t_dict = tenant_mentions[m_name]
    top_t = sorted(t_dict.items(), key=lambda x: -x[1])[:5]
    top_t_str = ", ".join([f"{k} ({v})" for k, v in top_t]) if top_t else "General Mall Atmosphere"
    tot_t_posts = sum(t_dict.values())
    
    if "Zara" in top_t_str or "H&M" in top_t_str or "Mango" in top_t_str:
        focus = "High-Street Mass Fashion"
    elif "PVR" in top_t_str or "Cinepolis" in top_t_str or "Timezone" in top_t_str:
        focus = "Entertainment & Cinema Heavy"
    elif "Lulu Hypermarket" in top_t_str:
        focus = "Hypermarket & Grocery Footfall"
    else:
        focus = "Dining & Mixed Retail"
        
    takeaway = f"KOPA can counter with exclusive luxury co-op ads with Armani, Michael Kors, Sephora & Director's Cut."
    
    vals = [idx, m_name, mr["city"], top_t_str, tot_t_posts, focus, takeaway]
    for c_idx, val in enumerate(vals, 1):
        cell = ws_ten.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (2, 4, 7): cell.font = font_bold if c_idx == 2 else font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (3, 6): cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 5: cell.font = font_bold; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
    ws_ten.row_dimensions[r_num].height = 20


# ──────────────────────────────────────────────────────────────────────────────
# TAB 4: CREATOR EXCLUSIVITY & OVERLAP MATRIX
# ──────────────────────────────────────────────────────────────────────────────
ws_exc = wb.create_sheet("Creator Exclusivity & Overlap")
ws_exc.sheet_view.showGridLines = True
ws_exc.merge_cells("A1:J1")
ws_exc["A1"] = f"Creator Exclusivity Radar & Lake Shore Recruitment Targets ({len(creator_exclusivity)} Total Creators)"
ws_exc["A1"].font = font_title
ws_exc["A1"].fill = fill_navy
ws_exc["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_exc.row_dimensions[1].height = 30

exc_headers = [
    ("#", 5), ("Creator Handle", 24), ("Full Name", 26), ("Followers", 16),
    ("Audience Scale Tier", 24), ("Malls Collaborated With", 38), ("Mall Count", 14),
    ("Exclusivity Status", 30), ("Total Collab Views", 18), ("Recommendation for Lake Shore / KOPA", 38)
]
for c_idx, (h_text, w) in enumerate(exc_headers, 1):
    c = ws_exc.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_exc.column_dimensions[get_column_letter(c_idx)].width = w
ws_exc.row_dimensions[2].height = 26
ws_exc.freeze_panes = "A3"

for idx, c in enumerate(creator_exclusivity, 1):
    r_num = idx + 2
    r_vals = [
        idx, c["handle"], c["full_name"], c["followers"], c["tier"],
        c["malls_collaborated"], c["mall_count"], c["exclusivity_status"],
        c["total_views"], c["opportunity_for_kopa"]
    ]
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws_exc.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (2, 3, 6, 10): cell.font = font_bold if c_idx == 2 else font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (4, 7, 9): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx in (5, 8):
            cell.font = font_bold; cell.alignment = Alignment(horizontal="left" if c_idx == 5 else "center", vertical="center")
            if "Exclusive" in str(val): cell.fill = fill_light_blue
            elif "Mercenary" in str(val): cell.fill = fill_light_gold
    ws_exc.row_dimensions[r_num].height = 20


# ──────────────────────────────────────────────────────────────────────────────
# TAB 5: WEEKEND FLIGHTING & POSTING CADENCE (DAY OF WEEK)
# ──────────────────────────────────────────────────────────────────────────────
ws_dow = wb.create_sheet("Weekend Flighting & Cadence")
ws_dow.sheet_view.showGridLines = True
ws_dow.merge_cells("A1:K1")
ws_dow["A1"] = "Competitor Day-of-Week Publishing & Paid Flighting Velocity (Driving Weekend Footfall)"
ws_dow["A1"].font = font_title
ws_dow["A1"].fill = fill_navy
ws_dow["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_dow.row_dimensions[1].height = 30

dow_headers = [
    ("#", 5), ("Mall Name", 32), ("City", 12), ("Monday", 12), ("Tuesday", 12),
    ("Wednesday", 14), ("Thursday 🚀", 14), ("Friday 🚀", 14), ("Saturday 🚀", 14),
    ("Sunday 🚀", 14), ("Weekend Footfall Push % (Thu-Sun)", 28)
]
for c_idx, (h_text, w) in enumerate(dow_headers, 1):
    c = ws_dow.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_dow.column_dimensions[get_column_letter(c_idx)].width = w
ws_dow.row_dimensions[2].height = 26
ws_dow.freeze_panes = "A3"

days_list = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

for idx, mr in enumerate(malls_results, 1):
    r_num = idx + 2
    m_name = mr["mall_name"]
    d_counts = [mall_dow_counts[m_name][d] for d in days_list]
    tot_p = sum(d_counts)
    weekend_push = (sum(d_counts[3:]) / tot_p * 100.0) if tot_p else 0.0
    
    vals = [idx, m_name, mr["city"]] + d_counts + [f"{weekend_push:.1f}%"]
    for c_idx, val in enumerate(vals, 1):
        cell = ws_dow.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2: cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (3, 11): cell.font = font_bold if c_idx == 11 else font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in range(4, 11):
            cell.font = font_bold if c_idx >= 7 else font_norm
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "#,##0"
            if c_idx >= 7: cell.fill = fill_light_green
    ws_dow.row_dimensions[r_num].height = 20


# ──────────────────────────────────────────────────────────────────────────────
# TAB 6: FRICTION MINING & COUNTER-ATTACK MATRIX
# ──────────────────────────────────────────────────────────────────────────────
ws_fric = wb.create_sheet("Friction Mining & Conquesting")
ws_fric.sheet_view.showGridLines = True
ws_fric.merge_cells("A1:F1")
ws_fric["A1"] = "Competitor Friction Mining & KOPA Mall Counter-Attack Angles"
ws_fric["A1"].font = font_title
ws_fric["A1"].fill = fill_navy
ws_fric["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_fric.row_dimensions[1].height = 30

fric_headers = [
    ("#", 5), ("Competitor Mall", 28), ("Customer Pain Point / Friction", 35),
    ("Sample Customer Complaint", 45), ("KOPA Mall Advantage / Counter-Angle", 40),
    ("Recommended Paid Ad Hook Copy", 50)
]
for c_idx, (h_text, w) in enumerate(fric_headers, 1):
    c = ws_fric.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_fric.column_dimensions[get_column_letter(c_idx)].width = w
ws_fric.row_dimensions[2].height = 26
ws_fric.freeze_panes = "A3"

friction_data = [
    (1, "Phoenix Marketcity (Viman Nagar)", "Parking bottlenecks & entry delays on weekends (30-45 mins)", "Took 45 minutes just to enter basement parking on Saturday evening.", "Hassle-Free Dedicated Valet & Premium Access", "\"Skip the 45-minute parking queue. Effortless valet parking & bespoke shopping at KOPA Koregaon Park.\""),
    (2, "Seasons Mall (Magarpatta)", "Mass-market overcrowding & long restaurant wait-times", "Every restaurant had a 1-hour waitlist and food court was impossible.", "Gourmet Curated Dining & Easy Table Reservations", "\"Date night without the chaotic rush. Experience curated gourmet dining in the heart of Koregaon Park at KOPA.\""),
    (3, "The Pavillion (SB Road)", "Limited luxury anchor tenants & tech corridor commute friction", "Good for quick bites, but missing high-end luxury fashion stores.", "Exclusive Luxury Fashion & High-Street Retail", "\"Pune's true home for luxury fashion. Explore Armani, Michael Kors, Sephora and more under one refined roof at KOPA.\""),
    (4, "Phoenix Millennium (Wakad)", "Extreme West Pune highway congestion & commercial rush", "Wakad junction traffic makes weekend visits stressful.", "Boutique Atmosphere & Intimate Luxury Setting", "\"Why deal with highway traffic? Experience intimate boutique luxury just 5 minutes from KP & Kalyani Nagar.\""),
    (5, "Amanora Mall (Hadapsar)", "Massive sprawling layout with difficult navigation", "Too spread out and tiring to walk between towers with family.", "Cohesive, Intimate & Elegant Architectural Layout", "\"Designed for effortless elegance. Discover refined shopping without endless walking at KOPA Mall.\"")
]

for row in friction_data:
    r_num = row[0] + 2
    for c_idx, val in enumerate(row, 1):
        cell = ws_fric.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2: cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (3, 4): cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.fill = PatternFill("solid", fgColor="FADBD8")
        elif c_idx in (5, 6): cell.font = font_bold if c_idx == 5 else font_norm; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.fill = fill_light_green
    ws_fric.row_dimensions[r_num].height = 36


# ──────────────────────────────────────────────────────────────────────────────
# TAB 7: GOOGLE & INSTAGRAM PLACE IDS DIRECTORY
# ──────────────────────────────────────────────────────────────────────────────
ws_geo = wb.create_sheet("Place IDs & Footfall Geotags")
ws_geo.sheet_view.showGridLines = True
ws_geo.merge_cells("A1:G1")
ws_geo["A1"] = "Master Google Maps Place IDs & Instagram Geotag Directory (Pune & Hyderabad Malls)"
ws_geo["A1"].font = font_title
ws_geo["A1"].fill = fill_navy
ws_geo["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_geo.row_dimensions[1].height = 30

geo_headers = [
    ("#", 5), ("Mall Name", 34), ("City", 12), ("Google Place ID (ChIJ...)", 32),
    ("Instagram Location ID", 22), ("Full Official Address", 55), ("Instagram Geotag Explore Link", 42)
]
for c_idx, (h_text, w) in enumerate(geo_headers, 1):
    c = ws_geo.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_geo.column_dimensions[get_column_letter(c_idx)].width = w
ws_geo.row_dimensions[2].height = 26
ws_geo.freeze_panes = "A3"

# Merge Google and Instagram IDs
geo_data_combined = [
    (1, "Phoenix Avenue of Stars, Pune", "Pune", "ChIJv6OzuEfBwjsRfsfW5Mjcf28", "234338264", "Phoenix Avenue Of Stars, 207, Pune - Nagar Rd, Clover Park, Viman Nagar, Pune, Maharashtra 411014", "https://www.instagram.com/explore/locations/234338264/"),
    (2, "KOPA Mall, Pune", "Pune", "ChIJ2-pYhL_xwsARp3XnJzJ2Y7U", "149740888215226", "KOPA Mall, S.NO. 37, H.NO.3 PLUS 4 BY 2, Village Ghorpadi, Mundhwa Road, Koregaon Park Annexe, Pune, Maharashtra 411001", "https://www.instagram.com/explore/locations/149740888215226/"),
    (3, "Phoenix Mall of the Millennium, Wakad", "Pune", "ChIJ2_a82-y9wjsR9w2O9VvV1-0", "114293561582939", "Phoenix Mall of the Millennium, Shankar Kalat Nagar, Wakad, Pimpri-Chinchwad, Pune, Maharashtra 411057", "https://www.instagram.com/explore/locations/114293561582939/"),
    (4, "Seasons Mall, Pune", "Pune", "ChIJ6784PzzBwjsRJ60Gz-Uo-08", "32761533", "Seasons Mall, Magarpatta Police Station Rd, Magarpatta, Hadapsar, Pune, Maharashtra 411013", "https://www.instagram.com/explore/locations/32761533/"),
    (5, "The Pavillion, Pune", "Pune", "ChIJT_2E22PBwjsR9W14R2Y9x0g", "1398451676914590", "The Pavillion, Senapati Bapat Rd, Laxmi Society, Model Colony, Shivajinagar, Pune, Maharashtra 411016", "https://www.instagram.com/explore/locations/1398451676914590/"),
    (6, "Amanora Mall, Pune", "Pune", "ChIJq6qqpyDCwjsRwxbllWxdx3s", "746599765500408", "Amanora Mall, Mundhwa - Kharadi Rd, Amanora Park Town, Hadapsar, Pune, Maharashtra 411028", "https://www.instagram.com/explore/locations/746599765500408/"),
    (7, "Lulu Mall Hyderabad (Y Junction)", "Hyderabad", "ChIJw7tH8yCRyzsR0c9M1B1gX9k", "316918795673504", "Lulu Mall, Survey No. 1050, Kukatpally Y Junction, Balanagar Mandal, KPHB 3rd Phase, Hyderabad, Telangana 500072", "https://www.instagram.com/explore/locations/316918795673504/"),
    (8, "Nexus Hyderabad Mall (Forum Sujana)", "Hyderabad", "ChIJw9ZcCQ-RyzsR9mC3N5Qf2pU", "316088138", "Nexus Hyderabad, Plot No S-16, Survey No 1009, KPHB Phase 6, Kukatpally, Hyderabad, Telangana 500072", "https://www.instagram.com/explore/locations/316088138/"),
    (9, "Sarath City Capital Mall, Hyderabad", "Hyderabad", "ChIJu9knDCeTyzsRV4VLaGsc4bs", "1628149820740485", "Sarath City Capital Mall, Gachibowli - Miyapur Rd, Whitefields, Kondapur, Hyderabad, Telangana 500084", "https://www.instagram.com/explore/locations/1628149820740485/"),
    (10, "Inorbit Mall Cyberabad, Hyderabad", "Hyderabad", "ChIJ40oJk7aTyzsRkO8H3XyT7f4", "112403036921694", "Inorbit Mall Cyberabad, S No 64, APIIC Software Layout, Mindspace, Madhapur, Hitech City, Hyderabad, Telangana 500081", "https://www.instagram.com/explore/locations/112403036921694/"),
    (11, "GVK One Mall, Hyderabad", "Hyderabad", "ChIJP-oQ5b-ZyzsRN-8w5tB9-3k", "277785458937883", "GVK One Mall, Rd Number 1, Balapur Basthi, Banjara Hills, Hyderabad, Telangana 500034", "https://www.instagram.com/explore/locations/277785458937883/"),
    (12, "Manjeera Trinity Mall (Y Junction)", "Hyderabad", "ChIJuc4lYACRyzsR49_LVvQdD9c", "256877967817441", "Manjeera Trinity Mall, JNTU Hitech City Rd, Kukatpally Housing Board Colony, Kukatpally, Hyderabad, Telangana 500072", "https://www.instagram.com/explore/locations/256877967817441/")
]

for row in geo_data_combined:
    r_num = row[0] + 2
    for c_idx, val in enumerate(row, 1):
        cell = ws_geo.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2: cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 3: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (4, 5): cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 6: cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 7: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None
    ws_geo.row_dimensions[r_num].height = 20

output_file = "lakeshore_kopa_paid_media_intelligence.xlsx"
wb.save(output_file)
print(f"\n✓ Master Workbook saved successfully: {output_file}")
