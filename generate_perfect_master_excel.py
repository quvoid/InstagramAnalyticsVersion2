"""
Generate Master Excel (jewellery_brands_master_analysis.xlsx) with:
1. Executive Summary tab breaking down unique creators and posts per tier across all 36 brands.
2. Master Hierarchy tab with all 1,179 posts grouped by Tier 1 -> Tier 2 -> Tier 3 -> Tier 4.
3. 36 Individual Brand Tabs, each internally grouped by Tier 1 -> Tier 2 -> Tier 3 -> Tier 4.
"""

import sys, json, os, csv, openpyxl
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

with open("unified_master_dataset.json", encoding="utf-8") as f:
    master_records = json.load(f)

# Sort master records strictly by Tier ascending, then Views descending
master_records.sort(key=lambda x: (x["tier"], -x["views"]))

# Group by brand
brand_records = defaultdict(list)
for r in master_records:
    brand_records[r["brand"]].append(r)

# ─────────────────────────────────────────────────────────────
# OpenPyXL Workbook Setup
# ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active) # remove default sheet

# Shared Styling Definitions
font_title = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
font_sub = Font(name="Calibri", size=10, italic=True, color="E0E0E0")
font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_norm = Font(name="Calibri", size=10, bold=False, color="000000")
font_mute = Font(name="Calibri", size=9, bold=False, color="5D6D7E")
font_link = Font(name="Calibri", size=10, bold=False, color="0563C1", underline="single")

# Tier Styles
# Tier 1: Bold Mint Green
fill_t1_banner = PatternFill("solid", fgColor="145A32") # Dark Emerald
fill_t1_row = PatternFill("solid", fgColor="D4EFDF") # Mint Green
font_t1_bold = Font(name="Calibri", size=10, bold=True, color="0E6251")
font_t1_link = Font(name="Calibri", size=10, bold=True, color="0B5345", underline="single")

# Tier 2: Sage / Seafoam Green
fill_t2_banner = PatternFill("solid", fgColor="1E8449") # Forest Green
fill_t2_row = PatternFill("solid", fgColor="EAFAF1") # Soft Green
font_t2_bold = Font(name="Calibri", size=10, bold=True, color="196F3D")
font_t2_link = Font(name="Calibri", size=10, bold=False, color="145A32", underline="single")

# Tier 3: Soft Amber / Gold / Peach
fill_t3_banner = PatternFill("solid", fgColor="B7950B") # Dark Gold
fill_t3_row = PatternFill("solid", fgColor="FEF9E7") # Light Warm Gold
font_t3_bold = Font(name="Calibri", size=10, bold=True, color="7D6608")
font_t3_link = Font(name="Calibri", size=10, bold=False, color="9A7D0A", underline="single")

# Tier 4: Light Gray / White (Noise)
fill_t4_banner = PatternFill("solid", fgColor="566573") # Slate Gray
fill_t4_row = PatternFill("solid", fgColor="FFFFFF")
font_t4_norm = Font(name="Calibri", size=10, bold=False, color="2C3E50")
font_t4_link = Font(name="Calibri", size=10, bold=False, color="2980B9", underline="single")

# Borders
thin_line = Side(style="thin", color="D5D8DC")
border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
thick_bottom = Border(bottom=Side(style="medium", color="2C3E50"))

# ─────────────────────────────────────────────────────────────
# 1. TAB 1: EXECUTIVE SUMMARY
# ─────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet("Executive Summary")
ws_sum.sheet_view.showGridLines = True

# Title Banner
ws_sum.merge_cells("A1:N1")
ws_sum["A1"] = "Executive Summary — 4-Tier Paid Creator Hierarchy across 36 Indian Jewellery Brands"
ws_sum["A1"].font = font_title
ws_sum["A1"].fill = PatternFill("solid", fgColor="0B2240") # Navy Blue
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 32

sum_headers = [
    ("#", 5),
    ("Brand Name", 28),
    ("Total Collab Posts", 16),
    ("Total Unique Creators", 18),
    ("🟢 Tier 1: Toggle ON + Boosted\n(Posts / Creators)", 24),
    ("🟢 Tier 2: Toggle ON + Organic\n(Posts / Creators)", 24),
    ("🚀 Tier 3: Toggle OFF + Boosted\n(Posts / Creators)", 24),
    ("⚪ Tier 4: Toggle OFF + Organic (Noise)\n(Posts / Creators)", 28),
    ("💎 Total High-Intent Paid\n(Tiers 1+2+3 Posts)", 22),
    ("High-Intent Paid %\n(Tiers 1+2+3)", 18),
    ("Avg Views / Post", 16),
    ("Avg Creator Followers", 20),
    ("Avg Creator ER%", 15),
    ("Top Creator / Ambassador Samples", 45)
]

for col_idx, (h_text, w) in enumerate(sum_headers, 1):
    c = ws_sum.cell(row=2, column=col_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="1B2631")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_cell
    ws_sum.column_dimensions[get_column_letter(col_idx)].width = w
ws_sum.row_dimensions[2].height = 36
ws_sum.freeze_panes = "A3"

# Calculate summary stats per brand
brand_summary_data = []
for b_name, posts in brand_records.items():
    tot_p = len(posts)
    tot_c = len(set(p["handle"].lower() for p in posts))
    
    t1_posts = [p for p in posts if p["tier"] == 1]
    t1_creators = len(set(p["handle"].lower() for p in t1_posts))
    
    t2_posts = [p for p in posts if p["tier"] == 2]
    t2_creators = len(set(p["handle"].lower() for p in t2_posts))
    
    t3_posts = [p for p in posts if p["tier"] == 3]
    t3_creators = len(set(p["handle"].lower() for p in t3_posts))
    
    t4_posts = [p for p in posts if p["tier"] == 4]
    t4_creators = len(set(p["handle"].lower() for p in t4_posts))
    
    high_intent_posts = len(t1_posts) + len(t2_posts) + len(t3_posts)
    high_intent_pct = high_intent_posts / tot_p if tot_p > 0 else 0.0
    
    views_list = [p["views"] for p in posts if p["views"] > 0]
    avg_v = int(sum(views_list) / len(views_list)) if views_list else 0
    
    fols_list = [p["followers"] for p in posts if p["followers"] > 0]
    avg_f = int(sum(fols_list) / len(fols_list)) if fols_list else 0
    
    ers_list = [p["er_pct"] for p in posts if p["er_pct"] > 0]
    avg_e = round(sum(ers_list) / len(ers_list), 2) if ers_list else 0.0
    
    # Top creators
    top_c = list(dict.fromkeys([p["handle"] for p in posts if p["tier"] in (1, 2, 3)]))[:4]
    if not top_c:
        top_c = list(dict.fromkeys([p["handle"] for p in posts]))[:4]
        
    brand_summary_data.append({
        "brand": b_name,
        "posts": tot_p,
        "creators": tot_c,
        "t1_p": len(t1_posts), "t1_c": t1_creators,
        "t2_p": len(t2_posts), "t2_c": t2_creators,
        "t3_p": len(t3_posts), "t3_c": t3_creators,
        "t4_p": len(t4_posts), "t4_c": t4_creators,
        "high_intent_p": high_intent_posts,
        "high_intent_pct": high_intent_pct,
        "avg_views": avg_v,
        "avg_followers": avg_f,
        "avg_er": avg_e,
        "top_creators": ", ".join(top_c)
    })

# Sort summary by High-Intent Paid posts descending, then total posts
brand_summary_data.sort(key=lambda x: (x["high_intent_p"], x["posts"]), reverse=True)

for idx, b in enumerate(brand_summary_data, 1):
    row_num = idx + 2
    row_vals = [
        idx,
        b["brand"],
        b["posts"],
        b["creators"],
        f"{b['t1_p']} posts ({b['t1_c']} creators)" if b["t1_p"] > 0 else "—",
        f"{b['t2_p']} posts ({b['t2_c']} creators)" if b["t2_p"] > 0 else "—",
        f"{b['t3_p']} posts ({b['t3_c']} creators)" if b["t3_p"] > 0 else "—",
        f"{b['t4_p']} posts ({b['t4_c']} creators)" if b["t4_p"] > 0 else "—",
        b["high_intent_p"],
        b["high_intent_pct"],
        b["avg_views"],
        b["avg_followers"],
        b["avg_er"] / 100 if b["avg_er"] else 0.0,
        b["top_creators"]
    ]
    
    for col_idx, val in enumerate(row_vals, 1):
        cell = ws_sum.cell(row=row_num, column=col_idx, value=val)
        cell.border = border_cell
        
        # Alignments and Formats
        if col_idx == 1:
            cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 2:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif col_idx in (3, 4):
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"
            cell.fill = PatternFill("solid", fgColor="EBF5FB")
        elif col_idx == 5: # Tier 1
            cell.font = font_t1_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = fill_t1_row
        elif col_idx == 6: # Tier 2
            cell.font = font_t2_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = fill_t2_row
        elif col_idx == 7: # Tier 3
            cell.font = font_t3_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = fill_t3_row
        elif col_idx == 8: # Tier 4
            cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = PatternFill("solid", fgColor="F8F9F9")
        elif col_idx == 9: # Total High Intent
            cell.font = Font(name="Calibri", size=10, bold=True, color="1B4F72"); cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"
            if val > 0: cell.fill = PatternFill("solid", fgColor="D6EAF8")
        elif col_idx == 10: # High Intent %
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.0%"
        elif col_idx in (11, 12):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif col_idx == 13:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
        elif col_idx == 14:
            cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
            
    ws_sum.row_dimensions[row_num].height = 22


# ─────────────────────────────────────────────────────────────
# 2. HELPER FUNCTION TO RENDER TIERED ROWS
# ─────────────────────────────────────────────────────────────
master_cols = [
    ("#", 5),
    ("Hierarchy Tier", 30),
    ("Brand Name", 26),
    ("Creator Handle", 24),
    ("Followers", 14),
    ("Views / Plays", 16),
    ("Likes", 14),
    ("Comments", 12),
    ("Like-to-View %", 15),
    ("Creator ER%", 13),
    ("Post Date", 13),
    ("Direct Instagram URL", 48),
    ("Boost Classification & Reason", 35),
    ("Caption Preview", 65)
]

def render_tiered_sheet(ws, records, sheet_title):
    ws.sheet_view.showGridLines = True
    
    # Title Banner
    ws.merge_cells("A1:N1")
    ws["A1"] = sheet_title
    ws["A1"].font = font_title
    ws["A1"].fill = PatternFill("solid", fgColor="0B2240")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    
    # Header Row
    for col_idx, (h_text, w) in enumerate(master_cols, 1):
        c = ws.cell(row=2, column=col_idx, value=h_text)
        c.font = font_hdr
        c.fill = PatternFill("solid", fgColor="1F2D3D")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_cell
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 25
    ws.freeze_panes = "A3"
    
    # Group by tier
    records_by_tier = {1: [], 2: [], 3: [], 4: []}
    for r in records:
        records_by_tier[r["tier"]].append(r)
        
    current_row = 3
    global_index = 1
    
    tier_meta = [
        (1, "🟢 TIER 1: TOGGLE ON + 🚀 BOOSTED (Formal Paid Partnership Label + Paid Ad Spend)", fill_t1_banner, fill_t1_row, font_t1_bold, font_t1_link),
        (2, "🟢 TIER 2: TOGGLE ON + ⚪ ORGANIC (Formal Paid Partnership Label + Organic Reach Only)", fill_t2_banner, fill_t2_row, font_t2_bold, font_t2_link),
        (3, "🚀 TIER 3: TOGGLE OFF + 🚀 BOOSTED (Co-Author Collab + Heavy Paid Ad Spend Detected)", fill_t3_banner, fill_t3_row, font_t3_bold, font_t3_link),
        (4, "⚪ TIER 4: TOGGLE OFF + ⚪ ORGANIC (Standard Collab / Low Organic Reach / Noise)", fill_t4_banner, fill_t4_row, font_t4_norm, font_t4_link)
    ]
    
    for t_id, banner_text, fill_banner, fill_row, font_b, font_l in tier_meta:
        t_records = records_by_tier[t_id]
        if not t_records:
            continue
            
        # Insert Section Banner
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        banner_cell = ws.cell(row=current_row, column=1, value=f"{banner_text} — {len(t_records)} Posts ({len(set(p['handle'].lower() for p in t_records))} Unique Creators)")
        banner_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        banner_cell.fill = fill_banner
        banner_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[current_row].height = 24
        current_row += 1
        
        # Insert records for this tier
        for p in t_records:
            vals = [
                global_index,
                p["tier_name"],
                p["brand"],
                p["handle"],
                p["followers"],
                p["views"],
                p["likes"],
                p["comments"],
                p["like_rate_pct"] / 100 if p["like_rate_pct"] else 0.0,
                p["er_pct"] / 100 if p["er_pct"] else 0.0,
                p["post_date"],
                p["url"],
                f"{p['boost_status']}: {p['reason']}" if p['reason'] else p['boost_status'],
                p["caption"]
            ]
            
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=current_row, column=c_idx, value=val)
                cell.border = border_cell
                cell.fill = fill_row
                
                if c_idx == 1:
                    cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 2:
                    cell.font = font_b; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx in (3, 4):
                    cell.font = font_b; cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx in (5, 6, 7, 8):
                    cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
                elif c_idx in (9, 10):
                    cell.font = font_b; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
                elif c_idx == 11:
                    cell.font = font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 12:
                    cell.font = font_l; cell.alignment = Alignment(horizontal="left", vertical="center")
                    if val: cell.hyperlink = val
                elif c_idx == 13:
                    cell.font = font_b; cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx == 14:
                    cell.font = font_mute; cell.alignment = Alignment(horizontal="left", vertical="center")
                    
            ws.row_dimensions[current_row].height = 21
            current_row += 1
            global_index += 1


# ─────────────────────────────────────────────────────────────
# 3. TAB 2: ALL BRANDS - MASTER HIERARCHY
# ─────────────────────────────────────────────────────────────
ws_all = wb.create_sheet("All Brands - Master Hierarchy")
render_tiered_sheet(
    ws_all,
    master_records,
    f"Master Consolidated Paid Creator Database across 36 Brands — 1,179 Total Posts (Tiers 1–4)"
)

# ─────────────────────────────────────────────────────────────
# 4. TABS 3 TO 38: 36 INDIVIDUAL BRAND SHEETS
# ─────────────────────────────────────────────────────────────
for b_name in sorted(brand_records.keys()):
    b_posts = brand_records[b_name]
    # Sanitize sheet name for excel (max 31 chars, no special characters)
    safe_name = b_name.replace("/", "-").replace(":", " ")[:31]
    ws_brand = wb.create_sheet(safe_name)
    render_tiered_sheet(
        ws_brand,
        b_posts,
        f"{b_name} — Paid Creator Partnerships & Boost Hierarchy ({len(b_posts)} Total Posts)"
    )

wb.save("jewellery_brands_master_analysis.xlsx")
print(f"\n✅ Master Excel Generated Successfully!")
print(f"   • File: jewellery_brands_master_analysis.xlsx")
print(f"   • Total Sheets: {len(wb.sheetnames)}")
print(f"   • Sheet 1: Executive Summary")
print(f"   • Sheet 2: All Brands - Master Hierarchy (1,179 posts)")
print(f"   • Sheets 3–38: 36 Individual Brand Tabs")
