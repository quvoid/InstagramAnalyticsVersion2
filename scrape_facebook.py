"""
Facebook Brand & Creator Partnership Posts Scraper
===================================================
Scrapes Facebook posts from brand pages, detects paid partnership / collaboration posts,
extracts engagement metrics (Reactions, Comments, Shares, Date, URLs, Captions),
and generates a clean Excel report.
"""

import sys, io, re, html, time, random, json, os
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from curl_cffi import requests

sys.stdout.reconfigure(encoding="utf-8")

# ════════════════════════════════════════════════════════════════
#  CONFIG
# ════════════════════════════════════════════════════════════════

PAGE_USERNAME = "grtjewellers"
OUTPUT_FILENAME = "grt_facebook_partnership_posts.xlsx"

# Facebook browser headers
HEADERS = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    "accept-language": "en-US,en;q=0.9",
    "sec-ch-ua": '"Chromium";v="122", "Not(A:Brand";v="24", "Google Chrome";v="122"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "document",
    "sec-fetch-mode": "navigate",
    "sec-fetch-site": "none",
    "sec-fetch-user": "?1",
    "upgrade-insecure-requests": "1",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
}


# ── Scraping Core ──────────────────────────────────────────────

def fetch_facebook_page_data(page_name: str):
    session = requests.Session(impersonate="chrome120")
    
    print(f"\n{'═'*65}")
    print(f"  Facebook Scraper — Target: https://www.facebook.com/{page_name}/")
    print(f"{'═'*65}\n")
    
    url = f"https://www.facebook.com/{page_name}/"
    r = session.get(url, headers=HEADERS, timeout=20)
    
    if r.status_code != 200:
        print(f"❌ Failed to load page: HTTP {r.status_code}")
        return [], {}
    
    html_text = r.text
    print(f"✓ Successfully fetched Facebook page ({len(html_text):,} bytes)")
    
    # 1. Extract Page Info
    page_id_m = re.search(r'"pageID":"(\d+)"', html_text) or re.search(r'"page_id":"(\d+)"', html_text) or re.search(r'"delegate_page":\{"id":"(\d+)"', html_text)
    page_id = page_id_m.group(1) if page_id_m else "N/A"
    
    name_m = re.search(r'<meta property="og:title" content="([^"]+)"', html_text)
    page_title = name_m.group(1) if name_m else page_name
    
    desc_m = re.search(r'<meta name="description" content="([^"]+)"', html_text)
    page_desc = desc_m.group(1) if desc_m else ""
    
    page_info = {
        "page_name": page_title,
        "page_id": page_id,
        "username": page_name,
        "url": url,
        "description": page_desc
    }
    print(f"  Page Name: {page_title} | ID: {page_id}")
    
    # 2. Extract Embedded SSR Script JSONs
    scripts = re.findall(r'<script type="application/json"[^>]*>(.*?)</script>', html_text, re.DOTALL)
    print(f"  Analyzing {len(scripts)} JSON data nodes...")
    
    posts = []
    
    def extract_nodes(obj):
        if isinstance(obj, dict):
            # Check for comet story structure
            if "node" in obj and isinstance(obj["node"], dict) and ("comet_sections" in obj["node"] or "creation_time" in obj["node"]):
                parse_story_node(obj["node"])
            elif "story" in obj and isinstance(obj["story"], dict):
                parse_story_node(obj["story"])
            for k, v in obj.items():
                extract_nodes(v)
        elif isinstance(obj, list):
            for item in obj:
                extract_nodes(item)

    def parse_story_node(node):
        post_id = node.get("post_id") or node.get("id") or (node.get("feedback") or {}).get("id")
        
        # Creation timestamp
        ts = node.get("creation_time") or 0
        if not ts and "comet_sections" in node:
            try:
                ts = node["comet_sections"]["context_layout"]["story"]["comet_sections"]["metadata"][0]["story"]["creation_time"]
            except Exception:
                ts = 0
                
        # Message text
        msg = ""
        msg_obj = node.get("message") or (node.get("comet_sections") or {}).get("content", {}).get("story", {}).get("message")
        if isinstance(msg_obj, dict):
            msg = msg_obj.get("text", "")
        elif isinstance(msg_obj, str):
            msg = msg_obj
            
        # Engagement Metrics
        feedback = node.get("feedback") or {}
        reacts = (feedback.get("reaction_count") or {}).get("count", 0) if isinstance(feedback.get("reaction_count"), dict) else feedback.get("reaction_count", 0)
        comments = (feedback.get("comments_count") or {}).get("total_count", 0) if isinstance(feedback.get("comments_count"), dict) else feedback.get("comment_count", 0)
        shares = (feedback.get("share_count") or {}).get("count", 0) if isinstance(feedback.get("share_count"), dict) else feedback.get("share_count", 0)
        
        # Actors / Co-Authors / Sponsors
        actors = node.get("actors", []) or (node.get("comet_sections") or {}).get("header", {}).get("story", {}).get("actors", [])
        actor_names = [a.get("name") for a in actors if isinstance(a, dict) and a.get("name")]
        
        sponsors = node.get("sponsor_tags") or node.get("branded_content_sponsor_relationship") or []
        sponsor_names = [s.get("sponsor", {}).get("name") for s in sponsors if isinstance(s, dict)]
        
        # Check partnership keywords
        msg_lower = msg.lower()
        has_partner_keyword = any(k in msg_lower for k in [
            "with", "paid partnership", "collab", "#ad", "ambassador", "celebrating", "featuring",
            "ashwin", "prithi", "athulya", "siri", "roshni", "megha", "shankar", "niharika"
        ])
        
        is_partnership = bool(sponsors) or len(actor_names) > 1 or has_partner_keyword
        
        if msg or reacts or ts:
            date_str = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d") if ts else "N/A"
            post_url = node.get("url") or f"https://www.facebook.com/{page_name}/posts/{post_id}" if post_id else f"https://www.facebook.com/{page_name}"
            
            partner_label = ", ".join(actor_names[1:]) if len(actor_names) > 1 else (", ".join(sponsor_names) if sponsor_names else ("Creator Mention in Caption" if has_partner_keyword else "None"))
            
            posts.append({
                "post_id": str(post_id),
                "date": date_str,
                "post_url": post_url,
                "caption": msg,
                "partner": partner_label,
                "is_partnership": is_partnership,
                "reactions": reacts or 0,
                "comments": comments or 0,
                "shares": shares or 0,
            })

    for s in scripts:
        try:
            d = json.loads(s)
            extract_nodes(d)
        except Exception:
            pass

    # Deduplicate by post_id
    dedup = {}
    for p in posts:
        pid = p["post_id"] or p["post_url"]
        if pid not in dedup:
            dedup[pid] = p

    return list(dedup.values()), page_info


# ── Excel Builder ─────────────────────────────────────────────

def build_excel_report(posts: list, page_info: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facebook Partnerships"
    ws.sheet_view.showGridLines = True
    
    thin = Side(style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1877F2") # Facebook Blue
    hdr_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    bf = Font(name="Calibri", bold=True, size=10, color="000000")
    nf = Font(name="Calibri", bold=False, size=10, color="000000")
    lnk = Font(name="Calibri", bold=False, size=10, color="0563C1", underline="single")
    
    # Title Banner
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Facebook Partnership & Collab Posts — {page_info.get('page_name', 'Brand Page')} ({len(posts)} Posts)"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0E3E85")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    
    headers = [
        ("#", 5),
        ("Date", 13),
        ("Partner / Creator", 25),
        ("Partnership Type", 22),
        ("Post URL", 45),
        ("Reactions", 12),
        ("Comments", 12),
        ("Shares", 12),
        ("Caption / Post Copy", 65),
    ]
    
    for col_idx, (h_text, width) in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx, value=h_text)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A3"
    
    for idx, p in enumerate(posts, 1):
        row = idx + 2
        p_type = "Paid / Boosted Collab" if p["is_partnership"] else "Organic Brand Post"
        
        vals = [
            idx,
            p["date"],
            p["partner"],
            p_type,
            p["post_url"],
            p["reactions"],
            p["comments"],
            p["shares"],
            p["caption"],
        ]
        
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.border = bdr
            if col_idx == 1:
                c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 3:
                c.font = bf; c.alignment = Alignment(horizontal="left", vertical="center")
                if p["is_partnership"]:
                    c.fill = PatternFill("solid", fgColor="E7F3FF")
            elif col_idx == 4:
                c.font = bf; c.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 5:
                c.font = lnk; c.alignment = Alignment(horizontal="left", vertical="center")
                if val: c.hyperlink = val
            elif col_idx in (6, 7, 8):
                c.font = nf; c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = "#,##0"
            elif col_idx == 9:
                c.font = nf; c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Main ──────────────────────────────────────────────────────

def main():
    posts, page_info = fetch_facebook_page_data(PAGE_USERNAME)
    
    if not posts:
        print("⚠ No posts extracted from public Facebook feed.")
        return
    
    print(f"\n✓ Extracted {len(posts)} total Facebook posts.")
    partner_posts = [p for p in posts if p["is_partnership"]]
    print(f"  • Partnership / Collab Posts: {len(partner_posts)}")
    print(f"  • Organic Brand Posts: {len(posts) - len(partner_posts)}")
    
    print("\n📊 Building Excel report...")
    excel_bytes = build_excel_report(posts, page_info)
    
    with open(OUTPUT_FILENAME, "wb") as f:
        f.write(excel_bytes)
        
    print(f"\n{'═'*65}")
    print(f"  ✅ Saved: {OUTPUT_FILENAME}")
    print(f"  Total Facebook posts analysed: {len(posts)}")
    print(f"{'═'*65}\n")


if __name__ == "__main__":
    main()
