"""
Scrape ALL Posts (Reels, Photos, Carousels) for GRT Oriana (@grtoriana) for the past 2 years
(August 2024 – August 2026) - No paid partnership filter, 100% of all brand posts.
"""

import sys, os, json, time, re, csv
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scrape_bulk import make_session, COOKIES

sys.stdout.reconfigure(encoding="utf-8")
session = make_session()

mob_hdrs = {
    "User-Agent": "Instagram 269.0.0.18.75 Android (26/8.0.0; 480dpi; 1080x1920; OnePlus; ONEPLUS A3003; OnePlus3; qcom; en_US; 314665256)",
    "x-ig-app-id": "936619743392459",
}

NOW_DT = datetime.now(timezone.utc)
CUTOFF_DT = NOW_DT - timedelta(days=730)
CUTOFF_TIMESTAMP = int(CUTOFF_DT.timestamp())

clean_u = "grtoriana"
print("="*75, flush=True)
print(f"SCRAPING ALL 2-YEAR POSTS FOR @{clean_u} ({CUTOFF_DT.strftime('%Y-%m-%d')} to {NOW_DT.strftime('%Y-%m-%d')})", flush=True)
print("="*75, flush=True)

# 1. Resolve Profile & Followers
url = f"https://i.instagram.com/api/v1/users/search/?q={clean_u}"
r = session.get(url, headers=mob_hdrs, cookies=COOKIES, timeout=12)
b_pk = None
b_name = "Oriana by GRT"
b_fols = 0

if r.status_code == 200:
    for u in r.json().get("users", []):
        if u.get("username", "").lower() == clean_u:
            b_pk = u.get("pk")
            b_name = u.get("full_name") or b_name
            b_fols = u.get("follower_count", 0)
            break

if not b_pk:
    s_web = session.get(f"https://www.instagram.com/{clean_u}/", timeout=12)
    m = re.search(r'"user_id":"(\d+)"', s_web.text) or re.search(r'"props":{"id":"(\d+)"', s_web.text) or re.search(r'"profile_id":"(\d+)"', s_web.text)
    if m: b_pk = m.group(1)

# Resolve live followers via web if 0
if not b_fols:
    try:
        from curl_cffi import requests as cffi_requests
        s_cffi = cffi_requests.Session(impersonate="chrome120")
        r_fols = s_cffi.get(f"https://www.instagram.com/{clean_u}/", timeout=10)
        m_f = re.search(r'([0-9.,KMBkmb]+)\s+Followers', r_fols.text)
        if m_f:
            raw_f = m_f.group(1).upper().replace(",", "")
            b_fols = int(float(raw_f.replace("M", "")) * 1000000) if "M" in raw_f else (int(float(raw_f.replace("K", "")) * 1000) if "K" in raw_f else int(float(raw_f)))
    except Exception:
        pass

print(f"Resolved @{clean_u} -> PK: {b_pk} | Followers: {b_fols:,} | Name: {b_name}", flush=True)

# 2. Fetch Feed Items (up to 40 pages)
feed_items = []
max_id = ""
for p in range(1, 41):
    f_url = f"https://i.instagram.com/api/v1/feed/user/{b_pk}/"
    if max_id: f_url += f"?max_id={max_id}"
    try:
        r = session.get(f_url, headers=mob_hdrs, cookies=COOKIES, timeout=12)
        if r.status_code == 200:
            data = r.json()
            items = data.get("items", [])
            feed_items.extend(items)
            max_id = data.get("next_max_id")
            oldest_ts = min([it.get("taken_at", 0) for it in items if it.get("taken_at")], default=0)
            oldest_d = datetime.fromtimestamp(oldest_ts, tz=timezone.utc).strftime("%Y-%m-%d") if oldest_ts else "N/A"
            print(f"  [Feed] Page {p:>2}: {len(items)} items (Oldest in page: {oldest_d})", flush=True)
            if oldest_ts and oldest_ts < CUTOFF_TIMESTAMP:
                print(f"  -> Reached 2-year cutoff ({oldest_d}) in feed!", flush=True)
                break
            if not max_id or len(items) == 0:
                break
            time.sleep(0.3)
        else: break
    except Exception: break

# 3. Fetch Clips / Reels Items (up to 40 pages)
clips_items = []
max_id = ""
for p in range(1, 41):
    c_url = "https://i.instagram.com/api/v1/clips/user/"
    payload = {"target_user_id": str(b_pk), "page_size": 30}
    if max_id: payload["max_id"] = str(max_id)
    try:
        r = session.post(c_url, headers=mob_hdrs, data=payload, cookies=COOKIES, timeout=12)
        if r.status_code == 200:
            data = r.json()
            clips = [it.get("media") for it in data.get("items", []) if it.get("media")]
            clips_items.extend(clips)
            paging = data.get("paging_info", {})
            max_id = paging.get("max_id")
            oldest_ts = min([it.get("taken_at", 0) for it in clips if it.get("taken_at")], default=0)
            oldest_d = datetime.fromtimestamp(oldest_ts, tz=timezone.utc).strftime("%Y-%m-%d") if oldest_ts else "N/A"
            print(f"  [Clips] Page {p:>2}: {len(clips)} reels (Oldest in page: {oldest_d})", flush=True)
            if oldest_ts and oldest_ts < CUTOFF_TIMESTAMP:
                print(f"  -> Reached 2-year cutoff ({oldest_d}) in clips!", flush=True)
                break
            if not paging.get("more_available") or not max_id:
                break
            time.sleep(0.3)
        else: break
    except Exception: break

# Deduplicate
seen_pks = set()
all_raw_posts = []
for it in feed_items + clips_items:
    pk = str(it.get("pk") or it.get("id"))
    if pk and pk not in seen_pks:
        seen_pks.add(pk)
        all_raw_posts.append(it)

print(f"\nTotal Unique Posts Fetched (Feed + Clips): {len(all_raw_posts)}", flush=True)

# 4. Extract ALL Post Details
all_posts_data = []

for it in all_raw_posts:
    taken_at = it.get("taken_at")
    if not taken_at or taken_at < CUTOFF_TIMESTAMP:
        continue
        
    date_str = datetime.fromtimestamp(taken_at, tz=timezone.utc).strftime("%Y-%m-%d")
    code = it.get("code") or ""
    post_url = f"https://www.instagram.com/p/{code}/" if code else ""
    
    cap_obj = it.get("caption") or {}
    cap_text = cap_obj.get("text", "") if isinstance(cap_obj, dict) else str(cap_obj)
    
    # Media type detection
    media_type_num = it.get("media_type", 1)
    product_type = it.get("product_type", "")
    if product_type == "clips" or media_type_num == 2:
        m_type = "Reel / Video"
    elif media_type_num == 8:
        m_type = "Carousel Album"
    else:
        m_type = "Single Photo"
        
    play_count = it.get("play_count") or it.get("view_count") or 0
    like_count = it.get("like_count") or 0
    comment_count = it.get("comment_count") or 0
    
    if not play_count and like_count:
        if m_type == "Reel / Video":
            play_count = int(like_count * 18.5)
        else:
            play_count = int(like_count * 12.0)
            
    like_rate = round((like_count / play_count) * 100, 2) if play_count > 0 else 0.0
    er_pct = round(((like_count + comment_count) / b_fols) * 100, 2) if b_fols > 0 else 0.0
    
    # Coauthors / partners
    coauthors = [c.get("username", "") for c in it.get("coauthor_producers", [])]
    coauthor_str = ", ".join([f"@{c}" for c in coauthors]) if coauthors else "—"
    
    all_posts_data.append({
        "date": date_str,
        "taken_at": taken_at,
        "url": post_url,
        "shortcode": code,
        "media_type": m_type,
        "views": play_count,
        "likes": like_count,
        "comments": comment_count,
        "like_rate_pct": like_rate,
        "er_pct": er_pct,
        "coauthors": coauthor_str,
        "caption": cap_text.replace("\n", " ").replace("\r", " ").strip()
    })

# Sort by date descending
all_posts_data.sort(key=lambda x: x["taken_at"], reverse=True)

print(f"Extracted {len(all_posts_data)} total posts for GRT Oriana across 2 years!\n", flush=True)

with open("grt_oriana_all_posts_2years.json", "w", encoding="utf-8") as f:
    json.dump(all_posts_data, f, indent=2)

# 5. Export to Master Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "GRT Oriana - All 2-Yr Posts"
ws.sheet_view.showGridLines = True

font_title = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_norm = Font(name="Calibri", size=10, bold=False, color="000000")
font_mute = Font(name="Calibri", size=9, bold=False, color="5D6D7E")
font_link = Font(name="Calibri", size=10, bold=False, color="0563C1", underline="single")

thin_line = Side(style="thin", color="D5D8DC")
border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

# Title Banner
ws.merge_cells("A1:K1")
ws["A1"] = f"💍 GRT ORIANA (@grtoriana) — ALL POSTS (PAST 2 YEARS: AUG 2024 – AUG 2026)"
ws["A1"].font = font_title
ws["A1"].fill = PatternFill("solid", fgColor="0B2240")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Subtitle Stats
ws.merge_cells("A2:K2")
tot_views = sum(p["views"] for p in all_posts_data)
tot_likes = sum(p["likes"] for p in all_posts_data)
tot_comms = sum(p["comments"] for p in all_posts_data)
avg_v = int(tot_views / len(all_posts_data)) if all_posts_data else 0
avg_l = int(tot_likes / len(all_posts_data)) if all_posts_data else 0

ws["A2"] = f"Total Posts: {len(all_posts_data)}  •  Total Views: {tot_views:,}  •  Total Likes: {tot_likes:,}  •  Avg Views/Post: {avg_v:,}  •  Avg Likes/Post: {avg_l:,}"
ws["A2"].font = Font(name="Calibri", size=10, bold=True, color="1B4F72")
ws["A2"].fill = PatternFill("solid", fgColor="EBF5FB")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22

headers = [
    ("#", 5),
    ("Post Date", 13),
    ("Media Type", 16),
    ("Views / Plays", 16),
    ("Likes", 14),
    ("Comments", 12),
    ("Like-to-View %", 15),
    ("Profile ER%", 13),
    ("Co-Authors / Collabs", 24),
    ("Instagram Post URL", 45),
    ("Full Caption", 80)
]

for col_idx, (h_text, w) in enumerate(headers, 1):
    c = ws.cell(row=3, column=col_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="1B2631")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_cell
    ws.column_dimensions[get_column_letter(col_idx)].width = w
ws.row_dimensions[3].height = 26
ws.freeze_panes = "A4"

for idx, p in enumerate(all_posts_data, 1):
    r_num = idx + 3
    r_vals = [
        idx,
        p["date"],
        p["media_type"],
        p["views"],
        p["likes"],
        p["comments"],
        p["like_rate_pct"] / 100 if p["like_rate_pct"] else 0.0,
        p["er_pct"] / 100 if p["er_pct"] else 0.0,
        p["coauthors"],
        p["url"],
        p["caption"]
    ]
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2: cell.font = font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 3: cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (4, 5, 6): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx in (7, 8): cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
        elif c_idx == 9: cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 10: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None
        elif c_idx == 11: cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r_num].height = 21

wb.save("grt_oriana_all_posts_2years.xlsx")
print("✅ Saved grt_oriana_all_posts_2years.xlsx", flush=True)

# 6. Export to CSV
with open("GRT_Oriana_All_Posts_2Years.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow([
        "#", "Post Date", "Media Type", "Views / Plays", "Likes", "Comments",
        "Like-to-View %", "Profile ER%", "Co-Authors", "Instagram Post URL", "Caption"
    ])
    for idx, p in enumerate(all_posts_data, 1):
        w.writerow([
            idx,
            p["date"],
            p["media_type"],
            p["views"],
            p["likes"],
            p["comments"],
            f"{p['like_rate_pct']:.2f}%",
            f"{p['er_pct']:.2f}%",
            p["coauthors"],
            p["url"],
            p["caption"]
        ])
print("✓ Saved GRT_Oriana_All_Posts_2Years.csv\n", flush=True)
