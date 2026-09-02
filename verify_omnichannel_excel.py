"""
Verify lakeshore_omnichannel_competitor_master.xlsx
"""

import openpyxl

wb = openpyxl.load_workbook("lakeshore_omnichannel_competitor_master.xlsx", read_only=True)
print("="*70)
print("WORKBOOK VERIFICATION: lakeshore_omnichannel_competitor_master.xlsx")
print("="*70)
print(f"Total Sheets: {len(wb.sheetnames)}\n")

for name in wb.sheetnames:
    ws = wb[name]
    print(f"  • Tab: {name:<38} | Total Rows: {ws.max_row}")

print("="*70)
