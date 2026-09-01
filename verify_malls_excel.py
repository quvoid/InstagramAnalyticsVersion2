"""
Verify pune_hyderabad_malls_master_analysis.xlsx
"""

import openpyxl

wb = openpyxl.load_workbook("pune_hyderabad_malls_master_analysis.xlsx", read_only=True)
print("="*70)
print("WORKBOOK VERIFICATION: pune_hyderabad_malls_master_analysis.xlsx")
print("="*70)
print(f"Total Sheets: {len(wb.sheetnames)}\n")

for name in wb.sheetnames:
    ws = wb[name]
    print(f"  • Tab: {name:<35} | Total Rows: {ws.max_row}")

print("="*70)
