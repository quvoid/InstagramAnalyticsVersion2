"""
Format All Brands - Paid Collabs in Excel and Google Sheets:
- Mark the entire row in soft green for posts with Toggle ON
- Format all columns cleanly
"""

import sys, openpyxl, json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

WORKBOOK_PATH = "jewellery_brands_consolidated_analysis.xlsx"
wb = openpyxl.load_workbook(WORKBOOK_PATH)

ws = wb["All Brands - Paid Collabs"]

thin = Side(style="thin", color="CCCCCC")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

# Full row green fill for Toggle ON rows
row_green_fill = PatternFill("solid", fgColor="D4EFDF") # Mint Green
on_font_bold = Font(name="Calibri", bold=True, size=10, color="145A32")
on_font_reg = Font(name="Calibri", bold=False, size=10, color="145A32")
on_lnk_font = Font(name="Calibri", bold=True, size=10, color="0B5345", underline="single")

toggle_on_indices = []

for r in range(3, ws.max_row + 1):
    toggle_val = str(ws.cell(row=r, column=4).value or "")
    if "ON" in toggle_val:
        toggle_on_indices.append(r)
        # Highlight entire row green
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = row_green_fill
            cell.border = bdr
            if c in (2, 3, 4):
                cell.font = on_font_bold
            elif c == 10 and cell.value:
                cell.font = on_lnk_font
            else:
                cell.font = on_font_reg

wb.save(WORKBOOK_PATH)
print(f"✓ Highlighted {len(toggle_on_indices)} full rows in green in {WORKBOOK_PATH}")
for r in toggle_on_indices:
    brand = ws.cell(row=r, column=2).value
    handle = ws.cell(row=r, column=3).value
    url = ws.cell(row=r, column=10).value
    print(f"  • Row {r}: {brand} | {handle} | {url}")
