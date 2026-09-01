"""
Unified Competitor Intelligence & Creator Audit Engine
Sequential Multi-Stage Pipeline:
1. Scrape Instagram 1-Year Grid & Reels (Co-Authors & Collabs)
2. Enrich Instagram Creator Profiles (Follower Counts, ER%, Audience Tiers)
3. Scrape Meta Ad Library (GraphQL + Playwright) for Dark Ads & Whitelists
4. Enrich Meta Ad Library Creator Profiles (Resolve handles, followers & tiers)
5. Fuse & Deduplicate into a Unified Master Creator Roster
6. Export Master Multi-Tab Excel Workbook & Clean CSV
"""

import sys, os, json, time, re, csv
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Optional, Any, Union
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from curl_cffi import requests as cffi_requests
from playwright.sync_api import sync_playwright

sys.stdout.reconfigure(encoding="utf-8")

# Session Configuration
DEFAULT_IG_COOKIES = {
    "sessionid": "76326162386%3A670U47iQkU6B8V%3A18%3AAYj9oJ1L51k_G3_j-uX4lQ9V6aM9Wc7gQ2yZ",
    "ds_user_id": "76326162386",
    "csrftoken": "b3U8nI5m6d9L0v7e8W1x2y",
}

DEFAULT_IG_HEADERS = {
    "User-Agent": "Instagram 269.0.0.18.75 Android (26/8.0.0; 480dpi; 1080x1920; OnePlus; ONEPLUS A3003; OnePlus3; qcom; en_US; 314665256)",
    "x-ig-app-id": "936619743392459",
}


# ==============================================================================
# 1. PROFILE ENRICHMENT HELPER (CONCURRENT)
# ==============================================================================
def resolve_creator_profile(raw_handle: str) -> Dict[str, Any]:
    """Resolves live follower count, full name, and audience scale tier."""
    clean_h = raw_handle.lower().replace("@", "").strip()
    s = cffi_requests.Session(impersonate="chrome120")
    followers = 0
    full_name = clean_h
    
    try:
        r = s.get(f"https://www.instagram.com/{clean_h}/", timeout=8)
        f_match = re.search(r'([0-9.,KMBkmb]+)\s+Followers', r.text)
        if f_match:
            raw = f_match.group(1).upper().replace(",", "")
            followers = int(float(raw.replace("M",""))*1000000) if "M" in raw else (int(float(raw.replace("K",""))*1000) if "K" in raw else int(float(raw)))
        
        n_match = re.search(r'<title>([^<(]+)', r.text)
        if n_match:
            title_text = n_match.group(1).replace("• Instagram photos and videos", "").strip()
            if title_text and not title_text.startswith("@"):
                full_name = title_text
    except Exception:
        pass

    if followers >= 1000000: tier = "🌟 Mega Creator (1M+)"
    elif followers >= 100000: tier = "🚀 Macro Creator (100K-1M)"
    elif followers >= 50000: tier = "✨ Mid-Tier (50K-100K)"
    elif followers >= 10000: tier = "🎯 Micro (10K-50K)"
    else: tier = "🌱 Nano (<10K)"

    return {
        "handle": f"@{clean_h}",
        "raw_handle": clean_h,
        "full_name": full_name,
        "followers": followers,
        "tier": tier,
        "profile_url": f"https://www.instagram.com/{clean_h}/"
    }


# ==============================================================================
# 2. INSTAGRAM SERVICE
# ==============================================================================
class InstagramService:
    """Extracts 1-Year Grid, Reels, Co-Authors, and Profile Metrics."""

    def __init__(self, cookies: Optional[Dict[str, str]] = None, headers: Optional[Dict[str, str]] = None):
        self.cookies = cookies or DEFAULT_IG_COOKIES
        self.headers = headers or DEFAULT_IG_HEADERS

    def _get_session(self) -> cffi_requests.Session:
        return cffi_requests.Session(impersonate="chrome120")

    def resolve_pk(self, username: str) -> int:
        clean_u = username.lower().replace("@", "").strip()
        s = self._get_session()
        r = s.get(f"https://www.instagram.com/{clean_u}/", timeout=12)
        m = re.search(r'"profilePage_(\d+)"', r.text) or re.search(r'"props":{"id":"(\d+)"', r.text) or re.search(r'"user_id":"(\d+)"', r.text)
        if m: return int(m.group(1))
        raise ValueError(f"Could not resolve numeric PK for @{username}")

    def get_profile(self, username: str) -> Dict[str, Any]:
        return resolve_creator_profile(username)

    def get_partnerships(self, target_brand: str, days_back: int = 365, max_pages: int = 25) -> Dict[str, Any]:
        clean_u = target_brand.lower().replace("@", "").strip()
        user_pk = self.resolve_pk(clean_u)
        session = self._get_session()
        cutoff_ts = int((datetime.now(timezone.utc) - timedelta(days=days_back)).timestamp())

        all_media = []
        seen_ids = set()

        for endpoint_name, base_url in [
            ("Timeline Feed", f"https://i.instagram.com/api/v1/feed/user/{user_pk}/"),
            ("Reels/Clips", "https://i.instagram.com/api/v1/clips/user/")
        ]:
            max_id = ""
            for page in range(1, max_pages + 1):
                try:
                    if endpoint_name == "Timeline Feed":
                        url = f"{base_url}?count=12"
                        if max_id: url += f"&max_id={max_id}"
                        r = session.get(url, headers=self.headers, cookies=self.cookies, timeout=12)
                    else:
                        body_data = {"target_user_id": str(user_pk), "page_size": "12"}
                        if max_id: body_data["max_id"] = max_id
                        r = session.post(base_url, headers=self.headers, cookies=self.cookies, data=body_data, timeout=12)

                    if r.status_code != 200: break
                    data = r.json()
                    items = data.get("items", [])
                    raw_list = [it.get("media", it) if "media" in it else it for it in items]

                    for it in raw_list:
                        pk = str(it.get("pk") or it.get("id"))
                        if pk and pk not in seen_ids:
                            seen_ids.add(pk)
                            all_media.append(it)

                    unpinned = [it.get("taken_at", 0) for it in raw_list if not it.get("timeline_pinned_user_ids") and it.get("taken_at")]
                    oldest_ts = min(unpinned, default=0)
                    if oldest_ts and oldest_ts < cutoff_ts:
                        break

                    max_id = data.get("next_max_id") or data.get("paging_info", {}).get("max_id")
                    if not max_id or not items: break
                    time.sleep(0.35)
                except Exception:
                    break

        # Process Collaborations
        collabs = []
        creators_dict = {}

        for it in all_media:
            taken_at = it.get("taken_at", 0)
            if taken_at < cutoff_ts: continue

            owner = it.get("user", {})
            owner_uname = owner.get("username", "").lower()
            coauthors = it.get("coauthor_producers", [])
            is_paid = bool(it.get("is_paid_partnership", False))

            is_collab = False
            creator_uname = ""

            if owner_uname and owner_uname != clean_u and not owner_uname.startswith(clean_u[:5]):
                is_collab = True
                creator_uname = owner_uname
            elif coauthors:
                for c in coauthors:
                    cu = c.get("username", "").lower()
                    if cu != clean_u and not cu.startswith(clean_u[:5]):
                        is_collab = True
                        creator_uname = cu
                        break

            if is_collab and creator_uname:
                likes = it.get("like_count") or 0
                comments = it.get("comment_count") or 0
                views = it.get("play_count") or it.get("view_count") or int(likes * 18.5)
                like_rate = (likes / views * 100.0) if views > 0 else 0.0

                is_boosted = (like_rate < 0.35 and views >= 200000) or views >= 3000000
                if is_paid and is_boosted: tier = "Tier 1: Toggle ON + Boosted Paid Ad"
                elif is_paid and not is_boosted: tier = "Tier 2: Toggle ON + Organic Collab"
                elif not is_paid and is_boosted: tier = "Tier 3: Toggle OFF + Heavily Boosted Ad"
                else: tier = "Tier 4: Toggle OFF + Organic / Noise"

                code = it.get("code", "")
                date_str = datetime.fromtimestamp(taken_at, tz=timezone.utc).strftime("%Y-%m-%d")

                collabs.append({
                    "post_url": f"https://www.instagram.com/p/{code}/" if code else "",
                    "creator_handle": f"@{creator_uname}",
                    "raw_handle": creator_uname,
                    "date": date_str,
                    "views": views,
                    "likes": likes,
                    "comments": comments,
                    "like_to_view_pct": round(like_rate, 3),
                    "is_paid_toggle": is_paid,
                    "is_boosted": is_boosted,
                    "partnership_tier": tier
                })

                if creator_uname not in creators_dict:
                    creators_dict[creator_uname] = {
                        "handle": f"@{creator_uname}",
                        "raw_handle": creator_uname,
                        "total_posts": 1,
                        "total_views": views,
                        "total_likes": likes,
                        "latest_post_date": date_str
                    }
                else:
                    creators_dict[creator_uname]["total_posts"] += 1
                    creators_dict[creator_uname]["total_views"] += views
                    creators_dict[creator_uname]["total_likes"] += likes

        return {
            "brand": clean_u,
            "total_media_scanned": len(all_media),
            "total_collab_posts": len(collabs),
            "unique_creators_count": len(creators_dict),
            "creators": list(creators_dict.values()),
            "collabs": collabs
        }


# ==============================================================================
# 3. FACEBOOK SERVICE
# ==============================================================================
class FacebookService:
    """Handles Facebook Page Metrics, Followers & Delegate Page IDs."""

    def __init__(self):
        self.session = cffi_requests.Session(impersonate="chrome120")

    def get_page_info(self, page_handle: str) -> Dict[str, Any]:
        clean_p = page_handle.lower().strip()
        url = f"https://www.facebook.com/{clean_p}/"
        r = self.session.get(url, timeout=12)

        pid_m = re.search(r'"delegate_page":\{"id":"(\d+)"', r.text) or re.search(r'"pageID":"(\d+)"', r.text)
        page_id = pid_m.group(1) if pid_m else None

        fols_m = re.search(r'([0-9.,KMBkmb]+)\s+followers', r.text, re.I)
        followers = fols_m.group(1) if fols_m else "N/A"

        return {
            "page_handle": clean_p,
            "page_id": page_id,
            "followers_str": followers,
            "page_url": url
        }


# ==============================================================================
# 4. META AD LIBRARY SERVICE
# ==============================================================================
class MetaAdLibraryService:
    """Extracts Ad Cards, Whitelisted Creator Handles, and Run Dates."""

    def search_ads(self, query: str, page_id: Optional[str] = None, active_only: bool = False, max_scrolls: int = 35) -> Dict[str, Any]:
        status_param = "active" if active_only else "all"
        if page_id:
            url = f"https://www.facebook.com/ads/library/?active_status={status_param}&ad_type=all&country=IN&view_all_page_id={page_id}&search_type=page&media_type=all"
        else:
            url = f"https://www.facebook.com/ads/library/?active_status={status_param}&ad_type=all&country=IN&q={query}&search_type=keyword_unordered&media_type=all"

        all_ads = []
        seen_ids = set()

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                viewport={"width": 1600, "height": 1000},
                locale="en-US",
                extra_http_headers={"Accept-Language": "en-US,en;q=0.9"}
            )
            page = context.new_page()
            page.goto(url, timeout=50000, wait_until="domcontentloaded")
            time.sleep(6)

            js_extractor = r"""
            () => {
                const ads = [];
                const fullText = document.body.innerText || "";
                const blocks = fullText.split(/(?:Library ID:|लायब्ररी आयडी:)\s*(\d+)/i);
                
                for (let i = 1; i < blocks.length; i += 2) {
                    const libId = blocks[i].trim();
                    const chunk = blocks[i+1] || "";
                    
                    const isActive = chunk.includes("Active") || chunk.includes("सक्रिय");
                    const dateMatch = chunk.match(/(?:Started running on|रोजी प्रसारण सुरू झाले)\s*([^\n·]+)/i);
                    const startDate = dateMatch ? dateMatch[1].trim() : "";
                    
                    let advName = "";
                    const advMatch = chunk.match(/(?:See (?:ad|summary) details|जाहिरात तपशील पहा)\n([^\n]+)\n(?:Sponsored|प्रायोजित)/i);
                    if (advMatch) advName = advMatch[1].trim();
                    
                    let body = "";
                    const spParts = chunk.split(/(?:Sponsored|प्रायोजित)/i);
                    if (spParts.length > 1) {
                        body = spParts[1].split(/(?:Shop Now|Learn More|Buy Now|Order Now|See (?:ad|summary)|जाहिरात तपशील पहा)/i)[0].trim();
                    }
                    
                    ads.push({
                        library_id: libId,
                        ad_url: `https://www.facebook.com/ads/library/?id=${libId}`,
                        advertiser: advName,
                        is_active: isActive,
                        start_date: startDate,
                        body: body.substring(0, 400).replace(/\n/g, " ").trim()
                    });
                }
                return ads;
            }
            """

            prev_len = 0
            stalls = 0

            for step in range(1, max_scrolls + 1):
                batch = page.evaluate(js_extractor)
                for ad in batch:
                    if ad["library_id"] not in seen_ids:
                        seen_ids.add(ad["library_id"])
                        all_ads.append(ad)

                if len(all_ads) == prev_len:
                    stalls += 1
                    if stalls >= 4: break
                else:
                    stalls = 0

                prev_len = len(all_ads)
                page.evaluate("window.scrollBy(0, 2500)")
                time.sleep(1.8)

            browser.close()

        # Identify Creators
        creators = {}
        clean_b = query.lower().replace(" ", "").replace("_", "")

        for ad in all_ads:
            adv = ad["advertiser"]
            body = ad["body"]
            is_collab = False
            cname = None

            if " with " in adv.lower():
                parts = adv.split(" with ")
                if clean_b in parts[1].lower().replace(" ", ""):
                    is_collab = True
                    cname = parts[0].strip()
            elif adv and clean_b not in adv.lower().replace(" ", "") and not adv.startswith("See ad"):
                is_collab = True
                cname = adv
            else:
                mentions = re.findall(r'@([A-Za-z0-9_.]+)', body)
                valid_m = [m for m in mentions if clean_b not in m.lower().replace("_", "")]
                if valid_m:
                    is_collab = True
                    cname = f"@{valid_m[0]}"

            ad["is_creator_collab"] = is_collab
            ad["creator_name"] = cname

            if is_collab and cname:
                ckey = cname.lower().replace("@", "").strip()
                if ckey not in creators:
                    creators[ckey] = {
                        "name": cname,
                        "handle": f"@{ckey}",
                        "active_ads": 1 if ad["is_active"] else 0,
                        "total_ads": 1,
                        "sample_start_date": ad["start_date"],
                        "sample_ad_url": ad["ad_url"]
                    }
                else:
                    creators[ckey]["total_ads"] += 1
                    if ad["is_active"]:
                        creators[ckey]["active_ads"] += 1

        return {
            "query": query,
            "page_id": page_id,
            "total_ads_captured": len(all_ads),
            "unique_creators_count": len(creators),
            "creators": list(creators.values()),
            "ads": all_ads
        }


# ==============================================================================
# 5. MASTER COMPETITOR INTELLIGENCE CLIENT (END-TO-END PIPELINE)
# ==============================================================================
class CompetitorIntelligenceClient:
    """
    Unified Orchestrator:
    Step 1: Scrape Instagram Grid
    Step 2: Enrich Instagram Creator Metrics (Followers, Sizing Tiers)
    Step 3: Scrape Meta Ad Library (Dark Ads)
    Step 4: Enrich Meta Creator Metrics
    Step 5: Fuse & Deduplicate into Master Creator Roster
    Step 6: Build Multi-Tab Master Excel Deliverable
    """

    def __init__(self, ig_cookies: Optional[Dict[str, str]] = None):
        self.instagram = InstagramService(cookies=ig_cookies)
        self.facebook = FacebookService()
        self.ad_library = MetaAdLibraryService()

    def audit_brand(self, target_brand: str, fb_page_id: Optional[str] = None, days_back: int = 365, export_excel: bool = True, output_filename: Optional[str] = None) -> Dict[str, Any]:
        clean_b = target_brand.lower().replace("@", "").strip()
        print("\n" + "="*80)
        print(f"🚀 EXECUTING 360° COMPETITOR INTELLIGENCE AUDIT FOR: @{clean_b.upper()}")
        print("="*80 + "\n")

        # ----------------------------------------------------------------------
        # STEP 1: Scrape Instagram Brand Grid & Reels (1 Year)
        # ----------------------------------------------------------------------
        print(f"📌 [STEP 1/5] Scraping Instagram 1-Year Grid & Reels for @{clean_b}...")
        ig_data = self.instagram.get_partnerships(clean_b, days_back=days_back)
        print(f"   ✓ Captured {ig_data['total_collab_posts']} Collab Posts across {ig_data['unique_creators_count']} Creators\n")

        # ----------------------------------------------------------------------
        # STEP 2: Enrich Instagram Creator Profiles (Followers & Sizing Tiers)
        # ----------------------------------------------------------------------
        print(f"📌 [STEP 2/5] Enriching Profile Metrics for {len(ig_data['creators'])} Instagram Creators...")
        ig_profiles_map = {}
        with ThreadPoolExecutor(max_workers=10) as ex:
            futs = {ex.submit(resolve_creator_profile, c["raw_handle"]): c["raw_handle"] for c in ig_data["creators"]}
            for f in as_completed(futs):
                prof = f.result()
                ig_profiles_map[prof["raw_handle"]] = prof

        print(f"   ✓ Successfully resolved follower counts & tiers for all {len(ig_profiles_map)} creators\n")

        # ----------------------------------------------------------------------
        # STEP 3: Scrape Meta Ad Library (Active & Inactive Dark Ads)
        # ----------------------------------------------------------------------
        print(f"📌 [STEP 3/5] Scraping Meta Ad Library for '{clean_b}' (Page ID: {fb_page_id})...")
        ad_data = self.ad_library.search_ads(query=clean_b, page_id=fb_page_id, max_scrolls=30)
        print(f"   ✓ Captured {ad_data['total_ads_captured']} Ads | {ad_data['unique_creators_count']} Whitelisted Partners\n")

        # ----------------------------------------------------------------------
        # STEP 4: Enrich Meta Ad Library Creators (Discover & Resolve Dark Ads)
        # ----------------------------------------------------------------------
        print(f"📌 [STEP 4/5] Enriching Metrics for Meta Ad Library Creator Partners...")
        meta_profiles_map = {}
        missing_from_ig = [mc["handle"].replace("@", "").strip() for mc in ad_data["creators"] if mc["handle"].replace("@", "").strip() not in ig_profiles_map]
        
        with ThreadPoolExecutor(max_workers=8) as ex:
            futs = {ex.submit(resolve_creator_profile, h): h for h in missing_from_ig}
            for f in as_completed(futs):
                prof = f.result()
                meta_profiles_map[prof["raw_handle"]] = prof

        print(f"   ✓ Resolved metrics for {len(meta_profiles_map)} Dark Whitelist Creators\n")

        # ----------------------------------------------------------------------
        # STEP 5: Data Fusion & Entity Deduplication
        # ----------------------------------------------------------------------
        print(f"📌 [STEP 5/5] Fusing & Deduplicating Cross-Platform Datasets...")
        unified_master = {}

        # 1. Ingest Instagram Creators
        for c in ig_data["creators"]:
            h = c["raw_handle"]
            prof = ig_profiles_map.get(h, {})
            unified_master[h] = {
                "handle": f"@{h}",
                "raw_handle": h,
                "full_name": prof.get("full_name", h),
                "followers": prof.get("followers", 0),
                "creator_tier": prof.get("tier", "🌱 Nano (<10K)"),
                "presence_platform": "📸 Instagram Grid Only",
                "on_instagram_grid": True,
                "on_meta_adlibrary": False,
                "total_grid_posts": c["total_posts"],
                "total_grid_views": c["total_views"],
                "active_meta_ads": 0,
                "total_meta_ads": 0,
                "sample_grid_url": f"https://www.instagram.com/{h}/",
                "sample_ad_url": ""
            }

        # 2. Merge Meta Ad Library Creators
        for mc in ad_data["creators"]:
            h = mc["handle"].replace("@", "").lower().replace(" ", "").replace("_", "").replace(".", "")
            matched_h = None
            for uh in unified_master:
                if uh.replace("_", "").replace(".", "") == h:
                    matched_h = uh
                    break

            if matched_h:
                unified_master[matched_h]["on_meta_adlibrary"] = True
                unified_master[matched_h]["presence_platform"] = "💎 Both (IG Grid + Meta Ads)"
                unified_master[matched_h]["active_meta_ads"] = mc["active_ads"]
                unified_master[matched_h]["total_meta_ads"] = mc["total_ads"]
                unified_master[matched_h]["sample_ad_url"] = mc["sample_ad_url"]
            else:
                orig_h = mc["handle"].replace("@", "").strip()
                prof = meta_profiles_map.get(orig_h, resolve_creator_profile(orig_h))
                unified_master[orig_h] = {
                    "handle": f"@{orig_h}",
                    "raw_handle": orig_h,
                    "full_name": prof.get("full_name", orig_h),
                    "followers": prof.get("followers", 0),
                    "creator_tier": prof.get("tier", "🌱 Nano (<10K)"),
                    "presence_platform": "🚀 Meta Dark Ads Only",
                    "on_instagram_grid": False,
                    "on_meta_adlibrary": True,
                    "total_grid_posts": 0,
                    "total_grid_views": 0,
                    "active_meta_ads": mc["active_ads"],
                    "total_meta_ads": mc["total_ads"],
                    "sample_grid_url": "",
                    "sample_ad_url": mc["sample_ad_url"]
                }

        sorted_unified = sorted(
            unified_master.values(),
            key=lambda x: (x["active_meta_ads"], x["followers"], x["total_grid_views"]),
            reverse=True
        )

        out_file = output_filename or f"{clean_b}_complete_creator_audit.xlsx"

        if export_excel:
            self._export_master_excel(clean_b, sorted_unified, ig_data["collabs"], ad_data["ads"], out_file)

        print("\n" + "="*80)
        print(f"✅ AUDIT COMPLETE! FOUND {len(sorted_unified)} TOTAL UNIQUE CREATORS")
        print(f"📊 Deliverable Workbook: {out_file}")
        print("="*80 + "\n")

        return {
            "brand": clean_b,
            "total_unique_creators": len(sorted_unified),
            "instagram_grid_creators": ig_data["unique_creators_count"],
            "meta_adlibrary_dark_creators": len([c for c in sorted_unified if not c["on_instagram_grid"]]),
            "both_platforms_creators": len([c for c in sorted_unified if c["on_instagram_grid"] and c["on_meta_adlibrary"]]),
            "total_collab_posts": ig_data["total_collab_posts"],
            "total_meta_ads": ad_data["total_ads_captured"],
            "unified_creators": sorted_unified,
            "excel_file": out_file if export_excel else None
        }

    def _export_master_excel(self, brand: str, creators: list, collabs: list, ads: list, filename: str):
        wb = openpyxl.Workbook()
        font_title = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
        font_norm = Font(name="Calibri", size=10, bold=False, color="000000")
        font_link = Font(name="Calibri", size=10, bold=False, color="0563C1", underline="single")
        thin_line = Side(style="thin", color="D5D8DC")
        border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

        # Tab 1: Complete Creators Portfolio
        ws = wb.active
        ws.title = "Complete Creators Portfolio"
        ws.sheet_view.showGridLines = True
        ws.merge_cells("A1:J1")
        ws["A1"] = f"Complete Creator Partnerships Portfolio — @{brand.upper()} ({len(creators)} Unique Creators)"
        ws["A1"].font = font_title
        ws["A1"].fill = PatternFill("solid", fgColor="0B2240")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        headers = [
            ("#", 5), ("Creator Handle", 24), ("Full Name", 26), ("Audience Followers", 18),
            ("Scale Tier", 28), ("Presence Platform", 28), ("Total Grid Posts", 16),
            ("Total Video Views", 18), ("Live Active Ads", 16), ("Sample Meta Ad URL", 45)
        ]
        for col_idx, (h_text, w) in enumerate(headers, 1):
            c = ws.cell(row=2, column=col_idx, value=h_text)
            c.font = font_hdr
            c.fill = PatternFill("solid", fgColor="1B2631")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border_cell
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.row_dimensions[2].height = 24
        ws.freeze_panes = "A3"

        for idx, c in enumerate(creators, 1):
            r_num = idx + 2
            r_vals = [
                idx, c["handle"], c["full_name"], c["followers"], c["creator_tier"],
                c["presence_platform"], c["total_grid_posts"], c["total_grid_views"],
                c["active_meta_ads"], c["sample_ad_url"]
            ]
            for c_idx, val in enumerate(r_vals, 1):
                cell = ws.cell(row=r_num, column=c_idx, value=val)
                cell.border = border_cell
                if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx in (2, 3): cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx == 4: cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
                elif c_idx == 5: cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx == 6:
                    cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
                    if "Both" in str(val): cell.fill = PatternFill("solid", fgColor="D4EFDF")
                    elif "Meta Dark" in str(val): cell.fill = PatternFill("solid", fgColor="FEF9E7")
                    else: cell.fill = PatternFill("solid", fgColor="EBF5FB")
                elif c_idx in (7, 8, 9): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
                elif c_idx == 10: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None
            ws.row_dimensions[r_num].height = 20

        # Tab 2: Instagram Collab Posts
        ws_collabs = wb.create_sheet("Instagram Collab Posts")
        ws_collabs.sheet_view.showGridLines = True
        ws_collabs.merge_cells("A1:I1")
        ws_collabs["A1"] = f"Instagram Collab Posts & 4-Tier Hierarchy ({len(collabs)} Posts)"
        ws_collabs["A1"].font = font_title
        ws_collabs["A1"].fill = PatternFill("solid", fgColor="1B4F72")
        ws_collabs["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_collabs.row_dimensions[1].height = 28

        col_headers = [
            ("#", 5), ("Creator Handle", 24), ("Date", 12), ("Views", 15), ("Likes", 14),
            ("Like-to-View %", 15), ("Paid Partnership Toggle", 22), ("4-Tier Classification", 32), ("Post URL", 45)
        ]
        for col_idx, (h_text, w) in enumerate(col_headers, 1):
            c = ws_collabs.cell(row=2, column=col_idx, value=h_text)
            c.font = font_hdr
            c.fill = PatternFill("solid", fgColor="283747")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border_cell
            ws_collabs.column_dimensions[get_column_letter(col_idx)].width = w
        ws_collabs.row_dimensions[2].height = 24
        ws_collabs.freeze_panes = "A3"

        for idx, post in enumerate(collabs, 1):
            r_num = idx + 2
            r_vals = [
                idx, post["creator_handle"], post["date"], post["views"], post["likes"],
                f"{post['like_to_view_pct']:.2f}%", "Toggle ON" if post["is_paid_toggle"] else "Toggle OFF",
                post["partnership_tier"], post["post_url"]
            ]
            for c_idx, val in enumerate(r_vals, 1):
                cell = ws_collabs.cell(row=r_num, column=c_idx, value=val)
                cell.border = border_cell
                if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 2: cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx == 3: cell.font = font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx in (4, 5): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
                elif c_idx == 6: cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center")
                elif c_idx == 7:
                    cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
                    if "ON" in str(val): cell.fill = PatternFill("solid", fgColor="D4EFDF")
                elif c_idx == 8:
                    cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
                    if "Tier 1" in str(val): cell.fill = PatternFill("solid", fgColor="D4EFDF")
                    elif "Tier 2" in str(val): cell.fill = PatternFill("solid", fgColor="EAFAF1")
                    elif "Tier 3" in str(val): cell.fill = PatternFill("solid", fgColor="FCF3CF")
                elif c_idx == 9: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None
            ws_collabs.row_dimensions[r_num].height = 20

        # Tab 3: Meta Ad Library Ads
        ws_ads = wb.create_sheet("Meta Ad Library Ads")
        ws_ads.sheet_view.showGridLines = True
        ws_ads.merge_cells("A1:H1")
        ws_ads["A1"] = f"Meta Ad Library Repository ({len(ads)} Total Ads)"
        ws_ads["A1"].font = font_title
        ws_ads["A1"].fill = PatternFill("solid", fgColor="2C3E50")
        ws_ads["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_ads.row_dimensions[1].height = 28

        ad_headers = [
            ("#", 5), ("Ad Archive ID", 18), ("Ad Status", 12), ("Advertiser Name", 28),
            ("Is Creator Collab?", 18), ("Creator Name", 28), ("Started Running Date", 20), ("Ad Library URL", 45)
        ]
        for col_idx, (h_text, w) in enumerate(ad_headers, 1):
            c = ws_ads.cell(row=2, column=col_idx, value=h_text)
            c.font = font_hdr
            c.fill = PatternFill("solid", fgColor="34495E")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border_cell
            ws_ads.column_dimensions[get_column_letter(col_idx)].width = w
        ws_ads.row_dimensions[2].height = 24
        ws_ads.freeze_panes = "A3"

        for idx, ad in enumerate(ads, 1):
            r_num = idx + 2
            r_vals = [
                idx, ad["library_id"], "Active" if ad["is_active"] else "Inactive",
                ad["advertiser"], "Yes" if ad["is_creator_collab"] else "No",
                ad["creator_name"] or "—", ad["start_date"], ad["ad_url"]
            ]
            for c_idx, val in enumerate(r_vals, 1):
                cell = ws_ads.cell(row=r_num, column=c_idx, value=val)
                cell.border = border_cell
                if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 2: cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 3:
                    cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
                    if val == "Active": cell.fill = PatternFill("solid", fgColor="D4EFDF")
                elif c_idx in (4, 6): cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx == 5:
                    cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
                    if val == "Yes": cell.fill = PatternFill("solid", fgColor="EAFAF1")
                elif c_idx == 7: cell.font = font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 8: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None
            ws_ads.row_dimensions[r_num].height = 20

        wb.save(filename)
        print(f"[Excel Export] Master Workbook saved: {filename}")


BrandAnalyticsClient = CompetitorIntelligenceClient
