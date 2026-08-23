"""
Add 'Paid Partnership Toggle' (ON / OFF) column to 'All Brands - Paid Collabs'
and update 'Brand-Creator Summary' with toggle adoption stats.
"""

import sys, openpyxl, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

WORKBOOK_PATH = "jewellery_brands_consolidated_analysis.xlsx"

wb = openpyxl.load_workbook(WORKBOOK_PATH)

thin = Side(style="thin", color="CCCCCC")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
bf = Font(name="Calibri", bold=True, size=10, color="000000")
nf = Font(name="Calibri", bold=False, size=10, color="000000")

# Styles for Toggle ON vs OFF
toggle_on_fill = PatternFill("solid", fgColor="D4EFDF") # Soft Mint Green
toggle_on_font = Font(name="Calibri", bold=True, size=10, color="145A32") # Dark Green

toggle_off_fill = PatternFill("solid", fgColor="EAECEE") # Neutral Soft Grey
toggle_off_font = Font(name="Calibri", bold=False, size=10, color="5D6D7E") # Slate Grey

# ── 1. Update 'All Brands - Paid Collabs' ──────────────────────
ws_con = wb["All Brands - Paid Collabs"]

# Rebuild the sheet with the new 'Paid Partnership Toggle' column
# We will read all rows, determine toggle status, and rewrite with proper formatting

raw_data = []
for r in range(3, ws_con.max_row + 1):
    brand = ws_con.cell(row=r, column=2).value
    handle = ws_con.cell(row=r, column=3).value
    fol = ws_con.cell(row=r, column=4).value
    likes = ws_con.cell(row=r, column=5).value
    com = ws_con.cell(row=r, column=6).value
    er = ws_con.cell(row=r, column=7).value
    date = ws_con.cell(row=r, column=8).value
    url = ws_con.cell(row=r, column=9).value
    via = ws_con.cell(row=r, column=10).value
    cap = str(ws_con.cell(row=r, column=11).value or "")
    
    if not handle:
        continue
        
    cap_lower = cap.lower()
    # Check if formal paid partnership label / hashtag disclosure is present
    has_toggle = any(w in cap_lower for w in [
        "#ad", "#sponsored", "#paidpartnership", "#collab", "#brandambassador",
        "paid partnership", "paid collab", "#ad|", "#collab|", "#ad "
    ])
    
    # Also check known creators with strict ASCI disclosure
    if handle.lower() in ["@chaitra_vasudevan_official_", "@rashwin99", "@prithinarayanan"]:
        has_toggle = True
        
    toggle_val = "ON (Formal Label)" if has_toggle else "OFF (Collab Only)"
    
    raw_data.append({
        "brand": brand,
        "handle": handle,
        "toggle": toggle_val,
        "is_on": has_toggle,
        "followers": fol,
        "avg_likes": likes,
        "avg_comments": com,
        "avg_er": er,
        "date": date,
        "url": url,
        "via": via,
        "caption": cap
    })

# Clear and rewrite
wb.remove(ws_con)
ws_con = wb.create_sheet("All Brands - Paid Collabs", 1)
ws_con.sheet_view.showGridLines = True

# Title row
ws_con.merge_cells("A1:L1")
on_total = sum(1 for p in raw_data if p["is_on"])
off_total = len(raw_data) - on_total
ws_con["A1"] = f"Consolidated Creator-Owned Paid Partnerships ('Post owned by partner') — {len(raw_data)} Total Posts (Toggle ON: {on_total} · Toggle OFF: {off_total})"
ws_con["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
ws_con["A1"].fill = PatternFill("solid", fgColor="1E5631")
ws_con["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_con.row_dimensions[1].height = 34

con_headers = [
    ("#", 5),
    ("Brand Name", 28),
    ("Creator Handle", 25),
    ("Paid Partnership Toggle", 22),
    ("Followers", 14),
    ("Avg Likes/Post", 14),
    ("Avg Comments/Post", 14),
    ("Avg ER%", 12),
    ("Post Date", 13),
    ("Post URL", 48),
    ("Detection Method", 28),
    ("Caption Preview", 65),
]

for col_idx, (h_text, width) in enumerate(con_headers, 1):
    c = ws_con.cell(row=2, column=col_idx, value=h_text)
    c.font = hdr_font
    c.fill = PatternFill("solid", fgColor="1E5631")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bdr
    ws_con.column_dimensions[get_column_letter(col_idx)].width = width
ws_con.row_dimensions[2].height = 24
ws_con.freeze_panes = "A3"

lnk = Font(name="Calibri", bold=False, size=10, color="0563C1", underline="single")

for idx, p in enumerate(raw_data, 1):
    row = idx + 2
    vals = [
        idx,
        p["brand"],
        p["handle"],
        p["toggle"],
        p["followers"],
        p["avg_likes"],
        p["avg_comments"],
        p["avg_er"],
        p["date"],
        p["url"],
        p["via"],
        p["caption"],
    ]
    
    for col_idx, val in enumerate(vals, 1):
        c = ws_con.cell(row=row, column=col_idx, value=val)
        c.border = bdr
        if col_idx == 1:
            c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 2:
            c.font = bf; c.alignment = Alignment(horizontal="left", vertical="center")
        elif col_idx == 3:
            c.font = bf; c.alignment = Alignment(horizontal="left", vertical="center")
            c.fill = PatternFill("solid", fgColor="F9EBEA")
        elif col_idx == 4: # Toggle column
            if p["is_on"]:
                c.font = toggle_on_font
                c.fill = toggle_on_fill
            else:
                c.font = toggle_off_font
                c.fill = toggle_off_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx in (5, 6, 7):
            c.font = nf; c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "#,##0"
        elif col_idx == 8:
            c.font = bf; c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "0.00%"
        elif col_idx == 9:
            c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 10:
            c.font = lnk; c.alignment = Alignment(horizontal="left", vertical="center")
            if val: c.hyperlink = val
        elif col_idx == 11:
            c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 12:
            c.font = nf; c.alignment = Alignment(horizontal="left", vertical="center")
        else:
            c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
    ws_con.row_dimensions[row].height = 22


# ── 2. Update 'Brand-Creator Summary' with Toggle Stats ───────
ws_sum = wb["Brand-Creator Summary"]

# Calculate per-brand toggle stats
brand_toggle_stats = {}
for p in raw_data:
    b = p["brand"]
    if b not in brand_toggle_stats:
        brand_toggle_stats[b] = {"on": 0, "off": 0, "total": 0}
    brand_toggle_stats[b]["total"] += 1
    if p["is_on"]:
        brand_toggle_stats[b]["on"] += 1
    else:
        brand_toggle_stats[b]["off"] += 1

# Rebuild Summary sheet to include Toggle ON/OFF columns
raw_sum_rows = []
for r in range(3, ws_sum.max_row):
    brand_name = ws_sum.cell(row=r, column=2).value
    if not brand_name or brand_name == "Total / Overall":
        continue
    posts = ws_sum.cell(row=r, column=3).value or 0
    u_creators = ws_sum.cell(row=r, column=4).value or 0
    avg_fol = ws_sum.cell(row=r, column=5).value or 0
    avg_er = ws_sum.cell(row=r, column=6).value or 0.0
    latest_dt = ws_sum.cell(row=r, column=7).value or "N/A"
    top_c = ws_sum.cell(row=r, column=8).value or ""
    
    t_info = brand_toggle_stats.get(brand_name, {"on": 0, "off": 0, "total": posts})
    
    raw_sum_rows.append({
        "brand": brand_name,
        "posts": posts,
        "creators": u_creators,
        "toggle_on": t_info["on"],
        "toggle_off": t_info["off"],
        "toggle_pct": t_info["on"] / posts if posts else 0.0,
        "avg_followers": avg_fol,
        "avg_er": avg_er,
        "latest_date": latest_dt,
        "top_creators": top_c
    })

wb.remove(ws_sum)
ws_sum = wb.create_sheet("Brand-Creator Summary", 0)
ws_sum.sheet_view.showGridLines = True

ws_sum.merge_cells("A1:J1")
ws_sum["A1"] = f"Executive Summary — Paid Creator Partnerships by Jewellery Brand ({len(raw_data)} Posts · {on_total} Toggle ON / {off_total} Toggle OFF)"
ws_sum["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
ws_sum["A1"].fill = PatternFill("solid", fgColor="0B2240")
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 34

sum_headers_new = [
    ("#", 5),
    ("Brand Name", 28),
    ("Paid Collab Posts", 17),
    ("Unique Creators", 16),
    ("Toggle ON", 13),
    ("Toggle OFF", 13),
    ("Toggle ON %", 14),
    ("Avg Creator Followers", 21),
    ("Avg Creator ER%", 15),
    ("Latest Collab Date", 17),
    ("Top Creator / Partner Samples", 42),
]

for col_idx, (h_text, width) in enumerate(sum_headers_new, 1):
    c = ws_sum.cell(row=2, column=col_idx, value=h_text)
    c.font = hdr_font
    c.fill = PatternFill("solid", fgColor="1F2D3D")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bdr
    ws_sum.column_dimensions[get_column_letter(col_idx)].width = width
ws_sum.row_dimensions[2].height = 24
ws_sum.freeze_panes = "A3"

for idx, b in enumerate(raw_sum_rows, 1):
    row = idx + 2
    vals = [
        idx,
        b["brand"],
        b["posts"],
        b["creators"],
        b["toggle_on"],
        b["toggle_off"],
        b["toggle_pct"],
        b["avg_followers"],
        b["avg_er"],
        b["latest_date"],
        b["top_creators"]
    ]
    for col_idx, val in enumerate(vals, 1):
        c = ws_sum.cell(row=row, column=col_idx, value=val)
        c.border = bdr
        if col_idx == 1:
            c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 2:
            c.font = bf; c.alignment = Alignment(horizontal="left", vertical="center")
        elif col_idx in (3, 4):
            c.font = bf; c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "#,##0"
            if val > 0: c.fill = PatternFill("solid", fgColor="EBF5FB")
        elif col_idx == 5: # Toggle ON
            c.font = toggle_on_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "#,##0"
            if val > 0: c.fill = toggle_on_fill
        elif col_idx == 6: # Toggle OFF
            c.font = toggle_off_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "#,##0"
            if val > 0: c.fill = toggle_off_fill
        elif col_idx == 7: # Toggle %
            c.font = bf; c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "0.0%"
        elif col_idx == 8:
            c.font = nf; c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "#,##0"
        elif col_idx == 9:
            c.font = bf; c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "0.00%"
        elif col_idx == 10:
            c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 11:
            c.font = nf; c.alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[row].height = 22

# Add Total Row
tot_r = len(raw_sum_rows) + 3
ws_sum.cell(row=tot_r, column=1, value="Total / Overall").font = Font(name="Calibri", bold=True, size=10)
ws_sum.cell(row=tot_r, column=1).fill = PatternFill("solid", fgColor="D6EAF8")
ws_sum.merge_cells(start_row=tot_r, start_column=1, end_row=tot_r, end_column=2)

tot_posts = sum(b["posts"] for b in raw_sum_rows)
tot_creators = len(set(p["handle"].lower() for p in raw_data))
tot_on = sum(b["toggle_on"] for b in raw_sum_rows)
tot_off = sum(b["toggle_off"] for b in raw_sum_rows)
tot_pct = tot_on / tot_posts if tot_posts else 0.0

ws_sum.cell(row=tot_r, column=3, value=tot_posts).font = bf; ws_sum.cell(row=tot_r, column=3).alignment = Alignment(horizontal="center", vertical="center"); ws_sum.cell(row=tot_r, column=3).number_format = "#,##0"; ws_sum.cell(row=tot_r, column=3).fill = PatternFill("solid", fgColor="D6EAF8"); ws_sum.cell(row=tot_r, column=3).border = bdr
ws_sum.cell(row=tot_r, column=4, value=tot_creators).font = bf; ws_sum.cell(row=tot_r, column=4).alignment = Alignment(horizontal="center", vertical="center"); ws_sum.cell(row=tot_r, column=4).number_format = "#,##0"; ws_sum.cell(row=tot_r, column=4).fill = PatternFill("solid", fgColor="D6EAF8"); ws_sum.cell(row=tot_r, column=4).border = bdr
ws_sum.cell(row=tot_r, column=5, value=tot_on).font = toggle_on_font; ws_sum.cell(row=tot_r, column=5).alignment = Alignment(horizontal="center", vertical="center"); ws_sum.cell(row=tot_r, column=5).number_format = "#,##0"; ws_sum.cell(row=tot_r, column=5).fill = PatternFill("solid", fgColor="D6EAF8"); ws_sum.cell(row=tot_r, column=5).border = bdr
ws_sum.cell(row=tot_r, column=6, value=tot_off).font = toggle_off_font; ws_sum.cell(row=tot_r, column=6).alignment = Alignment(horizontal="center", vertical="center"); ws_sum.cell(row=tot_r, column=6).number_format = "#,##0"; ws_sum.cell(row=tot_r, column=6).fill = PatternFill("solid", fgColor="D6EAF8"); ws_sum.cell(row=tot_r, column=6).border = bdr
ws_sum.cell(row=tot_r, column=7, value=tot_pct).font = bf; ws_sum.cell(row=tot_r, column=7).alignment = Alignment(horizontal="center", vertical="center"); ws_sum.cell(row=tot_r, column=7).number_format = "0.0%"; ws_sum.cell(row=tot_r, column=7).fill = PatternFill("solid", fgColor="D6EAF8"); ws_sum.cell(row=tot_r, column=7).border = bdr

for col_idx in [8, 9, 10, 11]:
    c = ws_sum.cell(row=tot_r, column=col_idx, value="")
    c.fill = PatternFill("solid", fgColor="D6EAF8")
    c.border = bdr
ws_sum.row_dimensions[tot_r].height = 24

wb.save(WORKBOOK_PATH)
print(f"\n✅ Successfully updated '{WORKBOOK_PATH}' with Paid Partnership Toggle column and stats!")
print(f"   • Total Posts: {len(raw_data)}")
print(f"   • Toggle ON (Formal Label): {on_total} ({on_total/len(raw_data)*100:.1f}%)")
print(f"   • Toggle OFF (Collab Only): {off_total} ({off_total/len(raw_data)*100:.1f}%)")
