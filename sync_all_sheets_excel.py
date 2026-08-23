"""
Ensure all 34 original brand sheets + GIVA & Palmonas are in jewellery_brands_master_analysis.xlsx
"""

import openpyxl, sys

sys.stdout.reconfigure(encoding="utf-8")

wb_orig = openpyxl.load_workbook("jewellery_brands_full_export.xlsx", data_only=True)
wb_master = openpyxl.load_workbook("jewellery_brands_master_analysis.xlsx")

overview_sheets = {"State-Brand Overview", "Creators", "All Brands - Paid Collabs", "Brand-Creator Summary"}
orig_brand_sheets = [s for s in wb_orig.sheetnames if s not in overview_sheets]

for s in orig_brand_sheets:
    if s not in wb_master.sheetnames:
        ws_src = wb_orig[s]
        ws_dst = wb_master.create_sheet(s)
        ws_dst.sheet_view.showGridLines = True
        for r in range(1, ws_src.max_row + 1):
            for c in range(1, ws_src.max_column + 1):
                ws_dst.cell(row=r, column=c, value=ws_src.cell(row=r, column=c).value)
        for col_l in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws_dst.column_dimensions[col_l].width = ws_src.column_dimensions[col_l].width or 20

wb_master.save("jewellery_brands_master_analysis.xlsx")
print(f"✓ Master Excel now contains {len(wb_master.sheetnames)} total sheets!")
