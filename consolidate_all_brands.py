"""
Consolidate 'Post owned by partner' across all 34 brand sheets in the workbook
and create:
1. 'All Brands - Paid Collabs' (consolidated creator-owned posts)
2. 'Brand-Creator Summary' (unique creator and post counts per brand)
"""

import sys, json, openpyxl, re
from collections import Counter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

INPUT_FILE = "jewellery_brands_full_export.xlsx"
OUTPUT_FILE = "jewellery_brands_consolidated_analysis.xlsx"

wb = openpyxl.load_workbook(INPUT_FILE)
all_sheets = wb.sheetnames

overview_sheets = {"State-Brand Overview", "Creators", "All Brands - Paid Collabs", "Brand-Creator Summary"}
brand_sheets = [s for s in all_sheets if s not in overview_sheets]

print(f"Total Brand Sheets found: {len(brand_sheets)}")

# Known self-referential / sub-brand / media accounts to filter out from creator lists
NON_CREATOR_HANDLES = {
    # Sub-brands
    "@grt.diamonds", "@grt.silverjewellery", "@grt.silverarticles", "@grtoriana",
    "@kalyan_jewellers_uk", "@kalyanjewellersusa", "@laksyah_", "@bluestone_india",
    "@tanishqjewellery", "@mia_by_tanishq", "@joyalukkasindia", "@joyalukkas_uk",
    "@sencogoldanddiamonds", "@malabargoldanddiamonds",
    # Media & Agencies
    "@elleindia", "@chennaitimestoi", "@eventart_india", "@coresocial05",
    "@platinumevara", "@platinumdaysoflove", "@menofplatinum"
}

all_paid_posts = []
brand_stats = {}

for sheet_name in brand_sheets:
    ws = wb[sheet_name]
    # Check title row / brand name
    title_val = ws.cell(row=1, column=1).value or sheet_name
    # Extract clean brand name
    brand_clean = sheet_name
    
    brand_posts = []
    brand_creators = set()
    
    # Data starts at row 3 (row 1 is title, row 2 is header)
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 11)]
        h = str(vals[0] or "").strip()
        fol = vals[1]
        l = vals[2]
        c = vals[3]
        er = vals[4]
        date = vals[5]
        url = vals[6]
        via = str(vals[7] or "").strip()
        cap = vals[8]
        stat = vals[9]
        
        if not h or not url or str(url) == "None":
            continue
            
        # Filter strictly for 'Post owned by partner (collab)'
        if "owned by partner" in via.lower():
            if h.lower() in {x.lower() for x in NON_CREATOR_HANDLES}:
                continue
                
            brand_creators.add(h.lower())
            post_entry = {
                "brand": brand_clean,
                "handle": h,
                "followers": fol,
                "avg_likes": l,
                "avg_comments": c,
                "avg_er": er,
                "date": str(date)[:10] if date else "N/A",
                "url": url,
                "via": via,
                "caption": str(cap or "")[:150],
                "status": stat or "Scraped OK"
            }
            brand_posts.append(post_entry)
            all_paid_posts.append(post_entry)
            
    # Calculate stats for this brand
    # Numeric follower & ER calculation
    fol_vals = []
    er_vals = []
    for p in brand_posts:
        # clean followers
        f_raw = str(p["followers"] or "").replace(",", "").replace("%", "").strip()
        try:
            fol_vals.append(float(f_raw))
        except Exception:
            pass
            
        # clean ER
        er_raw = str(p["avg_er"] or "").replace("%", "").strip()
        try:
            er_vals.append(float(er_raw))
        except Exception:
            pass
            
    avg_fol = int(sum(fol_vals) / len(fol_vals)) if fol_vals else 0
    avg_er = round(sum(er_vals) / len(er_vals), 2) if er_vals else 0.0
    latest_date = max([p["date"] for p in brand_posts if p["date"] != "N/A"], default="N/A")
    
    # Top creator handles
    top_creators = list(dict.fromkeys([p["handle"] for p in brand_posts]))[:4]
    top_creators_str = ", ".join(top_creators) if top_creators else "None detected"
    
    brand_stats[brand_clean] = {
        "brand": brand_clean,
        "total_posts": len(brand_posts),
        "unique_creators": len(brand_creators),
        "avg_followers": avg_fol,
        "avg_er": avg_er,
        "top_creators": top_creators_str,
        "latest_post_date": latest_date
    }
    print(f"{brand_clean:<32} | Collab Posts: {len(brand_posts):>3} | Unique Creators: {len(brand_creators):>3} | Latest: {latest_date}")

print(f"\n{'='*75}")
print(f"Total Consolidated 'Post Owned by Partner' Collabs: {len(all_paid_posts)}")
unique_overall_creators = len(set(p['handle'].lower() for p in all_paid_posts))
print(f"Total Unique Creators across ALL Brands: {unique_overall_creators}")
print(f"{'='*75}\n")


# ── 1. Create 'Brand-Creator Summary' Sheet ────────────────────
SUM_SHEET_NAME = "Brand-Creator Summary"
if SUM_SHEET_NAME in wb.sheetnames:
    del wb[SUM_SHEET_NAME]

ws_sum = wb.create_sheet(SUM_SHEET_NAME, 0) # Insert at beginning
ws_sum.sheet_view.showGridLines = True

thin = Side(style="thin", color="CCCCCC")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="1F2D3D")
hdr_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
bf = Font(name="Calibri", bold=True, size=10, color="000000")
nf = Font(name="Calibri", bold=False, size=10, color="000000")

# Title row
ws_sum.merge_cells("A1:G1")
ws_sum["A1"] = f"Executive Summary — Paid Creator Partnerships by Jewellery Brand ({len(all_paid_posts)} Total Posts · {unique_overall_creators} Unique Creators)"
ws_sum["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
ws_sum["A1"].fill = PatternFill("solid", fgColor="0B2240")
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 34

sum_headers = [
    ("#", 5),
    ("Brand Name", 30),
    ("Paid Collab Posts", 18),
    ("Unique Creators", 18),
    ("Avg Creator Followers", 22),
    ("Avg Creator ER%", 16),
    ("Latest Collab Date", 18),
    ("Top Creator / Partner Samples", 45),
]

for col_idx, (h_text, width) in enumerate(sum_headers, 1):
    c = ws_sum.cell(row=2, column=col_idx, value=h_text)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bdr
    ws_sum.column_dimensions[get_column_letter(col_idx)].width = width
ws_sum.row_dimensions[2].height = 24
ws_sum.freeze_panes = "A3"

# Sort brands by total paid posts descending
sorted_brands = sorted(brand_stats.values(), key=lambda x: x["total_posts"], reverse=True)

for idx, b in enumerate(sorted_brands, 1):
    row = idx + 2
    vals = [
        idx,
        b["brand"],
        b["total_posts"],
        b["unique_creators"],
        b["avg_followers"],
        b["avg_er"] / 100 if b["avg_er"] else 0.0,
        b["latest_post_date"],
        b["top_creators"]
    ]
    for col_idx, val in enumerate(vals, 1):
        c = ws_sum.cell(row=row, column=col_idx, value=val)
        c.border = bdr
        if col_idx == 1:
            c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 2:
            c.font = bf; c.alignment = Alignment(horizontal="left", vertical="center")
        elif col_idx in (3, 4):
            c.font = bf; c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "#,##0"
            if val > 0:
                c.fill = PatternFill("solid", fgColor="EBF5FB")
        elif col_idx == 5:
            c.font = nf; c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "#,##0"
        elif col_idx == 6:
            c.font = bf; c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "0.00%"
        elif col_idx == 7:
            c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 8:
            c.font = nf; c.alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[row].height = 22

# Summary Totals Row
tot_row = len(sorted_brands) + 3
ws_sum.cell(row=tot_row, column=1, value="Total / Overall").font = Font(name="Calibri", bold=True, size=10)
ws_sum.cell(row=tot_row, column=1).fill = PatternFill("solid", fgColor="D6EAF8")
ws_sum.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=2)

ws_sum.cell(row=tot_row, column=3, value=len(all_paid_posts)).font = Font(name="Calibri", bold=True, size=10)
ws_sum.cell(row=tot_row, column=3).number_format = "#,##0"
ws_sum.cell(row=tot_row, column=3).fill = PatternFill("solid", fgColor="D6EAF8")
ws_sum.cell(row=tot_row, column=3).alignment = Alignment(horizontal="center", vertical="center")
ws_sum.cell(row=tot_row, column=3).border = bdr

ws_sum.cell(row=tot_row, column=4, value=unique_overall_creators).font = Font(name="Calibri", bold=True, size=10)
ws_sum.cell(row=tot_row, column=4).number_format = "#,##0"
ws_sum.cell(row=tot_row, column=4).fill = PatternFill("solid", fgColor="D6EAF8")
ws_sum.cell(row=tot_row, column=4).alignment = Alignment(horizontal="center", vertical="center")
ws_sum.cell(row=tot_row, column=4).border = bdr

for col_idx in [5, 6, 7, 8]:
    c = ws_sum.cell(row=tot_row, column=col_idx, value="")
    c.fill = PatternFill("solid", fgColor="D6EAF8")
    c.border = bdr
ws_sum.row_dimensions[tot_row].height = 24


# ── 2. Create 'All Brands - Paid Collabs' Sheet ────────────────
CON_SHEET_NAME = "All Brands - Paid Collabs"
if CON_SHEET_NAME in wb.sheetnames:
    del wb[CON_SHEET_NAME]

ws_con = wb.create_sheet(CON_SHEET_NAME, 1) # Insert second
ws_con.sheet_view.showGridLines = True

ws_con.merge_cells("A1:K1")
ws_con["A1"] = f"Consolidated Creator-Owned Paid Partnerships ('Post owned by partner') — {len(all_paid_posts)} Total Posts"
ws_con["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
ws_con["A1"].fill = PatternFill("solid", fgColor="1E5631") # Green Header
ws_con["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_con.row_dimensions[1].height = 34

con_headers = [
    ("#", 5),
    ("Brand Name", 28),
    ("Creator Handle", 25),
    ("Followers", 14),
    ("Avg Likes/Post", 14),
    ("Avg Comments/Post", 14),
    ("Avg ER%", 12),
    ("Post Date", 13),
    ("Post URL", 48),
    ("Detection Method", 28),
    ("Caption Preview", 65),
]

for col_idx, (h_text, width) in enumerate(con_headers, 1):
    c = ws_con.cell(row=2, column=col_idx, value=h_text)
    c.font = hdr_font
    c.fill = PatternFill("solid", fgColor="1E5631")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bdr
    ws_con.column_dimensions[get_column_letter(col_idx)].width = width
ws_con.row_dimensions[2].height = 24
ws_con.freeze_panes = "A3"

# Sort by Date descending
all_paid_posts.sort(key=lambda x: x["date"], reverse=True)

lnk = Font(name="Calibri", bold=False, size=10, color="0563C1", underline="single")

for idx, p in enumerate(all_paid_posts, 1):
    row = idx + 2
    
    # parse followers and ER to numeric
    fol_clean = p["followers"]
    try:
        if isinstance(fol_clean, str):
            fol_clean = int(fol_clean.replace(",", ""))
    except Exception:
        pass
        
    likes_clean = p["avg_likes"]
    try:
        if isinstance(likes_clean, str):
            likes_clean = int(likes_clean.replace(",", ""))
    except Exception:
        pass
        
    com_clean = p["avg_comments"]
    try:
        if isinstance(com_clean, str):
            com_clean = int(com_clean.replace(",", ""))
    except Exception:
        pass
        
    er_clean = p["avg_er"]
    try:
        if isinstance(er_clean, str):
            er_clean = float(er_clean.replace("%", "")) / 100
        elif isinstance(er_clean, (int, float)) and er_clean > 1:
            er_clean = er_clean / 100
    except Exception:
        pass

    vals = [
        idx,
        p["brand"],
        p["handle"],
        fol_clean,
        likes_clean,
        com_clean,
        er_clean,
        p["date"],
        p["url"],
        p["via"],
        p["caption"],
    ]
    
    for col_idx, val in enumerate(vals, 1):
        c = ws_con.cell(row=row, column=col_idx, value=val)
        c.border = bdr
        if col_idx == 1:
            c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 2:
            c.font = bf; c.alignment = Alignment(horizontal="left", vertical="center")
        elif col_idx == 3:
            c.font = bf; c.alignment = Alignment(horizontal="left", vertical="center")
            c.fill = PatternFill("solid", fgColor="F9EBEA")
        elif col_idx in (4, 5, 6):
            c.font = nf; c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = "#,##0"
        elif col_idx == 7:
            c.font = bf; c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "0.00%"
        elif col_idx == 8:
            c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 9:
            c.font = lnk; c.alignment = Alignment(horizontal="left", vertical="center")
            if val: c.hyperlink = val
        elif col_idx == 10:
            c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 11:
            c.font = nf; c.alignment = Alignment(horizontal="left", vertical="center")
        else:
            c.font = nf; c.alignment = Alignment(horizontal="center", vertical="center")
    ws_con.row_dimensions[row].height = 22

# Save updated workbook
wb.save(OUTPUT_FILE)
print(f"\n✅ Created and saved: {OUTPUT_FILE}")
print(f"   • Sheet 1: '{SUM_SHEET_NAME}' ({len(sorted_brands)} brands summary)")
print(f"   • Sheet 2: '{CON_SHEET_NAME}' ({len(all_paid_posts)} total creator-owned posts)")
