"""
Instagram Bulk Profile Analyzer — Standalone Script
===================================================
✏️  HOW TO USE:
    1. Fill in USERNAMES below with a list of Instagram handles.
    2. Update COOKIES with your Instagram session cookies.
    3. Run:  python scrape_profiles.py
    4. Excel report will be saved as "instagram_profile_analysis.xlsx".

Install dependencies:
    pip install curl_cffi openpyxl requests
"""

# ════════════════════════════════════════════════════════════════
#  ✏️  STEP 1: ADD INSTAGRAM USERNAMES HERE
# ════════════════════════════════════════════════════════════════

USERNAMES = [
    "therajivmakhni",
    "nandupatilofficial",
    # Add more usernames here (e.g. "MohitVerma", "AkshayJadhav")
]

# ════════════════════════════════════════════════════════════════
#  ✏️  STEP 2: PASTE YOUR INSTAGRAM COOKIES HERE
#  Chrome → F12 → Application → Cookies → instagram.com
# ════════════════════════════════════════════════════════════════

COOKIES = {
    "sessionid":  "25113411270%3AyQFaao428g6Xb9%3A0%3AAYgcV3Qe5uiiJD1FyRkSD2vULzcMkEegPbBVeDmWYQ",
    "csrftoken":  "3gJbkGDZp99lA8QQ0brobyoHzOreuu8f",
    "mid":        "afyCbwALAAFRStE-k17-dfO5_jfa",
    "ds_user_id": "25113411270",
}

# ════════════════════════════════════════════════════════════════
#  ✏️  STEP 3: CONFIGURATION
# ════════════════════════════════════════════════════════════════

POSTS_PER_USER   = 20     # Number of recent posts to analyze per user
OUTPUT_FILENAME  = "instagram_profile_analysis.xlsx"

# ════════════════════════════════════════════════════════════════
#  BELOW: Core Scraping and Excel Generation Logic
# ════════════════════════════════════════════════════════════════

import sys, io, re, html, time, random, json
from datetime import datetime, timezone

sys.stdout.reconfigure(encoding="utf-8")

try:
    from curl_cffi import requests as cffi_requests
    USE_CURL_CFFI = True
    print("✓ curl_cffi found — using Chrome TLS fingerprint")
except ImportError:
    import requests as cffi_requests
    USE_CURL_CFFI = False
    print("⚠ curl_cffi not found — using standard requests")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IG_APP_ID = "936619743392459"

BASE_HEADERS = {
    "accept":             "*/*",
    "accept-language":    "en-US,en;q=0.9",
    "user-agent":         (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "x-ig-app-id":        IG_APP_ID,
    "x-ig-www-claim":     "0",
    "x-requested-with":   "XMLHttpRequest",
    "x-csrftoken":        COOKIES.get("csrftoken", ""),
    "origin":             "https://www.instagram.com",
    "referer":            "https://www.instagram.com/",
}


# ── Session & Network Utilities ────────────────────────────────

def make_session():
    if USE_CURL_CFFI:
        s = cffi_requests.Session(impersonate="chrome120")
    else:
        import requests
        s = requests.Session()
    s.headers.update(BASE_HEADERS)
    s.cookies.update(COOKIES)
    return s

def delay(a=1.0, b=2.5):
    time.sleep(random.uniform(a, b))

def fetch_profile_info(username: str, session) -> dict:
    url = f"https://www.instagram.com/api/v1/users/web_profile_info/?username={username}"
    hdrs = {**BASE_HEADERS, "referer": f"https://www.instagram.com/{username}/"}
    for attempt in range(1, 4):
        try:
            r = session.get(url, headers=hdrs, cookies=COOKIES, timeout=15)
            if r.status_code == 200:
                return r.json()
            elif r.status_code == 429:
                wait = 30 * attempt
                print(f"  ⏳ Rate limited — waiting {wait}s...")
                time.sleep(wait)
            elif r.status_code == 404:
                print(f"  ❌ Username '{username}' not found (404).")
                return {}
        except Exception as e:
            print(f"  ⚠ Network/parsing error: {e}")
            delay(2, 4)
    return {}

def fetch_posts_v2(user_id: str, session, count=20) -> list:
    url = f"https://www.instagram.com/api/v1/feed/user/{user_id}/?count={count}"
    for attempt in range(1, 4):
        try:
            r = session.get(url, headers=BASE_HEADERS, cookies=COOKIES, timeout=15)
            if r.status_code == 200:
                return r.json().get("items", [])
            elif r.status_code == 429:
                time.sleep(30 * attempt)
        except Exception:
            delay(2, 4)
    return []

def parse_post_v2(item) -> dict:
    likes = item.get("like_count", 0) or 0
    comments = item.get("comment_count", 0) or 0
    cap_obj = item.get("caption")
    caption = (cap_obj.get("text", "") if isinstance(cap_obj, dict) else "") or ""
    mt_map = {1: "Image", 2: "Video", 8: "Carousel"}
    ts = item.get("taken_at", 0)
    date_str = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d") if ts else "N/A"
    code = item.get("code", item.get("shortcode", ""))
    lines = [l.strip() for l in caption.splitlines() if l.strip()]
    first_line = lines[0] if lines else "Instagram Post"
    return {
        "shortcode": code,
        "url": f"https://www.instagram.com/p/{code}/",
        "title": first_line[:100],
        "type": mt_map.get(item.get("media_type", 1), "Image"),
        "date": date_str,
        "likes": likes,
        "comments": comments,
        "video_views": item.get("view_count", 0) or item.get("play_count", 0) or 0,
        "caption": caption,
    }


# ── Excel Export Styling Utilities ──────────────────────────────

def build_excel_report(profiles_data: list) -> bytes:
    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="D0D0D0")
    tbdr = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="888888"))
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(color_hex): return PatternFill("solid", fgColor=color_hex)
    def bf(size=10, color="000000"): return Font(name="Calibri", bold=True, size=size, color=color)
    def nf(size=10, color="000000"): return Font(name="Calibri", bold=False, size=size, color=color)
    def lnk(size=10): return Font(name="Calibri", bold=False, size=size, color="0563C1", underline="single")
    def ctr(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    def lft(wrap=True):  return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)
    def rgt(wrap=False): return Alignment(horizontal="right",  vertical="center", wrap_text=wrap)

    # ══════════════════════════════════════════════
    # Sheet 1: Profiles Overview
    # ══════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Profiles Overview"
    ws1.sheet_view.showGridLines = True
    
    ws1.merge_cells("A1:K1")
    ws1["A1"] = "Instagram Creators Profiles Summary"
    ws1["A1"].font = bf(14); ws1["A1"].alignment = ctr(); ws1.row_dimensions[1].height = 32

    headers1 = [
        "#", "Username", "Full Name", "Verified", "Business", 
        "Total Followers", "Following", "Total Posts Count", 
        "Avg Likes / Post", "Avg Comments / Post", "Avg ER%"
    ]
    widths1 = [5, 18, 24, 10, 11, 16, 12, 18, 16, 18, 12]

    for col, (h, w) in enumerate(zip(headers1, widths1), 1):
        cell = ws1.cell(row=2, column=col, value=h)
        cell.font = bf(10); cell.fill = fill("F2F2F2"); cell.alignment = ctr(); cell.border = tbdr
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[2].height = 24

    for i, p in enumerate(profiles_data, 1):
        row = i + 2
        vals = [
            i,
            f"@{p['username']}",
            p["full_name"],
            "Yes" if p["verified"] else "No",
            "Yes" if p["is_business"] else "No",
            p["followers"],
            p["following"],
            p["total_posts"],
            p["avg_likes"],
            p["avg_comments"],
            p["avg_er"] / 100 # stored as decimal for excel percentage format
        ]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=row, column=col, value=val); c.border = bdr
            if col == 2:
                c.font = bf(10); c.fill = fill("FFF8F0"); c.alignment = lft()
            elif col in (3,):
                c.font = nf(10); c.alignment = lft()
            elif col == 11:
                c.number_format = '0.00%'; c.font = bf(10); c.alignment = ctr()
            elif col in (6, 7, 8, 9, 10):
                c.number_format = "#,##0"; c.font = nf(10); c.alignment = rgt()
            else:
                c.font = nf(10); c.alignment = ctr()
        ws1.row_dimensions[row].height = 24

    # ══════════════════════════════════════════════
    # Sheet 2: Detailed Post Metrics
    # ══════════════════════════════════════════════
    ws2 = wb.create_sheet("Detailed Post Metrics")
    ws2.sheet_view.showGridLines = True
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:L1")
    ws2["A1"] = f"Instagram Detailed Post Analytics (Last {POSTS_PER_USER} Posts / Account)"
    ws2["A1"].font = bf(14); ws2["A1"].alignment = ctr(); ws2.row_dimensions[1].height = 30

    headers2 = [
        "#", "Username", "Creator Name", "Title", "Full Description / Caption", 
        "Date", "Type", "Post URL", "Likes", "Comments", "Video Views", "ER%"
    ]
    widths2 = [5, 18, 20, 28, 50, 13, 11, 45, 12, 12, 13, 10]

    for col, (h, w) in enumerate(zip(headers2, widths2), 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = bf(10); cell.fill = fill("F2F2F2"); cell.alignment = ctr(); cell.border = tbdr
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[2].height = 24

    post_idx = 1
    for p in profiles_data:
        for pst in p["posts"]:
            row = post_idx + 2
            vals = [
                post_idx,
                f"@{p['username']}",
                p["full_name"],
                pst["title"],
                pst["caption"],
                pst["date"],
                pst["type"],
                pst["url"],
                pst["likes"],
                pst["comments"],
                pst["video_views"],
                pst["er_percent"] / 100
            ]
            for col, val in enumerate(vals, 1):
                c = ws2.cell(row=row, column=col, value=val); c.border = bdr
                if col == 2:
                    c.font = bf(10); c.fill = fill("FFF8F0"); c.alignment = lft()
                elif col in (4, 5):
                    c.font = nf(10); c.alignment = lft(wrap=True)
                elif col == 8:
                    c.font = lnk(); c.alignment = lft(wrap=False)
                    if val: c.hyperlink = val
                elif col == 12:
                    c.number_format = '0.00%'; c.font = bf(10); c.alignment = ctr()
                elif col in (9, 10, 11):
                    c.number_format = "#,##0"; c.font = nf(10); c.alignment = rgt()
                else:
                    c.font = nf(10); c.alignment = ctr()
            ws2.row_dimensions[row].height = 28
            post_idx += 1

    # ══════════════════════════════════════════════
    # Sheet 3: Captions
    # ══════════════════════════════════════════════
    ws3 = wb.create_sheet("Captions & Descriptions")
    ws3.sheet_view.showGridLines = True
    for col, w in zip("ABCDEFG", [5, 18, 28, 45, 65, 12, 10]):
        ws3.column_dimensions[col].width = w

    ws3.merge_cells("A1:G1")
    ws3["A1"] = "Post Titles & Full Captions List"
    ws3["A1"].font = bf(14); ws3["A1"].alignment = ctr(); ws3.row_dimensions[1].height = 30

    for col, h in enumerate(["#", "Username", "Title", "URL", "Full Description / Caption", "Likes", "ER%"], 1):
        c = ws3.cell(row=2, column=col, value=h)
        c.font = bf(10); c.fill = fill("F2F2F2"); c.alignment = ctr(); c.border = tbdr
    ws3.row_dimensions[2].height = 24

    caption_idx = 1
    for p in profiles_data:
        for pst in p["posts"]:
            row = caption_idx + 2
            vals = [
                caption_idx,
                f"@{p['username']}",
                pst["title"],
                pst["url"],
                pst["caption"],
                pst["likes"],
                pst["er_percent"] / 100
            ]
            for col, val in enumerate(vals, 1):
                c = ws3.cell(row=row, column=col, value=val); c.border = bdr
                if col == 2:
                    c.font = bf(10); c.fill = fill("FFF8F0"); c.alignment = lft()
                elif col in (3, 5):
                    c.font = nf(10); c.alignment = lft(wrap=True)
                elif col == 4:
                    c.font = lnk(); c.alignment = lft(wrap=False)
                    if val: c.hyperlink = val
                elif col == 7:
                    c.number_format = '0.00%'; c.font = bf(10); c.alignment = ctr()
                elif col == 6:
                    c.number_format = "#,##0"; c.font = nf(10); c.alignment = rgt()
                else:
                    c.font = nf(10); c.alignment = ctr()
            ws3.row_dimensions[row].height = 28
            caption_idx += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Main Runner ────────────────────────────────────────────────

def main():
    if not USERNAMES:
        print("❌ No usernames found in scrape_profiles.py configuration.")
        return

    print(f"\n{'═'*65}")
    print("  Instagram Bulk Profile Analyzer")
    print(f"  Analyzing: {len(USERNAMES)} profiles  |  Posts per profile: {POSTS_PER_USER}")
    print(f"{'═'*65}\n")

    session = make_session()
    profiles_data = []

    for idx, raw_uname in enumerate(USERNAMES, 1):
        uname = raw_uname.strip().lstrip("@")
        if not uname:
            continue
        print(f"[{idx}/{len(USERNAMES)}] Scraping @{uname}...")

        # 1. Fetch Profile info
        raw_info = fetch_profile_info(uname, session)
        user_data = raw_info.get("data", {}).get("user")
        if not user_data:
            print(f"         ⚠ Skipped: Could not fetch profile data for @{uname}.")
            continue

        followers = user_data.get("edge_followed_by", {}).get("count", 0)
        profile = {
            "username":    uname,
            "full_name":   user_data.get("full_name", "N/A"),
            "followers":   followers,
            "following":   user_data.get("edge_follow", {}).get("count", 0),
            "total_posts": user_data.get("edge_owner_to_timeline_media", {}).get("count", 0),
            "bio":         user_data.get("biography", ""),
            "verified":    user_data.get("is_verified", False),
            "is_business": user_data.get("is_business_account", False),
            "user_id":     user_data.get("id", ""),
            "posts":       []
        }

        print(f"         Followers: {followers:,} | Total Posts: {profile['total_posts']:,}")

        # 2. Fetch recent 20 posts
        if profile["user_id"]:
            print(f"         Fetching last {POSTS_PER_USER} posts...")
            raw_posts = fetch_posts_v2(profile["user_id"], session, count=POSTS_PER_USER)
            
            total_likes = 0
            total_comments = 0
            for item in raw_posts[:POSTS_PER_USER]:
                p = parse_post_v2(item)
                p["er_percent"] = round((p["likes"] + p["comments"]) / followers * 100, 4) if followers else 0
                total_likes += p["likes"]
                total_comments += p["comments"]
                profile["posts"].append(p)

            count_fetched = len(profile["posts"])
            if count_fetched > 0:
                profile["avg_likes"] = total_likes // count_fetched
                profile["avg_comments"] = total_comments // count_fetched
                profile["avg_er"] = round((total_likes + total_comments) / count_fetched / followers * 100, 4) if followers else 0
            else:
                profile["avg_likes"] = 0
                profile["avg_comments"] = 0
                profile["avg_er"] = 0

            print(f"         ✓ Analyzed {count_fetched} posts. Avg ER%: {profile['avg_er']:.2f}% | Avg Likes: {profile['avg_likes']:,}")
        
        profiles_data.append(profile)
        delay(1.5, 3.0)
        print()

    if not profiles_data:
        print("❌ No profile data scraped successfully.")
        return

    print("📊 Generating Excel spreadsheet...")
    excel_bytes = build_excel_report(profiles_data)

    with open(OUTPUT_FILENAME, "wb") as f:
        f.write(excel_bytes)

    print(f"\n{'═'*65}")
    print(f"  ✅ Complete! Saved: {OUTPUT_FILENAME}")
    print(f"  Profiles analyzed: {len(profiles_data)}")
    print("  Sheets: Profiles Overview · Detailed Post Metrics · Captions & Descriptions")
    print(f"{'═'*65}\n")


if __name__ == "__main__":
    main()
