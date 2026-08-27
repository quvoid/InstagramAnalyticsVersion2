"""
Deep 2-Year Scraper & 4-Tier Creator Intelligence Engine for GRT Oriana (@grtoriana)
(Aug 27, 2024 – Aug 27, 2026)
"""

import sys, os, json, time, re, csv
from datetime import datetime, timezone, timedelta
from collections import defaultdict, Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from curl_cffi import requests as cffi_requests

from scrape_bulk import make_session, extract_shortcode, shortcode_to_id, COOKIES

sys.stdout.reconfigure(encoding="utf-8")

session = make_session()

mob_hdrs = {
    "User-Agent": "Instagram 269.0.0.18.75 Android (26/8.0.0; 480dpi; 1080x1920; OnePlus; ONEPLUS A3003; OnePlus3; qcom; en_US; 314665256)",
    "x-ig-app-id": "936619743392459",
}

NOW_DT = datetime.now(timezone.utc)
CUTOFF_DT = NOW_DT - timedelta(days=730)
CUTOFF_TIMESTAMP = int(CUTOFF_DT.timestamp())

BRAND_INFO = {
    "name": "GRT Oriana",
    "username": "grtoriana",
    "state": "Pan-India / Tamil Nadu (HQ: Chennai, GRT Jewellers)"
}

INTERNAL_ACCOUNTS = {
    "grtoriana", "grt_oriana", "orianabygrt", "grtjewellers", "grt_jewellers", "grtjewellery", "grt_live"
}

print("="*75, flush=True)
print(f"GRT ORIANA 2-YEAR CREATOR AUDIT: {CUTOFF_DT.strftime('%Y-%m-%d')} to {NOW_DT.strftime('%Y-%m-%d')}", flush=True)
print("="*75, flush=True)

# 1. Resolve Profile
clean_u = "grtoriana"
url = f"https://i.instagram.com/api/v1/users/search/?q={clean_u}"
r = session.get(url, headers=mob_hdrs, cookies=COOKIES, timeout=12)
b_pk = None
b_name = "Oriana by GRT"
b_fols = 0

if r.status_code == 200:
    for u in r.json().get("users", []):
        if u.get("username", "").lower() == clean_u:
            b_pk = u.get("pk")
            b_name = u.get("full_name") or b_name
            b_fols = u.get("follower_count", 0)
            break

if not b_pk:
    # Try web profile
    s_web = session.get(f"https://www.instagram.com/{clean_u}/", timeout=12)
    m = re.search(r'"user_id":"(\d+)"', s_web.text) or re.search(r'"props":{"id":"(\d+)"', s_web.text) or re.search(r'"profile_id":"(\d+)"', s_web.text)
    if m:
        b_pk = m.group(1)

print(f"Resolved @{clean_u} -> PK: {b_pk} | Followers: {b_fols:,} | Name: {b_name}", flush=True)

# 2. Fetch 2-Year Feed & Clips
feed_items = []
max_id = ""
for p in range(1, 41):
    f_url = f"https://i.instagram.com/api/v1/feed/user/{b_pk}/"
    if max_id: f_url += f"?max_id={max_id}"
    try:
        r = session.get(f_url, headers=mob_hdrs, cookies=COOKIES, timeout=12)
        if r.status_code == 200:
            data = r.json()
            items = data.get("items", [])
            feed_items.extend(items)
            max_id = data.get("next_max_id")
            oldest_ts = min([it.get("taken_at", 0) for it in items if it.get("taken_at")], default=0)
            oldest_d = datetime.fromtimestamp(oldest_ts, tz=timezone.utc).strftime("%Y-%m-%d") if oldest_ts else "N/A"
            print(f"  [Feed] Page {p:>2}: {len(items)} items (Oldest in page: {oldest_d})", flush=True)
            if oldest_ts and oldest_ts < CUTOFF_TIMESTAMP:
                print(f"  -> Reached 2-year cutoff ({oldest_d}) in feed!", flush=True)
                break
            if not max_id or len(items) == 0:
                break
            time.sleep(0.35)
        else:
            break
    except Exception:
        break

clips_items = []
max_id = ""
for p in range(1, 41):
    c_url = "https://i.instagram.com/api/v1/clips/user/"
    payload = {"target_user_id": str(b_pk), "page_size": 30}
    if max_id: payload["max_id"] = str(max_id)
    try:
        r = session.post(c_url, headers=mob_hdrs, data=payload, cookies=COOKIES, timeout=12)
        if r.status_code == 200:
            data = r.json()
            clips = [it.get("media") for it in data.get("items", []) if it.get("media")]
            clips_items.extend(clips)
            paging = data.get("paging_info", {})
            max_id = paging.get("max_id")
            oldest_ts = min([it.get("taken_at", 0) for it in clips if it.get("taken_at")], default=0)
            oldest_d = datetime.fromtimestamp(oldest_ts, tz=timezone.utc).strftime("%Y-%m-%d") if oldest_ts else "N/A"
            print(f"  [Clips] Page {p:>2}: {len(clips)} reels (Oldest in page: {oldest_d})", flush=True)
            if oldest_ts and oldest_ts < CUTOFF_TIMESTAMP:
                print(f"  -> Reached 2-year cutoff ({oldest_d}) in clips!", flush=True)
                break
            if not paging.get("more_available") or not max_id:
                break
            time.sleep(0.35)
        else:
            break
    except Exception:
        break

seen_pks = set()
unique_items = []
for it in feed_items + clips_items:
    pk = str(it.get("pk") or it.get("id"))
    if pk and pk not in seen_pks:
        seen_pks.add(pk)
        unique_items.append(it)

print(f"\nTotal unique items fetched in 2 years: {len(unique_items)}", flush=True)

# 3. Extract Genuine Creator Collaborations
collab_posts = []
for it in unique_items:
    taken_at = it.get("taken_at")
    if not taken_at or taken_at < CUTOFF_TIMESTAMP:
        continue
        
    date_str = datetime.fromtimestamp(taken_at, tz=timezone.utc).strftime("%Y-%m-%d")
    owner = it.get("user", {}).get("username", "").lower()
    coauthors = [c.get("username", "").lower() for c in it.get("coauthor_producers", [])]
    is_paid = bool(it.get("is_paid_partnership", False))
    code = it.get("code") or ""
    post_url = f"https://www.instagram.com/p/{code}/" if code else ""
    cap_obj = it.get("caption") or {}
    cap_text = cap_obj.get("text", "") if isinstance(cap_obj, dict) else str(cap_obj)
    play_count = it.get("play_count") or it.get("view_count") or 0
    like_count = it.get("like_count") or 0
    comment_count = it.get("comment_count") or 0
    if not play_count and like_count:
        play_count = int(like_count * 18.5)
        
    creator_handle = ""
    if owner != clean_u:
        if owner not in INTERNAL_ACCOUNTS:
            creator_handle = f"@{owner}"
    elif coauthors:
        ext = [c for c in coauthors if c not in INTERNAL_ACCOUNTS and c != clean_u]
        if ext: creator_handle = f"@{ext[0]}"
        
    if creator_handle:
        collab_posts.append({
            "brand": BRAND_INFO["name"],
            "state": BRAND_INFO["state"],
            "handle": creator_handle,
            "raw_handle": creator_handle.replace("@", ""),
            "url": post_url,
            "shortcode": code,
            "media_id": str(it.get("pk") or ""),
            "date": date_str,
            "taken_at": taken_at,
            "is_paid_partnership": is_paid,
            "views": play_count,
            "likes": like_count,
            "comments": comment_count,
            "caption": cap_text[:250].replace("\n", " ").replace("\r", " ")
        })

print(f"Extracted {len(collab_posts)} genuine creator collaboration posts for GRT Oriana!\n", flush=True)

# 4. Resolve Creator Profile Metrics & Audience Sizing
unique_creators = sorted(list(set(p["raw_handle"].lower() for p in collab_posts)))
print(f"[+] Resolving profile metrics for {len(unique_creators)} unique creators...", flush=True)

def fetch_creator_opengraph(handle):
    s = cffi_requests.Session(impersonate="chrome120")
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }
    fols = 0; fing = 0; posts_cnt = 0; fn = handle; ver = False
    try:
        r = s.get(f"https://www.instagram.com/{handle}/", headers=headers, timeout=10)
        if r.status_code == 200:
            m = re.search(r'([0-9.,KMBkmb]+)\s+Followers,\s*([0-9.,KMBkmb]+)\s+Following,\s*([0-9.,KMBkmb]+)\s+Posts', r.text)
            if m:
                raw_fols = m.group(1).upper().replace(",", "")
                raw_fing = m.group(2).upper().replace(",", "")
                raw_posts = m.group(3).upper().replace(",", "")
                fols = int(float(raw_fols.replace("M", "")) * 1000000) if "M" in raw_fols else (int(float(raw_fols.replace("K", "")) * 1000) if "K" in raw_fols else int(float(raw_fols)))
                fing = int(float(raw_fing.replace("M", "")) * 1000000) if "M" in raw_fing else (int(float(raw_fing.replace("K", "")) * 1000) if "K" in raw_fing else int(float(raw_fing)))
                posts_cnt = int(float(raw_posts.replace("M", "")) * 1000000) if "M" in raw_posts else (int(float(raw_posts.replace("K", "")) * 1000) if "K" in raw_posts else int(float(raw_posts)))
            m_title = re.search(r'<title>([^(<]+)\s*\(@', r.text)
            if m_title:
                fn = m_title.group(1).strip()
    except Exception:
        pass
    return {
        "handle": f"@{handle}",
        "raw_handle": handle,
        "brands": "GRT Oriana",
        "full_name": fn,
        "followers": fols,
        "following": fing,
        "total_posts": posts_cnt,
        "verified": ver,
        "profile_url": f"https://www.instagram.com/{handle}/"
    }

profiles_data = []
with ThreadPoolExecutor(max_workers=10) as executor:
    futures = {executor.submit(fetch_creator_opengraph, h): h for h in unique_creators}
    done = 0
    for fut in as_completed(futures):
        res = fut.result()
        profiles_data.append(res)
        done += 1
        if done % 10 == 0 or done == len(unique_creators):
            print(f"  Progress: {done:>2}/{len(unique_creators)} profiles scraped", flush=True)

def get_pure_tier(followers):
    if followers >= 1000000: return "🌟 Mega Creator / Celebrity (1M+)"
    elif followers >= 100000: return "🚀 Macro Creator (100K - 1M)"
    elif followers >= 50000: return "✨ Mid-Tier Creator (50K - 100K)"
    elif followers >= 10000: return "🎯 Micro Creator (10K - 50K)"
    else: return "🌱 Nano Creator (<10K)"

# Compute profile stats
creator_posts_map = defaultdict(list)
for p in collab_posts:
    creator_posts_map[p["raw_handle"].lower()].append(p)

for p in profiles_data:
    rh = p["raw_handle"].lower()
    p["creator_tier"] = get_pure_tier(p["followers"])
    c_posts = creator_posts_map.get(rh, [])
    if c_posts:
        likes_l = [cp["likes"] for cp in c_posts]
        comms_l = [cp["comments"] for cp in c_posts]
        p["avg_likes"] = int(sum(likes_l) / len(likes_l))
        p["avg_comments"] = int(sum(comms_l) / len(comms_l))
        p["avg_er"] = round(((p["avg_likes"] + p["avg_comments"]) / p["followers"]) * 100, 2) if p["followers"] > 0 else 0.0
    else:
        p["avg_likes"] = 0; p["avg_comments"] = 0; p["avg_er"] = 0.0

profiles_data.sort(key=lambda x: x["followers"], reverse=True)

with open("grt_oriana_creators_profile_metrics.json", "w", encoding="utf-8") as f:
    json.dump(profiles_data, f, indent=2)

print(f"✓ Saved grt_oriana_creators_profile_metrics.json ({len(profiles_data)} profiles)", flush=True)

# 5. NLP Video Content Genre Classification & 4-Tier Evaluation
def detect_jewellery_video_genre(handle, caption):
    h = handle.lower()
    c = caption.lower()
    
    if any(w in c for w in ["styling", "how to style", "outfit", "ootd", "lookbook", "grwm", "fit check", "drip", "wardrobe"]):
        return "👗 Jewellery Styling & OOTD / GRWM"
    if any(w in c for w in ["bridal", "wedding", "bride", "trousseau", "haldi", "mehendi", "sangeet", "muhuratham", "reception"]):
        return "👰 Bridal & Festive Jewellery Edit"
    if any(w in c for w in ["unboxing", "unbox", "try on", "haul", "first look", "collection review", "packaging", "delivery"]):
        return "📦 Unboxing, Try-On & Product Review"
    if any(w in c for w in ["diamond", "gold", "craft", "design", "hallmarked", "certified", "solitaire", "lightweight", "workwear", "everyday"]):
        return "💎 Modern Gold & Diamond Craft Lore"
    if any(w in c for w in ["store", "showroom", "launch", "visit", "shopping", "walkthrough", "outlet", "chennai", "experience"]):
        return "🏬 Store Walkthrough & Showroom Experience"
    if any(w in c for w in ["celebrity", "actor", "ambassador", "exclusive", "campaign", "tvc", "dvc"]):
        return "🌟 Celebrity & Ambassador Campaign"
    if any(w in c for w in ["festival", "akshaya tritiya", "diwali", "varamahalakshmi", "pongal", "dhanteras", "offer", "discount"]):
        return "🎉 Festival & Auspicious Muhurat Drops"
    return "✨ Contemporary Jewellery & Lifestyle"

prof_map = {p["raw_handle"].lower(): p for p in profiles_data}

for p in collab_posts:
    rh = p["raw_handle"].lower()
    prof = prof_map.get(rh, {})
    fols = prof.get("followers", 0)
    p["followers"] = fols
    
    views = p["views"]
    likes = p["likes"]
    comments = p["comments"]
    cap_lower = p["caption"].lower()
    
    has_disclosure = any(t in cap_lower for t in ["#ad", "#paidpartnership", "#sponsored", "#collab", "#brandpartner", "paid partnership"])
    is_formal_paid = p.get("is_paid_partnership", False) or has_disclosure
    
    like_rate = round((likes / views) * 100, 2) if views > 0 else 0.0
    view_mult = round(views / fols, 2) if fols > 0 else 0.0
    er = round(((likes + comments) / fols) * 100, 2) if fols > 0 else 0.0
    
    p["like_rate_pct"] = like_rate
    p["view_multiplier"] = view_mult
    p["er_pct"] = er
    p["video_genre"] = detect_jewellery_video_genre(p["handle"], p["caption"])
    
    is_boosted = False
    if views >= 500000 and like_rate < 0.35:
        is_boosted = True
        boost_status = "🚀 Heavily Boosted (Paid Ad Spend)"
        reason = f"High view count ({views:,}) with sub-0.35% like rate ({like_rate}%) indicates paid video ads campaign"
    elif view_mult >= 5.0 and like_rate < 0.70:
        is_boosted = True
        boost_status = "🚀 Boosted (Paid Ad Spend)"
        reason = f"High view multiplier ({view_mult}x followers) combined with low engagement rate ({like_rate}%)"
    elif view_mult >= 3.0 and like_rate < 1.00 and views >= 80000:
        is_boosted = True
        boost_status = "🔍 Likely Boosted (Targeted Ad)"
        reason = f"Disproportionate views ({views:,}) relative to likes ({likes:,})"
    elif er >= 4.0 and like_rate >= 2.00:
        is_boosted = False
        boost_status = "📈 Viral Organic Reach"
        reason = f"High organic views ({views:,}) with strong like rate ({like_rate}%)"
    else:
        is_boosted = False
        boost_status = "⚪ Standard Organic"
        reason = "Baseline organic collab reach"
        
    if is_formal_paid and is_boosted:
        tier = 1; tier_name = "Tier 1: Toggle ON + Boosted"
    elif is_formal_paid and not is_boosted:
        tier = 2; tier_name = "Tier 2: Toggle ON + Organic"
    elif not is_formal_paid and is_boosted:
        tier = 3; tier_name = "Tier 3: Toggle OFF + Boosted"
    else:
        tier = 4; tier_name = "Tier 4: Toggle OFF + Organic (Noise)"
        
    p["tier"] = tier
    p["tier_name"] = tier_name
    p["is_boosted"] = is_boosted
    p["boost_status"] = boost_status
    p["boost_reason"] = reason

collab_posts.sort(key=lambda x: (x["tier"], -x["views"]))

with open("grt_oriana_2year_4tier_dataset.json", "w", encoding="utf-8") as f:
    json.dump(collab_posts, f, indent=2)

print(f"✓ Saved grt_oriana_2year_4tier_dataset.json ({len(collab_posts)} posts)", flush=True)

# 6. Build Master Excel Workbook & CSVs
wb = openpyxl.Workbook()
wb.remove(wb.active)

font_title = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_norm = Font(name="Calibri", size=10, bold=False, color="000000")
font_mute = Font(name="Calibri", size=9, bold=False, color="5D6D7E")
font_link = Font(name="Calibri", size=10, bold=False, color="0563C1", underline="single")

fill_t1_banner = PatternFill("solid", fgColor="145A32")
fill_t1_row = PatternFill("solid", fgColor="D4EFDF")
font_t1_bold = Font(name="Calibri", size=10, bold=True, color="0E6251")
font_t1_link = Font(name="Calibri", size=10, bold=True, color="0B5345", underline="single")

fill_t2_banner = PatternFill("solid", fgColor="1E8449")
fill_t2_row = PatternFill("solid", fgColor="EAFAF1")
font_t2_bold = Font(name="Calibri", size=10, bold=True, color="196F3D")
font_t2_link = Font(name="Calibri", size=10, bold=False, color="145A32", underline="single")

fill_t3_banner = PatternFill("solid", fgColor="B7950B")
fill_t3_row = PatternFill("solid", fgColor="FEF9E7")
font_t3_bold = Font(name="Calibri", size=10, bold=True, color="7D6608")
font_t3_link = Font(name="Calibri", size=10, bold=False, color="9A7D0A", underline="single")

fill_t4_banner = PatternFill("solid", fgColor="566573")
fill_t4_row = PatternFill("solid", fgColor="FFFFFF")
font_t4_norm = Font(name="Calibri", size=10, bold=False, color="2C3E50")
font_t4_link = Font(name="Calibri", size=10, bold=False, color="2980B9", underline="single")

thin_line = Side(style="thin", color="D5D8DC")
border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

# Tab 1: Executive Summary
ws_sum = wb.create_sheet("Executive Summary")
ws_sum.sheet_view.showGridLines = True

ws_sum.merge_cells("A1:O1")
ws_sum["A1"] = "Executive Summary — 2-Year Paid Creator Collab Hierarchy (GRT Oriana) [Aug 2024 – Aug 2026]"
ws_sum["A1"].font = font_title
ws_sum["A1"].fill = PatternFill("solid", fgColor="0B2240")
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 32

sum_headers = [
    ("#", 5),
    ("Brand Name", 24),
    ("State / Origin (HQ)", 35),
    ("Total Collab Posts (2-Yr)", 18),
    ("Total Unique Creators", 18),
    ("🟢 Tier 1: Toggle ON + Boosted\n(Posts / Creators)", 24),
    ("🟢 Tier 2: Toggle ON + Organic\n(Posts / Creators)", 24),
    ("🚀 Tier 3: Toggle OFF + Boosted\n(Posts / Creators)", 24),
    ("⚪ Tier 4: Toggle OFF + Organic (Noise)\n(Posts / Creators)", 28),
    ("💎 Total High-Intent Paid\n(Tiers 1+2+3 Posts)", 22),
    ("High-Intent Paid %\n(Tiers 1+2+3)", 18),
    ("Avg Views / Post", 16),
    ("Avg Creator Followers", 20),
    ("Avg Creator ER%", 15),
    ("Top Creator / Ambassador Samples", 48)
]

for col_idx, (h_text, w) in enumerate(sum_headers, 1):
    c = ws_sum.cell(row=2, column=col_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="1B2631")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_cell
    ws_sum.column_dimensions[get_column_letter(col_idx)].width = w
ws_sum.row_dimensions[2].height = 36
ws_sum.freeze_panes = "A3"

tot_p = len(collab_posts)
tot_c = len(unique_creators)
t1_posts = [p for p in collab_posts if p["tier"] == 1]
t2_posts = [p for p in collab_posts if p["tier"] == 2]
t3_posts = [p for p in collab_posts if p["tier"] == 3]
t4_posts = [p for p in collab_posts if p["tier"] == 4]
high_intent = len(t1_posts) + len(t2_posts) + len(t3_posts)
high_intent_pct = high_intent / tot_p if tot_p > 0 else 0.0
views_list = [p["views"] for p in collab_posts if p["views"] > 0]
avg_v = int(sum(views_list) / len(views_list)) if views_list else 0
fols_list = [p["followers"] for p in collab_posts if p["followers"] > 0]
avg_f = int(sum(fols_list) / len(fols_list)) if fols_list else 0
ers_list = [p["er_pct"] for p in collab_posts if p["er_pct"] > 0]
avg_e = round(sum(ers_list) / len(ers_list), 2) if ers_list else 0.0
top_c = list(dict.fromkeys([p["handle"] for p in collab_posts if p["tier"] in (1, 2, 3)]))[:5]
if not top_c: top_c = list(dict.fromkeys([p["handle"] for p in collab_posts]))[:5]

sum_vals = [
    1,
    BRAND_INFO["name"],
    BRAND_INFO["state"],
    tot_p,
    tot_c,
    f"{len(t1_posts)} posts ({len(set(p['handle'].lower() for p in t1_posts))} creators)" if t1_posts else "—",
    f"{len(t2_posts)} posts ({len(set(p['handle'].lower() for p in t2_posts))} creators)" if t2_posts else "—",
    f"{len(t3_posts)} posts ({len(set(p['handle'].lower() for p in t3_posts))} creators)" if t3_posts else "—",
    f"{len(t4_posts)} posts ({len(set(p['handle'].lower() for p in t4_posts))} creators)" if t4_posts else "—",
    high_intent,
    high_intent_pct,
    avg_v,
    avg_f,
    avg_e / 100 if avg_e else 0.0,
    ", ".join(top_c)
]

for col_idx, val in enumerate(sum_vals, 1):
    cell = ws_sum.cell(row=3, column=col_idx, value=val)
    cell.border = border_cell
    if col_idx == 1: cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
    elif col_idx == 2: cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
    elif col_idx == 3: cell.font = Font(name="Calibri", size=9, bold=True, color="2C3E50"); cell.alignment = Alignment(horizontal="left", vertical="center"); cell.fill = PatternFill("solid", fgColor="F4F6F6")
    elif col_idx in (4, 5): cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"; cell.fill = PatternFill("solid", fgColor="EBF5FB")
    elif col_idx == 6:
        cell.font = font_t1_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
        if "posts" in str(val): cell.fill = fill_t1_row
    elif col_idx == 7:
        cell.font = font_t2_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
        if "posts" in str(val): cell.fill = fill_t2_row
    elif col_idx == 8:
        cell.font = font_t3_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
        if "posts" in str(val): cell.fill = fill_t3_row
    elif col_idx == 9:
        cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
        if "posts" in str(val): cell.fill = PatternFill("solid", fgColor="F8F9F9")
    elif col_idx == 10: cell.font = Font(name="Calibri", size=10, bold=True, color="1B4F72"); cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"; cell.fill = PatternFill("solid", fgColor="D6EAF8")
    elif col_idx == 11: cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.0%"
    elif col_idx in (12, 13): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
    elif col_idx == 14: cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
    elif col_idx == 15: cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
ws_sum.row_dimensions[3].height = 24

# Tab 2: Creators Profile Metrics
ws_prof = wb.create_sheet("Creators Profile Metrics", index=1)
ws_prof.sheet_view.showGridLines = True

ws_prof.merge_cells("A1:M1")
ws_prof["A1"] = f"Deduped Creator Profiles & Tier Classification ({len(profiles_data)} Creators for GRT Oriana)"
ws_prof["A1"].font = font_title
ws_prof["A1"].fill = PatternFill("solid", fgColor="1B4F72")
ws_prof["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_prof.row_dimensions[1].height = 30

prof_headers = [
    ("#", 5),
    ("Creator Handle", 22),
    ("Creator Tier / Size", 30),
    ("Brands Collaborated With", 26),
    ("Full Name", 26),
    ("Verified", 10),
    ("Business / Pro", 14),
    ("Total Followers", 16),
    ("Following", 12),
    ("Total Posts", 14),
    ("Avg Likes / Post", 16),
    ("Avg Comments / Post", 18),
    ("Avg Profile ER%", 14),
]

for col_idx, (h_text, w) in enumerate(prof_headers, 1):
    c = ws_prof.cell(row=2, column=col_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="283747")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_cell
    ws_prof.column_dimensions[get_column_letter(col_idx)].width = w
ws_prof.row_dimensions[2].height = 25
ws_prof.freeze_panes = "A3"

tier_fills = {
    "🌟 Mega Creator / Celebrity (1M+)": PatternFill("solid", fgColor="E8F8F5"),
    "🚀 Macro Creator (100K - 1M)": PatternFill("solid", fgColor="FEF9E7"),
    "✨ Mid-Tier Creator (50K - 100K)": PatternFill("solid", fgColor="EBF5FB"),
    "🎯 Micro Creator (10K - 50K)": PatternFill("solid", fgColor="F4F6F7"),
    "🌱 Nano Creator (<10K)": PatternFill("solid", fgColor="FFFFFF"),
}

for idx, p in enumerate(profiles_data, 1):
    r_num = idx + 2
    r_vals = [
        idx,
        p["handle"],
        p["creator_tier"],
        p["brands"],
        p["full_name"],
        "Yes" if p["verified"] else "No",
        "Yes" if p.get("is_business") else "No",
        p["followers"],
        p["following"],
        p["total_posts"],
        p.get("avg_likes", 0),
        p.get("avg_comments", 0),
        p.get("avg_er", 0.0) / 100
    ]
    tier_fill = tier_fills.get(p["creator_tier"], PatternFill("solid", fgColor="FFFFFF"))
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws_prof.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2: cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = p["profile_url"]
        elif c_idx == 3: cell.font = Font(name="Calibri", size=10, bold=True, color="1B4F72"); cell.alignment = Alignment(horizontal="left", vertical="center"); cell.fill = tier_fill
        elif c_idx in (4, 5): cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (6, 7):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
            if val == "Yes": cell.fill = PatternFill("solid", fgColor="EAFAF1")
        elif c_idx in (8, 9, 10, 11, 12): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx == 13: cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
    ws_prof.row_dimensions[r_num].height = 21

# Tab 3: All 4-Tier Collaborations
ws_all = wb.create_sheet("GRT Oriana - Master Hierarchy")
ws_all.sheet_view.showGridLines = True

ws_all.merge_cells("A1:O1")
ws_all["A1"] = f"💍 GRT ORIANA  |  📍 STATE / REGION: {BRAND_INFO['state']}"
ws_all["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
ws_all["A1"].fill = PatternFill("solid", fgColor="0B2240")
ws_all["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_all.row_dimensions[1].height = 28

ws_all.merge_cells("A2:O2")
ws_all["A2"] = f"2-Year Collab Posts: {tot_p}  •  Unique Creators: {tot_c}  •  💎 High-Intent Paid: {high_intent} (T1: {len(t1_posts)} | T2: {len(t2_posts)} | T3: {len(t3_posts)})  •  ⚪ Noise/Unboosted (T4): {len(t4_posts)}"
ws_all["A2"].font = Font(name="Calibri", size=10, bold=True, color="1B4F72")
ws_all["A2"].fill = PatternFill("solid", fgColor="EBF5FB")
ws_all["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_all.row_dimensions[2].height = 22

table_cols = [
    ("#", 5),
    ("Hierarchy Tier", 30),
    ("Brand Name", 20),
    ("Creator Handle", 24),
    ("Followers", 14),
    ("Video Content Genre", 35),
    ("Views / Plays", 16),
    ("Likes", 14),
    ("Comments", 12),
    ("Like-to-View %", 15),
    ("Creator ER%", 13),
    ("Post Date", 13),
    ("Direct Instagram URL", 48),
    ("Boost Classification & Reason", 38),
    ("Caption Preview", 65)
]

for col_idx, (h_text, w) in enumerate(table_cols, 1):
    c = ws_all.cell(row=3, column=col_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="1F2D3D")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_cell
    ws_all.column_dimensions[get_column_letter(col_idx)].width = w
ws_all.row_dimensions[3].height = 25
ws_all.freeze_panes = "A4"

records_by_tier = {1: [], 2: [], 3: [], 4: []}
for r in collab_posts: records_by_tier[r["tier"]].append(r)

current_row = 4
global_index = 1

tier_meta = [
    (1, "🟢 TIER 1: TOGGLE ON + 🚀 BOOSTED (Formal Paid Partnership Label + Paid Ad Spend)", fill_t1_banner, fill_t1_row, font_t1_bold, font_t1_link),
    (2, "🟢 TIER 2: TOGGLE ON + ⚪ ORGANIC (Formal Paid Partnership Label + Organic Reach Only)", fill_t2_banner, fill_t2_row, font_t2_bold, font_t2_link),
    (3, "🚀 TIER 3: TOGGLE OFF + 🚀 BOOSTED (Co-Author Collab + Heavy Paid Ad Spend Detected)", fill_t3_banner, fill_t3_row, font_t3_bold, font_t3_link),
    (4, "⚪ TIER 4: TOGGLE OFF + ⚪ ORGANIC (Standard Collab / Low Organic Reach / Noise)", fill_t4_banner, fill_t4_row, font_t4_norm, font_t4_link)
]

for t_id, banner_text, fill_banner, fill_row, font_b, font_l in tier_meta:
    t_records = records_by_tier[t_id]
    if not t_records: continue
    ws_all.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
    banner_cell = ws_all.cell(row=current_row, column=1, value=f"{banner_text} — {len(t_records)} Posts ({len(set(p['handle'].lower() for p in t_records))} Unique Creators)")
    banner_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    banner_cell.fill = fill_banner
    banner_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_all.row_dimensions[current_row].height = 23
    current_row += 1
    
    for p in t_records:
        vals = [
            global_index,
            p["tier_name"],
            p["brand"],
            p["handle"],
            p["followers"],
            p["video_genre"],
            p["views"],
            p["likes"],
            p["comments"],
            p["like_rate_pct"] / 100 if p["like_rate_pct"] else 0.0,
            p["er_pct"] / 100 if p["er_pct"] else 0.0,
            p["date"],
            p["url"],
            f"{p['boost_status']}: {p['boost_reason']}" if p.get('boost_reason') else p['boost_status'],
            p["caption"]
        ]
        for c_idx, val in enumerate(vals, 1):
            cell = ws_all.cell(row=current_row, column=c_idx, value=val)
            cell.border = border_cell
            cell.fill = fill_row
            if c_idx == 1: cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx == 2: cell.font = font_b; cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in (3, 4): cell.font = font_b; cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx == 5: cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
            elif c_idx == 6: cell.font = Font(name="Calibri", size=10, bold=True, color="1A5276"); cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx in (7, 8, 9): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
            elif c_idx in (10, 11): cell.font = font_b; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
            elif c_idx == 12: cell.font = font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx == 13: cell.font = font_l; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None
            elif c_idx == 14: cell.font = font_b; cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx == 15: cell.font = font_mute; cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_all.row_dimensions[current_row].height = 21
        current_row += 1
        global_index += 1

wb.save("grt_oriana_2year_master_analysis.xlsx")
print("\n✅ Saved grt_oriana_2year_master_analysis.xlsx", flush=True)

# 7. CSV Exports
with open("GRT_Oriana_Creator_Profile_Metrics.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow([
        "#", "Creator Handle", "Creator Tier / Size", "Brands Collaborated With", "Full Name",
        "Verified", "Business / Professional", "Total Followers", "Following", "Total Posts",
        "Avg Likes / Post", "Avg Comments / Post", "Avg Profile ER%", "Instagram Profile URL"
    ])
    for idx, p in enumerate(profiles_data, 1):
        w.writerow([
            idx,
            p["handle"],
            p["creator_tier"],
            p["brands"],
            p["full_name"],
            "Yes" if p["verified"] else "No",
            "Yes" if p.get("is_business") else "No",
            p["followers"],
            p["following"],
            p["total_posts"],
            p.get("avg_likes", 0),
            p.get("avg_comments", 0),
            f"{p.get('avg_er', 0.0):.2f}%",
            p["profile_url"]
        ])
print("✓ Saved GRT_Oriana_Creator_Profile_Metrics.csv", flush=True)

with open("GRT_Oriana_All_4Tier_Master.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow([
        "#", "Hierarchy Tier", "Brand Name", "State / Origin (HQ)", "Creator Handle", "Followers",
        "Video Content Genre", "Views / Plays", "Likes", "Comments", "Like-to-View %", "Creator ER%",
        "Post Date", "Direct Instagram URL", "Boost Status & Reason", "Caption Preview"
    ])
    for idx, p in enumerate(collab_posts, 1):
        w.writerow([
            idx,
            p["tier_name"],
            p["brand"],
            p["state"],
            p["handle"],
            p["followers"],
            p["video_genre"],
            p["views"],
            p["likes"],
            p["comments"],
            f"{p['like_rate_pct']:.2f}%",
            f"{p['er_pct']:.2f}%",
            p["date"],
            p["url"],
            f"{p['boost_status']}: {p['boost_reason']}" if p.get('boost_reason') else p['boost_status'],
            p["caption"]
        ])
print("✓ Saved GRT_Oriana_All_4Tier_Master.csv", flush=True)

with open("GRT_Oriana_Summary.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow([
        "#", "Brand Name", "State / Origin (HQ)", "Total Collab Posts (2-Yr)", "Total Unique Creators",
        "Tier 1 (Toggle ON + Boosted)", "Tier 2 (Toggle ON + Organic)",
        "Tier 3 (Toggle OFF + Boosted)", "Tier 4 (Noise / Unboosted)",
        "Total High-Intent Paid Posts", "High-Intent Paid %",
        "Avg Estimated Views", "Avg Creator Followers", "Avg Creator ER%", "Top Creator Samples"
    ])
    w.writerow([
        1,
        BRAND_INFO["name"],
        BRAND_INFO["state"],
        tot_p,
        tot_c,
        f"{len(t1_posts)} posts ({len(set(p['handle'].lower() for p in t1_posts))} creators)" if t1_posts else "—",
        f"{len(t2_posts)} posts ({len(set(p['handle'].lower() for p in t2_posts))} creators)" if t2_posts else "—",
        f"{len(t3_posts)} posts ({len(set(p['handle'].lower() for p in t3_posts))} creators)" if t3_posts else "—",
        f"{len(t4_posts)} posts ({len(set(p['handle'].lower() for p in t4_posts))} creators)" if t4_posts else "—",
        high_intent,
        f"{high_intent_pct*100:.1f}%",
        avg_v,
        avg_f,
        f"{avg_e:.2f}%",
        ", ".join(top_c)
    ])
print("✓ Saved GRT_Oriana_Summary.csv\n", flush=True)
