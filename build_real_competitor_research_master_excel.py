"""
Build Comprehensive Master Excel from Real Instagram Comments and Real Meta Ads:
Workbook: lakeshore_real_competitor_research_master.xlsx
"""

import sys, json, os, openpyxl
from collections import defaultdict
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# 1. Load Real Comments Dataset (1,524 Comments)
with open("real_mall_comments_dataset.json", encoding="utf-8") as f:
    comments_data = json.load(f)

# 2. Load Real Meta Ads Dataset (425 Ads)
with open("real_mall_meta_ads_dataset.json", encoding="utf-8") as f:
    meta_ads_data = json.load(f)

# 3. Load 1-Year Malls Dataset (654 Collabs / 3,668 Posts)
with open("pune_hyderabad_malls_1year_dataset.json", encoding="utf-8") as f:
    malls_dataset = json.load(f)

comments_list = comments_data["comments"]
ads_list = meta_ads_data["ads"]
malls_results = malls_dataset["malls_results"]

print(f"Loaded {len(comments_list)} Real Comments and {len(ads_list)} Real Meta Ads.")

# Aggregate Real Comments Sentiment per Mall
mall_comments_stats = defaultdict(lambda: defaultdict(int))
for c in comments_list:
    m = c["mall_name"]
    mall_comments_stats[m]["total_comments"] += 1
    cat = c["intent_category"]
    mall_comments_stats[m][cat] += 1
    sent = c["sentiment"]
    mall_comments_stats[m][sent] += 1

# Aggregate Meta Ads Stats per Mall
mall_ads_stats = defaultdict(lambda: {"active": 0, "inactive": 0, "total": 0, "ctas": defaultdict(int)})
for ad in ads_list:
    m = ad["target_mall"]
    mall_ads_stats[m]["total"] += 1
    if ad["is_active"]:
        mall_ads_stats[m]["active"] += 1
    else:
        mall_ads_stats[m]["inactive"] += 1
    mall_ads_stats[m]["ctas"][ad.get("cta", "Learn More")] += 1

# Workbook Setup
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Styles
font_title = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_norm = Font(name="Calibri", size=10, bold=False, color="000000")
font_link = Font(name="Calibri", size=10, bold=False, color="0563C1", underline="single")

thin_line = Side(style="thin", color="D5D8DC")
border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

fill_navy = PatternFill("solid", fgColor="0B2240")
fill_dark = PatternFill("solid", fgColor="1B2631")
fill_friction = PatternFill("solid", fgColor="FADBD8") # Soft Red
fill_intent = PatternFill("solid", fgColor="D4EFDF") # Soft Mint Green
fill_positive = PatternFill("solid", fgColor="EAFAF1") # Soft Green
fill_active = PatternFill("solid", fgColor="D4EFDF") # Mint Green


# ==============================================================================
# TAB 1: EXECUTIVE RESEARCH DASHBOARD
# ==============================================================================
ws_sum = wb.create_sheet("Executive Research Dashboard")
ws_sum.sheet_view.showGridLines = True
ws_sum.merge_cells("A1:L1")
ws_sum["A1"] = "Executive Research Dashboard — Real Customer Comments & Meta Ad Intelligence (10 Competitors vs KOPA Mall)"
ws_sum["A1"].font = font_title
ws_sum["A1"].fill = fill_navy
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 30

sum_headers = [
    ("#", 5), ("Mall Name", 32), ("City", 12), ("Real Comments Scraped", 18),
    ("🚗 Parking & Traffic Friction", 22), ("👥 Overcrowding & Queue Friction", 24),
    ("🛍️ Store / Brand Queries", 20), ("🎟️ Event / Pass Queries", 20),
    ("🍽️ Dining Inquiries", 18), ("Live Active Meta Ads", 18),
    ("Total Meta Ads Scraped", 18), ("Top Ad Call-to-Action (CTA)", 24)
]
for c_idx, (h_text, w) in enumerate(sum_headers, 1):
    c = ws_sum.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_sum.column_dimensions[get_column_letter(c_idx)].width = w
ws_sum.row_dimensions[2].height = 28
ws_sum.freeze_panes = "A3"

all_malls_list = [
    ("Phoenix Avenue of Stars Pune", "Pune"),
    ("Phoenix Mall of the Millennium Wakad", "Pune"),
    ("The Pavillion Pune", "Pune"),
    ("Seasons Mall Pune", "Pune"),
    ("Amanora Mall Pune", "Pune"),
    ("KOPA Mall Pune", "Pune"),
    ("Lulu Mall Hyderabad", "Hyderabad"),
    ("Nexus Hyderabad Mall", "Hyderabad"),
    ("Sarath City Capital Mall", "Hyderabad"),
    ("Inorbit Mall Cyberabad", "Hyderabad"),
    ("GVK One Mall Hyderabad", "Hyderabad"),
]

for idx, (m_name, city) in enumerate(all_malls_list, 1):
    r_num = idx + 2
    c_stats = mall_comments_stats[m_name]
    a_stats = mall_ads_stats[m_name]
    
    top_cta = max(a_stats["ctas"].keys(), key=lambda k: a_stats["ctas"][k]) if a_stats["ctas"] else "Learn More"
    
    vals = [
        idx, m_name, city, c_stats["total_comments"],
        c_stats["🚗 Parking & Traffic Friction"], c_stats["👥 Overcrowding & Queue Friction"],
        c_stats["🛍️ Store & Brand Location Query"], c_stats["🎟️ Event Ticket & Entry Inquiry"],
        c_stats["🍽️ Dining & F&B Inquiry"], a_stats["active"], a_stats["total"], top_cta
    ]
    
    for c_idx, val in enumerate(vals, 1):
        cell = ws_sum.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2: cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (3, 12): cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in range(4, 12):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
            if c_idx in (5, 6) and val > 0: cell.fill = fill_friction
            elif c_idx in (7, 8, 9) and val > 0: cell.fill = fill_intent
            elif c_idx == 10 and val > 0: cell.fill = fill_active

    ws_sum.row_dimensions[r_num].height = 20


# ==============================================================================
# TAB 2: REAL CUSTOMER COMMENTS STREAM (1,524 COMMENTS)
# ==============================================================================
ws_comm = wb.create_sheet("Real Customer Comments (1,524)")
ws_comm.sheet_view.showGridLines = True
ws_comm.merge_cells("A1:K1")
ws_comm["A1"] = f"Raw Verified Instagram Comments Stream — Real User Feedback ({len(comments_list)} Scraped Comments)"
ws_comm["A1"].font = font_title
ws_comm["A1"].fill = fill_navy
ws_comm["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_comm.row_dimensions[1].height = 30

comm_headers = [
    ("#", 5), ("City", 12), ("Mall Name", 30), ("Commenter Handle", 22),
    ("Real Scraped Comment Text", 65), ("Likes on Comment", 15), ("Timestamp", 18),
    ("Intent / Topic Category", 30), ("Sentiment", 22), ("Intent Type", 24), ("Reel URL", 40)
]
for c_idx, (h_text, w) in enumerate(comm_headers, 1):
    c = ws_comm.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_comm.column_dimensions[get_column_letter(c_idx)].width = w
ws_comm.row_dimensions[2].height = 26
ws_comm.freeze_panes = "A3"

for idx, c in enumerate(comments_list, 1):
    r_num = idx + 2
    r_vals = [
        idx, c["city"], c["mall_name"], c["username"], c["text"], c["likes"],
        c["date"], c["intent_category"], c["sentiment"], c["intent_type"], c["reel_url"]
    ]
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws_comm.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (2, 7): cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (3, 4): cell.font = font_bold if c_idx == 4 else font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 5: cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 6: cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx in (8, 9, 10):
            cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
            if "Negative" in str(val) or "Friction" in str(val) or "Pain" in str(val): cell.fill = fill_friction
            elif "Intent" in str(val) or "Query" in str(val): cell.fill = fill_intent
            elif "Positive" in str(val) or "Proof" in str(val): cell.fill = fill_positive
        elif c_idx == 11: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None

    ws_comm.row_dimensions[r_num].height = 20


# ==============================================================================
# TAB 3: REAL FRICTION & PAIN-POINT MINING
# ==============================================================================
ws_fric = wb.create_sheet("Friction & Pain-Point Mining")
ws_fric.sheet_view.showGridLines = True
ws_fric.merge_cells("A1:H1")

friction_comments = [c for c in comments_list if "Negative" in c.get("sentiment", "") or "Friction" in c.get("intent_category", "") or "Complaint" in c.get("intent_category", "")]
ws_fric["A1"] = f"Real Customer Friction Quotes & Lake Shore Counter-Strategy ({len(friction_comments)} Friction Comments)"
ws_fric["A1"].font = font_title
ws_fric["A1"].fill = fill_navy
ws_fric["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_fric.row_dimensions[1].height = 30

fric_headers = [
    ("#", 5), ("Competitor Mall", 30), ("Friction Category", 28),
    ("Real Scraped User Comment Quote", 65), ("Commenter", 20),
    ("Likes", 10), ("KOPA Mall Advantage", 35), ("Counter Paid Ad Hook for KOPA", 50)
]
for c_idx, (h_text, w) in enumerate(fric_headers, 1):
    c = ws_fric.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_fric.column_dimensions[get_column_letter(c_idx)].width = w
ws_fric.row_dimensions[2].height = 26
ws_fric.freeze_panes = "A3"

for idx, c in enumerate(friction_comments, 1):
    r_num = idx + 2
    cat = c["intent_category"]
    
    if "Parking" in cat:
        adv = "Dedicated Valet Parking & Zero Traffic Congestion in KP"
        hook = "\"Skip the 40-minute parking nightmare. Effortless valet parking & luxury shopping at KOPA Koregaon Park.\""
    elif "Overcrowding" in cat:
        adv = "Boutique, Tranquil & Curated Luxury Setting"
        hook = "\"Escape the weekend chaos. Discover refined shopping and fine dining in an intimate setting at KOPA.\""
    elif "Pricing" in cat:
        adv = "True Luxury & Genuine Value for HNIs"
        hook = "\"Where quality matches the experience. Explore curated high-end fashion and gastronomy at KOPA.\""
    else:
        adv = "Seamless Guest Experience & Fast Service"
        hook = "\"Designed for those who value their time. Experience seamless luxury at KOPA Mall.\""

    r_vals = [idx, c["mall_name"], cat, c["text"], c["username"], c["likes"], adv, hook]
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws_fric.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2: cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 3: cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.fill = fill_friction
        elif c_idx == 4: cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 5: cell.font = font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 6: cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx in (7, 8): cell.font = font_bold if c_idx == 7 else font_norm; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.fill = fill_intent
    ws_fric.row_dimensions[r_num].height = 24


# ==============================================================================
# TAB 4: SHOPPER HIGH-INTENT INQUIRIES
# ==============================================================================
ws_intent = wb.create_sheet("Shopper High-Intent Inquiries")
ws_intent.sheet_view.showGridLines = True
ws_intent.merge_cells("A1:H1")

intent_comments = [c for c in comments_list if "Intent" in c.get("sentiment", "") or "Query" in c.get("intent_category", "") or "Inquiry" in c.get("intent_category", "")]
ws_intent["A1"] = f"Real In-Market Shopper Queries — Brand Locations, Tickets & Dining ({len(intent_comments)} Inquiries)"
ws_intent["A1"].font = font_title
ws_intent["A1"].fill = fill_navy
ws_intent["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_intent.row_dimensions[1].height = 30

intent_headers = [
    ("#", 5), ("Mall Name", 30), ("Inquiry Category", 30),
    ("Real Shopper Question", 65), ("Commenter", 20),
    ("Likes", 10), ("Shopper Conversion Intent", 26), ("Reel Source Link", 40)
]
for c_idx, (h_text, w) in enumerate(intent_headers, 1):
    c = ws_intent.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_intent.column_dimensions[get_column_letter(c_idx)].width = w
ws_intent.row_dimensions[2].height = 26
ws_intent.freeze_panes = "A3"

for idx, c in enumerate(intent_comments, 1):
    r_num = idx + 2
    r_vals = [idx, c["mall_name"], c["intent_category"], c["text"], c["username"], c["likes"], c["sentiment"], c["reel_url"]]
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws_intent.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2: cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (3, 7): cell.font = font_bold; cell.alignment = Alignment(horizontal="left" if c_idx == 3 else "center", vertical="center"); cell.fill = fill_intent
        elif c_idx == 4: cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 5: cell.font = font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 6: cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx == 8: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None
    ws_intent.row_dimensions[r_num].height = 20


# ==============================================================================
# TAB 5: REAL META ADS REPOSITORY (425 ADS)
# ==============================================================================
ws_ads = wb.create_sheet("Real Meta Ads Repository (425)")
ws_ads.sheet_view.showGridLines = True
ws_ads.merge_cells("A1:J1")
ws_ads["A1"] = f"Real Meta Ad Library Repository — Competitors & KOPA Mall ({len(ads_list)} Total Scraped Ads)"
ws_ads["A1"].font = font_title
ws_ads["A1"].fill = fill_navy
ws_ads["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_ads.row_dimensions[1].height = 30

ads_headers = [
    ("#", 5), ("City", 12), ("Target Mall Name", 32), ("Advertiser Name", 28),
    ("Ad Library ID", 18), ("Ad Status", 14), ("Started Running Date", 22),
    ("Call-to-Action (CTA)", 20), ("Ad Copy / Primary Text", 65), ("Ad Library Link", 42)
]
for c_idx, (h_text, w) in enumerate(ads_headers, 1):
    c = ws_ads.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_ads.column_dimensions[get_column_letter(c_idx)].width = w
ws_ads.row_dimensions[2].height = 26
ws_ads.freeze_panes = "A3"

for idx, ad in enumerate(ads_list, 1):
    r_num = idx + 2
    r_vals = [
        idx, ad.get("city", "Pune"), ad["target_mall"], ad["advertiser"], ad["library_id"],
        "Active" if ad["is_active"] else "Inactive", ad["start_date"],
        ad.get("cta", "Learn More"), ad["body"], ad["ad_url"]
    ]
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws_ads.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (2, 5, 7, 8): cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (3, 4): cell.font = font_bold if c_idx == 3 else font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 6:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
            if val == "Active": cell.fill = fill_active
        elif c_idx == 9: cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 10: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None
    ws_ads.row_dimensions[r_num].height = 20

output_master = "lakeshore_real_competitor_research_master.xlsx"
wb.save(output_master)
print(f"\n✓ Master Workbook saved successfully: {output_master}")
