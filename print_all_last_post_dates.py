"""
Extract the exact latest post dates from both datasets:
1. Latest Partnership Post Date detected per Brand (from the brand audit logs)
2. Latest Creator Post Date sampled per Brand (from the creator profiles analysis)
"""

import openpyxl, json, sys

sys.stdout.reconfigure(encoding="utf-8")

def parse_workbook_collabs(filename):
    results = {}
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        if "Raw Partnership Posts" in wb.sheetnames:
            ws = wb["Raw Partnership Posts"]
            for r in range(3, ws.max_row + 1):
                b = ws.cell(row=r, column=1).value
                u = ws.cell(row=r, column=2).value
                d = ws.cell(row=r, column=3).value
                p = ws.cell(row=r, column=5).value
                c = ws.cell(row=r, column=6).value
                if b and d and str(d) != "None":
                    d_str = str(d)[:10]
                    if b not in results or d_str > results[b]["date"]:
                        results[b] = {
                            "date": d_str,
                            "url": u,
                            "partner": str(p),
                            "caption": str(c)[:60] if c else ""
                        }
    except Exception as e:
        print(f"Error reading {filename}: {e}")
    return results

print("=== 1. LUXURY / PEN BRANDS ===")
pens_collabs = parse_workbook_collabs("competitor_paid_media_analysis.xlsx")
for brand, data in pens_collabs.items():
    print(f"Brand: {brand:<18} | Latest Collab Post: {data['date']} | Partner: {data['partner']:<25} | URL: {data['url']}")

print("\n=== 2. ELECTRONICS RETAIL BRANDS ===")
retail_collabs = parse_workbook_collabs("retail_brands_competitor_analysis.xlsx")
for brand, data in retail_collabs.items():
    print(f"Brand: {brand:<18} | Latest Collab Post: {data['date']} | Partner: {data['partner']:<25} | URL: {data['url']}")

# 3. Creator Profile 20-post date ranges from cache
print("\n=== 3. CREATOR POOL LATEST POST DATES BY BRAND ===")
try:
    with open("brand_partnership_analysis_cache.json", encoding="utf-8") as f:
        pen_cache = json.load(f)
    for b, plist in pen_cache["brands"].items():
        dates = [pst["date"] for p in plist for pst in p.get("posts", []) if pst.get("date") and pst.get("date") != "N/A"]
        dates.sort(reverse=True)
        if dates:
            print(f"Pen Brand: {b:<15} | Latest Partner Activity: {dates[0]} | Sampled Posts: {len(dates)}")
except Exception as e:
    print(f"Note pen cache: {e}")

try:
    with open("retail_creators_cache.json", encoding="utf-8") as f:
        ret_cache = json.load(f)
    from scrape_retail_creators import CREATORS
    for b, ulist in CREATORS.items():
        dates = []
        for u in ulist:
            norm = u.strip().lstrip("@").lower()
            p = ret_cache.get(norm)
            if p and p.get("posts_count", 0) > 0:
                pass
except Exception as e:
    pass
