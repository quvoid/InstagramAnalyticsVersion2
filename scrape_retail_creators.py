"""
Instagram Creator Metrics Scraper & Excel Updater
for Retail Brands: Vijay Sales, Croma, Reliance Digital
=========================================================
"""

import sys, io, re, html, time, random, json, os
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# ── Brand Creator Lists ────────────────────────────────────────

CREATORS = {
    "VIJAY SALES": [
        "evalive.in", "agasthya.shah", "rakheevaswani", "saint_in_baggy", "ansh_rajpal_",
        "theuntamedlenss", "baadshahig", "thedeverakonda", "theblogginghogger", "hetthakkar_06",
        "riapednekar03", "bandedinacandid", "niftystretch", "bmp_online", "konshrutik",
        "shivank_panwanda", "elementec", "biswakalyanrath", "ravish_shetty_", "wokenupsid_",
        "techmumbaikar", "neha1584", "crazy_busy_mom", "bbuzzz08", "garimasgoodlife",
        "naisuonly", "techpocket", "gopihappy2", "aarzoo_dahiya26", "whattechno", "hyderabad_talkies"
    ],
    "CROMA": [
        "jaipunjabii", "niswoman", "flyingboysz", "soniyasolosign", "yashwanth_shettyy",
        "vyom25", "funcho", "astuti_mw", "cookininshort", "mihirjoshimusic",
        "whos.devuu", "kuaurek", "dhirajjjjj_", "jashwanth_bopanna", "kaizzzzlol",
        "justaiming", "rheagurnanii", "kanchi.sharma", "mohit_hiranandani93", "chirayu_m",
        "harshaisavailable341", "pari_dua", "smrutiandonkar", "sushant_ghadge_", "sara_nareen",
        "vishal_d_o_p", "vivekjadoo", "realkajalchauhann", "deeshakatkar", "naisuonly",
        "anubhaviiiiiii", "deepavedha", "rjdhruviiiii", "vaishali_and_anand", "rjkisnaa",
        "officialrohitsaluja", "gj_krish", "mahathalli", "ankur_agarwal_vines", "khushaal_pawaar",
        "or.junn", "_spindia_", "ajaywhines", "theuntamedlenss", "just_neel_things",
        "hustlingrajan", "chandnimimic", "thatkushalmistry", "voxhub.rehan", "thesunshineladki",
        "knishkk", "eattripclick", "himalaya.wala", "becausewhynotofficial", "vikshitha_v",
        "shettybrotherofficial", "focusedindian", "ijaybhanushali", "mavi.harman", "whereareyoupraveen",
        "divyagupta7811", "prannayjoshi", "rizzaalee", "cookwithajay", "sahilbulla",
        "antophilip", "tanejamainhoon", "amn.agrwl", "jogipet_ratnam", "maroofculmen",
        "kenavita", "aditinng", "vishakha_divesh", "bisuonthego", "deekay_boss",
        "avi7x._", "niralichoksi_", "nitesh_shetty99", "thevishnukaushal", "sani_trakari",
        "trulynomadly", "ramyaramapriya", "haram_khor_", "realsudeshlehri", "rjnaved",
        "its_jamielever", "viraj_ghelani", "chaitaliandharshil", "siddyshetty", "vikasmarwah.official",
        "instafunny_manan", "villagecookingboy_official", "lets_eat_with_prachi", "shreecookingacademy", "nikitarawlani",
        "coffeemonkk", "adityagoswami04", "ipsashahh", "alka_rajput_patil", "siddhant_sarfare",
        "mangaajjii", "aanalkotak"
    ],
    "RELIANCE DIGITAL": [
        "viraajita", "vivekjadoo", "pavanwaghulkar", "abhijeetkain", "purewalparamjit",
        "i._.gujarati", "ganeshkaranth", "saikatdey_24", "saikat_ishan_", "twishaaapatel",
        "muktigautam", "onkarrandhawa", "vijayviruz", "funny_pratik", "hey_bhagyashree",
        "_srishittt", "ravimantrii", "selexsj", "flavorsoflucknow", "mistertikku",
        "hyderabad.food.diaries", "juhigodambe", "sejalkumar1195", "therajivmakhni", "manasvivashist",
        "aditiprabhudeva", "parvathy_ayyappadas", "rejoy_thomas_lml", "akhil.jackson", "navneetrandhey",
        "wondermunna", "mangaajjii", "hariimuniyappan", "_spindia_", "krushna30",
        "katariaaryann", "arshgoyalyt", "jogipet_ratnam", "naisuonly", "vinesofmonu",
        "raghu_vinestore_official", "rahuldey", "thajmola", "sayliraut_", "thevanquishment",
        "beyounick", "ramesh_somani"
    ]
}

# ── Cookies & Headers ──────────────────────────────────────────

COOKIES = {
    "sessionid":  "25113411270%3AyQFaao428g6Xb9%3A0%3AAYjwS_M5UIDcS-i41tvdVFu8WKSqfzfgEhlZLGMvFg",
    "csrftoken":  "3gJbkGDZp99lA8QQ0brobyoHzOreuu8f",
    "mid":        "afyCbwALAAFRStE-k17-dfO5_jfa",
    "ds_user_id": "25113411270",

}

IG_APP_ID = "936619743392459"
BASE_HEADERS = {
    "accept":           "*/*",
    "accept-language":  "en-US,en;q=0.9",
    "user-agent":       (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "x-ig-app-id":      IG_APP_ID,
    "x-ig-www-claim":   "0",
    "x-requested-with": "XMLHttpRequest",
    "x-csrftoken":      COOKIES.get("csrftoken", ""),
    "origin":           "https://www.instagram.com",
    "referer":          "https://www.instagram.com/",
}

POSTS_PER_USER = 20
CACHE_FILE = "retail_creators_cache.json"
EXCEL_FILE = "retail_brands_competitor_analysis.xlsx"

try:
    from curl_cffi import requests as cffi_requests
    USE_CURL_CFFI = True
    print("✓ curl_cffi found — Chrome TLS fingerprint active")
except ImportError:
    import requests as cffi_requests
    USE_CURL_CFFI = False
    print("⚠ curl_cffi not found — using standard requests")


# ── Scraping Helpers ──────────────────────────────────────────

def make_session():
    if USE_CURL_CFFI:
        s = cffi_requests.Session(impersonate="chrome120")
    else:
        import requests
        s = requests.Session()
    s.headers.update(BASE_HEADERS)
    s.cookies.update(COOKIES)
    return s

def delay(a=1.2, b=2.5):
    time.sleep(random.uniform(a, b))

def fetch_profile_info(username: str, session) -> dict:
    url = f"https://www.instagram.com/api/v1/users/web_profile_info/?username={username}"
    hdrs = {**BASE_HEADERS, "referer": f"https://www.instagram.com/{username}/"}
    for attempt in range(1, 4):
        try:
            r = session.get(url, headers=hdrs, cookies=COOKIES, timeout=18)
            if r.status_code == 200:
                return r.json()
            elif r.status_code == 429:
                wait = 30 * attempt
                print(f"    ⏳ Rate-limited on @{username} — sleeping {wait}s...")
                time.sleep(wait)
            elif r.status_code in (401, 403):
                print(f"    ❌ Auth error ({r.status_code}) on @{username} — cookies may need refresh.")
                return {"_auth_error": r.status_code}
            elif r.status_code == 404:
                print(f"    ❌ @{username} not found (404).")
                return {"_not_found": 404}
            else:
                delay(2, 4)
        except Exception as e:
            print(f"    ⚠ Request error: {e}")
            delay(2, 4)
    return {}

def fetch_posts_v2(user_id: str, session, count=20) -> list:
    url = f"https://www.instagram.com/api/v1/feed/user/{user_id}/?count={count}"
    for attempt in range(1, 4):
        try:
            r = session.get(url, headers=BASE_HEADERS, cookies=COOKIES, timeout=18)
            if r.status_code == 200:
                return r.json().get("items", [])
            elif r.status_code == 429:
                time.sleep(30 * attempt)
            else:
                delay(2, 4)
        except Exception:
            delay(2, 4)
    return []

def scrape_one_creator(username: str, session) -> dict | None:
    raw = fetch_profile_info(username, session)
    if "_auth_error" in raw or "_not_found" in raw:
        return None
    user = raw.get("data", {}).get("user")
    if not user:
        return None

    followers = user.get("edge_followed_by", {}).get("count", 0)
    profile = {
        "username":    username,
        "full_name":   user.get("full_name", "N/A"),
        "followers":   followers,
        "following":   user.get("edge_follow", {}).get("count", 0),
        "total_posts": user.get("edge_owner_to_timeline_media", {}).get("count", 0),
        "verified":    user.get("is_verified", False),
        "is_business": user.get("is_business_account", False),
        "user_id":     user.get("id", ""),
        "avg_likes":   0,
        "avg_comments":0,
        "avg_er":      0.0,
        "posts_count": 0,
    }

    if profile["user_id"]:
        raw_posts = fetch_posts_v2(profile["user_id"], session, count=POSTS_PER_USER)
        total_l = 0
        total_c = 0
        post_items = raw_posts[:POSTS_PER_USER]
        for p in post_items:
            total_l += p.get("like_count", 0) or 0
            total_c += p.get("comment_count", 0) or 0
        n = len(post_items)
        profile["posts_count"] = n
        if n > 0:
            profile["avg_likes"]    = total_l // n
            profile["avg_comments"] = total_c // n
            profile["avg_er"]       = round((total_l + total_c) / n / followers * 100, 2) if followers else 0.0

    return profile


# ── Main Runner & Excel Updater ────────────────────────────────

def run_pipeline():
    # 1. Deduplicate usernames across all brands
    unique_map = {} # normalized_username -> original_username
    for brand, u_list in CREATORS.items():
        for u in u_list:
            norm = u.strip().lstrip("@").lower()
            if norm and norm not in unique_map:
                unique_map[norm] = u.strip().lstrip("@")

    total_unique = len(unique_map)
    print(f"\n{'═'*65}")
    print(f"  Retail Brands Creator Metrics Scraper")
    print(f"  Brands: {len(CREATORS)} | Unique Creators: {total_unique}")
    print(f"{'═'*65}\n")

    # 2. Load existing cache if any
    cache = {}
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                cache = json.load(f)
            print(f"✓ Loaded {len(cache)} existing cached creator profiles from {CACHE_FILE}")
        except Exception as e:
            print(f"⚠ Could not read cache: {e}")

    # 3. Scrape remaining
    session = make_session()
    to_scrape = [(norm, orig) for norm, orig in unique_map.items() if norm not in cache]
    print(f"  Creators to scrape now: {len(to_scrape)} / {total_unique}")

    for idx, (norm, orig) in enumerate(to_scrape, 1):
        print(f"[{idx}/{len(to_scrape)}] @{orig} ...", end=" ", flush=True)
        res = scrape_one_creator(orig, session)
        if res:
            cache[norm] = res
            print(f"✓ fol: {res['followers']:,} | likes: {res['avg_likes']:,} | com: {res['avg_comments']:,} | ER: {res['avg_er']:.2f}%")
        else:
            cache[norm] = None
            print(f"⚠ skipped / failed")

        # Save cache every 5 scrapes or on finish
        if idx % 5 == 0 or idx == len(to_scrape):
            with open(CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(cache, f, ensure_ascii=False, indent=2)

        delay(1.2, 2.4)

    # Final cache save
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

    valid_scrapes = sum(1 for v in cache.values() if v is not None)
    print(f"\n✅ Scraping finished. {valid_scrapes}/{total_unique} creators have metrics.")

    # 4. Update Excel Sheet
    print(f"\n📊 Updating '{EXCEL_FILE}' [Instagram Partners]...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Instagram Partners"]

    # Formatting styles matching existing sheet
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F2D3D")
    hdr_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    data_font = Font(name="Calibri", bold=False, size=10, color="000000")
    num_font  = Font(name="Calibri", bold=False, size=10, color="000000")
    er_font   = Font(name="Calibri", bold=True, size=10, color="000000")

    # Set Column Headers on row 2
    headers = {
        7: "Followers",
        8: "Avg Likes/Post",
        9: "Avg Comments/Post",
        10: "Avg ER%"
    }
    widths = {7: 15, 8: 15, 9: 15, 10: 12}

    for col_idx, h_text in headers.items():
        cell = ws.cell(row=2, column=col_idx, value=h_text)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx]

    updated_count = 0
    skipped_count = 0

    for row_idx in range(3, ws.max_row + 1):
        handle_raw = ws.cell(row=row_idx, column=2).value
        if not handle_raw:
            continue

        handle_str = str(handle_raw).strip()
        norm_h = handle_str.lstrip("@").lower()

        if norm_h in cache and cache[norm_h] is not None:
            p = cache[norm_h]
            # Col G: Followers
            cG = ws.cell(row=row_idx, column=7, value=p["followers"])
            cG.font = num_font
            cG.number_format = "#,##0"
            cG.alignment = Alignment(horizontal="right", vertical="center")
            cG.border = bdr

            # Col H: Avg Likes/Post
            cH = ws.cell(row=row_idx, column=8, value=p["avg_likes"])
            cH.font = num_font
            cH.number_format = "#,##0"
            cH.alignment = Alignment(horizontal="right", vertical="center")
            cH.border = bdr

            # Col I: Avg Comments/Post
            cI = ws.cell(row=row_idx, column=9, value=p["avg_comments"])
            cI.font = num_font
            cI.number_format = "#,##0"
            cI.alignment = Alignment(horizontal="right", vertical="center")
            cI.border = bdr

            # Col J: Avg ER% (plain number, e.g. 2.35)
            cJ = ws.cell(row=row_idx, column=10, value=p["avg_er"])
            cJ.font = er_font
            cJ.number_format = "0.00"
            cJ.alignment = Alignment(horizontal="center", vertical="center")
            cJ.border = bdr

            updated_count += 1
        else:
            skipped_count += 1

    wb.save(EXCEL_FILE)
    print(f"✅ Workbook saved successfully!")
    print(f"   Rows updated with metrics: {updated_count}")
    print(f"   Rows untouched (OEM/business/skipped): {skipped_count}")


if __name__ == "__main__":
    run_pipeline()
