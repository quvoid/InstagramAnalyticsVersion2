"""
Instagram Paid-Partnership / Collab Account Finder — Standalone Script
========================================================================
Scans a target Instagram account's recent posts and pulls out every
OTHER account that shows up as a collaborator / paid-partnership tag —
i.e. posts where Instagram shows a dual byline like

    "apkikashishh and officialzivame"  ...  #ad #paidpartnership

Detection signals, per post:
  1. `owner != target`        — the post is actually AUTHORED by the
      partner account and shows on the target's feed via Instagram's
      "collab" / co-author feature (this is what powers the dual
      byline in the screenshot — the overwhelming majority of hits).
  2. `coauthor_producers`     — structured collab-tag field, for posts
      authored BY the target that tag another account as co-author.
  3. `is_paid_partnership`    — Instagram's structured paid-partnership
      flag on the media item.
  4. Caption hashtags         — fallback regex for #ad / #paidpartnership
      / #sponsored disclosures, pulling @mentions out of the caption.

NOTE on @officialzivame specifically: Instagram's newer
`web_profile_info` REST endpoint currently 400s for this account
("...ig_business_category_subvertical has been deleted...", an IG-side
bug unrelated to this script/your cookies). To work around it, the
target's numeric user_id is resolved by scraping the `profilePage_<id>`
marker embedded in the profile page's own HTML instead.

✏️  HOW TO USE:
    1. Set TARGET_USERNAME below (default: officialzivame).
    2. Cookies are reused from scrape_profiles.py in this same folder.
       Edit COOKIES there (or set OVERRIDE_COOKIES below) with a fresh
       logged-in Instagram session if requests start failing.
    3. Run:  python find_partnership_accounts.py
    4. Results are printed to console AND saved to
       "partnership_accounts.xlsx".

Install dependencies:
    pip install curl_cffi openpyxl requests
"""

# ════════════════════════════════════════════════════════════════
#  ✏️  STEP 1: TARGET ACCOUNT
# ════════════════════════════════════════════════════════════════

TARGET_USERNAME = "officialzivame"

# ════════════════════════════════════════════════════════════════
#  ✏️  STEP 2: SESSION COOKIES
#  Reused from scrape_profiles.py — edit the COOKIES dict there,
#  or set OVERRIDE_COOKIES below to use different ones for this script.
# ════════════════════════════════════════════════════════════════

OVERRIDE_COOKIES = None   # e.g. {"sessionid": "...", "csrftoken": "...", "mid": "...", "ds_user_id": "..."}

# ════════════════════════════════════════════════════════════════
#  ✏️  STEP 3: CONFIGURATION
# ════════════════════════════════════════════════════════════════

MAX_POSTS        = 600    # How many recent posts to scan (paginated)
PAGE_SIZE        = 12     # Posts per API page request
OUTPUT_FILENAME  = "partnership_accounts.xlsx"

# ════════════════════════════════════════════════════════════════
#  BELOW: Core Scraping and Analysis Logic
# ════════════════════════════════════════════════════════════════

import sys, io, re, time, random
from datetime import datetime, timezone

sys.stdout.reconfigure(encoding="utf-8")

try:
    from curl_cffi import requests as cffi_requests
    USE_CURL_CFFI = True
    print("✓ curl_cffi found — using Chrome TLS fingerprint")
except ImportError:
    import requests as cffi_requests
    USE_CURL_CFFI = False
    print("⚠ curl_cffi not found — using standard requests")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Pull cookies from scrape_profiles.py so credentials live in one place.
if OVERRIDE_COOKIES:
    COOKIES = OVERRIDE_COOKIES
else:
    try:
        from scrape_profiles import COOKIES
    except Exception:
        COOKIES = {}

IG_APP_ID = "936619743392459"

BASE_HEADERS = {
    "accept":           "*/*",
    "accept-language":  "en-US,en;q=0.9",
    "user-agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "x-ig-app-id":       IG_APP_ID,
    "x-ig-www-claim":    "0",
    "x-requested-with":  "XMLHttpRequest",
    "x-csrftoken":       COOKIES.get("csrftoken", ""),
    "origin":            "https://www.instagram.com",
    "referer":           "https://www.instagram.com/",
}

HASHTAG_PATTERN = re.compile(
    r"#ad\b|#paidpartnership|#sponsored|#collab|#partnership|paid partnership with",
    re.IGNORECASE,
)
MENTION_PATTERN = re.compile(r"@([A-Za-z0-9_.]+)")


def make_session():
    if USE_CURL_CFFI:
        s = cffi_requests.Session(impersonate="chrome120")
    else:
        s = cffi_requests.Session()
    s.headers.update(BASE_HEADERS)
    s.cookies.update(COOKIES)
    return s


def delay(a=1.0, b=2.5):
    time.sleep(random.uniform(a, b))


def resolve_user_id(username: str, session) -> str:
    """
    Resolve a username -> numeric user_id.
    Tries the standard web_profile_info API first; falls back to
    scraping the profilePage_<id> marker out of the raw profile HTML
    (works around IG's occasional 400 on that endpoint for certain
    business accounts).
    """
    hdrs = {**BASE_HEADERS, "referer": f"https://www.instagram.com/{username}/"}

    # Attempt 1: structured API
    try:
        url = f"https://www.instagram.com/api/v1/users/web_profile_info/?username={username}"
        r = session.get(url, headers=hdrs, cookies=COOKIES, timeout=15)
        if r.status_code == 200:
            user_data = r.json().get("data", {}).get("user") or {}
            if user_data.get("id"):
                return user_data["id"]
        else:
            print(f"  ⚠ web_profile_info returned {r.status_code} — falling back to HTML scrape.")
    except Exception as e:
        print(f"  ⚠ web_profile_info error: {e} — falling back to HTML scrape.")

    # Attempt 2: raw profile page HTML
    try:
        r = session.get(f"https://www.instagram.com/{username}/", headers=hdrs, cookies=COOKIES, timeout=15)
        if r.status_code == 200:
            m = re.search(r'"profilePage_(\d+)"', r.text)
            if m:
                return m.group(1)
    except Exception as e:
        print(f"  ⚠ HTML fallback error: {e}")

    return ""


def fetch_posts_page(user_id: str, session, count=12, max_id=None) -> dict:
    url = f"https://www.instagram.com/api/v1/feed/user/{user_id}/?count={count}"
    if max_id:
        url += f"&max_id={max_id}"
    for attempt in range(1, 4):
        try:
            r = session.get(url, headers=BASE_HEADERS, cookies=COOKIES, timeout=15)
            if r.status_code == 200:
                return r.json()
            elif r.status_code == 429:
                wait = 30 * attempt
                print(f"  ⏳ Rate limited — waiting {wait}s...")
                time.sleep(wait)
            else:
                print(f"  ⚠ Unexpected status {r.status_code} fetching posts page.")
                return {}
        except Exception as e:
            print(f"  ⚠ Network/parsing error: {e}")
            delay(2, 4)
    return {}


def extract_partnership_info(item: dict, target_username: str) -> dict:
    """
    Inspect one feed item and return partner usernames + how they were
    detected, or None if this post shows no partnership signal.
    """
    partners = {}   # username(lower) -> (method, display_username)

    owner_username = (item.get("user", {}) or {}).get("username", "").strip()
    if owner_username and owner_username.lower() != target_username.lower():
        partners[owner_username.lower()] = ("Post owned by partner (collab)", owner_username)

    for coauthor in item.get("coauthor_producers") or []:
        uname = (coauthor.get("username") or "").strip()
        if uname and uname.lower() != target_username.lower():
            partners.setdefault(uname.lower(), ("Coauthor tag", uname))

    is_paid = bool(item.get("is_paid_partnership"))

    cap_obj = item.get("caption")
    caption = (cap_obj.get("text", "") if isinstance(cap_obj, dict) else "") or ""
    caption_flagged = bool(HASHTAG_PATTERN.search(caption))
    if caption_flagged:
        for m in MENTION_PATTERN.findall(caption):
            if m.lower() != target_username.lower():
                partners.setdefault(m.lower(), ("Caption hashtag + @mention", m))

    if not partners and not is_paid:
        return None

    code = item.get("code", item.get("shortcode", ""))
    ts = item.get("taken_at", 0)
    date_str = (
        datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d")
        if ts else "N/A"
    )
    lines = [l.strip() for l in caption.splitlines() if l.strip()]
    first_line = lines[0] if lines else ""

    return {
        "url": f"https://www.instagram.com/p/{code}/",
        "date": date_str,
        "is_paid_partnership_flag": is_paid,
        "caption_flagged": caption_flagged,
        "caption_preview": first_line[:150],
        "partners": partners,   # {username_lower: (method, display_username)}
    }


def build_excel_report(target_username: str, partnership_posts: list, unique_partners: dict):
    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="D0D0D0")
    tbdr = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="888888"))
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(color_hex): return PatternFill("solid", fgColor=color_hex)
    def bf(size=10): return Font(name="Calibri", bold=True, size=size)
    def nf(size=10): return Font(name="Calibri", bold=False, size=size)
    def lnk(size=10): return Font(name="Calibri", bold=False, size=size, color="0563C1", underline="single")
    def ctr(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    def lft(wrap=True):  return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)

    # Sheet 1: Partnership Posts
    ws1 = wb.active
    ws1.title = "Partnership Posts"
    ws1.merge_cells("A1:F1")
    ws1["A1"] = f"@{target_username} — Detected Partnership / Collab Posts"
    ws1["A1"].font = Font(name="Calibri", bold=True, size=14)
    ws1["A1"].alignment = ctr()
    ws1.row_dimensions[1].height = 30

    headers1 = ["#", "Post URL", "Date", "Detected Via", "Partner Username(s)", "Caption Preview"]
    widths1 = [5, 42, 12, 30, 30, 60]
    for col, (h, w) in enumerate(zip(headers1, widths1), 1):
        c = ws1.cell(row=2, column=col, value=h)
        c.font = bf(); c.fill = fill("F2F2F2"); c.alignment = ctr(); c.border = tbdr
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[2].height = 24

    for i, post in enumerate(partnership_posts, 1):
        row = i + 2
        methods = ", ".join(sorted({m for m, _ in post["partners"].values()})) or (
            "Paid partnership flag" if post["is_paid_partnership_flag"] else ""
        )
        names = ", ".join(f"@{u}" for _, u in post["partners"].values())
        vals = [i, post["url"], post["date"], methods, names, post["caption_preview"]]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=row, column=col, value=val); c.border = bdr
            if col == 2:
                c.font = lnk(); c.alignment = lft(wrap=False)
                if val: c.hyperlink = val
            elif col in (5, 6):
                c.font = nf(); c.alignment = lft(wrap=True)
            else:
                c.font = nf(); c.alignment = ctr()
        ws1.row_dimensions[row].height = 26

    # Sheet 2: Unique Partner Usernames
    ws2 = wb.create_sheet("Unique Partner Usernames")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "Unique Partnership / Collab Accounts"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14)
    ws2["A1"].alignment = ctr()
    ws2.row_dimensions[1].height = 30

    headers2 = ["#", "Username", "# Posts Detected"]
    widths2 = [6, 30, 18]
    for col, (h, w) in enumerate(zip(headers2, widths2), 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = bf(); c.fill = fill("F2F2F2"); c.alignment = ctr(); c.border = tbdr
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[2].height = 24

    sorted_partners = sorted(unique_partners.items(), key=lambda kv: -kv[1]["count"])
    for i, (uname_lower, info) in enumerate(sorted_partners, 1):
        row = i + 2
        vals = [i, f"@{info['display']}", info["count"]]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=row, column=col, value=val); c.border = bdr
            c.font = bf() if col == 2 else nf()
            c.alignment = lft(wrap=False) if col == 2 else ctr()
        ws2.row_dimensions[row].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def main():
    if not COOKIES.get("sessionid"):
        print("❌ No Instagram session cookies found. Set them in scrape_profiles.py "
              "(COOKIES dict) or via OVERRIDE_COOKIES in this script.")
        return

    uname = TARGET_USERNAME.strip().lstrip("@")
    print(f"\n{'═'*65}")
    print(f"  Instagram Partnership / Collab Account Finder")
    print(f"  Target: @{uname}  |  Scanning up to {MAX_POSTS} posts")
    print(f"{'═'*65}\n")

    session = make_session()

    print(f"[1/2] Resolving user_id for @{uname}...")
    user_id = resolve_user_id(uname, session)
    if not user_id:
        print(f"❌ Could not resolve user_id for @{uname}. Check cookies / username.")
        return
    print(f"      User ID: {user_id}\n")

    print(f"[2/2] Paginating posts and scanning for partnership signals...")
    partnership_posts = []
    unique_partners = {}   # username_lower -> {"display": str, "count": int}
    scanned = 0
    max_id = None

    while scanned < MAX_POSTS:
        page = fetch_posts_page(user_id, session, count=PAGE_SIZE, max_id=max_id)
        items = page.get("items", [])
        if not items:
            break

        for item in items:
            scanned += 1
            info = extract_partnership_info(item, uname)
            if info:
                partnership_posts.append(info)
                for uname_lower, (_, display) in info["partners"].items():
                    entry = unique_partners.setdefault(uname_lower, {"display": display, "count": 0})
                    entry["count"] += 1
            if scanned >= MAX_POSTS:
                break

        print(f"      Scanned {scanned} posts... found {len(partnership_posts)} partnership posts so far.")

        max_id = page.get("next_max_id")
        if not page.get("more_available") or not max_id:
            break
        delay(1.2, 2.5)

    print(f"\n✓ Done. Scanned {scanned} posts total. Detected {len(partnership_posts)} partnership posts.\n")

    if not unique_partners:
        print("No partnership / collab accounts detected.")
        return

    print(f"{'─'*65}")
    print(f"  PARTNERSHIP ACCOUNT USERNAME LIST ({len(unique_partners)} unique)")
    print(f"{'─'*65}")
    for uname_lower, info in sorted(unique_partners.items(), key=lambda kv: -kv[1]["count"]):
        print(f"  @{info['display']}   ({info['count']} post{'s' if info['count'] != 1 else ''})")
    print(f"{'─'*65}\n")

    print("📊 Generating Excel report...")
    excel_bytes = build_excel_report(uname, partnership_posts, unique_partners)
    with open(OUTPUT_FILENAME, "wb") as f:
        f.write(excel_bytes)

    print(f"\n{'═'*65}")
    print(f"  ✅ Complete! Saved: {OUTPUT_FILENAME}")
    print(f"  Sheets: Partnership Posts · Unique Partner Usernames")
    print(f"{'═'*65}\n")


if __name__ == "__main__":
    main()
