"""
1. Download all new Toggle ON creatives (GIVA & Palmonas) into video/
2. Update Master Excel (jewellery_brands_master_analysis.xlsx)
3. Update Master CSVs (All_Brands_Paid_Collabs.csv, Brand_Creator_Summary.csv)
"""

import sys, os, json, time, re, csv, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scrape_bulk import make_session, COOKIES

sys.stdout.reconfigure(encoding="utf-8")

VIDEO_DIR = os.path.join(os.getcwd(), "video")
os.makedirs(VIDEO_DIR, exist_ok=True)

with open("giva_palmonas_scraped.json", encoding="utf-8") as f:
    new_posts = json.load(f)

print(f"Total new posts from GIVA & Palmonas: {len(new_posts)}")
new_toggle_on = [p for p in new_posts if p["is_paid_partnership"]]
print(f"New Toggle ON creatives to download: {len(new_toggle_on)}")

# Download new Toggle ON creatives
mob_hdrs = {
    "User-Agent": "Instagram 269.0.0.18.75 Android (26/8.0.0; 480dpi; 1080x1920; OnePlus; ONEPLUS A3003; OnePlus3; qcom; en_US; 314665256)",
    "x-ig-app-id": "936619743392459",
}
session = make_session()

# Find highest existing index in video folder
existing_files = os.listdir(VIDEO_DIR)
existing_indices = [int(f.split("_")[0]) for f in existing_files if re.match(r"^\d+_", f)]
start_idx = max(existing_indices, default=0) + 1

print(f"\n[+] Downloading {len(new_toggle_on)} new Toggle ON creatives starting at index {start_idx}...")

for idx_offset, p in enumerate(new_toggle_on):
    curr_idx = start_idx + idx_offset
    mid = p["media_id"]
    sc = p["shortcode"]
    brand_clean = re.sub(r'[^a-zA-Z0-9]', '_', p["brand"]).strip('_')
    handle_clean = re.sub(r'[^a-zA-Z0-9]', '_', p["raw_handle"]).strip('_')
    
    print(f"[{idx_offset+1:>2}/{len(new_toggle_on)}] Fetching {p['brand']} | @{p['raw_handle']} ({sc})...", end=" ", flush=True)
    try:
        r = session.get(f"https://i.instagram.com/api/v1/media/{mid}/info/", headers=mob_hdrs, cookies=COOKIES, timeout=12)
        if r.status_code == 200:
            item = r.json().get("items", [])[0]
            vids = item.get("video_versions", [])
            if vids:
                v_url = vids[0]["url"]
                fn = f"{curr_idx:02d}_{brand_clean}_{handle_clean}_{sc}.mp4"
                fp = os.path.join(VIDEO_DIR, fn)
                r_dl = session.get(v_url, timeout=30)
                if r_dl.status_code == 200:
                    with open(fp, "wb") as f_out:
                        f_out.write(r_dl.content)
                    sz_mb = len(r_dl.content) / (1024*1024)
                    print(f"✓ Video saved ({sz_mb:.1f} MB) -> {fn}")
            else:
                imgs = item.get("image_versions2", {}).get("candidates", [])
                if imgs:
                    img_url = imgs[0]["url"]
                    fn = f"{curr_idx:02d}_{brand_clean}_{handle_clean}_{sc}.jpg"
                    fp = os.path.join(VIDEO_DIR, fn)
                    r_dl = session.get(img_url, timeout=30)
                    if r_dl.status_code == 200:
                        with open(fp, "wb") as f_out:
                            f_out.write(r_dl.content)
                        print(f"✓ Photo saved -> {fn}")
        else:
            print(f"⚠ API HTTP {r.status_code}")
    except Exception as e:
        print(f"⚠ Error: {e}")
    time.sleep(0.4)

# ── 2. Update Master Dataset and CSVs ──────────────────────────
CSV_COLLABS = "All_Brands_Paid_Collabs.csv"
CSV_SUMMARY = "Brand_Creator_Summary.csv"
EXCEL_OUTPUT = "jewellery_brands_master_analysis.xlsx"

# Read existing collab posts
with open(CSV_COLLABS, encoding="utf-8-sig") as f_csv:
    existing_collab_rows = list(csv.reader(f_csv))[1:]

# Format new posts to match CSV structure
formatted_new_rows = []
for p in new_posts:
    formatted_new_rows.append([
        0, # placeholder index
        p["brand"],
        p["handle"],
        p["toggle_str"],
        p["followers"],
        p["likes"],
        p["comments"],
        f"{p['er']:.2f}%",
        p["date"],
        p["url"],
        p["via"],
        p["caption"]
    ])

# Merge and re-sort by Date descending
all_rows_merged = existing_collab_rows + formatted_new_rows
all_rows_merged.sort(key=lambda x: str(x[8]), reverse=True)

# Re-index
for i, r in enumerate(all_rows_merged, 1):
    r[0] = i

# Write updated All_Brands_Paid_Collabs.csv
with open(CSV_COLLABS, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow([
        "#", "Brand Name", "Creator Handle", "Paid Partnership Toggle",
        "Followers", "Avg Likes/Post", "Avg Comments/Post", "Avg ER%",
        "Post Date", "Post URL", "Detection Method", "Caption Preview"
    ])
    for r in all_rows_merged:
        writer.writerow(r)
print(f"\n✓ Updated {CSV_COLLABS} (Total Rows: {len(all_rows_merged)})")


# ── 3. Rebuild Brand Summary CSV & Excel ───────────────────────
# Calculate stats per brand across the complete merged dataset
from collections import defaultdict
brand_groups = defaultdict(list)
for r in all_rows_merged:
    brand_groups[r[1]].append(r)

brand_summary_list = []
for b_name, posts in brand_groups.items():
    u_creators = len(set(p[2].lower() for p in posts))
    on_cnt = sum(1 for p in posts if "ON" in p[3])
    off_cnt = len(posts) - on_cnt
    on_pct = on_cnt / len(posts) if posts else 0.0
    
    fols = [int(str(p[4]).replace(",","") or 0) for p in posts if int(str(p[4]).replace(",","") or 0) > 0]
    ers = [float(str(p[7]).replace("%","") or 0) for p in posts if float(str(p[7]).replace("%","") or 0) > 0]
    
    avg_f = int(sum(fols)/len(fols)) if fols else 0
    avg_e = round(sum(ers)/len(ers), 2) if ers else 0.0
    latest_d = max([p[8] for p in posts if p[8] != "N/A"], default="N/A")
    top_c = list(dict.fromkeys([p[2] for p in posts]))[:4]
    
    brand_summary_list.append({
        "brand": b_name,
        "posts": len(posts),
        "creators": u_creators,
        "toggle_on": on_cnt,
        "toggle_off": off_cnt,
        "toggle_pct": on_pct,
        "avg_followers": avg_f,
        "avg_er": avg_e,
        "latest_date": latest_d,
        "top_creators": ", ".join(top_c)
    })

brand_summary_list.sort(key=lambda x: x["posts"], reverse=True)

with open(CSV_SUMMARY, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow([
        "#", "Brand Name", "Paid Collab Posts", "Unique Creators",
        "Toggle ON", "Toggle OFF", "Toggle ON %",
        "Avg Creator Followers", "Avg Creator ER%", "Latest Collab Date", "Top Creator / Partner Samples"
    ])
    for idx, b in enumerate(brand_summary_list, 1):
        writer.writerow([
            idx, b["brand"], b["posts"], b["creators"],
            b["toggle_on"], b["toggle_off"], f"{b['toggle_pct']*100:.1f}%",
            b["avg_followers"], f"{b['avg_er']:.2f}%", b["latest_date"], b["top_creators"]
        ])
print(f"✓ Updated {CSV_SUMMARY} (Total Brands: {len(brand_summary_list)})")


# ── 4. Rebuild Formatted Master Excel ──────────────────────────
wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

thin = Side(style="thin", color="D0D3D4")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

f_title = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
f_hdr = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
f_bold = Font(name="Calibri", bold=True, size=10, color="000000")
f_norm = Font(name="Calibri", bold=False, size=10, color="000000")
f_link = Font(name="Calibri", bold=False, size=10, color="0563C1", underline="single")

fill_green_row = PatternFill("solid", fgColor="D4EFDF") # Mint Green for Toggle ON
f_green_bold = Font(name="Calibri", bold=True, size=10, color="145A32")
f_green_norm = Font(name="Calibri", bold=False, size=10, color="145A32")
f_green_link = Font(name="Calibri", bold=True, size=10, color="0B5345", underline="single")

fill_off_pill = PatternFill("solid", fgColor="F2F4F4")
f_off_pill = Font(name="Calibri", bold=False, size=10, color="5D6D7E")

# Sheet 1: Brand-Creator Summary
ws1 = wb_out.create_sheet("Brand-Creator Summary")
ws1.sheet_view.showGridLines = True

tot_posts_all = len(all_rows_merged)
tot_creators_all = len(set(r[2].lower() for r in all_rows_merged))
tot_on_all = sum(1 for r in all_rows_merged if "ON" in r[3])
tot_off_all = tot_posts_all - tot_on_all

ws1.merge_cells("A1:K1")
ws1["A1"] = f"Executive Summary — Paid Creator Partnerships by Brand ({tot_posts_all} Total Posts · {tot_on_all} Toggle ON · {tot_off_all} Toggle OFF)"
ws1["A1"].font = f_title
ws1["A1"].fill = PatternFill("solid", fgColor="0B2240")
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 32

sum_cols = [
    ("#", 5), ("Brand Name", 28), ("Paid Collab Posts", 17), ("Unique Creators", 16),
    ("Toggle ON", 13), ("Toggle OFF", 13), ("Toggle ON %", 14),
    ("Avg Creator Followers", 21), ("Avg Creator ER%", 15), ("Latest Collab Date", 17),
    ("Top Creator / Partner Samples", 45)
]

for col_idx, (h_text, width) in enumerate(sum_cols, 1):
    c = ws1.cell(row=2, column=col_idx, value=h_text)
    c.font = f_hdr
    c.fill = PatternFill("solid", fgColor="1F2D3D")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bdr
    ws1.column_dimensions[get_column_letter(col_idx)].width = width
ws1.row_dimensions[2].height = 24
ws1.freeze_panes = "A3"

for idx, b in enumerate(brand_summary_list, 1):
    r = idx + 2
    vals = [
        idx, b["brand"], b["posts"], b["creators"],
        b["toggle_on"], b["toggle_off"], b["toggle_pct"],
        b["avg_followers"], b["avg_er"] / 100 if b["avg_er"] else 0.0, b["latest_date"], b["top_creators"]
    ]
    for c_idx, val in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=c_idx, value=val)
        cell.border = bdr
        if c_idx == 1:
            cell.font = f_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2:
            cell.font = f_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (3, 4):
            cell.font = f_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"
            if val > 0: cell.fill = PatternFill("solid", fgColor="EBF5FB")
        elif c_idx == 5:
            cell.font = f_green_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"
            if val > 0: cell.fill = fill_green_row
        elif c_idx == 6:
            cell.font = f_off_pill; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"
            if val > 0: cell.fill = fill_off_pill
        elif c_idx == 7:
            cell.font = f_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.0%"
        elif c_idx == 8:
            cell.font = f_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx == 9:
            cell.font = f_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
        elif c_idx == 10:
            cell.font = f_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 11:
            cell.font = f_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[r].height = 22


# Sheet 2: All Brands - Paid Collabs
ws2 = wb_out.create_sheet("All Brands - Paid Collabs")
ws2.sheet_view.showGridLines = True

ws2.merge_cells("A1:L1")
ws2["A1"] = f"Consolidated Creator-Owned Paid Partnerships ('Post owned by partner') — {tot_posts_all} Total Posts (Toggle ON: {tot_on_all} · Toggle OFF: {tot_off_all})"
ws2["A1"].font = f_title
ws2["A1"].fill = PatternFill("solid", fgColor="1E5631")
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

collab_cols = [
    ("#", 5), ("Brand Name", 28), ("Creator Handle", 25), ("Paid Partnership Toggle", 25),
    ("Followers", 14), ("Avg Likes/Post", 14), ("Avg Comments/Post", 14), ("Avg ER%", 12),
    ("Post Date", 13), ("Post URL", 48), ("Detection Method", 28), ("Caption Preview", 65)
]

for col_idx, (h_text, width) in enumerate(collab_cols, 1):
    c = ws2.cell(row=2, column=col_idx, value=h_text)
    c.font = f_hdr
    c.fill = PatternFill("solid", fgColor="1E5631")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bdr
    ws2.column_dimensions[get_column_letter(col_idx)].width = width
ws2.row_dimensions[2].height = 24
ws2.freeze_panes = "A3"

for idx, r_data in enumerate(all_rows_merged, 1):
    r = idx + 2
    is_on = "ON" in r_data[3]
    
    fol_clean = int(str(r_data[4]).replace(",", "") or 0)
    likes_clean = int(str(r_data[5]).replace(",", "") or 0)
    com_clean = int(str(r_data[6]).replace(",", "") or 0)
    er_raw = str(r_data[7]).replace("%", "").strip()
    er_clean = float(er_raw)/100 if er_raw else 0.0
    
    vals = [
        idx, r_data[1], r_data[2], r_data[3],
        fol_clean, likes_clean, com_clean, er_clean,
        r_data[8], r_data[9], r_data[10], r_data[11]
    ]
    for c_idx, val in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=c_idx, value=val)
        cell.border = bdr
        
        if is_on:
            cell.fill = fill_green_row
            if c_idx in (2, 3, 4):
                cell.font = f_green_bold
            elif c_idx == 10:
                cell.font = f_green_link
                if val: cell.hyperlink = val
            else:
                cell.font = f_green_norm
        else:
            if c_idx == 1: cell.font = f_norm
            elif c_idx == 2: cell.font = f_bold
            elif c_idx == 3: cell.font = f_bold; cell.fill = PatternFill("solid", fgColor="FDF2E9")
            elif c_idx == 4: cell.font = f_off_pill; cell.fill = fill_off_pill
            elif c_idx == 10:
                cell.font = f_link
                if val: cell.hyperlink = val
            else: cell.font = f_norm
            
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (2, 3): cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 4: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (5, 6, 7): cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx == 8: cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
        elif c_idx == 9: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 10: cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 11: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 12: cell.alignment = Alignment(horizontal="left", vertical="center")
        
    ws2.row_dimensions[r].height = 22

# Sheet 3: GIVA Jewellery
ws_giva = wb_out.create_sheet("GIVA Jewellery")
ws_giva.sheet_view.showGridLines = True
giva_posts = [p for p in new_posts if p["brand"] == "GIVA Jewellery"]
ws_giva.append(["Creator Handle", "Followers", "Avg Likes/Post", "Avg Comments/Post", "Avg ER%", "Post Date", "Post URL", "Detection Method", "Caption Preview", "Boost Status"])
for p in giva_posts:
    ws_giva.append([p["handle"], p["followers"], p["likes"], p["comments"], f"{p['er']:.2f}%", p["date"], p["url"], p["via"], p["caption"], p["boost_status"]])

# Sheet 4: Palmonas
ws_pal = wb_out.create_sheet("Palmonas")
ws_pal.sheet_view.showGridLines = True
pal_posts = [p for p in new_posts if p["brand"] == "Palmonas"]
ws_pal.append(["Creator Handle", "Followers", "Avg Likes/Post", "Avg Comments/Post", "Avg ER%", "Post Date", "Post URL", "Detection Method", "Caption Preview", "Boost Status"])
for p in pal_posts:
    ws_pal.append([p["handle"], p["followers"], p["likes"], p["comments"], f"{p['er']:.2f}%", p["date"], p["url"], p["via"], p["caption"], p["boost_status"]])

wb_out.save(EXCEL_OUTPUT)
print(f"\n✅ Master analysis successfully updated:")
print(f"   • Total Collab Posts across ALL Brands: {tot_posts_all}")
print(f"   • Total Unique Creators: {tot_creators_all}")
print(f"   • Total Toggle ON: {tot_on_all} (Highlighted in green)")
print(f"   • Total Toggle OFF: {tot_off_all}")
print(f"   • Excel File: {EXCEL_OUTPUT}")
print(f"   • CSV File: {CSV_COLLABS}")
