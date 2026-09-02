"""
Agency-Grade Paid Media Blueprint & Media Planning Workbook for Lake Shore (KOPA Pune & Y Junction Hyderabad)
File: lakeshore_agency_media_buying_blueprint.xlsx

Built strictly from the perspective of Performance Media Buyers, Creative Strategists, and Media Planners:
Tab 1: 📋 Campaign Architecture & Taxonomy (Naming conventions, objectives, budget pacing, bid strategies)
Tab 2: 🎯 Precision Targeting & Audiences (Interest stacks, LALs, Geofence coordinates, Pin codes, Exclusions)
Tab 3: 🎨 Creative Ad Matrix & Scripts (0-3s Hook, 3-8s Hold, 8-15s CTA, Ad copy, Aspect ratios, Asset types)
Tab 4: 🤝 Creator Whitelisting (Branded Content Ads, Creator handles, Rate benchmarks, Permissions)
Tab 5: 📊 Media Budget & Unit Economics (Daily spend, CPM, CPC, CTR, Cost per Store Direction, CPL, Footfall proxy)
Tab 6: 🔗 UTM & Tracking Governance (Full UTM taxonomy, QR valet tracking, Concierge Wi-Fi pixel capture)
"""

import sys, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Color Palette & Typography
font_title = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
font_sec = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_norm = Font(name="Calibri", size=10, bold=False, color="111111")
font_code = Font(name="Consolas", size=9, bold=False, color="1B4F72")
font_italic = Font(name="Calibri", size=9, italic=True, color="555555")

thin_line = Side(style="thin", color="D5D8DC")
border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

fill_navy = PatternFill("solid", fgColor="0B2240")
fill_blue_hdr = PatternFill("solid", fgColor="1B4F72")
fill_dark = PatternFill("solid", fgColor="212F3D")
fill_client = PatternFill("solid", fgColor="D4EFDF") # Soft Emerald
fill_highlight = PatternFill("solid", fgColor="FCF3CF") # Soft Yellow
fill_neg = PatternFill("solid", fgColor="FADBD8") # Soft Red
fill_pos = PatternFill("solid", fgColor="EAFAF1") # Soft Green
fill_accent = PatternFill("solid", fgColor="EBF5FB") # Soft Blue
fill_purple = PatternFill("solid", fgColor="F4ECF7") # Soft Purple

def set_row(ws, r_num, vals, font=font_norm, fill=None, align_center_cols=None, align_right_cols=None, code_cols=None, height=22, wrap=True):
    align_center_cols = align_center_cols or []
    align_right_cols = align_right_cols or []
    code_cols = code_cols or []
    for c_idx, val in enumerate(vals, 1):
        c = ws.cell(row=r_num, column=c_idx, value=val)
        c.font = font_code if c_idx in code_cols else font
        c.border = border_cell
        if fill: c.fill = fill
        if c_idx in align_center_cols:
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        elif c_idx in align_right_cols:
            c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=wrap)
        else:
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    ws.row_dimensions[r_num].height = height


# ──────────────────────────────────────────────────────────────────────────────
# TAB 1: CAMPAIGN ARCHITECTURE & TAXONOMY (META & GOOGLE ADS)
# ──────────────────────────────────────────────────────────────────────────────
ws1 = wb.create_sheet("1. Campaign Architecture")
ws1.sheet_view.showGridLines = True
ws1.merge_cells("A1:I1")
ws1["A1"] = "LAKE SHORE PAID MEDIA ENGINE — FULL CAMPAIGN ARCHITECTURE & NAMING TAXONOMY"
ws1["A1"].font = font_title; ws1["A1"].fill = fill_navy; ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 32

headers1 = [
    ("#", 5), ("Campaign Name (Taxonomy)", 38), ("Funnel Tier", 16), ("Buying Objective", 20),
    ("Optimization Goal", 22), ("Bid Strategy", 18), ("Dayparting / Flighting Rule", 32),
    ("Placement Matrix", 28), ("Monthly Budget Share", 20)
]
for c_idx, (h_text, w) in enumerate(headers1, 1):
    c = ws1.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.column_dimensions[get_column_letter(c_idx)].width = w
ws1.row_dimensions[2].height = 28
ws1.freeze_panes = "A3"

campaigns_data = [
    (1, "LS_KOPA_META_TOF_REACH_QUIETAESTHETIC_V1", "Top of Funnel (TOF)", "Awareness (Reach / Video)", "ThruPlay / 2s Continuous View", "Lowest Cost (Auto)", "Always-On (Mon-Sun 10:00 - 23:00)", "Instagram Reels (9:16), Feed (4:5)", "25% of Total Spend"),
    (2, "LS_KOPA_META_MOF_CONQUEST_PHOENIX_PARKING", "Mid Funnel (MOF)", "Traffic / Engagement", "Landing Page Views / Direction Clicks", "Cost Cap / Lowest Cost", "Dayparted: Fri-Sun 16:00 - 21:30 (Peak Rush)", "IG Stories, Reels, FB Feed (Mobile Only)", "30% of Total Spend"),
    (3, "LS_KOPA_META_MOF_CONQUEST_SEASONS_CROWD", "Mid Funnel (MOF)", "Traffic / Engagement", "Store Saves / Direction Clicks", "Lowest Cost", "Dayparted: Fri-Sun 15:00 - 21:00", "IG Reels, Stories (9:16)", "15% of Total Spend"),
    (4, "LS_KOPA_META_BOF_DINING_RESERVATIONS_CORP", "Bottom of Funnel (BOF)", "Leads / Conversions", "Lead Form Submit / Link Clicks (Zomato)", "Cost per Lead (CPL Cap ₹180)", "Flighted: Thu 14:00 to Sat 20:00", "IG Feed (1:1), Stories (9:16)", "15% of Total Spend"),
    (5, "LS_KOPA_META_BOF_PVR_DIRECTORSCUT_VIP", "Bottom of Funnel (BOF)", "Sales / Traffic", "Outbound Clicks (BookMyShow)", "Lowest Cost", "Flighted: Wed 18:00 to Sun 22:00", "IG Reels, Stories, In-Stream Video", "10% of Total Spend"),
    (6, "LS_KOPA_META_TOF_HOTEL_GEOFENCE_EXPATS", "Top of Funnel (TOF)", "Reach / Store Traffic", "Impressions / Direction Clicks", "Lowest Cost", "Always-On (Geofenced strictly to 5-Star Hotels)", "IG Reels, Stories, Feed (English Only)", "5% of Total Spend")
]

for row in campaigns_data:
    r = row[0] + 2
    set_row(ws1, r, [row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8]], 
            font=font_norm, fill=fill_client if "CONQUEST" in row[1] else (fill_accent if "DINING" in row[1] else None),
            align_center_cols=[1, 3, 4, 6, 9], code_cols=[2], height=32)


# ──────────────────────────────────────────────────────────────────────────────
# TAB 2: PRECISION TARGETING & AUDIENCE STACKS
# ──────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("2. Audience Targeting Matrix")
ws2.sheet_view.showGridLines = True
ws2.merge_cells("A1:H1")
ws2["A1"] = "TARGETING PARAMETERS: GEO-COORDINATES, PIN CODES, INTEREST STACKS & EXCLUSIONS"
ws2["A1"].font = font_title; ws2["A1"].fill = fill_navy; ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

headers2 = [
    ("#", 5), ("Audience Segment Name", 32), ("Geo-Radius / Pin Codes", 40),
    ("Demographics (Age/Gender)", 24), ("Detailed Interests & Behaviors", 50),
    ("Custom / Lookalike Audiences", 35), ("Strict Exclusions", 30), ("Strategic Purpose", 35)
]
for c_idx, (h_text, w) in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.column_dimensions[get_column_letter(c_idx)].width = w
ws2.row_dimensions[2].height = 28
ws2.freeze_panes = "A3"

targeting_data = [
    (1, "AUD_KOPA_CORE_CATCHMENT_HNI", 
     "Koregaon Park (411001), Kalyani Nagar (411006), Boat Club Rd (411001), Sopan Baug (411001) + 4km radius", 
     "All Gender | Age 26 - 55", 
     "Luxury Goods, Fine Dining, Armani, Sephora, Jo Malone, Single Malt Whisky, High-end fashion, Frequent International Travelers", 
     "1% LAL of IG Engagers (365d) + 1% LAL of Lake Shore Website Visitors", 
     "Exclude: Bargain Hunters, Job Seekers, Delivery drivers", 
     "Primary affluent catchment driving everyday high-ticket retail & gourmet dining."),
    
    (2, "AUD_KOPA_CONQUEST_PHOENIX_CATCHMENT", 
     "Viman Nagar, Clover Park, Nagar Road (18.5622° N, 73.9166° E) + 3km Radius around Phoenix Marketcity", 
     "All Gender | Age 24 - 48", 
     "Shopping Malls, Zara, H&M, Mango, Fine Dining, Coffeehouses (Starbucks, Third Wave), Lifestyle", 
     "Retargeting: Engaged with KOPA Reels in past 60 days", 
     "Exclude: Residents located >8km away", 
     "Intercepts shoppers in Phoenix catchment facing 40-min parking queues on weekend evenings."),
    
    (3, "AUD_KOPA_TECH_CORP_DINING_NIGHTLIFE", 
     "EON IT Park Kharadi (411014), World Trade Center, Magarpatta City (411028), Cybercity", 
     "All Gender | Age 27 - 45", 
     "Software engineering, IT Management, Craft Beer, Rooftop Bars, Zomato Gold, Date Night, Cocktails, Premium Dining", 
     "Custom List: Past Dining Lead form openers & Website table booking clickers", 
     "Exclude: College Students, Non-working individuals", 
     "Generates Thursday-Saturday corporate happy hour & weekend date-night restaurant bookings."),
    
    (4, "AUD_KOPA_HOTEL_GEOFENCE_EXPATS", 
     "Strict 500m Geofence around: Ritz-Carlton Pune, Conrad Pune, JW Marriott SB Rd, The Westin KP", 
     "All Gender | Age 30 - 62", 
     "Frequent Business Travelers, Luxury Travelers, English Speakers (Device Language = English)", 
     "Broad within Geofence (Location: 'People recently in this location')", 
     "Exclude: Permanent residential staff (exclude home location = geofence)", 
     "Captures affluent business executives & hotel guests with high discretionary spend."),
    
    (5, "AUD_KOPA_VIP_CINEPHILES_COUPLES", 
     "Boat Club Rd, KP, Kalyani Nagar, Model Colony, Prabhat Rd (411004)", 
     "Couples & Movie Enthusiasts | Age 25 - 50", 
     "PVR IMAX, Director's Cut, Film Festivals, Fine Wine, Gourmet Snacks, Movie Connoisseurs", 
     "Engaged with PVR Director's Cut Video Ads", 
     "Exclude: Standard multiplex coupon seekers", 
     "Fills high-margin PVR Director's Cut recliner seats with in-cinema dining packages.")
]

for row in targeting_data:
    r = row[0] + 2
    set_row(ws2, r, [row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]], 
            font=font_norm, fill=fill_highlight if "CONQUEST" in row[1] else (fill_accent if "HOTEL" in row[1] else None),
            align_center_cols=[1, 4], code_cols=[2], height=36)


# ──────────────────────────────────────────────────────────────────────────────
# TAB 3: CREATIVE AD MATRIX & SCRIPT BLUEPRINTS
# ──────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("3. Creative Ad Matrix & Scripts")
ws3.sheet_view.showGridLines = True
ws3.merge_cells("A1:I1")
ws3["A1"] = "CREATIVE AD MATRIX: 0-3S HOOK, 3-8S HOLD, 8-15S CTA, SCRIPT & COPY DIRECTIVES"
ws3["A1"].font = font_title; ws3["A1"].fill = fill_navy; ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 32

headers3 = [
    ("#", 5), ("Creative Code", 26), ("Format / Ratio", 16), ("0-3s Visual Hook", 35),
    ("0-3s Audio Hook / Voiceover", 35), ("3-8s Body Value Proposition", 45),
    ("8-15s Call to Action (CTA)", 32), ("Primary Text (Ad Copy)", 50), ("CTA Button", 16)
]
for c_idx, (h_text, w) in enumerate(headers3, 1):
    c = ws3.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws3.column_dimensions[get_column_letter(c_idx)].width = w
ws3.row_dimensions[2].height = 28
ws3.freeze_panes = "A3"

creative_data = [
    (1, "CR_KOPA_VID_VALET_CONQUEST_916", "9:16 Reel (Video)", 
     "Split screen: Red car stuck in chaotic basement queue vs Black Mercedes smoothly pulling into KOPA valet.", 
     "\"Still waiting in basement level 4?\"", 
     "Show seamless key handover to valet, walking into fragrant, calm, sunlit KOPA lobby with Armani/Sephora in backdrop.", 
     "On-screen: \"5 mins from Viman Nagar. Valet ready.\" Button pops up.", 
     "Weekend shopping shouldn't feel like a traffic jam. Skip the 40-minute parking queue — experience dedicated valet parking & curated luxury at KOPA Koregaon Park.", "Get Directions"),
    
    (2, "CR_KOPA_VID_DINING_DATE_NIGHT_916", "9:16 Reel (Video)", 
     "Close-up pour of smoked cocktail under warm ambient rooftop lighting with live jazz audio.", 
     "\"Date night without the 1-hour waitlist.\"", 
     "Quick cuts: artisanal sushi platter, candlelit outdoor table, couple laughing, curated dessert plating.", 
     "\"Reserve your table for tonight at KOPA's rooftop dining lounges.\"", 
     "Tired of noisy food courts and endless waitlists? Discover Pune's most sophisticated dining sanctuary. Reserve your table at KOPA's curated gourmet restaurants in Koregaon Park.", "Book Now"),
    
    (3, "CR_KOPA_VID_DIRECTORSCUT_VIP_916", "9:16 Reel (Video)", 
     "Pushing the plush leather recliner button in PVR Director's Cut as butler serves truffle popcorn.", 
     "\"This is how movies were meant to be watched.\"", 
     "Show 4K laser projection, private lounge access, bespoke dinner menu served at seat.", 
     "\"Experience cinema without compromises. Book PVR Director's Cut at KOPA.\"", 
     "Reimagine cinema. Luxury recliners, in-theatre gourmet dining, and personalized butler service. PVR Director's Cut — exclusively at KOPA Pune.", "Book Tickets"),
    
    (4, "CR_KOPA_CAROUSEL_LUXURY_BRANDS_11", "1:1 Carousel (6 Slides)", 
     "Slide 1 Cover: Sleek architectural shot with bold text: \"Pune's Curated Luxury Enclave.\"", 
     "N/A (Static Carousel)", 
     "Slide 2: Sephora & Tira beauty counters | Slide 3: Armani & Mango lookbooks | Slide 4: Rooftop Dining | Slide 5: PVR Director's Cut", 
     "Slide 6: \"Your weekend itinerary is ready. Visit KOPA Koregaon Park.\"", 
     "Discover the finest global fashion, beauty, and culinary brands under one roof. No crowds, no chaos — just pure luxury in Koregaon Park.", "Learn More"),
    
    (5, "CR_KOPA_VID_HOTEL_CONCIERGE_169", "16:9 In-Feed (Video)", 
     "Cinematic slow-motion pan of KOPA's open-air atrium with natural sunlight and fountain.", 
     "\"Welcome to Pune. Your private luxury destination awaits.\"", 
     "Highlights proximity to Ritz-Carlton/Conrad, personal shopper availability, quiet coffee lounges for business meetings.", 
     "\"Just 5 minutes from your hotel suite. Open daily 11 AM - 11 PM.\"", 
     "Visiting Pune for business or leisure? Experience curated global luxury fashion and world-class dining — located just minutes from Pune's 5-star hotels at KOPA.", "Get Directions")
]

for row in creative_data:
    r = row[0] + 2
    set_row(ws3, r, [row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8]], 
            font=font_norm, fill=fill_client if "VALET" in row[1] else (fill_highlight if "DINING" in row[1] else None),
            align_center_cols=[1, 3, 9], code_cols=[2], height=42)


# ──────────────────────────────────────────────────────────────────────────────
# TAB 4: CREATOR WHITELISTING & PARTNERSHIP ADS
# ──────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("4. Creator Whitelisting Matrix")
ws4.sheet_view.showGridLines = True
ws4.merge_cells("A1:H1")
ws4["A1"] = "CREATOR WHITELISTING: META PARTNERSHIP ADS (RUNNING DARK ADS FROM CREATOR HANDLES)"
ws4["A1"].font = font_title; ws4["A1"].fill = fill_navy; ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 32

headers4 = [
    ("#", 5), ("Creator Tier & Category", 28), ("Recommended Pune Creators", 32),
    ("Follower Scale", 18), ("Campaign Type", 24), ("Branded Content Meta Permission", 32),
    ("Ad Spend Multiplier Rule", 28), ("Why Whitelisting Wins over Brand Post", 40)
]
for c_idx, (h_text, w) in enumerate(headers4, 1):
    c = ws4.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws4.column_dimensions[get_column_letter(c_idx)].width = w
ws4.row_dimensions[2].height = 28
ws4.freeze_panes = "A3"

creator_whitelisting = [
    (1, "Macro Fashion / Luxury Stylist", "@rupaliirathore, @theclassycloud, @pune_lifestyle_diva", "250K - 850K", "Meta Partnership Ad (Dark Post)", "Grant Partnership Ad Permission via Meta Business Suite", "Allocate 3x ad spend behind creator post vs brand handle", "Creator faces generate 2.8x higher Hook Rate and 40% lower CPMs than corporate mall handles."),
    (2, "Mid-Tier Food & Cocktails Critic", "@punefoodhunt, @thecravingtales, @bhukkad_pune", "60K - 180K", "Instagram Reel + Boosted Whitelist", "Branded Content Tag + 'Allow business partner to boost'", "Flight ₹15,000 behind high-performing review reel over Thu-Sat", "Food lovers trust local food reviewers for date night recommendations over sponsored brand accounts."),
    (3, "Micro Lifestyle & Couple Creators", "@koregaon_diaries, @explorewith_us, @punecouplegoals", "15K - 45K", "Dark Video Ad targeting KP/Kalyani Nagar", "Meta Ads Manager Partnership Access Token", "Allocate ₹8,000 per creator for hyper-local 3km geofence", "Authentic, non-commercial aesthetic delivers highest comment sentiment and 'Save Post' rate.")
]

for row in creator_whitelisting:
    r = row[0] + 2
    set_row(ws4, r, [row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]], 
            font=font_norm, fill=fill_accent if row[0]%2==0 else None,
            align_center_cols=[1, 4, 5], code_cols=[3], height=38)


# ──────────────────────────────────────────────────────────────────────────────
# TAB 5: MEDIA BUDGET & UNIT ECONOMICS
# ──────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("5. Budget & Unit Economics")
ws5.sheet_view.showGridLines = True
ws5.merge_cells("A1:I1")
ws5["A1"] = "MEDIA BUYING UNIT ECONOMICS: DAILY PACING, BENCHMARKS & CONVERSION PROXIES"
ws5["A1"].font = font_title; ws5["A1"].fill = fill_navy; ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 32

headers5 = [
    ("#", 5), ("Campaign Stream", 32), ("Monthly Budget (₹)", 20), ("Daily Pacing (₹)", 18),
    ("Benchmark CPM (₹)", 18), ("Target CTR %", 15), ("Expected CPC / ThruPlay (₹)", 24),
    ("Cost per Direction / Lead (₹)", 26), ("Estimated Monthly Footfall Impact", 35)
]
for c_idx, (h_text, w) in enumerate(headers5, 1):
    c = ws5.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws5.column_dimensions[get_column_letter(c_idx)].width = w
ws5.row_dimensions[2].height = 28
ws5.freeze_panes = "A3"

budget_data = [
    (1, "1. Competitor Parking Conquesting (Dark Ads)", 180000, 6000, "₹ 65 - 85", "1.95%", "₹ 3.80 / click", "₹ 18 - 25 per Direction Click", "7,200 - 9,500 High-Intent Footfall Visitors"),
    (2, "2. Gourmet Dining & Rooftop Reservations", 90000, 3000, "₹ 80 - 110", "2.40%", "₹ 4.50 / click", "₹ 140 - 180 per Table Lead", "550 - 650 Verified Dining Table Bookings"),
    (3, "3. PVR Director's Cut VIP Entertainment", 60000, 2000, "₹ 75 - 95", "1.80%", "₹ 4.20 / click", "₹ 28 - 35 per Ticket Clickout", "1,700 - 2,100 High-Ticket Cinema Patrons"),
    (4, "4. 4K Aesthetic Quiet Luxury (Awareness)", 120000, 4000, "₹ 45 - 60", "1.45%", "₹ 0.18 / ThruPlay", "₹ 55 per Store Profile Save", "1.8M - 2.4M Verified HNI Brand Impressions"),
    (5, "5. 5-Star Hotel Expats & Business Geofence", 30000, 1000, "₹ 120 - 150", "2.10%", "₹ 6.50 / click", "₹ 35 - 45 per Hotel Direction Click", "650 - 850 High-AOV Luxury Tourists"),
    (6, "6. Retargeting (Warm Audience / Profile Saves)", 60000, 2000, "₹ 90 - 120", "3.20%", "₹ 3.10 / click", "₹ 14 per Direction / Voucher Claim", "3,800 - 4,500 Re-Engaged Luxury Shoppers"),
]

for row in budget_data:
    r = row[0] + 2
    set_row(ws5, r, [row[0], row[1], f"₹ {row[2]:,}", f"₹ {row[3]:,}", row[4], row[5], row[6], row[7], row[8]], 
            font=font_norm, fill=fill_client if row[0]==1 else (fill_highlight if row[0]==2 else None),
            align_center_cols=[1, 5, 6], align_right_cols=[3, 4, 7, 8], height=28)

# Total Row
tot_r = len(budget_data) + 3
set_row(ws5, tot_r, ["TOTAL", "Integrated Monthly Paid Engine", "₹ 540,000", "₹ 18,000 / day", "Blended ₹ 74", "Blended 2.15%", "Blended ₹ 3.65", "Blended CPDC ₹ 22", "14,500+ Verified Direct Footfall Conversions"],
        font=font_bold, fill=fill_dark, align_center_cols=[1, 5, 6], align_right_cols=[3, 4, 7, 8], height=30)
ws5.cell(row=tot_r, column=1).font = font_hdr
ws5.cell(row=tot_r, column=2).font = font_hdr
ws5.cell(row=tot_r, column=3).font = font_hdr
ws5.cell(row=tot_r, column=4).font = font_hdr
ws5.cell(row=tot_r, column=5).font = font_hdr
ws5.cell(row=tot_r, column=6).font = font_hdr
ws5.cell(row=tot_r, column=7).font = font_hdr
ws5.cell(row=tot_r, column=8).font = font_hdr
ws5.cell(row=tot_r, column=9).font = font_hdr


# ──────────────────────────────────────────────────────────────────────────────
# TAB 6: UTM TAXONOMY & ATTRIBUTION GOVERNANCE
# ──────────────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("6. UTM & Attribution Governance")
ws6.sheet_view.showGridLines = True
ws6.merge_cells("A1:G1")
ws6["A1"] = "ATTRIBUTION GOVERNANCE: UTM TAXONOMY, QR CODES & VALET PIXEL CAPTURE"
ws6["A1"].font = font_title; ws6["A1"].fill = fill_navy; ws6["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 32

headers6 = [
    ("#", 5), ("Tracking Channel / Touchpoint", 28), ("UTM Parameter Specification", 55),
    ("Physical On-Ground Proxy Metric", 35), ("Pixel Event Triggered", 24),
    ("Lookback Attribution Window", 25), ("Measurement Tool", 25)
]
for c_idx, (h_text, w) in enumerate(headers6, 1):
    c = ws6.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr; c.fill = fill_dark; c.border = border_cell
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws6.column_dimensions[get_column_letter(c_idx)].width = w
ws6.row_dimensions[2].height = 28
ws6.freeze_panes = "A3"

tracking_data = [
    (1, "Meta Dark Ads (Parking Conquest)", "utm_source=meta&utm_medium=paid_social&utm_campaign=conquest_phoenix&utm_content=valet_vs_queue_v1", "Google Maps Directions Click + In-Mall Wi-Fi Login within 4 hrs", "Lead / FindLocation", "7-Day Click, 1-Day View", "Meta Pixel + Google Analytics 4"),
    (2, "Creator Whitelist Reels (Dining)", "utm_source=creator_handle&utm_medium=branded_content&utm_campaign=dining_datenight&utm_term=influencer_id", "Zomato Table Booking Clickout + Valet QR Scan", "Contact / Schedule", "7-Day Click", "GA4 + Zomato Affiliate Tracker"),
    (3, "Hotel Geofenced Ads (Expats)", "utm_source=meta_hotel&utm_medium=geofence&utm_campaign=hotel_expats&utm_content=ritz_conrad", "Concierge Desk Scan / High-ticket Store Purchase Scan", "ViewContent / Purchase", "1-Day View, 7-Day Click", "Meta Ads Manager Offline Events"),
    (4, "On-Ground Valet Counter QR", "utm_source=offline_qr&utm_medium=valet_stand&utm_campaign=valet_loyalty_signup", "Valet Ticket digital claim & Wi-Fi opt-in", "CompleteRegistration", "Instant Offline Event", "Offline Conversions API (CAPI)")
]

for row in tracking_data:
    r = row[0] + 2
    set_row(ws6, r, [row[0], row[1], row[2], row[3], row[4], row[5], row[6]], 
            font=font_norm, fill=fill_accent if row[0]%2==0 else None,
            align_center_cols=[1, 5, 6], code_cols=[3], height=34)

output_file = "lakeshore_agency_media_buying_blueprint.xlsx"
wb.save(output_file)
print(f"✓ Agency Media Buying Blueprint created: {output_file}")
