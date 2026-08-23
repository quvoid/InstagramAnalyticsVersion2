"""
Generate a high-res Line Chart for Google Trends search interest & embed native Excel LineChart.
"""

import os, sys
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
import openpyxl
from openpyxl.chart import LineChart, Reference, Series

sys.stdout.reconfigure(encoding="utf-8")

WORKBOOK_PATH = "competitor_paid_media_analysis.xlsx"
OUTPUT_IMAGE_NAME = "google_trends_line_chart.png"
ARTIFACT_DIR = r"C:\Users\omkar\.gemini\antigravity\brain\d196c916-bdf1-40a6-835a-3259489a070c"
ARTIFACT_IMG_PATH = os.path.join(ARTIFACT_DIR, OUTPUT_IMAGE_NAME)
LOCAL_IMG_PATH = os.path.join(".", OUTPUT_IMAGE_NAME)

# 1. Load Excel Data
wb = openpyxl.load_workbook(WORKBOOK_PATH)
ws = wb["Google Trends"]

data = []
# Row 13 is header: ['Week', 'Montblanc', 'Parker Pens', 'Sheaffer', 'Lamy', 'Makoba', 'Submarine Pens']
header = [ws.cell(row=13, column=col).value for col in range(1, 8)]

for r in range(14, ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=col).value for col in range(1, 8)]
    if row_vals[0]:
        data.append(row_vals)

df = pd.DataFrame(data, columns=header)
df["Week"] = pd.to_datetime(df["Week"])

# Fill or drop None
brand_cols = [c for c in header[1:] if c != "Sheaffer"] # Sheaffer was not available
for col in brand_cols:
    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

# 2. Render Beautiful Matplotlib Line Chart
plt.figure(figsize=(14, 7), dpi=300)
plt.style.use("seaborn-v0_8-whitegrid" if "seaborn-v0_8-whitegrid" in plt.style.available else "default")

colors = {
    "Montblanc": "#2B2D42",       # Sleek Obsidian / Charcoal
    "Parker Pens": "#0066CC",     # Royal Blue
    "Lamy": "#2E7D32",            # Forest Green
    "Makoba": "#8E24AA",          # Violet / Purple
    "Submarine Pens": "#E65100",  # Coral / Deep Orange
}

# Plot lines
for brand in brand_cols:
    if brand in colors:
        # Plot raw line with lower alpha
        plt.plot(df["Week"], df[brand], color=colors[brand], alpha=0.35, linewidth=1.2, linestyle="--")
        # Plot smoothed 3-week rolling average
        smooth = df[brand].rolling(window=3, min_periods=1, center=True).mean()
        plt.plot(df["Week"], smooth, label=f"{brand} (3-wk avg)", color=colors[brand], linewidth=2.8)

# Styling
plt.title("Google Search Interest Trends in India — Pen Brands (12 Months)", fontsize=16, fontweight="bold", pad=20, color="#1A1A1A")
plt.xlabel("Timeline (Weeks)", fontsize=12, labelpad=12, fontweight="bold", color="#333333")
plt.ylabel("Search Interest Index (0–100)", fontsize=12, labelpad=12, fontweight="bold", color="#333333")
plt.ylim(-2, 108)

# Formatting Date Axis
ax = plt.gca()
ax.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %Y'))
plt.xticks(rotation=0, fontsize=10)
plt.yticks(fontsize=10)

# Annotate Notable Peaks
# Makoba recent spike
makoba_max_idx = df["Makoba"].idxmax()
makoba_max_date = df.loc[makoba_max_idx, "Week"]
makoba_max_val = df.loc[makoba_max_idx, "Makoba"]
if makoba_max_val > 50:
    ax.annotate(
        f'Makoba Surge ({int(makoba_max_val)})',
        xy=(makoba_max_date, makoba_max_val),
        xytext=(makoba_max_date - pd.Timedelta(days=40), makoba_max_val + 4),
        arrowprops=dict(facecolor='#8E24AA', shrink=0.08, width=1.5, headwidth=6),
        fontsize=9, fontweight='bold', color='#8E24AA',
        bbox=dict(boxstyle="round,pad=0.3", fc="#F3E5F5", ec="#8E24AA", lw=1)
    )

# Parker peak
parker_max_idx = df["Parker Pens"].idxmax()
parker_max_date = df.loc[parker_max_idx, "Week"]
parker_max_val = df.loc[parker_max_idx, "Parker Pens"]
ax.annotate(
    f'Parker Peak ({int(parker_max_val)})',
    xy=(parker_max_date, parker_max_val),
    xytext=(parker_max_date + pd.Timedelta(days=10), parker_max_val - 5),
    arrowprops=dict(facecolor='#0066CC', shrink=0.08, width=1.5, headwidth=6),
    fontsize=9, fontweight='bold', color='#0066CC',
    bbox=dict(boxstyle="round,pad=0.3", fc="#E3F2FD", ec="#0066CC", lw=1)
)

# Grid & Legend
plt.grid(True, linestyle=":", alpha=0.6, color="#CCCCCC")
plt.legend(frameon=True, facecolor="white", edgecolor="#E0E0E0", framealpha=0.95, fontsize=10, loc="upper left", bbox_to_anchor=(1.01, 1))

plt.tight_layout()

# Save image locally & to artifacts
plt.savefig(LOCAL_IMG_PATH, bbox_inches="tight")
plt.savefig(ARTIFACT_IMG_PATH, bbox_inches="tight")
plt.close()
print(f"✓ Line chart saved to {LOCAL_IMG_PATH} and {ARTIFACT_IMG_PATH}")

# 3. Add Native Excel LineChart to openpyxl sheet
# Check if chart already exists
if not hasattr(ws, "_charts") or len(ws._charts) == 0:
    chart = LineChart()
    chart.title = "Google Search Interest Index (India - 12 Months)"
    chart.style = 13
    chart.y_axis.title = "Interest Index (0-100)"
    chart.x_axis.title = "Week"
    chart.width = 24
    chart.height = 14

    # Data references
    # Row 13 is headers, col 1 is Week, cols 2-7 are brands. Rows 14 to ws.max_row are data
    data_ref = Reference(ws, min_col=2, min_row=13, max_col=7, max_row=ws.max_row)
    cats_ref = Reference(ws, min_col=1, min_row=14, max_row=ws.max_row)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Smooth the lines
    for series in chart.series:
        series.smooth = True
        series.graphicalProperties.line.width = 25000

    ws.add_chart(chart, "I3")
    wb.save(WORKBOOK_PATH)
    print("✓ Native Excel LineChart embedded onto sheet 'Google Trends' at cell I3")
else:
    print("✓ Excel already contains charts, saved image only.")
