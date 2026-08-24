"""
Run Full Creator Profile Metrics & Populate Master Excel & CSVs for Footwear Brands:
1. Skechers India (@skechersindia)
2. Gully Labs (@gullylabs)
3. Comet (@thecometuniverse)

Fetches:
- Followers, Following, Total Posts, Full Name, Verified, Business
- Avg Likes/Post, Avg Comments/Post, Avg ER% (from recent 12 posts)
- Updates Master Excel (with Creators Profile Metrics sheet) & CSV exports.
"""

import sys, os, json, time, re, csv, random
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scrape_bulk import make_session, COOKIES

sys.stdout.reconfigure(encoding="utf-8")

session = make_session()

web_hdrs = {
    "accept": "*/*",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "x-ig-app-id": "936619743392459",
}

mob_hdrs = {
    "User-Agent": "Instagram 269.0.0.18.75 Android (26/8.0.0; 480dpi; 1080x1920; OnePlus; ONEPLUS A3003; OnePlus3; qcom; en_US; 314665256)",
    "x-ig-app-id": "936619743392459",
}

with open("footwear_1year_4tier_dataset.json", encoding="utf-8") as f:
    posts = json.load(f)

# Map creators to the brands they collaborated with
creator_brands_map = defaultdict(set)
for p in posts:
    creator_brands_map[p["raw_handle"].lower()].add(p["brand"])

unique_creators = sorted(list(set(p["raw_handle"] for p in posts)))
print(f"Total Unique Creator Accounts to Scrape: {len(unique_creators)}\n")

def fetch_single_creator_profile(username):
    clean_user = username.replace("@", "").strip().lower()
    sess = make_session()
    
    url = f"https://www.instagram.com/api/v1/users/web_profile_info/?username={clean_user}"
    hdrs = {**web_hdrs, "referer": f"https://www.instagram.com/{clean_user}/"}
    
    profile_data = {
        "handle": f"@{clean_user}",
        "raw_handle": clean_user,
        "brands": ", ".join(sorted(list(creator_brands_map.get(clean_user, [])))),
        "full_name": clean_user,
        "followers": 0,
        "following": 0,
        "total_posts": 0,
        "verified": False,
        "is_business": False,
        "avg_likes": 0,
        "avg_comments": 0,
        "avg_er": 0.0,
        "profile_url": f"https://www.instagram.com/{clean_user}/",
        "scrape_status": "Failed"
    }
    
    for attempt in range(1, 3):
        try:
            r = sess.get(url, headers=hdrs, cookies=COOKIES, timeout=12)
            if r.status_code == 200:
                u = r.json().get("data", {}).get("user", {})
                if u:
                    fols = u.get("edge_followed_by", {}).get("count", 0)
                    fing = u.get("edge_follow", {}).get("count", 0)
                    posts_cnt = u.get("edge_owner_to_timeline_media", {}).get("count", 0)
                    fn = u.get("full_name") or clean_user
                    ver = bool(u.get("is_verified", False))
                    is_biz = bool(u.get("is_business_account", False) or u.get("is_professional_account", False))
                    
                    # Recent 12 posts engagement
                    edges = u.get("edge_owner_to_timeline_media", {}).get("edges", [])
                    likes_list = []
                    comments_list = []
                    for e in edges:
                        node = e.get("node", {})
                        l = node.get("edge_liked_by", {}).get("count") or node.get("edge_media_preview_like", {}).get("count") or 0
                        c = node.get("edge_media_to_comment", {}).get("count") or 0
                        likes_list.append(l)
                        comments_list.append(c)
                        
                    avg_l = int(sum(likes_list) / len(likes_list)) if likes_list else 0
                    avg_c = int(sum(comments_list) / len(comments_list)) if comments_list else 0
                    avg_er = round(((avg_l + avg_c) / fols) * 100, 2) if fols > 0 else 0.0
                    
                    profile_data.update({
                        "full_name": fn,
                        "followers": fols,
                        "following": fing,
                        "total_posts": posts_cnt,
                        "verified": ver,
                        "is_business": is_biz,
                        "avg_likes": avg_l,
                        "avg_comments": avg_c,
                        "avg_er": avg_er,
                        "scrape_status": "Scraped OK"
                    })
                    return profile_data
            elif r.status_code == 404:
                profile_data["scrape_status"] = "User Not Found (404)"
                return profile_data
        except Exception:
            time.sleep(0.5)
            
    # Fallback to mobile search API if web fails
    try:
        r2 = sess.get(f"https://i.instagram.com/api/v1/users/search/?q={clean_user}", headers=mob_hdrs, cookies=COOKIES, timeout=8)
        if r2.status_code == 200:
            for u in r2.json().get("users", []):
                if u.get("username", "").lower() == clean_user:
                    fols = u.get("follower_count", 0)
                    fn = u.get("full_name") or clean_user
                    ver = bool(u.get("is_verified", False))
                    profile_data.update({
                        "full_name": fn,
                        "followers": fols,
                        "verified": ver,
                        "scrape_status": "Scraped OK (Mobile API)"
                    })
                    return profile_data
    except Exception:
        pass
        
    return profile_data

print("Fetching profile metrics for all 91 creators with multithreading...")
t0 = time.time()
profiles_results = []

with ThreadPoolExecutor(max_workers=10) as executor:
    futures = {executor.submit(fetch_single_creator_profile, u): u for u in unique_creators}
    done = 0
    for fut in as_completed(futures):
        res = fut.result()
        profiles_results.append(res)
        done += 1
        if done % 15 == 0 or done == len(unique_creators):
            print(f"  Progress: {done:>2}/{len(unique_creators)} profiles scraped ({time.time()-t0:.1f}s)")

# Sort profiles by followers descending
profiles_results.sort(key=lambda x: x["followers"], reverse=True)

with open("footwear_creators_profile_metrics.json", "w", encoding="utf-8") as f:
    json.dump(profiles_results, f, indent=2)

print(f"\n✓ Saved footwear_creators_profile_metrics.json ({len(profiles_results)} creators)")

# Map metrics back to posts
prof_map = {p["raw_handle"].lower(): p for p in profiles_results}

for p in posts:
    rh = p["raw_handle"].lower()
    if rh in prof_map:
        prof = prof_map[rh]
        fols = prof["followers"]
        p["followers"] = fols
        
        views = p["views"]
        likes = p["likes"]
        comments = p["comments"]
        
        like_rate = round((likes / views) * 100, 2) if views > 0 else 0.0
        view_mult = round(views / fols, 2) if fols > 0 else 0.0
        er = round(((likes + comments) / fols) * 100, 2) if fols > 0 else 0.0
        
        p["like_rate_pct"] = like_rate
        p["view_multiplier"] = view_mult
        p["er_pct"] = er

# Save updated dataset
posts.sort(key=lambda x: (x["tier"], -x["views"]))
with open("footwear_1year_4tier_dataset.json", "w", encoding="utf-8") as f:
    json.dump(posts, f, indent=2)

print("✓ Updated footwear_1year_4tier_dataset.json with verified follower counts and ER%\n")

# ─────────────────────────────────────────────────────────────
# BUILD MASTER EXCEL WITH CREATORS PROFILE METRICS SHEET
# ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Styles
font_title = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_norm = Font(name="Calibri", size=10, bold=False, color="000000")
font_mute = Font(name="Calibri", size=9, bold=False, color="5D6D7E")
font_link = Font(name="Calibri", size=10, bold=False, color="0563C1", underline="single")

fill_t1_banner = PatternFill("solid", fgColor="145A32") # Dark Emerald
fill_t1_row = PatternFill("solid", fgColor="D4EFDF")    # Mint Green
font_t1_bold = Font(name="Calibri", size=10, bold=True, color="0E6251")
font_t1_link = Font(name="Calibri", size=10, bold=True, color="0B5345", underline="single")

fill_t2_banner = PatternFill("solid", fgColor="1E8449") # Forest Green
fill_t2_row = PatternFill("solid", fgColor="EAFAF1")    # Sage Green
font_t2_bold = Font(name="Calibri", size=10, bold=True, color="196F3D")
font_t2_link = Font(name="Calibri", size=10, bold=False, color="145A32", underline="single")

fill_t3_banner = PatternFill("solid", fgColor="B7950B") # Dark Gold
fill_t3_row = PatternFill("solid", fgColor="FEF9E7")    # Warm Soft Gold
font_t3_bold = Font(name="Calibri", size=10, bold=True, color="7D6608")
font_t3_link = Font(name="Calibri", size=10, bold=False, color="9A7D0A", underline="single")

fill_t4_banner = PatternFill("solid", fgColor="566573") # Slate Gray
fill_t4_row = PatternFill("solid", fgColor="FFFFFF")    # Clean White
font_t4_norm = Font(name="Calibri", size=10, bold=False, color="2C3E50")
font_t4_link = Font(name="Calibri", size=10, bold=False, color="2980B9", underline="single")

thin_line = Side(style="thin", color="D5D8DC")
border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

BRANDS = [
    {"name": "Skechers India", "username": "skechersindia", "state": "Pan-India / Maharashtra (HQ: Mumbai)"},
    {"name": "Gully Labs", "username": "gullylabs", "state": "Pan-India / Delhi NCR (HQ: New Delhi, D2C)"},
    {"name": "Comet", "username": "thecometuniverse", "state": "Pan-India / Karnataka (HQ: Bengaluru, D2C)"},
]

brand_records = defaultdict(list)
for r in posts:
    brand_records[r["brand"]].append(r)

# ─────────────────────────────────────────────────────────────
# 1. TAB 1: EXECUTIVE SUMMARY
# ─────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet("Executive Summary")
ws_sum.sheet_view.showGridLines = True

ws_sum.merge_cells("A1:O1")
ws_sum["A1"] = "Executive Summary — 1-Year Paid Creator Collab Hierarchy (Skechers India, Gully Labs, Comet) [Aug 2025 – Aug 2026]"
ws_sum["A1"].font = font_title
ws_sum["A1"].fill = PatternFill("solid", fgColor="0B2240")
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 32

sum_headers = [
    ("#", 5),
    ("Brand Name", 24),
    ("State / Origin (HQ)", 32),
    ("Total Collab Posts (1-Yr)", 18),
    ("Total Unique Creators", 18),
    ("🟢 Tier 1: Toggle ON + Boosted\n(Posts / Creators)", 24),
    ("🟢 Tier 2: Toggle ON + Organic\n(Posts / Creators)", 24),
    ("🚀 Tier 3: Toggle OFF + Boosted\n(Posts / Creators)", 24),
    ("⚪ Tier 4: Toggle OFF + Organic (Noise)\n(Posts / Creators)", 28),
    ("💎 Total High-Intent Paid\n(Tiers 1+2+3 Posts)", 22),
    ("High-Intent Paid %\n(Tiers 1+2+3)", 18),
    ("Avg Views / Post", 16),
    ("Avg Creator Followers", 20),
    ("Avg Creator ER%", 15),
    ("Top Creator / Ambassador Samples", 48)
]

for col_idx, (h_text, w) in enumerate(sum_headers, 1):
    c = ws_sum.cell(row=2, column=col_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="1B2631")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_cell
    ws_sum.column_dimensions[get_column_letter(col_idx)].width = w
ws_sum.row_dimensions[2].height = 36
ws_sum.freeze_panes = "A3"

brand_summary_data = []
for b in BRANDS:
    b_name = b["name"]
    b_posts = brand_records.get(b_name, [])
    tot_p = len(b_posts)
    tot_c = len(set(p["handle"].lower() for p in b_posts))
    
    t1_posts = [p for p in b_posts if p["tier"] == 1]
    t1_creators = len(set(p["handle"].lower() for p in t1_posts))
    
    t2_posts = [p for p in b_posts if p["tier"] == 2]
    t2_creators = len(set(p["handle"].lower() for p in t2_posts))
    
    t3_posts = [p for p in b_posts if p["tier"] == 3]
    t3_creators = len(set(p["handle"].lower() for p in t3_posts))
    
    t4_posts = [p for p in b_posts if p["tier"] == 4]
    t4_creators = len(set(p["handle"].lower() for p in t4_posts))
    
    high_intent = len(t1_posts) + len(t2_posts) + len(t3_posts)
    high_intent_pct = high_intent / tot_p if tot_p > 0 else 0.0
    
    views_list = [p["views"] for p in b_posts if p["views"] > 0]
    avg_v = int(sum(views_list) / len(views_list)) if views_list else 0
    
    fols_list = [p["followers"] for p in b_posts if p["followers"] > 0]
    avg_f = int(sum(fols_list) / len(fols_list)) if fols_list else 0
    
    ers_list = [p["er_pct"] for p in b_posts if p["er_pct"] > 0]
    avg_e = round(sum(ers_list) / len(ers_list), 2) if ers_list else 0.0
    
    top_c = list(dict.fromkeys([p["handle"] for p in b_posts if p["tier"] in (1, 2, 3)]))[:4]
    if not top_c:
        top_c = list(dict.fromkeys([p["handle"] for p in b_posts]))[:4]
        
    brand_summary_data.append({
        "brand": b_name,
        "state": b["state"],
        "posts": tot_p,
        "creators": tot_c,
        "t1_p": len(t1_posts), "t1_c": t1_creators,
        "t2_p": len(t2_posts), "t2_c": t2_creators,
        "t3_p": len(t3_posts), "t3_c": t3_creators,
        "t4_p": len(t4_posts), "t4_c": t4_creators,
        "high_intent_p": high_intent,
        "high_intent_pct": high_intent_pct,
        "avg_views": avg_v,
        "avg_followers": avg_f,
        "avg_er": avg_e,
        "top_creators": ", ".join(top_c)
    })

brand_summary_data.sort(key=lambda x: (x["high_intent_p"], x["posts"]), reverse=True)

for idx, b in enumerate(brand_summary_data, 1):
    row_num = idx + 2
    row_vals = [
        idx,
        b["brand"],
        b["state"],
        b["posts"],
        b["creators"],
        f"{b['t1_p']} posts ({b['t1_c']} creators)" if b["t1_p"] > 0 else "—",
        f"{b['t2_p']} posts ({b['t2_c']} creators)" if b["t2_p"] > 0 else "—",
        f"{b['t3_p']} posts ({b['t3_c']} creators)" if b["t3_p"] > 0 else "—",
        f"{b['t4_p']} posts ({b['t4_c']} creators)" if b["t4_p"] > 0 else "—",
        b["high_intent_p"],
        b["high_intent_pct"],
        b["avg_views"],
        b["avg_followers"],
        b["avg_er"] / 100 if b["avg_er"] else 0.0,
        b["top_creators"]
    ]
    
    for col_idx, val in enumerate(row_vals, 1):
        cell = ws_sum.cell(row=row_num, column=col_idx, value=val)
        cell.border = border_cell
        
        if col_idx == 1:
            cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 2:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif col_idx == 3:
            cell.font = Font(name="Calibri", size=9, bold=True, color="2C3E50"); cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = PatternFill("solid", fgColor="F4F6F6")
        elif col_idx in (4, 5):
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"
            cell.fill = PatternFill("solid", fgColor="EBF5FB")
        elif col_idx == 6:
            cell.font = font_t1_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = fill_t1_row
        elif col_idx == 7:
            cell.font = font_t2_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = fill_t2_row
        elif col_idx == 8:
            cell.font = font_t3_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = fill_t3_row
        elif col_idx == 9:
            cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
            if "posts" in str(val): cell.fill = PatternFill("solid", fgColor="F8F9F9")
        elif col_idx == 10:
            cell.font = Font(name="Calibri", size=10, bold=True, color="1B4F72"); cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "#,##0"
            if val > 0: cell.fill = PatternFill("solid", fgColor="D6EAF8")
        elif col_idx == 11:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.0%"
        elif col_idx in (12, 13):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif col_idx == 14:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
        elif col_idx == 15:
            cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
            
    ws_sum.row_dimensions[row_num].height = 22

# ─────────────────────────────────────────────────────────────
# 2. TAB 2: CREATORS PROFILE METRICS (DEDUPED)
# ─────────────────────────────────────────────────────────────
ws_prof = wb.create_sheet("Creators Profile Metrics")
ws_prof.sheet_view.showGridLines = True

ws_prof.merge_cells("A1:L1")
ws_prof["A1"] = f"Deduped Creator Profiles — Full Profile Metrics ({len(profiles_results)} Creators across Skechers India, Gully Labs, Comet)"
ws_prof["A1"].font = font_title
ws_prof["A1"].fill = PatternFill("solid", fgColor="1B4F72")
ws_prof["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_prof.row_dimensions[1].height = 30

prof_headers = [
    ("#", 5),
    ("Creator Handle", 22),
    ("Brands Collaborated With", 28),
    ("Full Name", 26),
    ("Verified", 10),
    ("Business / Pro", 14),
    ("Total Followers", 16),
    ("Following", 12),
    ("Total Posts", 14),
    ("Avg Likes / Post", 16),
    ("Avg Comments / Post", 18),
    ("Avg Profile ER%", 14),
]

for col_idx, (h_text, w) in enumerate(prof_headers, 1):
    c = ws_prof.cell(row=2, column=col_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="283747")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_cell
    ws_prof.column_dimensions[get_column_letter(col_idx)].width = w
ws_prof.row_dimensions[2].height = 25
ws_prof.freeze_panes = "A3"

for idx, p in enumerate(profiles_results, 1):
    r_num = idx + 2
    r_vals = [
        idx,
        p["handle"],
        p["brands"],
        p["full_name"],
        "Yes" if p["verified"] else "No",
        "Yes" if p["is_business"] else "No",
        p["followers"],
        p["following"],
        p["total_posts"],
        p["avg_likes"],
        p["avg_comments"],
        p["avg_er"] / 100 if p["avg_er"] else 0.0
    ]
    
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws_prof.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        
        if c_idx == 1:
            cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.hyperlink = p["profile_url"]
        elif c_idx in (3, 4):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (5, 6):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
            if val == "Yes": cell.fill = PatternFill("solid", fgColor="EAFAF1")
        elif c_idx in (7, 8, 9, 10, 11):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx == 12:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
            
    ws_prof.row_dimensions[r_num].height = 20

# ─────────────────────────────────────────────────────────────
# 3. HELPER FUNCTION FOR MASTER & BRAND TABS
# ─────────────────────────────────────────────────────────────
table_cols = [
    ("#", 5),
    ("Hierarchy Tier", 30),
    ("Brand Name", 22),
    ("Creator Handle", 24),
    ("Followers", 14),
    ("Views / Plays", 16),
    ("Likes", 14),
    ("Comments", 12),
    ("Like-to-View %", 15),
    ("Creator ER%", 13),
    ("Post Date", 13),
    ("Direct Instagram URL", 48),
    ("Boost Classification & Reason", 38),
    ("Caption Preview", 65)
]

def render_footwear_sheet(ws, records, sheet_title, state_str):
    ws.sheet_view.showGridLines = True
    
    tot_p = len(records)
    tot_c = len(set(p["handle"].lower() for p in records))
    t1_p = sum(1 for p in records if p["tier"] == 1)
    t2_p = sum(1 for p in records if p["tier"] == 2)
    t3_p = sum(1 for p in records if p["tier"] == 3)
    t4_p = sum(1 for p in records if p["tier"] == 4)
    high_intent = t1_p + t2_p + t3_p
    
    # Header Banner
    ws.merge_cells("A1:N1")
    ws["A1"] = f"👟 {sheet_title.upper()}  |  📍 STATE / REGION: {state_str}"
    ws["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0B2240")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    
    # Metadata Overview
    ws.merge_cells("A2:N2")
    ws["A2"] = f"1-Year Collab Posts: {tot_p}  •  Unique Creators: {tot_c}  •  💎 High-Intent Paid: {high_intent} (T1: {t1_p} | T2: {t2_p} | T3: {t3_p})  •  ⚪ Noise/Unboosted (T4): {t4_p}"
    ws["A2"].font = Font(name="Calibri", size=10, bold=True, color="1B4F72")
    ws["A2"].fill = PatternFill("solid", fgColor="EBF5FB")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22
    
    # Table Header Row
    for col_idx, (h_text, w) in enumerate(table_cols, 1):
        c = ws.cell(row=3, column=col_idx, value=h_text)
        c.font = font_hdr
        c.fill = PatternFill("solid", fgColor="1F2D3D")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_cell
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[3].height = 25
    ws.freeze_panes = "A4"
    
    records_by_tier = {1: [], 2: [], 3: [], 4: []}
    for r in records:
        records_by_tier[r["tier"]].append(r)
        
    current_row = 4
    global_index = 1
    
    tier_meta = [
        (1, "🟢 TIER 1: TOGGLE ON + 🚀 BOOSTED (Formal Paid Partnership Label + Paid Ad Spend)", fill_t1_banner, fill_t1_row, font_t1_bold, font_t1_link),
        (2, "🟢 TIER 2: TOGGLE ON + ⚪ ORGANIC (Formal Paid Partnership Label + Organic Reach Only)", fill_t2_banner, fill_t2_row, font_t2_bold, font_t2_link),
        (3, "🚀 TIER 3: TOGGLE OFF + 🚀 BOOSTED (Co-Author Collab + Heavy Paid Ad Spend Detected)", fill_t3_banner, fill_t3_row, font_t3_bold, font_t3_link),
        (4, "⚪ TIER 4: TOGGLE OFF + ⚪ ORGANIC (Standard Collab / Low Organic Reach / Noise)", fill_t4_banner, fill_t4_row, font_t4_norm, font_t4_link)
    ]
    
    for t_id, banner_text, fill_banner, fill_row, font_b, font_l in tier_meta:
        t_records = records_by_tier[t_id]
        if not t_records:
            continue
            
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        banner_cell = ws.cell(row=current_row, column=1, value=f"{banner_text} — {len(t_records)} Posts ({len(set(p['handle'].lower() for p in t_records))} Unique Creators)")
        banner_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        banner_cell.fill = fill_banner
        banner_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[current_row].height = 23
        current_row += 1
        
        for p in t_records:
            vals = [
                global_index,
                p["tier_name"],
                p["brand"],
                p["handle"],
                p["followers"],
                p["views"],
                p["likes"],
                p["comments"],
                p["like_rate_pct"] / 100 if p["like_rate_pct"] else 0.0,
                p["er_pct"] / 100 if p["er_pct"] else 0.0,
                p["date"],
                p["url"],
                f"{p['boost_status']}: {p['boost_reason']}" if p.get('boost_reason') else p['boost_status'],
                p["caption"]
            ]
            
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=current_row, column=c_idx, value=val)
                cell.border = border_cell
                cell.fill = fill_row
                
                if c_idx == 1:
                    cell.font = font_mute; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 2:
                    cell.font = font_b; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx in (3, 4):
                    cell.font = font_b; cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx in (5, 6, 7, 8):
                    cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
                elif c_idx in (9, 10):
                    cell.font = font_b; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.number_format = "0.00%"
                elif c_idx == 11:
                    cell.font = font_norm; cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 12:
                    cell.font = font_l; cell.alignment = Alignment(horizontal="left", vertical="center")
                    if val: cell.hyperlink = val
                elif c_idx == 13:
                    cell.font = font_b; cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c_idx == 14:
                    cell.font = font_mute; cell.alignment = Alignment(horizontal="left", vertical="center")
                    
            ws.row_dimensions[current_row].height = 21
            current_row += 1
            global_index += 1

# Tab 3: Master Hierarchy
ws_all = wb.create_sheet("All Brands - Master Hierarchy")
render_footwear_sheet(
    ws_all,
    posts,
    "CONSOLIDATED 3 SNEAKER & FOOTWEAR BRANDS",
    "PAST 1 YEAR (AUG 2025 – AUG 2026)"
)

# Tabs 4 to 6: Individual Brand sheets
for b in BRANDS:
    b_name = b["name"]
    b_posts = brand_records.get(b_name, [])
    ws_brand = wb.create_sheet(b_name[:31])
    render_footwear_sheet(
        ws_brand,
        b_posts,
        b_name,
        b["state"]
    )

wb.save("footwear_sneaker_brands_master_analysis.xlsx")
print("✅ Master Excel Saved: footwear_sneaker_brands_master_analysis.xlsx (6 Sheets)")

# ─────────────────────────────────────────────────────────────
# EXPORT CSVS
# ─────────────────────────────────────────────────────────────
with open("Footwear_Creator_Profile_Metrics.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow([
        "#", "Creator Handle", "Brands Collaborated With", "Full Name", "Verified", "Business / Professional",
        "Total Followers", "Following", "Total Posts", "Avg Likes / Post", "Avg Comments / Post", "Avg Profile ER%", "Instagram Profile URL"
    ])
    for idx, p in enumerate(profiles_results, 1):
        w.writerow([
            idx,
            p["handle"],
            p["brands"],
            p["full_name"],
            "Yes" if p["verified"] else "No",
            "Yes" if p["is_business"] else "No",
            p["followers"],
            p["following"],
            p["total_posts"],
            p["avg_likes"],
            p["avg_comments"],
            f"{p['avg_er']:.2f}%",
            p["profile_url"]
        ])
print("✓ Saved Footwear_Creator_Profile_Metrics.csv")

with open("Footwear_All_Brands_4Tier_Master.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow([
        "#", "Hierarchy Tier", "Brand Name", "State / Origin (HQ)", "Creator Handle", "Followers",
        "Views / Plays", "Likes", "Comments", "Like-to-View %", "Creator ER%",
        "Post Date", "Direct Instagram URL", "Boost Status & Reason", "Caption Preview"
    ])
    for idx, p in enumerate(posts, 1):
        w.writerow([
            idx,
            p["tier_name"],
            p["brand"],
            p["state"],
            p["handle"],
            p["followers"],
            p["views"],
            p["likes"],
            p["comments"],
            f"{p['like_rate_pct']:.2f}%",
            f"{p['er_pct']:.2f}%",
            p["date"],
            p["url"],
            f"{p['boost_status']}: {p['boost_reason']}" if p.get('boost_reason') else p['boost_status'],
            p["caption"]
        ])
print("✓ Saved Footwear_All_Brands_4Tier_Master.csv")

with open("Footwear_Brand_Summary.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow([
        "#", "Brand Name", "State / Origin (HQ)", "Total Collab Posts (1-Yr)", "Total Unique Creators",
        "Tier 1 (Toggle ON + Boosted)", "Tier 2 (Toggle ON + Organic)",
        "Tier 3 (Toggle OFF + Boosted)", "Tier 4 (Noise / Unboosted)",
        "Total High-Intent Paid Posts", "High-Intent Paid %",
        "Avg Estimated Views", "Avg Creator Followers", "Avg Creator ER%", "Top Creator Samples"
    ])
    for idx, b in enumerate(brand_summary_data, 1):
        w.writerow([
            idx,
            b["brand"],
            b["state"],
            b["posts"],
            b["creators"],
            f"{b['t1_p']} posts ({b['t1_c']} creators)" if b["t1_p"] > 0 else "—",
            f"{b['t2_p']} posts ({b['t2_c']} creators)" if b["t2_p"] > 0 else "—",
            f"{b['t3_p']} posts ({b['t3_c']} creators)" if b["t3_p"] > 0 else "—",
            f"{b['t4_p']} posts ({b['t4_c']} creators)" if b["t4_p"] > 0 else "—",
            b["high_intent_p"],
            f"{b['high_intent_pct']*100:.1f}%",
            b["avg_views"],
            b["avg_followers"],
            f"{b['avg_er']:.2f}%",
            b["top_creators"]
        ])
print("✓ Saved Footwear_Brand_Summary.csv")
