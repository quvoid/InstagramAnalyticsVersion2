"""
Instagram Bulk Metrics Scraper — Standalone Script
====================================================
✏️  HOW TO USE:
    1. Fill in POSTS below with ("Name / Label", "Instagram URL") pairs
    2. Update COOKIES with your Instagram session cookies
    3. Run:  python scrape_bulk.py
    4. Excel file will be saved next to this script

Install deps:
    pip install curl_cffi openpyxl requests
"""

# ════════════════════════════════════════════════════════════════
#  ✏️  STEP 1: ADD YOUR POSTS HERE
#  Format: ("Name / Campaign Label", "Instagram Post or Reel URL")
# ════════════════════════════════════════════════════════════════

POSTS = [
    # ("Name Placeholder",           "URL Placeholder"),
    ("Rajiv Makhni",               "https://www.instagram.com/reels/Dazdx9LRzIZ/"),
    ("Nandu Patil - Tech Marathi", "https://www.instagram.com/p/Da1n-orIKMG/"),
    # Add more rows below ↓
    # ("Another Creator",          "https://www.instagram.com/p/XXXXXXXX/"),
]

# ════════════════════════════════════════════════════════════════
#  ✏️  STEP 2: PASTE YOUR INSTAGRAM COOKIES HERE
#  Chrome → F12 → Application → Cookies → instagram.com
# ════════════════════════════════════════════════════════════════

COOKIES = {
    "sessionid":  "25113411270%3AyQFaao428g6Xb9%3A0%3AAYgcV3Qe5uiiJD1FyRkSD2vULzcMkEegPbBVeDmWYQ",
    "csrftoken":  "3gJbkGDZp99lA8QQ0brobyoHzOreuu8f",
    "mid":        "afyCbwALAAFRStE-k17-dfO5_jfa",
    "ds_user_id": "25113411270",
}

# ════════════════════════════════════════════════════════════════
#  ✏️  STEP 3: OPTIONS
# ════════════════════════════════════════════════════════════════

FETCH_COMMENTS   = True   # Set False to skip comment scraping
MAX_COMMENTS     = 50     # Max comments to fetch per post
OUTPUT_FILENAME  = "ig_bulk_metrics.xlsx"   # Output Excel filename

# ════════════════════════════════════════════════════════════════
#  BELOW: Don't touch unless you know what you're doing 🙂
# ════════════════════════════════════════════════════════════════

import sys, io, re, html, time, random, json
from datetime import datetime, timezone

sys.stdout.reconfigure(encoding="utf-8")

try:
    from curl_cffi import requests as cffi_requests
    USE_CURL_CFFI = True
    print("✓ curl_cffi found — using Chrome TLS fingerprint")
except ImportError:
    import requests as cffi_requests
    USE_CURL_CFFI = False
    print("⚠ curl_cffi not found — using standard requests")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IG_APP_ID = "936619743392459"

BASE_HEADERS = {
    "accept":             "*/*",
    "accept-language":    "en-US,en;q=0.9",
    "user-agent":         (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "x-ig-app-id":        IG_APP_ID,
    "x-ig-www-claim":     "0",
    "x-requested-with":   "XMLHttpRequest",
    "x-csrftoken":        COOKIES.get("csrftoken", ""),
    "origin":             "https://www.instagram.com",
    "referer":            "https://www.instagram.com/",
    "sec-fetch-dest":     "empty",
    "sec-fetch-mode":     "cors",
    "sec-fetch-site":     "same-origin",
}


# ── Utilities ─────────────────────────────────────────────────

def make_session():
    if USE_CURL_CFFI:
        s = cffi_requests.Session(impersonate="chrome120")
    else:
        import requests
        s = requests.Session()
    s.headers.update(BASE_HEADERS)
    s.cookies.update(COOKIES)
    return s

def extract_shortcode(url: str) -> str:
    m = re.search(r'/(?:p|reel|reels|tv|share/p)/([A-Za-z0-9_-]+)', url.strip())
    if m:
        return m.group(1)
    return url.strip().split("?")[0].strip("/").split("/")[-1]

def shortcode_to_id(sc: str) -> int:
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-_"
    n = 0
    for ch in sc:
        n = n * 64 + alphabet.index(ch)
    return n

def parse_num(s) -> int:
    if not s: return 0
    s = str(s).strip().replace(",", "").upper()
    try:
        if "M" in s: return int(float(s.replace("M","")) * 1_000_000)
        if "K" in s: return int(float(s.replace("K","")) * 1_000)
        return int(float(s))
    except Exception:
        return 0

def delay(a=0.8, b=2.0):
    time.sleep(random.uniform(a, b))

def fetch_profile(username: str, session) -> dict:
    url = f"https://www.instagram.com/api/v1/users/web_profile_info/?username={username}"
    hdrs = {**BASE_HEADERS, "referer": f"https://www.instagram.com/{username}/"}
    for attempt in range(1, 4):
        try:
            r = session.get(url, headers=hdrs, cookies=COOKIES, timeout=18)
            if r.status_code == 200:
                return r.json()
            elif r.status_code == 429:
                wait = 30 * attempt
                print(f"  ⏳ Rate limited — waiting {wait}s...")
                time.sleep(wait)
        except Exception as e:
            print(f"  ⚠ Profile fetch error: {e}")
            delay(2, 4)
    return {}


# ── Post Details Fetcher ──────────────────────────────────────

def fetch_post(label: str, url: str, session, profile_cache: dict) -> dict:
    sc      = extract_shortcode(url)
    mid     = shortcode_to_id(sc)
    clean   = f"https://www.instagram.com/p/{sc}/"
    hdrs    = {**BASE_HEADERS, "x-csrftoken": COOKIES.get("csrftoken","")}

    result = {
        "label":       label,
        "shortcode":   sc,
        "url":         clean,
        "username":    "N/A",
        "full_name":   "N/A",
        "title":       "N/A",
        "caption":     "",
        "type":        "Image",
        "date":        "N/A",
        "likes":       0,
        "comments":    0,
        "video_views": 0,
        "followers":   0,
        "er_percent":  0.0,
        "comments_list": [],
    }

    # Strategy 1: HTML page meta tags (full multi-line caption)
    try:
        r = session.get(clean, headers=hdrs, cookies=COOKIES, timeout=12)
        if r.status_code == 200:
            txt = r.text
            # Title / username from og:title
            og_t = re.search(r'<meta\s+(?:name|property)="og:title"\s+content="([^"]+)"', txt, re.I) or \
                   re.search(r'content="([^"]+)"\s+(?:name|property)="og:title"', txt, re.I)
            if og_t:
                v = html.unescape(og_t.group(1))
                if " on Instagram" in v:
                    result["full_name"] = v.split(" on Instagram")[0].strip()
                    um = re.search(r'\(@([A-Za-z0-9._]+)\)', v)
                    if um: result["username"] = um.group(1)

            # Full caption from og:description
            meta_m = (
                re.search(r'<meta\s+(?:name|property)="og:description"\s+content="(.*?)"\s*/?>', txt, re.DOTALL | re.I) or
                re.search(r'<meta\s+(?:name|property)="description"\s+content="(.*?)"\s*/?>', txt, re.DOTALL | re.I) or
                re.search(r'content="(.*?)"\s+(?:name|property)="(?:og:)?description"', txt, re.DOTALL | re.I)
            )
            if meta_m:
                raw = html.unescape(meta_m.group(1))
                lm = re.search(r'([\d,KkMm.]+)\s+likes', raw)
                cm = re.search(r'([\d,KkMm.]+)\s+comments', raw)
                um = re.search(r'-\s+([A-Za-z0-9._]+)\s+on\s+([^:]+):', raw) or re.search(r'([A-Za-z0-9._]+)\s+on\s+([^:]+):', raw)
                if lm: result["likes"]    = parse_num(lm.group(1))
                if cm: result["comments"] = parse_num(cm.group(1))
                if um:
                    if result["username"] == "N/A": result["username"] = um.group(1)
                    result["date"] = um.group(2).strip()
                cap_m = re.search(r':\s*[""](.*)', raw, re.DOTALL)
                if cap_m:
                    cap = cap_m.group(1).strip()
                    for trail in ['".', '"', '"']:
                        if cap.endswith(trail): cap = cap[:-len(trail)].strip(); break
                    result["caption"] = cap
    except Exception as e:
        print(f"  ⚠ HTML fetch error: {e}")

    # Strategy 2: Mobile API (backup for caption / username)
    if not result["caption"] or result["username"] == "N/A":
        mob_hdrs = {
            "User-Agent": "Instagram 269.0.0.18.75 Android (26/8.0.0; 480dpi; 1080x1920; OnePlus; ONEPLUS A3003; OnePlus3; qcom; en_US; 314665256)",
            "x-ig-app-id": IG_APP_ID,
        }
        try:
            r2 = session.get(f"https://i.instagram.com/api/v1/media/{mid}/info/", headers=mob_hdrs, cookies=COOKIES, timeout=10)
            if r2.status_code == 200:
                items = r2.json().get("items", [])
                if items:
                    item = items[0]
                    u    = item.get("user", {})
                    cap  = (item.get("caption") or {})
                    mt   = {1:"Image", 2:"Video", 8:"Carousel"}
                    if result["username"] == "N/A": result["username"]  = u.get("username","N/A")
                    if result["full_name"] == "N/A": result["full_name"] = u.get("full_name","N/A")
                    if not result["likes"]:    result["likes"]    = item.get("like_count",0) or 0
                    if not result["comments"]: result["comments"] = item.get("comment_count",0) or 0
                    result["video_views"] = item.get("view_count",0) or item.get("play_count",0) or 0
                    result["type"]        = mt.get(item.get("media_type",1),"Image")
                    if not result["caption"] and isinstance(cap, dict):
                        result["caption"] = cap.get("text","")
                    ts = item.get("taken_at",0)
                    if ts and result["date"] == "N/A":
                        result["date"] = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d")
        except Exception as e:
            print(f"  ⚠ Mobile API error: {e}")

    # Populate title from first line of caption
    if result["caption"]:
        lines = [l.strip() for l in result["caption"].splitlines() if l.strip()]
        if lines: result["title"] = lines[0][:100]
    if result["title"] == "N/A":
        result["title"] = f"Instagram Post ({sc})"

    if "/reel/" in url or "/reels/" in url:
        result["type"] = "Video"

    # Fetch follower count for ER%
    uname = result["username"]
    if uname != "N/A":
        if uname in profile_cache:
            result["followers"] = profile_cache[uname]
        else:
            try:
                pr = fetch_profile(uname, session)
                fc = pr.get("data",{}).get("user",{}).get("edge_followed_by",{}).get("count",0)
                profile_cache[uname] = fc
                result["followers"] = fc
            except Exception:
                pass

    if result["followers"] > 0 and result["likes"] > 0:
        result["er_percent"] = round((result["likes"] + result["comments"]) / result["followers"] * 100, 4)

    return result


# ── Comments Fetcher ──────────────────────────────────────────

def fetch_comments(url: str, session, max_count=50) -> list:
    sc  = extract_shortcode(url)
    mid = shortcode_to_id(sc)
    hdrs = {**BASE_HEADERS, "x-csrftoken": COOKIES.get("csrftoken","")}
    out = []
    has_more = True
    min_id = None

    while len(out) < max_count and has_more:
        q = f"https://www.instagram.com/api/v1/media/{mid}/comments/?can_support_threading=true"
        if min_id: q += f"&min_id={min_id}"
        try:
            r = session.get(q, headers=hdrs, cookies=COOKIES, timeout=12)
            if r.status_code == 200:
                data = r.json()
                for c in data.get("comments", []):
                    u  = c.get("user", {})
                    ts = c.get("created_at", 0)
                    dt = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S") if ts else "N/A"
                    out.append({
                        "commenter_username": u.get("username","N/A"),
                        "commenter_name":     u.get("full_name","N/A"),
                        "text":               c.get("text",""),
                        "date":               dt,
                        "likes":              c.get("comment_like_count",0) or 0,
                    })
                has_more = data.get("has_more_comments", False)
                min_id   = data.get("next_min_id")
                if not min_id: break
                delay(0.3, 1.0)
            else:
                break
        except Exception:
            break

    return out[:max_count]


# ── Excel Builder ─────────────────────────────────────────────

def build_excel(posts: list) -> bytes:
    wb   = openpyxl.Workbook()
    thin = Side(style="thin", color="D0D0D0")
    tbdr = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="AAAAAA"))
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(h): return PatternFill("solid", fgColor=h)
    def bf(sz=10, c="000000"): return Font(name="Calibri", bold=True,  size=sz, color=c)
    def nf(sz=10, c="000000"): return Font(name="Calibri", bold=False, size=sz, color=c)
    def lnk(sz=10): return Font(name="Calibri", bold=False, size=sz, color="0563C1", underline="single")
    def ctr(w=False): return Alignment(horizontal="center", vertical="center", wrap_text=w)
    def lft(w=True):  return Alignment(horizontal="left",   vertical="center", wrap_text=w)
    def rgt(w=False): return Alignment(horizontal="right",  vertical="center", wrap_text=w)

    n          = len(posts)
    total_l    = sum(p["likes"]    for p in posts)
    total_c    = sum(p["comments"] for p in posts)
    total_v    = sum(p["video_views"] for p in posts)
    valid_er   = [p["er_percent"] for p in posts if p["er_percent"] > 0]
    avg_er     = sum(valid_er) / len(valid_er) if valid_er else 0
    total_cmts = sum(len(p["comments_list"]) for p in posts)
    unique_u   = len(set(p["username"] for p in posts if p["username"] != "N/A"))

    # ══════════════════════════════════════════════
    # Sheet 1 — Post Metrics
    # ══════════════════════════════════════════════
    ws = wb.active
    ws.title = "Post Metrics"
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:P1")
    ws["A1"] = f"Instagram Metrics — Batch Post Analysis ({n} Posts)"
    ws["A1"].font = Font(name="Calibri", bold=True, size=15, color="000000")
    ws["A1"].alignment = ctr(); ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:P2")
    ws["A2"] = (f"Posts Analysed: {n}  |  Unique Accounts: {unique_u}  |  "
                f"Total Likes: {total_l:,}  |  Total Comments: {total_c:,}  |  "
                f"Comments Scraped: {total_cmts:,}  |  Avg ER%: {avg_er:.2f}%")
    ws["A2"].font = Font(name="Calibri", bold=False, size=9, color="444444")
    ws["A2"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18

    hdr_cols = [
        ("#",                  4),
        ("Name / Label",      20),
        ("Account",           16),
        ("Account Name",      20),
        ("Title",             30),
        ("Full Description",  55),
        ("Date",              13),
        ("Type",              11),
        ("Post URL",          45),
        ("Likes",             12),
        ("Comments",          12),
        ("Video Views",       13),
        ("Followers",         13),
        ("Saves",             12),
        ("Shares",            12),
        ("ER%",               10),
    ]
    for col, (h, w) in enumerate(hdr_cols, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = bf(10, "000000"); c.fill = fill("F2F2F2")
        c.alignment = ctr(); c.border = tbdr
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 24

    for i, p in enumerate(posts):
        row = i + 4
        fol = p["followers"]
        vals = [
            i+1,
            p["label"],
            f"@{p['username']}",
            p["full_name"],
            p["title"],
            p["caption"],
            p["date"],
            p["type"],
            p["url"],
            p["likes"],
            p["comments"],
            p["video_views"],
            fol if fol > 0 else "N/A",
            "N/A — private",
            "N/A — private",
            None,
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col); c.border = bdr
            if col == 2:                          # Name / Label  → bold, highlighted
                c.value = val; c.font = bf(10, "1A1A1A"); c.fill = fill("FFF8F0")
                c.alignment = lft(w=True)
            elif col in (5, 6):                   # Title / Description
                c.value = val; c.font = nf(10); c.alignment = lft(w=True)
            elif col == 9:                        # URL
                c.value = val; c.font = lnk(); c.alignment = lft(w=False)
                if val: c.hyperlink = val
            elif col == 16:                       # ER%
                if isinstance(fol, int) and fol > 0:
                    c.value = f"=(J{row}+K{row})/M{row}*100"
                    c.number_format = '0.00"%"'
                else:
                    c.value = f"{p['er_percent']:.2f}%" if p["er_percent"] else "—"
                c.font = bf(10); c.alignment = ctr()
            elif col == 8:                        # Type
                c.value = val; c.font = bf(10); c.alignment = ctr()
            elif col in (10, 11, 12, 13) and isinstance(val, int):
                c.value = val; c.number_format = "#,##0"; c.font = nf(10); c.alignment = rgt()
            elif col in (14, 15):
                c.value = val; c.font = nf(9, "888888"); c.alignment = ctr()
            else:
                c.value = val; c.font = nf(10); c.alignment = ctr()
        ws.row_dimensions[row].height = 28

    # Totals row
    tr = n + 4
    ws.merge_cells(f"A{tr}:I{tr}")
    ws[f"A{tr}"] = "TOTALS / AVERAGES"
    ws[f"A{tr}"].font = bf(10); ws[f"A{tr}"].fill = fill("F9F9F9"); ws[f"A{tr}"].alignment = ctr()
    for col, val in zip([10,11,12,13,14,15,16], [total_l, total_c, total_v, "","","", f"{avg_er:.2f}%"]):
        c = ws.cell(row=tr, column=col)
        c.fill = fill("F9F9F9"); c.border = bdr; c.font = bf(10)
        c.alignment = rgt() if isinstance(val, int) else ctr()
        c.value = val
        if isinstance(val, int): c.number_format = "#,##0"
    ws.row_dimensions[tr].height = 22

    # ══════════════════════════════════════════════
    # Sheet 2 — Full Captions & Descriptions
    # ══════════════════════════════════════════════
    ws2 = wb.create_sheet("Captions & Descriptions")
    ws2.sheet_view.showGridLines = True
    for col, w in zip("ABCDEFGH", [5, 20, 16, 30, 40, 70, 12, 12]):
        ws2.column_dimensions[col].width = w

    ws2.merge_cells("A1:H1")
    ws2["A1"] = "Post Labels, Titles, Captions & Full Descriptions"
    ws2["A1"].font = bf(14); ws2["A1"].alignment = ctr(); ws2.row_dimensions[1].height = 30

    for col, h in enumerate(["#","Name / Label","Account","Title","URL","Full Description / Caption","Likes","ER%"], 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = bf(10); c.fill = fill("F2F2F2"); c.alignment = ctr(); c.border = tbdr
    ws2.row_dimensions[2].height = 24

    for i, p in enumerate(posts):
        row = i + 3
        vals = [i+1, p["label"], f"@{p['username']}", p["title"], p["url"], p["caption"], p["likes"], p["er_percent"]]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=row, column=col, value=val); c.border = bdr
            if col == 2:
                c.font = bf(10); c.fill = fill("FFF8F0"); c.alignment = lft()
            elif col in (4, 6):
                c.font = nf(10); c.alignment = lft()
            elif col == 5:
                c.font = lnk(); c.alignment = lft(w=False)
                if val: c.hyperlink = val
            elif col == 8:
                c.number_format = '0.00"%"'; c.font = bf(10); c.alignment = ctr()
            elif col == 7:
                c.number_format = "#,##0"; c.font = nf(10); c.alignment = rgt()
            else:
                c.font = nf(10); c.alignment = ctr()
        ws2.row_dimensions[row].height = 28

    # ══════════════════════════════════════════════
    # Sheet 3 — Batch Analytics Summary
    # ══════════════════════════════════════════════
    ws3 = wb.create_sheet("Batch Analytics")
    ws3.sheet_view.showGridLines = True
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 55

    ws3.merge_cells("A1:B1")
    ws3["A1"] = "Batch Analytics Summary"
    ws3["A1"].font = bf(14); ws3["A1"].alignment = ctr(); ws3.row_dimensions[1].height = 30

    top_liked = max(posts, key=lambda x: x["likes"]) if posts else {}
    top_er    = max(posts, key=lambda x: x["er_percent"]) if posts else {}

    summary = [
        ("Total Posts",             n),
        ("Unique Accounts",         unique_u),
        ("Total Likes",             f"{total_l:,}"),
        ("Total Comments",          f"{total_c:,}"),
        ("Total Video Views",       f"{total_v:,}"),
        ("Comments Scraped",        f"{total_cmts:,}"),
        ("Avg Likes / Post",        f"{total_l // n:,}" if n else "0"),
        ("Avg Comments / Post",     f"{total_c // n:,}" if n else "0"),
        ("Avg ER%",                 f"{avg_er:.2f}%"),
        ("", ""),
        ("Most Liked — Name",       top_liked.get("label","N/A")),
        ("Most Liked — Account",    f"@{top_liked.get('username','N/A')} ({top_liked.get('likes',0):,} likes)"),
        ("Most Liked — URL",        top_liked.get("url","N/A")),
        ("", ""),
        ("Highest ER% — Name",      top_er.get("label","N/A")),
        ("Highest ER% — Value",     f"{top_er.get('er_percent',0):.2f}% by @{top_er.get('username','N/A')}"),
        ("", ""),
        ("Exported At",             datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (label, value) in enumerate(summary):
        row = i + 2
        if not label:
            ws3.row_dimensions[row].height = 8; continue
        lc = ws3.cell(row=row, column=1, value=label)
        lc.font = bf(10, "333333"); lc.fill = fill("F2F2F2"); lc.alignment = lft(w=False); lc.border = bdr
        vc = ws3.cell(row=row, column=2, value=value)
        vc.font = nf(10); vc.fill = fill("FFFFFF"); vc.alignment = lft(w=False); vc.border = bdr
        ws3.row_dimensions[row].height = 20

    # ══════════════════════════════════════════════
    # Sheet 4 — All Comments
    # ══════════════════════════════════════════════
    all_cmts = [(p, c) for p in posts for c in p["comments_list"]]
    if all_cmts:
        ws4 = wb.create_sheet("All Comments")
        ws4.sheet_view.showGridLines = True
        for col, w in zip("ABCDEFGHIJ", [5, 20, 16, 30, 40, 20, 22, 65, 20, 12]):
            ws4.column_dimensions[col].width = w

        ws4.merge_cells("A1:J1")
        ws4["A1"] = f"Extracted Comments — {len(all_cmts)} Total Comments"
        ws4["A1"].font = bf(14); ws4["A1"].alignment = ctr(); ws4.row_dimensions[1].height = 30

        for col, h in enumerate(["#","Name / Label","Post Account","Post Title","Post URL",
                                   "Commenter Username","Commenter Name","Comment Text","Comment Date","Likes"], 1):
            c = ws4.cell(row=2, column=col, value=h)
            c.font = bf(10); c.fill = fill("F2F2F2"); c.alignment = ctr(); c.border = tbdr
        ws4.row_dimensions[2].height = 24

        for i, (p, c) in enumerate(all_cmts, 1):
            row = i + 2
            vals = [i, p["label"], f"@{p['username']}", p["title"], p["url"],
                    f"@{c['commenter_username']}", c["commenter_name"],
                    c["text"], c["date"], c["likes"]]
            for col, val in enumerate(vals, 1):
                cell = ws4.cell(row=row, column=col, value=val); cell.border = bdr
                if col == 2:
                    cell.font = bf(10); cell.fill = fill("FFF8F0"); cell.alignment = lft()
                elif col in (4, 8):
                    cell.font = nf(10); cell.alignment = lft()
                elif col == 5:
                    cell.font = lnk(); cell.alignment = lft(w=False)
                    if val: cell.hyperlink = val
                elif col == 10:
                    cell.number_format = "#,##0"; cell.font = nf(10); cell.alignment = rgt()
                else:
                    cell.font = nf(10); cell.alignment = ctr()
            ws4.row_dimensions[row].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── MAIN RUNNER ───────────────────────────────────────────────

def main():
    if not POSTS:
        print("❌ No posts defined. Add entries to the POSTS list at the top of this file.")
        return

    print(f"\n{'═'*58}")
    print(f"  Instagram Bulk Metrics Scraper")
    print(f"  {len(POSTS)} posts  |  Comments: {'Yes' if FETCH_COMMENTS else 'No'}  |  Max per post: {MAX_COMMENTS}")
    print(f"{'═'*58}\n")

    session  = make_session()
    prof_cache = {}
    results  = []

    for idx, (label, url) in enumerate(POSTS, 1):
        print(f"[{idx}/{len(POSTS)}] Fetching: {label!r}")
        print(f"         URL: {url}")

        post = fetch_post(label, url, session, prof_cache)
        print(f"         ✓ @{post['username']} | {post['likes']:,} likes | {post['comments']:,} comments | ER: {post['er_percent']:.2f}%")

        if FETCH_COMMENTS:
            print(f"         💬 Scraping comments (max {MAX_COMMENTS})...")
            cmts = fetch_comments(url, session, max_count=MAX_COMMENTS)
            post["comments_list"] = cmts
            print(f"         ✓ {len(cmts)} comments fetched")

        results.append(post)
        if idx < len(POSTS):
            delay(1.0, 2.5)
        print()

    print("📊 Building Excel report...")
    excel_bytes = build_excel(results)

    with open(OUTPUT_FILENAME, "wb") as f:
        f.write(excel_bytes)

    print(f"\n{'═'*58}")
    print(f"  ✅ Done! Saved: {OUTPUT_FILENAME}")
    print(f"  Posts: {len(results)}  |  Total comments scraped: {sum(len(p['comments_list']) for p in results):,}")
    print(f"  Sheets: Post Metrics · Captions · Batch Analytics · All Comments")
    print(f"{'═'*58}\n")


if __name__ == "__main__":
    main()
