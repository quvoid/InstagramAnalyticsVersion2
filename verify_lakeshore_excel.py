"""
Verify lakeshore_kopa_paid_media_intelligence.xlsx
"""

import openpyxl

wb = openpyxl.load_workbook("lakeshore_kopa_paid_media_intelligence.xlsx", read_only=True)
print("="*70)
print("WORKBOOK VERIFICATION: lakeshore_kopa_paid_media_intelligence.xlsx")
print("="*70)
print(f"Total Sheets: {len(wb.sheetnames)}\n")

for name in wb.sheetnames:
    ws = wb[name]
    print(f"  • Tab: {name:<35} | Total Rows: {ws.max_row}")

print("="*70)
