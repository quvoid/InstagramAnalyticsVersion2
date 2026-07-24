"""
Instagram Metrics Scraper — Streamlit App
==========================================
Run with:  streamlit run app.py

Install:
    pip install streamlit curl_cffi requests openpyxl
"""

import streamlit as st
import time, random, json, io, re, html
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from curl_cffi import requests as cffi_requests
    USE_CURL_CFFI = True
except ImportError:
    import requests as cffi_requests
    USE_CURL_CFFI = False

# ── PAGE CONFIG ───────────────────────────────────────────────
st.set_page_config(
    page_title="Instagram Metrics",
    page_icon="📊",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ── CUSTOM CSS ────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&display=swap');

*, html, body, [class*="css"] {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
}

.stApp {
    background: #0A0A0A;
}

/* Hide entire header bar */
header[data-testid="stHeader"] {
    display: none !important;
}

/* Hide sidebar and collapsed controls */
[data-testid="stSidebar"],
[data-testid="collapsedControl"],
[data-testid="stSidebarCollapsedControl"],
section[data-testid="stSidebarCollapsedControl"],
[class*="collapsedControl"],
[class*="sidebarCollapsed"] {
    display: none !important;
    width: 0 !important;
    overflow: hidden !important;
}

/* Nuke stray Material Icons text fallback */
.material-symbols-rounded,
.material-icons {
    font-size: 0 !important;
    visibility: hidden !important;
}

/* Inputs & Textareas */
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea > div > div > textarea {
    background: #141414 !important;
    border: 1px solid #282828 !important;
    border-radius: 10px !important;
    color: #E8E8E8 !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-size: 14px !important;
    padding: 12px 16px !important;
}
.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
    border-color: #E8405A !important;
    box-shadow: 0 0 0 2px rgba(232,64,90,0.15) !important;
}
.stTextInput label, .stNumberInput label, .stTextArea label {
    color: #888 !important;
    font-size: 11px !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-weight: 600 !important;
    text-transform: uppercase;
    letter-spacing: 0.09em;
}

/* Radio Mode Switcher */
div[data-testid="stRadio"] > label {
    display: none !important;
}
div[data-testid="stRadio"] > div {
    background: #141414 !important;
    padding: 4px !important;
    border-radius: 12px !important;
    border: 1px solid #222 !important;
    display: flex !important;
    gap: 4px !important;
}
div[data-testid="stRadio"] label[data-baseweb="radio"] {
    flex: 1 !important;
    text-align: center !important;
    background: transparent !important;
    border-radius: 8px !important;
    padding: 10px 16px !important;
    cursor: pointer !important;
    transition: all 0.2s ease !important;
}
div[data-testid="stRadio"] label[data-baseweb="radio"] span {
    color: #888 !important;
    font-weight: 600 !important;
    font-size: 13px !important;
}

/* Buttons */
.stButton > button {
    background: linear-gradient(135deg, #E8405A, #FF6B35) !important;
    color: #fff !important;
    border: none !important;
    border-radius: 10px !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-weight: 700 !important;
    font-size: 14px !important;
    padding: 13px 28px !important;
    width: 100% !important;
    letter-spacing: 0.03em;
    transition: all 0.2s ease !important;
}
.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 8px 24px rgba(232,64,90,0.3) !important;
}

.stDownloadButton > button {
    background: linear-gradient(135deg, #1DB954, #17A449) !important;
    color: #fff !important;
    border: none !important;
    border-radius: 10px !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-weight: 700 !important;
    font-size: 14px !important;
    padding: 13px 28px !important;
    width: 100% !important;
    letter-spacing: 0.03em;
    transition: all 0.2s ease !important;
}
.stDownloadButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 8px 24px rgba(29,185,84,0.3) !important;
}

/* Expander */
.streamlit-expanderHeader {
    background: #141414 !important;
    border-radius: 10px !important;
    color: #CCC !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
}

/* Spinner */
.stSpinner > div { border-top-color: #E8405A !important; }

/* Scrollbar */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #0A0A0A; }
::-webkit-scrollbar-thumb { background: #2A2A2A; border-radius: 3px; }

/* Metric grid */
.metric-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 10px;
    margin: 18px 0;
}
.metric-card {
    background: #141414;
    border: 1px solid #1E1E1E;
    border-radius: 12px;
    padding: 18px 14px;
    text-align: center;
}
.metric-value {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 24px;
    font-weight: 800;
    color: #E8E8E8;
    line-height: 1;
    letter-spacing: -0.02em;
}
.metric-label {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 10px;
    font-weight: 600;
    color: #555;
    margin-top: 6px;
    text-transform: uppercase;
    letter-spacing: 0.1em;
}
.metric-er .metric-value { color: #E8405A; }
.metric-followers .metric-value { color: #FF6B35; }

/* Profile card */
.profile-card {
    background: #141414;
    border: 1px solid #202020;
    border-radius: 14px;
    padding: 22px 20px;
    margin: 18px 0;
}
.profile-name {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 20px;
    font-weight: 800;
    color: #F0F0F0;
    letter-spacing: -0.01em;
}
.profile-username {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 13px;
    font-weight: 500;
    color: #E8405A;
    margin-top: 3px;
}
.profile-bio {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 13px;
    font-weight: 400;
    color: #777;
    margin-top: 10px;
    line-height: 1.55;
}
.badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 11px;
    font-weight: 600;
    margin-right: 6px;
    margin-top: 8px;
}
.badge-verified  { background: #1a2d40; color: #4fc3f7; }
.badge-business  { background: #2d1b00; color: #FF8C42; }
.badge-category  { background: #1e1228; color: #B07FE8; }

/* Section label */
.section-label {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 11px;
    font-weight: 700;
    color: #444;
    text-transform: uppercase;
    letter-spacing: 0.14em;
    margin: 26px 0 10px;
}

/* Post table */
.post-table {
    background: #0E0E0E;
    border: 1px solid #1A1A1A;
    border-radius: 12px;
    overflow: hidden;
}
.post-row {
    display: grid;
    grid-template-columns: 32px 105px 125px 1fr 70px 65px 65px 65px 60px;
    align-items: center;
    padding: 10px 12px;
    border-bottom: 1px solid #161616;
    gap: 6px;
}
.post-row-profile {
    grid-template-columns: 36px 90px 80px 1fr 80px 80px 68px !important;
}
.post-row:last-child { border-bottom: none; }
.post-row-header {
    background: #141414;
    border-bottom: 1px solid #202020 !important;
}
.post-row-header span {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 10px;
    font-weight: 700;
    color: #555;
    text-transform: uppercase;
    letter-spacing: 0.1em;
}
.post-cell {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 12px;
    font-weight: 400;
    color: #BDBDBD;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.post-cell-title {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 12px;
    font-weight: 600;
    color: #E0E0E0;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.post-cell-desc {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 11px;
    font-weight: 400;
    color: #888888;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.post-cell-user {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 12px;
    font-weight: 600;
    color: #E8405A;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.post-cell-num {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 12px;
    font-weight: 600;
    color: #D0D0D0;
    text-align: center;
}
.post-cell-idx {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 11px;
    color: #444;
    text-align: center;
}
.post-cell-er {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 12px;
    font-weight: 700;
    color: #E8405A;
    text-align: center;
}
.type-image    { color: #4A9EFF; font-weight: 600; }
.type-video    { color: #A78BFA; font-weight: 600; }
.type-carousel { color: #FBBF24; font-weight: 600; }
.post-link {
    color: #4A9EFF;
    text-decoration: none;
    font-size: 12px;
    font-family: 'Plus Jakarta Sans', sans-serif;
}
.post-link:hover { text-decoration: underline; }
</style>
""", unsafe_allow_html=True)


# ── COOKIES ───────────────────────────────────────────────────
DEFAULT_COOKIES = {
    "sessionid":  "25113411270%3AyQFaao428g6Xb9%3A0%3AAYgcV3Qe5uiiJD1FyRkSD2vULzcMkEegPbBVeDmWYQ",
    "csrftoken":  "3gJbkGDZp99lA8QQ0brobyoHzOreuu8f",
    "mid":        "afyCbwALAAFRStE-k17-dfO5_jfa",
    "ds_user_id": "25113411270",
}

# Override from Streamlit Cloud Secrets if configured
try:
    if "sessionid" in st.secrets:  DEFAULT_COOKIES["sessionid"]  = st.secrets["sessionid"]
    if "csrftoken" in st.secrets:  DEFAULT_COOKIES["csrftoken"]  = st.secrets["csrftoken"]
    if "mid" in st.secrets:        DEFAULT_COOKIES["mid"]        = st.secrets["mid"]
    if "ds_user_id" in st.secrets: DEFAULT_COOKIES["ds_user_id"] = st.secrets["ds_user_id"]
except Exception:
    pass

IG_APP_ID = "936619743392459"

BASE_HEADERS = {
    "accept":             "*/*",
    "accept-language":    "en-US,en;q=0.9",
    "accept-encoding":    "gzip, deflate, br",
    "user-agent":         (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "x-ig-app-id":        IG_APP_ID,
    "x-ig-www-claim":     "0",
    "x-asbd-id":          "129477",
    "x-requested-with":   "XMLHttpRequest",
    "x-csrftoken":        "",
    "origin":             "https://www.instagram.com",
    "referer":            "https://www.instagram.com/",
    "sec-ch-ua":          '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
    "sec-ch-ua-mobile":   "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest":     "empty",
    "sec-fetch-mode":     "cors",
    "sec-fetch-site":     "same-origin",
}


# ── HELPERS ───────────────────────────────────────────────────

def extract_username(raw: str) -> str:
    raw = raw.strip().lstrip("@")
    if "instagram.com" in raw:
        parts = [p for p in raw.replace("https://", "").replace("http://", "").split("/") if p and "instagram.com" not in p]
        raw = parts[0] if parts else raw
    return raw.split("?")[0].strip()

def extract_shortcode(url: str) -> str:
    url = url.strip()
    match = re.search(r'/(?:p|reel|reels|tv|share/p)/([A-Za-z0-9_-]+)', url)
    if match:
        return match.group(1)
    clean = url.split("?")[0].strip("/").split("/")[-1]
    return clean

def extract_urls(raw_text: str) -> list[str]:
    lines = raw_text.strip().splitlines()
    urls = []
    for line in lines:
        line_str = line.strip()
        if not line_str:
            continue
        tokens = [t.strip(",; ") for t in line_str.split()]
        for tok in tokens:
            if "instagram.com" in tok or len(tok) >= 5:
                sc = extract_shortcode(tok)
                if sc and sc not in [extract_shortcode(u) for u in urls]:
                    urls.append(f"https://www.instagram.com/p/{sc}/")
    return urls

def shortcode_to_id(shortcode: str) -> int:
    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-_'
    media_id = 0
    for char in shortcode:
        media_id = (media_id * 64) + alphabet.index(char)
    return media_id

def parse_num(s: str) -> int:
    if not s: return 0
    s = s.strip().replace(',', '').upper()
    try:
        if 'M' in s:
            return int(float(s.replace('M', '')) * 1_000_000)
        if 'K' in s:
            return int(float(s.replace('K', '')) * 1_000)
        return int(float(s))
    except Exception:
        return 0

def human_delay(a=1.0, b=3.0):
    time.sleep(random.uniform(a, b))

def make_session(cookies: dict):
    headers = {**BASE_HEADERS, "x-csrftoken": cookies.get("csrftoken", "")}
    if USE_CURL_CFFI:
        s = cffi_requests.Session(impersonate="chrome120")
    else:
        import requests
        s = requests.Session()
    s.headers.update(headers)
    s.cookies.update(cookies)
    return s


# ── PROFILE SCRAPER FUNCTIONS ─────────────────────────────────

def fetch_profile(username, session, cookies, status_fn=None):
    urls = [
        f"https://www.instagram.com/api/v1/users/web_profile_info/?username={username}",
        f"https://i.instagram.com/api/v1/users/web_profile_info/?username={username}",
    ]
    headers = {**BASE_HEADERS, "x-csrftoken": cookies.get("csrftoken", ""),
               "referer": f"https://www.instagram.com/{username}/"}

    for url in urls:
        for attempt in range(1, 4):
            try:
                if status_fn: status_fn(f"Fetching profile (attempt {attempt}/3)...")
                if USE_CURL_CFFI:
                    r = session.get(url, headers=headers, cookies=cookies,
                                    timeout=20, impersonate="chrome120")
                else:
                    r = session.get(url, headers=headers, cookies=cookies, timeout=20)

                if r.status_code == 200:
                    return r.json()
                elif r.status_code in (401, 403, 404):
                    return {"_error": r.status_code}
                elif r.status_code == 429:
                    if status_fn: status_fn(f"Rate limited. Waiting {30*attempt}s...")
                    time.sleep(30 * attempt)
                else:
                    human_delay(2, 5)
            except Exception:
                human_delay(2, 4)
    return None


def fetch_posts_v2(user_id, session, cookies, count=12, status_fn=None):
    url = f"https://www.instagram.com/api/v1/feed/user/{user_id}/?count={count}"
    headers = {**BASE_HEADERS, "x-csrftoken": cookies.get("csrftoken", "")}
    for attempt in range(1, 4):
        try:
            human_delay(1, 2.5)
            if status_fn: status_fn(f"Fetching posts via fallback feed (attempt {attempt}/3)...")
            if USE_CURL_CFFI:
                r = session.get(url, headers=headers, cookies=cookies,
                                timeout=20, impersonate="chrome120")
            else:
                r = session.get(url, headers=headers, cookies=cookies, timeout=20)
            if r.status_code == 200:
                return [parse_post_v2(i) for i in r.json().get("items", [])]
            elif r.status_code == 429:
                time.sleep(30 * attempt)
            else:
                human_delay(3, 5)
        except Exception:
            human_delay(2, 4)
    return []


def fetch_posts_v2_paginated(user_id, session, cookies, max_id, count=12):
    url = f"https://www.instagram.com/api/v1/feed/user/{user_id}/?count={count}&max_id={max_id}"
    headers = {**BASE_HEADERS, "x-csrftoken": cookies.get("csrftoken", "")}
    try:
        human_delay(2, 4)
        if USE_CURL_CFFI:
            r = session.get(url, headers=headers, cookies=cookies,
                            timeout=20, impersonate="chrome120")
        else:
            r = session.get(url, headers=headers, cookies=cookies, timeout=20)
        if r.status_code == 200:
            data = r.json()
            posts = [parse_post_v2(i) for i in data.get("items", [])]
            next_id = data.get("next_max_id") if data.get("more_available") else None
            return posts, next_id
    except Exception:
        pass
    return [], None


def fetch_more_posts_gql(user_id, end_cursor, session, cookies, count=12):
    QUERY_HASHES = [
        "e7e2f4da4b02303f74f0841279e52d76",
        "69cba40317214236af40e7efa9ca7448",
    ]
    variables = json.dumps({"id": user_id, "first": count, "after": end_cursor},
                           separators=(",", ":"))
    headers = {**BASE_HEADERS, "x-csrftoken": cookies.get("csrftoken", "")}
    for qhash in QUERY_HASHES:
        url = f"https://www.instagram.com/graphql/query/?query_hash={qhash}&variables={variables}"
        try:
            human_delay(2, 4)
            if USE_CURL_CFFI:
                r = session.get(url, headers=headers, cookies=cookies,
                                timeout=20, impersonate="chrome120")
            else:
                r = session.get(url, headers=headers, cookies=cookies, timeout=20)
            if r.status_code == 200:
                return r.json()
        except Exception:
            pass
    return None


def parse_post(node):
    likes    = node.get("edge_media_preview_like", {}).get("count", 0) or 0
    comments = node.get("edge_media_to_comment", {}).get("count", 0) or 0
    edges    = node.get("edge_media_to_caption", {}).get("edges", [])
    caption  = edges[0]["node"]["text"] if edges else ""
    mt_map   = {"GraphImage": "Image", "GraphVideo": "Video", "GraphSidecar": "Carousel"}
    ts       = node.get("taken_at_timestamp", 0)
    date_str = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d") if ts else "N/A"
    lines    = [l.strip() for l in caption.splitlines() if l.strip()]
    first_line = lines[0] if lines else "Instagram Post"
    return {
        "shortcode":   node.get("shortcode", ""),
        "url":         f"https://www.instagram.com/p/{node.get('shortcode', '')}/",
        "title":       first_line[:100],
        "type":        mt_map.get(node.get("__typename", ""), "Image"),
        "date":        date_str,
        "likes":       likes,
        "comments":    comments,
        "video_views": node.get("video_view_count", 0) or 0,
        "caption":     caption,
    }


def parse_post_v2(item):
    likes    = item.get("like_count", 0) or 0
    comments = item.get("comment_count", 0) or 0
    cap_obj  = item.get("caption")
    caption  = (cap_obj.get("text", "") if isinstance(cap_obj, dict) else "") or ""
    mt_map   = {1: "Image", 2: "Video", 8: "Carousel"}
    ts       = item.get("taken_at", 0)
    date_str = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d") if ts else "N/A"
    code     = item.get("code", item.get("shortcode", ""))
    lines    = [l.strip() for l in caption.splitlines() if l.strip()]
    first_line = lines[0] if lines else "Instagram Post"
    return {
        "shortcode":   code,
        "url":         f"https://www.instagram.com/p/{code}/",
        "title":       first_line[:100],
        "type":        mt_map.get(item.get("media_type", 1), "Image"),
        "date":        date_str,
        "likes":       likes,
        "comments":    comments,
        "video_views": item.get("view_count", 0) or item.get("play_count", 0) or 0,
        "caption":     caption,
    }


def run_scrape(username, num_posts, cookies, status_fn=None):
    session = make_session(cookies)

    raw = fetch_profile(username, session, cookies, status_fn)
    if not raw:
        return None, None, "Could not reach Instagram. Check your internet connection."
    if isinstance(raw, dict) and "_error" in raw:
        code = raw["_error"]
        if code == 401: return None, None, "401 — Cookies expired. Refresh from Chrome DevTools."
        if code == 403: return None, None, "403 — Rate limited or forbidden."
        if code == 404: return None, None, f"404 — @{username} not found. Check the username."
        return None, None, f"HTTP {code} error."

    user = raw.get("data", {}).get("user")
    if not user:
        return None, None, "User not found or account is private."

    profile = {
        "username":    username,
        "full_name":   user.get("full_name", "N/A"),
        "followers":   user.get("edge_followed_by", {}).get("count", 0),
        "following":   user.get("edge_follow", {}).get("count", 0),
        "total_posts": user.get("edge_owner_to_timeline_media", {}).get("count", 0),
        "bio":         user.get("biography", ""),
        "verified":    user.get("is_verified", False),
        "user_id":     user.get("id", ""),
        "category":    user.get("category_name", "N/A"),
        "is_business": user.get("is_business_account", False),
        "external_url":user.get("external_url", ""),
    }

    timeline   = user.get("edge_owner_to_timeline_media", {})
    edges      = timeline.get("edges", [])
    page_info  = timeline.get("page_info", {})
    has_next   = page_info.get("has_next_page", False)
    end_cursor = page_info.get("end_cursor", "")
    posts      = [parse_post(e["node"]) for e in edges]

    if status_fn: status_fn(f"Got {len(posts)} posts from profile. Collecting more...")

    if len(posts) == 0 and profile["user_id"]:
        if status_fn: status_fn("Business/verified account — using fallback feed endpoint...")
        posts = fetch_posts_v2(profile["user_id"], session, cookies,
                               count=min(num_posts, 12), status_fn=status_fn)
        if posts:
            next_max_id = posts[-1]["shortcode"]
            while len(posts) < num_posts:
                if status_fn: status_fn(f"Paginating... {len(posts)}/{num_posts} posts")
                new_posts, next_max_id = fetch_posts_v2_paginated(
                    profile["user_id"], session, cookies,
                    max_id=next_max_id, count=min(12, num_posts - len(posts))
                )
                if not new_posts: break
                posts.extend(new_posts)
                if not next_max_id: break
    elif len(posts) < num_posts:
        while len(posts) < num_posts and has_next and end_cursor:
            if status_fn: status_fn(f"GraphQL pagination... {len(posts)}/{num_posts} posts")
            page_data = fetch_more_posts_gql(
                profile["user_id"], end_cursor, session, cookies,
                count=min(12, num_posts - len(posts))
            )
            if not page_data: break
            media      = page_data.get("data", {}).get("user", {}).get("edge_owner_to_timeline_media", {})
            new_edges  = media.get("edges", [])
            page_info  = media.get("page_info", {})
            has_next   = page_info.get("has_next_page", False)
            end_cursor = page_info.get("end_cursor", "")
            posts.extend([parse_post(e["node"]) for e in new_edges])

    posts = posts[:num_posts]

    if not posts:
        return profile, [], None

    followers = profile["followers"]
    for p in posts:
        p["er_percent"] = round((p["likes"] + p["comments"]) / followers * 100, 4) if followers else 0

    return profile, posts, None


# ── SINGLE / BATCH POST URL SCRAPER FUNCTIONS ────────────────

def fetch_post_details(url: str, session, cookies, profile_cache=None):
    if profile_cache is None:
        profile_cache = {}

    shortcode = extract_shortcode(url)
    clean_url = f"https://www.instagram.com/p/{shortcode}/"
    media_id = shortcode_to_id(shortcode)
    headers = {**BASE_HEADERS, 'x-csrftoken': cookies.get('csrftoken', '')}

    post_data = {
        "shortcode": shortcode,
        "url": clean_url,
        "username": "N/A",
        "full_name": "N/A",
        "title": "N/A",
        "caption": "",
        "type": "Image",
        "date": "N/A",
        "likes": 0,
        "comments": 0,
        "video_views": 0,
        "followers": 0,
        "er_percent": 0.0
    }

    # Step 1: HTML Page GET & Meta Tag Extraction (Gets FULL 100% Description with linebreaks & hashtags)
    try:
        r_page = session.get(clean_url, headers=headers, cookies=cookies, timeout=12)
        if r_page.status_code == 200:
            page_text = r_page.text
            
            # og:title -> Name & Username
            og_t = re.search(r'<meta\s+(?:name|property)="og:title"\s+content="([^"]+)"\s*/?>', page_text, re.IGNORECASE) or \
                   re.search(r'content="([^"]+)"\s+(?:name|property)="og:title"', page_text, re.IGNORECASE)
            if og_t:
                val = html.unescape(og_t.group(1))
                if " on Instagram" in val:
                    name_part = val.split(" on Instagram")[0].strip()
                    post_data["full_name"] = name_part
                    if "@" in val:
                        u_m = re.search(r'\(@([A-Za-z0-9._]+)\)', val)
                        if u_m: post_data["username"] = u_m.group(1)
                elif "• Instagram" in val:
                    post_data["full_name"] = val.split("•")[0].strip()

            # Full multi-line caption from meta description
            meta_m = re.search(r'<meta\s+(?:name|property)="og:description"\s+content="(.*?)"\s*/?>', page_text, re.DOTALL | re.IGNORECASE) or \
                     re.search(r'<meta\s+(?:name|property)="description"\s+content="(.*?)"\s*/?>', page_text, re.DOTALL | re.IGNORECASE) or \
                     re.search(r'content="(.*?)"\s+(?:name|property)="(?:og:)?description"', page_text, re.DOTALL | re.IGNORECASE)
                     
            if meta_m:
                raw_content = meta_m.group(1)
                unescaped = html.unescape(raw_content)

                l_m = re.search(r'([\d,KkMm.]+)\s+likes', unescaped)
                c_m = re.search(r'([\d,KkMm.]+)\s+comments', unescaped)
                u_m = re.search(r'-\s+([A-Za-z0-9._]+)\s+on\s+([^:]+):', unescaped) or re.search(r'([A-Za-z0-9._]+)\s+on\s+([^:]+):', unescaped)

                if l_m: post_data["likes"] = parse_num(l_m.group(1))
                if c_m: post_data["comments"] = parse_num(c_m.group(1))
                if u_m:
                    if post_data["username"] == "N/A": post_data["username"] = u_m.group(1)
                    post_data["date"] = u_m.group(2).strip()

                cap_match = re.search(r':\s*["“](.*)', unescaped, re.DOTALL)
                if cap_match:
                    cap_text = cap_match.group(1).strip()
                    if cap_text.endswith('".') or cap_text.endswith('". ') or cap_text.endswith('".'):
                        cap_text = cap_text[:-2].strip()
                    elif cap_text.endswith('"') or cap_text.endswith('”'):
                        cap_text = cap_text[:-1].strip()
                    post_data["caption"] = cap_text
    except Exception:
        pass

    # Step 2: Mobile Media Info API (Backup for likes/comments/caption if missing)
    if not post_data["caption"] or post_data["username"] == "N/A":
        mobile_headers = {
            'User-Agent': 'Instagram 269.0.0.18.75 Android (26/8.0.0; 480dpi; 1080x1920; OnePlus; ONEPLUS A3003; OnePlus3; qcom; en_US; 314665256)',
            'x-ig-app-id': '936619743392459',
        }
        try:
            r_info = session.get(f"https://i.instagram.com/api/v1/media/{media_id}/info/", headers=mobile_headers, cookies=cookies, timeout=10)
            if r_info.status_code == 200:
                items = r_info.json().get("items", [])
                if items:
                    item = items[0]
                    user_obj = item.get("user", {})
                    cap_obj = item.get("caption") or {}
                    mt_map = {1: "Image", 2: "Video", 8: "Carousel"}
                    
                    if post_data["username"] == "N/A": post_data["username"] = user_obj.get("username", "N/A")
                    if post_data["full_name"] == "N/A": post_data["full_name"] = user_obj.get("full_name", "N/A")
                    if not post_data["followers"]: post_data["followers"] = user_obj.get("follower_count", 0) or 0
                    if not post_data["likes"]: post_data["likes"] = item.get("like_count", 0) or 0
                    if not post_data["comments"]: post_data["comments"] = item.get("comment_count", 0) or 0
                    post_data["video_views"] = item.get("view_count", 0) or item.get("play_count", 0) or 0
                    post_data["type"] = mt_map.get(item.get("media_type", 1), "Image")
                    
                    cap_str = (cap_obj.get("text", "") if isinstance(cap_obj, dict) else "") or ""
                    if cap_str: post_data["caption"] = cap_str
                    
                    ts = item.get("taken_at", 0)
                    if ts and post_data["date"] == "N/A":
                        post_data["date"] = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d")
        except Exception:
            pass

    # Step 3: Embed Page Fallback
    if not post_data["caption"] or post_data["username"] == "N/A":
        try:
            emb_url = f"https://www.instagram.com/p/{shortcode}/embed/captioned/"
            r_emb = session.get(emb_url, headers=headers, cookies=cookies, timeout=10)
            if r_emb.status_code == 200:
                html_text = r_emb.text
                u_m = re.search(r'class="UsernameText"[^>]*>([^<]+)<', html_text) or re.search(r'instagram\.com/([^/"]+)', html_text)
                if u_m and post_data["username"] == "N/A":
                    post_data["username"] = u_m.group(1).strip()
                
                l_m = re.search(r'class="SocialProof[^"]*"[^>]*>([\d,KkMm.]+)', html_text) or re.search(r'([\d,KkMm.]+)\s+likes', html_text)
                if l_m and not post_data["likes"]:
                    post_data["likes"] = parse_num(l_m.group(1))
                    
                cap_m = re.search(r'class="Caption"[^>]*>(.*?)</div>', html_text, re.DOTALL)
                if cap_m and not post_data["caption"]:
                    clean_cap = re.sub(r'<[^>]+>', '', cap_m.group(1))
                    post_data["caption"] = html.unescape(clean_cap).strip()
                    
                if "Video" in html_text or "play" in html_text.lower():
                    post_data["type"] = "Video"
                elif "carousel" in html_text.lower():
                    post_data["type"] = "Carousel"
        except Exception:
            pass

    # Populate Title (First non-empty line of caption)
    if post_data["caption"]:
        lines = [l.strip() for l in post_data["caption"].splitlines() if l.strip()]
        if lines:
            post_data["title"] = lines[0][:100]
    if post_data["title"] == "N/A":
        post_data["title"] = f"Instagram Post ({shortcode})"

    # Check if URL is reel/video
    if "/reel/" in url or "/reels/" in url:
        post_data["type"] = "Video"

    # Fetch follower count for account if username is known
    uname = post_data["username"]
    if uname != "N/A":
        if uname in profile_cache:
            post_data["followers"] = profile_cache[uname]
        else:
            try:
                prof = fetch_profile(uname, session, cookies)
                if prof and "data" in prof:
                    f_count = prof["data"]["user"].get("edge_followed_by", {}).get("count", 0)
                    profile_cache[uname] = f_count
                    post_data["followers"] = f_count
            except Exception:
                pass

    if post_data["followers"] > 0 and post_data["likes"] > 0:
        post_data["er_percent"] = round((post_data["likes"] + post_data["comments"]) / post_data["followers"] * 100, 4)

    return post_data


def run_batch_scrape(urls: list[str], cookies: dict, status_fn=None, progress_fn=None):
    session = make_session(cookies)
    results = []
    profile_cache = {}

    total = len(urls)
    for i, url in enumerate(urls, 1):
        if status_fn:
            status_fn(f"Fetching post {i} of {total} ({url})...")
        if progress_fn:
            progress_fn(int((i / total) * 80))

        p_data = fetch_post_details(url, session, cookies, profile_cache)
        results.append(p_data)
        human_delay(0.8, 2.0)

    return results


# ── EXCEL BUILDERS (CLEAN WHITE THEME) ─────────────────────────

def build_excel(profile, posts) -> bytes:
    wb   = openpyxl.Workbook()
    thin = Side(style="thin", color="D0D0D0")
    thick_bottom = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="000000"))
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(h): return PatternFill("solid", fgColor=h)
    def bf(sz=10, c="000000"): return Font(name="Calibri", bold=True,  size=sz, color=c)
    def nf(sz=10, c="000000"): return Font(name="Calibri", bold=False, size=sz, color=c)
    def link_font(sz=10): return Font(name="Calibri", bold=False, size=sz, color="0563C1", underline="single")
    def ctr(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    def lft(wrap=True): return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)
    def rgt(wrap=False): return Alignment(horizontal="right",  vertical="center", wrap_text=wrap)

    followers      = profile["followers"]
    n              = len(posts)
    avg_er         = round(sum(p["er_percent"] for p in posts) / n, 4) if n else 0
    total_likes    = sum(p["likes"]       for p in posts)
    total_comments = sum(p["comments"]    for p in posts)
    total_views    = sum(p["video_views"] for p in posts)

    # ── Sheet 1: Post Metrics ──────────────────────────────
    ws = wb.active
    ws.title = "Post Metrics"
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:M1")
    ws["A1"] = f"Instagram Metrics — @{profile['username']}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=15, color="000000")
    ws["A1"].alignment = ctr(False); ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:M2")
    ws["A2"] = (f"Followers: {followers:,}  |  Posts: {n}  |  "
                f"Avg ER%: {avg_er:.2f}%  |  ER = (Likes+Comments)/Followers×100  "
                f"[Saves & Shares: not available via any public API]")
    ws["A2"].font = Font(name="Calibri", bold=False, size=9, color="444444")
    ws["A2"].alignment = Alignment(horizontal="right", vertical="center"); ws.row_dimensions[2].height = 20

    headers = ["#", "Date", "Type", "Title", "Full Description / Caption", "Post URL", "Likes", "Comments",
               "Video Views", "Followers", "Saves", "Shares", "ER%"]
    widths  = [5,   13,     11,     30,      55,                         45,         12,      12,
               13,          13,         12,       12,       10]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = bf(10, "000000"); c.fill = fill("F2F2F2")
        c.alignment = ctr(False); c.border = thick_bottom
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 24

    for i, p in enumerate(posts):
        row = i + 4
        values = [i+1, p["date"], p["type"], p.get("title","N/A"), p.get("caption","—"), p["url"],
                  p["likes"], p["comments"],
                  p["video_views"] or 0,
                  followers, "N/A — private", "N/A — private", None]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col)
            c.border = bdr
            if col in (4, 5):
                c.value = val; c.font = nf(10, "000000"); c.alignment = lft(wrap=True)
            elif col == 6:
                c.value = val; c.font = link_font(10); c.alignment = lft(wrap=False); c.hyperlink = val
            elif col == 13:
                c.value = f"=(G{row}+H{row})/J{row}*100"
                c.number_format = '0.00"%"'; c.font = bf(10, "000000"); c.alignment = ctr(False)
            elif col == 3:
                c.value = val; c.font = bf(10, "000000"); c.alignment = ctr(False)
            elif col in (7, 8, 9, 10):
                c.value = val; c.number_format = "#,##0"; c.font = nf(10, "000000"); c.alignment = rgt(False)
            elif col in (11, 12):
                c.value = val; c.font = nf(9, "777777"); c.alignment = ctr(False)
            else:
                c.value = val; c.font = nf(10, "000000"); c.alignment = ctr(False)
        ws.row_dimensions[row].height = 28

    tr = n + 4
    ws.merge_cells(f"A{tr}:F{tr}")
    ws[f"A{tr}"] = "TOTALS / AVERAGES"
    ws[f"A{tr}"].font = bf(10, "000000"); ws[f"A{tr}"].fill = fill("F9F9F9")
    ws[f"A{tr}"].alignment = ctr(False)
    for col, val in zip([7,8,9,10,11,12,13],
                        [total_likes, total_comments, total_views, followers, "", "", f"{avg_er:.2f}%"]):
        c = ws.cell(row=tr, column=col)
        c.fill = fill("F9F9F9"); c.border = bdr; c.alignment = ctr(False)
        c.font = bf(10, "000000")
        if col in (7,8,9,10) and isinstance(val, int):
            c.value = val; c.number_format = "#,##0"; c.alignment = rgt(False)
        else:
            c.value = val
    ws.row_dimensions[tr].height = 22

    # ── Sheet 2: Profile Summary ───────────────────────────
    ws2 = wb.create_sheet("Profile Summary")
    ws2.sheet_view.showGridLines = True
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 50

    ws2.merge_cells("A1:B1")
    ws2["A1"] = f"Profile — @{profile['username']}"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="000000")
    ws2["A1"].alignment = ctr(False); ws2.row_dimensions[1].height = 30

    rows2 = [
        ("Username",            f"@{profile['username']}"),
        ("Full Name",           profile["full_name"]),
        ("Category",            profile["category"]),
        ("Business Account",    "Yes" if profile["is_business"] else "No"),
        ("Verified",            "Yes" if profile["verified"] else "No"),
        ("Followers",           f"{profile['followers']:,}"),
        ("Following",           f"{profile['following']:,}"),
        ("Total Posts",         f"{profile['total_posts']:,}"),
        ("External URL",        profile["external_url"] or "—"),
        ("Bio",                 profile["bio"]),
        ("", ""),
        ("Posts Analysed",      n),
        ("Total Likes",         f"{total_likes:,}"),
        ("Total Comments",      f"{total_comments:,}"),
        ("Total Video Views",   f"{total_views:,}"),
        ("Avg Likes / Post",    f"{total_likes//n:,}" if n else "0"),
        ("Avg Comments / Post", f"{total_comments//n:,}" if n else "0"),
        ("Avg ER%",             f"{avg_er:.2f}%"),
        ("Saves Available",     "Not available — Instagram private"),
        ("Shares Available",    "Not available — Instagram private"),
        ("ER% Formula",         "(Likes + Comments) / Followers × 100"),
        ("Scrape Method",       "Instagram Internal API + Cookie Auth"),
        ("TLS Bypass",          "curl_cffi Chrome120" if USE_CURL_CFFI else "requests"),
        ("Scraped At",          datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (label, value) in enumerate(rows2):
        row = i + 2
        if not label:
            ws2.row_dimensions[row].height = 8; continue
        lc = ws2.cell(row=row, column=1, value=label)
        lc.font = bf(10, "333333"); lc.fill = fill("F2F2F2")
        lc.alignment = lft(False); lc.border = bdr
        vc = ws2.cell(row=row, column=2, value=value)
        vc.font = nf(10, "000000"); vc.fill = fill("FFFFFF")
        vc.alignment = lft(False); vc.border = bdr
        ws2.row_dimensions[row].height = 20

    # ── Sheet 3: Captions ──────────────────────────────────
    ws3 = wb.create_sheet("Captions")
    ws3.sheet_view.showGridLines = True
    ws3.column_dimensions["A"].width = 5
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 55
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 12
    ws3.column_dimensions["F"].width = 12

    ws3.merge_cells("A1:F1")
    ws3["A1"] = f"Post Captions — @{profile['username']}"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color="000000")
    ws3["A1"].alignment = ctr(False); ws3.row_dimensions[1].height = 30

    for col, h in enumerate(["#","Date","Caption","Likes","Comments","ER%"], 1):
        c = ws3.cell(row=2, column=col, value=h)
        c.font = bf(10, "000000"); c.fill = fill("F2F2F2")
        c.alignment = ctr(False); c.border = thick_bottom

    for i, p in enumerate(posts):
        row = i + 3
        vals = [i+1, p["date"], p["caption"] or "—", p["likes"], p["comments"], p["er_percent"]]
        for col, val in enumerate(vals, 1):
            c = ws3.cell(row=row, column=col, value=val)
            c.border = bdr
            if col == 3:
                c.alignment = lft(wrap=True); c.font = nf(10, "000000")
            elif col == 6:
                c.number_format = '0.00"%"'; c.font = bf(10, "000000"); c.alignment = ctr(False)
            elif col in (4, 5):
                c.number_format = "#,##0"; c.font = nf(10, "000000"); c.alignment = rgt(False)
            else:
                c.font = nf(10, "000000"); c.alignment = ctr(False)
        ws3.row_dimensions[row].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_batch_excel(posts) -> bytes:
    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="D0D0D0")
    thick_bottom = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="000000"))
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(h): return PatternFill("solid", fgColor=h)
    def bf(sz=10, c="000000"): return Font(name="Calibri", bold=True, size=sz, color=c)
    def nf(sz=10, c="000000"): return Font(name="Calibri", bold=False, size=sz, color=c)
    def link_font(sz=10): return Font(name="Calibri", bold=False, size=sz, color="0563C1", underline="single")
    def ctr(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    def lft(wrap=True): return Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    def rgt(wrap=False): return Alignment(horizontal="right", vertical="center", wrap_text=wrap)

    n = len(posts)
    unique_users = len(set(p.get("username", "") for p in posts if p.get("username") and p.get("username") != "N/A"))
    total_likes = sum(p.get("likes", 0) for p in posts)
    total_comments = sum(p.get("comments", 0) for p in posts)
    total_views = sum(p.get("video_views", 0) for p in posts)
    valid_er = [p["er_percent"] for p in posts if p.get("er_percent", 0) > 0]
    avg_er = sum(valid_er) / len(valid_er) if valid_er else 0

    # ── Sheet 1: Post Metrics ──────────────────────────────
    ws = wb.active
    ws.title = "Post Metrics"
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:O1")
    ws["A1"] = f"Instagram Metrics — Batch Post Analysis ({n} Posts)"
    ws["A1"].font = Font(name="Calibri", bold=True, size=15, color="000000")
    ws["A1"].alignment = ctr(False); ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:O2")
    ws["A2"] = (f"Posts Analysed: {n}  |  Unique Accounts: {unique_users}  |  "
                f"Total Likes: {total_likes:,}  |  Total Comments: {total_comments:,}  |  "
                f"Avg ER%: {avg_er:.2f}%  [ER = (Likes+Comments)/Followers×100]")
    ws["A2"].font = Font(name="Calibri", bold=False, size=9, color="444444")
    ws["A2"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 20

    headers = ["#", "Account", "Account Name", "Title", "Full Description / Caption", "Date", "Type", "Post URL",
               "Likes", "Comments", "Video Views", "Followers", "Saves", "Shares", "ER%"]
    widths  = [5,   16,        20,             32,      55,                         13,     11,     45,
               12,      12,         13,            13,          12,      12,       10]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = bf(10, "000000"); c.fill = fill("F2F2F2")
        c.alignment = ctr(False); c.border = thick_bottom
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 24

    for i, p in enumerate(posts):
        row = i + 4
        followers = p.get("followers", 0)
        
        values = [
            i+1,
            f"@{p.get('username', 'N/A')}",
            p.get("full_name", "N/A"),
            p.get("title", "N/A"),
            p.get("caption", "—"),
            p.get("date", "N/A"),
            p.get("type", "Image"),
            p.get("url", ""),
            p.get("likes", 0),
            p.get("comments", 0),
            p.get("video_views", 0),
            followers if followers > 0 else "N/A",
            "N/A — private",
            "N/A — private",
            None
        ]

        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col)
            c.border = bdr

            if col in (4, 5): # Title & Description
                c.value = val; c.font = nf(10, "000000"); c.alignment = lft(wrap=True)
            elif col == 8: # Post URL
                c.value = val; c.font = link_font(10); c.alignment = lft(wrap=False); c.hyperlink = val
            elif col == 15: # ER%
                if isinstance(followers, int) and followers > 0:
                    c.value = f"=(I{row}+J{row})/L{row}*100"
                    c.number_format = '0.00"%"'; c.font = bf(10, "000000")
                else:
                    c.value = f"{p.get('er_percent', 0.0):.2f}%" if p.get('er_percent') else "—"
                    c.font = bf(10, "000000")
                c.alignment = ctr(False)
            elif col in (9, 10, 11, 12) and isinstance(val, int): # Numbers
                c.value = val; c.number_format = "#,##0"; c.font = nf(10, "000000"); c.alignment = rgt(False)
            elif col == 7: # Type
                c.value = val; c.font = bf(10, "000000"); c.alignment = ctr(False)
            elif col in (13, 14): # Saves/Shares
                c.value = val; c.font = nf(9, "777777"); c.alignment = ctr(False)
            elif col in (2, 3): # Account & Name
                c.value = val; c.font = nf(10, "000000"); c.alignment = lft(wrap=True)
            else:
                c.value = val; c.font = nf(10, "000000"); c.alignment = ctr(False)

        ws.row_dimensions[row].height = 28

    # Totals Row
    tr = n + 4
    ws.merge_cells(f"A{tr}:H{tr}")
    ws[f"A{tr}"] = "TOTALS / AVERAGES"
    ws[f"A{tr}"].font = bf(10, "000000"); ws[f"A{tr}"].fill = fill("F9F9F9")
    ws[f"A{tr}"].alignment = ctr(False)

    total_vals = [total_likes, total_comments, total_views, "", "", "", f"{avg_er:.2f}%"]
    for col_idx, col in enumerate([9, 10, 11, 12, 13, 14, 15]):
        val = total_vals[col_idx]
        c = ws.cell(row=tr, column=col)
        c.fill = fill("F9F9F9"); c.border = bdr; c.alignment = ctr(False)
        c.font = bf(10, "000000")
        if isinstance(val, int):
            c.value = val; c.number_format = "#,##0"; c.alignment = rgt(False)
        else:
            c.value = val
    ws.row_dimensions[tr].height = 22

    # ── Sheet 2: Post Descriptions / Captions ──────────────
    ws2 = wb.create_sheet("Captions & Descriptions")
    ws2.sheet_view.showGridLines = True
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 28
    ws2.column_dimensions["D"].width = 40
    ws2.column_dimensions["E"].width = 70
    ws2.column_dimensions["F"].width = 12
    ws2.column_dimensions["G"].width = 12
    ws2.column_dimensions["H"].width = 12

    ws2.merge_cells("A1:H1")
    ws2["A1"] = "Post Titles, Captions & Full Descriptions"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="000000")
    ws2["A1"].alignment = ctr(False); ws2.row_dimensions[1].height = 30

    headers2 = ["#", "Account", "Title", "URL", "Full Description / Caption", "Likes", "Comments", "ER%"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = bf(10, "000000"); c.fill = fill("F2F2F2")
        c.alignment = ctr(False); c.border = thick_bottom

    for i, p in enumerate(posts):
        row = i + 3
        vals = [
            i+1,
            f"@{p.get('username', 'N/A')}",
            p.get("title", "N/A"),
            p.get("url", ""),
            p.get("caption", "—"),
            p.get("likes", 0),
            p.get("comments", 0),
            p.get("er_percent", 0.0)
        ]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=row, column=col, value=val)
            c.border = bdr
            if col in (3, 5):
                c.alignment = lft(wrap=True); c.font = nf(10, "000000")
            elif col == 4:
                c.alignment = lft(wrap=False); c.font = link_font(10); c.hyperlink = val
            elif col == 8:
                c.number_format = '0.00"%"'; c.font = bf(10, "000000"); c.alignment = ctr(False)
            elif col in (6, 7):
                c.number_format = "#,##0"; c.font = nf(10, "000000"); c.alignment = rgt(False)
            else:
                c.font = nf(10, "000000"); c.alignment = ctr(False)
        ws2.row_dimensions[row].height = 28

    # ── Sheet 3: Batch Analytics ──────────────────────────────
    ws3 = wb.create_sheet("Batch Analytics")
    ws3.sheet_view.showGridLines = True
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 50

    ws3.merge_cells("A1:B1")
    ws3["A1"] = "Batch Post Analytics Summary"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color="000000")
    ws3["A1"].alignment = ctr(False); ws3.row_dimensions[1].height = 32

    top_liked = max(posts, key=lambda x: x.get("likes", 0)) if posts else {}
    top_er    = max(posts, key=lambda x: x.get("er_percent", 0)) if posts else {}

    summary_rows = [
        ("Total Post URLs",       n),
        ("Unique Accounts",       unique_users),
        ("Total Likes",           f"{total_likes:,}"),
        ("Total Comments",        f"{total_comments:,}"),
        ("Total Video Views",     f"{total_views:,}"),
        ("Avg Likes / Post",      f"{total_likes // n:,}" if n else "0"),
        ("Avg Comments / Post",   f"{total_comments // n:,}" if n else "0"),
        ("Avg ER%",               f"{avg_er:.2f}%"),
        ("", ""),
        ("Most Liked Post Title", top_liked.get("title", "N/A")),
        ("Most Liked Account",    f"@{top_liked.get('username', 'N/A')} ({top_liked.get('likes', 0):,} likes)"),
        ("Most Liked Post URL",   top_liked.get("url", "N/A")),
        ("", ""),
        ("Highest ER% Post",      f"{top_er.get('er_percent', 0):.2f}% by @{top_er.get('username', 'N/A')}"),
        ("Highest ER% Post URL",  top_er.get("url", "N/A")),
        ("", ""),
        ("Export Method",         "Batch Post URL Scraping"),
        ("Export Timestamp",      datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    for i, (label, value) in enumerate(summary_rows):
        row = i + 2
        if not label:
            ws3.row_dimensions[row].height = 8; continue
        lc = ws3.cell(row=row, column=1, value=label)
        lc.font = bf(10, "333333"); lc.fill = fill("F2F2F2")
        lc.alignment = lft(False); lc.border = bdr
        vc = ws3.cell(row=row, column=2, value=value)
        vc.font = nf(10, "000000"); vc.fill = fill("FFFFFF")
        vc.alignment = lft(False); vc.border = bdr
        ws3.row_dimensions[row].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── UTILS ─────────────────────────────────────────────────────

def fmt_num(n):
    if n >= 1_000_000: return f"{n/1_000_000:.1f}M"
    if n >= 1_000:     return f"{n/1_000:.1f}K"
    return str(n)


# ── STREAMLIT UI ──────────────────────────────────────────────

def main():
    # ── Session state init ──────────────────────────────────
    if "mode" not in st.session_state:
        st.session_state.mode = "Dump Post URLs"
    if "result_profile" not in st.session_state:
        st.session_state.result_profile = None
    if "result_posts" not in st.session_state:
        st.session_state.result_posts = None
    if "result_excel" not in st.session_state:
        st.session_state.result_excel = None
    if "result_fname" not in st.session_state:
        st.session_state.result_fname = None
    if "batch_posts" not in st.session_state:
        st.session_state.batch_posts = None

    # ── Header ─────────────────────────────────────────────
    st.markdown("""
    <div style="padding: 36px 0 20px; text-align: center;">
      <div style="font-family:'Plus Jakarta Sans',sans-serif; font-size:11px; font-weight:700;
                  color:#E8405A; letter-spacing:0.22em; text-transform:uppercase; margin-bottom:10px;">
        Analytics Tool
      </div>
      <h1 style="font-family:'Plus Jakarta Sans',sans-serif; font-size:38px; font-weight:800;
                 color:#F0F0F0; margin:0; line-height:1.1; letter-spacing:-0.02em;">
        Instagram <span style="color:#E8405A;">Metrics</span>
      </h1>
      <p style="color:#555; font-size:13px; margin-top:10px; font-family:'Plus Jakarta Sans',sans-serif; font-weight:400;">
        Dump Post URLs or Scrape User Profiles &rarr; Download Title, Full Description & Metrics in Excel
      </p>
    </div>
    """, unsafe_allow_html=True)

    # ── Mode Switcher ──────────────────────────────────────
    mode = st.radio(
        "Select Mode",
        options=["Dump Post URLs", "Profile Analytics"],
        horizontal=True,
        index=0,
        key="mode_radio"
    )

    # ── Cookies Section ────────────────────────────────────
    st.markdown("""
    <div style="background:#111; border:1px solid #1E1E1E; border-radius:12px;
                padding:18px 20px 10px; margin-bottom:18px; margin-top:16px;">
      <div style="font-family:'Plus Jakarta Sans',sans-serif; font-size:12px; font-weight:700;
                  color:#E8405A; letter-spacing:0.12em; text-transform:uppercase; margin-bottom:4px;">
        Session Cookies
      </div>
      <div style="font-family:'Plus Jakarta Sans',sans-serif; font-size:12px; color:#555;
                  margin-bottom:14px; line-height:1.6;">
        Chrome DevTools (F12) &rarr; Application &rarr; Cookies &rarr; instagram.com
      </div>
    </div>
    """, unsafe_allow_html=True)

    ck1, ck2 = st.columns(2)
    with ck1:
        sessionid  = st.text_input("sessionid",  value=DEFAULT_COOKIES.get("sessionid",""),  type="password")
        ds_user_id = st.text_input("ds_user_id", value=DEFAULT_COOKIES.get("ds_user_id",""))
    with ck2:
        csrftoken  = st.text_input("csrftoken",  value=DEFAULT_COOKIES.get("csrftoken",""))
        mid        = st.text_input("mid",         value=DEFAULT_COOKIES.get("mid",""))

    cookies = {
        "sessionid":  sessionid,
        "csrftoken":  csrftoken,
        "ds_user_id": ds_user_id,
        "mid":        mid,
    }

    # ── MODE 1: DUMP POST URLS ──────────────────────────────
    if mode == "Dump Post URLs":
        raw_urls = st.text_area(
            "Dump Post URLs",
            placeholder="Paste your Instagram Post / Reel URLs here (one per line or separated by spaces):\n\nhttps://www.instagram.com/p/DbJML3hGVVJ/\nhttps://www.instagram.com/reels/Dazdx9LRzIZ/",
            height=140
        )

        go_batch = st.button("▶  Fetch Metrics for Dumped URLs", use_container_width=True)

        if go_batch:
            if not raw_urls.strip():
                st.error("Please dump at least one valid Instagram post URL.")
                return
            if not all([sessionid, csrftoken, ds_user_id, mid]):
                st.error("⚠️ Fill in all 4 cookies in the Session Cookies section above.")
                return

            urls = extract_urls(raw_urls)
            if not urls:
                st.error("No valid Instagram URLs or shortcodes detected. Check your pasted text.")
                return

            # Clear state
            st.session_state.result_profile = None
            st.session_state.result_posts   = None
            st.session_state.batch_posts    = None
            st.session_state.result_excel   = None

            status_box = st.empty()
            progress   = st.progress(0)

            def update_status(msg):
                status_box.info(f"⏳ {msg}")

            def update_progress(val):
                progress.progress(val)

            update_status(f"Found {len(urls)} URLs. Extracting Title, Full Description & Metrics...")
            progress.progress(10)

            batch_posts = run_batch_scrape(urls, cookies, status_fn=update_status, progress_fn=update_progress)

            progress.progress(85)
            update_status("Building Excel Report with Titles, Descriptions & Metrics...")

            excel_bytes = build_batch_excel(batch_posts)
            fname = f"ig_batch_metrics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            progress.progress(100)
            status_box.empty()
            progress.empty()

            st.session_state.batch_posts  = batch_posts
            st.session_state.result_excel = excel_bytes
            st.session_state.result_fname = fname

        # Render Batch Results
        b_posts = st.session_state.batch_posts
        if b_posts:
            n = len(b_posts)
            total_l = sum(p.get("likes", 0) for p in b_posts)
            total_c = sum(p.get("comments", 0) for p in b_posts)
            total_v = sum(p.get("video_views", 0) for p in b_posts)
            valid_er = [p["er_percent"] for p in b_posts if p.get("er_percent", 0) > 0]
            avg_er = sum(valid_er) / len(valid_er) if valid_er else 0

            # Success Banner
            st.markdown(f"""
            <div style="background:#0d1f16; border:1px solid #1a4d30; border-radius:12px;
                        padding:18px 22px; margin:14px 0;">
              <div style="font-family:'Plus Jakarta Sans',sans-serif; font-size:17px; font-weight:800;
                          color:#1DB954; letter-spacing:-0.01em;">✅ Done — Processed {n} Post URLs</div>
              <div style="font-family:'Plus Jakarta Sans',sans-serif; font-size:13px; color:#3d7a55; margin-top:3px;">
                Extracted Title, Full Description, Likes, Comments, Views & ER% into Excel format
              </div>
            </div>
            """, unsafe_allow_html=True)

            # Metric grid
            st.markdown(f"""
            <div class="metric-grid">
              <div class="metric-card metric-followers">
                <div class="metric-value">{n}</div>
                <div class="metric-label">Posts Processed</div>
              </div>
              <div class="metric-card">
                <div class="metric-value">{fmt_num(total_l)}</div>
                <div class="metric-label">Total Likes</div>
              </div>
              <div class="metric-card">
                <div class="metric-value">{fmt_num(total_c)}</div>
                <div class="metric-label">Total Comments</div>
              </div>
              <div class="metric-card">
                <div class="metric-value">{fmt_num(total_l // n if n else 0)}</div>
                <div class="metric-label">Avg Likes / Post</div>
              </div>
              <div class="metric-card">
                <div class="metric-value">{fmt_num(total_v)}</div>
                <div class="metric-label">Video Views</div>
              </div>
              <div class="metric-card metric-er">
                <div class="metric-value">{avg_er:.2f}%</div>
                <div class="metric-label">Avg ER%</div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            # Post Table
            st.markdown('<div class="section-label">Extracted Post Metrics & Descriptions</div>', unsafe_allow_html=True)

            type_cls = {"Image": "type-image", "Video": "type-video", "Carousel": "type-carousel"}
            rows_html = ""
            for i, p in enumerate(b_posts):
                tc        = type_cls.get(p["type"], "")
                shortcode = p.get("shortcode", "")
                link_text = shortcode[:10] + "↗" if shortcode else "—"
                title_txt = p.get("title", "Post Title")
                if len(title_txt) > 28: title_txt = title_txt[:28] + "..."
                
                desc_txt = p.get("caption", "—").replace("\n", " ")
                if len(desc_txt) > 40: desc_txt = desc_txt[:40] + "..."
                user_txt  = f"@{p.get('username', 'N/A')}"

                rows_html += (
                    f'<div class="post-row">'
                    f'  <div class="post-cell-idx">{i+1}</div>'
                    f'  <div class="post-cell-user">{user_txt}</div>'
                    f'  <div class="post-cell-title">{title_txt}</div>'
                    f'  <div class="post-cell-desc">{desc_txt}</div>'
                    f'  <div class="post-cell">{p.get("date","N/A")}</div>'
                    f'  <div class="post-cell {tc}">{p.get("type","Image")}</div>'
                    f'  <div class="post-cell-num">{p["likes"]:,}</div>'
                    f'  <div class="post-cell-num">{p["comments"]:,}</div>'
                    f'  <div class="post-cell-er">{p["er_percent"]:.2f}%</div>'
                    f'</div>'
                )

            st.markdown(
                f'<div class="post-table">'
                f'  <div class="post-row post-row-header">'
                f'    <span>#</span><span>Account</span><span>Title</span><span>Description</span><span>Date</span>'
                f'    <span>Type</span><span>Likes</span><span>Comments</span><span>ER%</span>'
                f'  </div>'
                f'  {rows_html}'
                f'</div>',
                unsafe_allow_html=True
            )

            # Expandable Titles & Full Descriptions View
            st.markdown('<div class="section-label">Titles & Full Descriptions</div>', unsafe_allow_html=True)
            for i, p in enumerate(b_posts):
                with st.expander(f"#{i+1} | @{p.get('username','N/A')} — {p.get('title','Post Title')[:60]}"):
                    st.markdown(f"**Account:** @{p.get('username','N/A')} ({p.get('full_name','N/A')})  |  **Date:** {p.get('date','N/A')}  |  **Type:** {p.get('type','Image')}")
                    st.markdown(f"**Title:** {p.get('title','N/A')}")
                    st.markdown(f"**Full Description / Caption:**\n```text\n{p.get('caption','No description available.')}\n```")
                    st.markdown(f"[Open Post on Instagram ↗]({p['url']})")

            # Download Excel
            st.markdown('<div class="section-label">Export Excel</div>', unsafe_allow_html=True)
            st.download_button(
                label="⬇  Download Batch Excel Report",
                data=st.session_state.result_excel,
                file_name=st.session_state.result_fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.markdown("""
            <div style="text-align:center; font-family:'Plus Jakarta Sans',sans-serif;
                        font-size:11px; color:#383838; margin-top:6px;">
              Includes 3 sheets: Post Metrics (with Full Description) · Captions & Descriptions · Batch Analytics
            </div>
            """, unsafe_allow_html=True)

    # ── MODE 2: PROFILE ANALYTICS ───────────────────────────
    else:
        col1, col2 = st.columns([3, 1])
        with col1:
            raw_input = st.text_input(
                "Instagram Username or URL",
                placeholder="username  or  https://www.instagram.com/username/",
            )
        with col2:
            num_posts = st.number_input("Posts", min_value=1, max_value=50, value=15, step=5)

        go_profile = st.button("▶  Fetch Profile Metrics", use_container_width=True)

        if go_profile:
            if not raw_input.strip():
                st.error("Please enter an Instagram username or URL.")
                return
            if not all([sessionid, csrftoken, ds_user_id, mid]):
                st.error("⚠️ Fill in all 4 cookies in the Session Cookies section above.")
                return

            st.session_state.result_profile = None
            st.session_state.result_posts   = None
            st.session_state.batch_posts    = None
            st.session_state.result_excel   = None

            username   = extract_username(raw_input)
            status_box = st.empty()
            progress   = st.progress(0)

            def update_status(msg):
                status_box.info(f"⏳ {msg}")

            update_status("Connecting to Instagram...")
            progress.progress(10)

            profile, posts, err = run_scrape(
                username, int(num_posts), cookies,
                status_fn=update_status
            )

            if err:
                status_box.empty(); progress.empty()
                st.error(f"❌ {err}")
                return

            progress.progress(80)
            update_status("Building Excel file...")

            if not posts:
                status_box.empty(); progress.empty()
                st.warning("⚠️ Profile found but no posts could be retrieved. "
                           "Cookies may be expired or the account is private.")
                return

            excel_bytes = build_excel(profile, posts)
            fname       = f"ig_metrics_{username}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

            progress.progress(100)
            status_box.empty()
            progress.empty()

            st.session_state.result_profile = profile
            st.session_state.result_posts   = posts
            st.session_state.result_excel   = excel_bytes
            st.session_state.result_fname   = fname

        # Render Profile Results
        profile = st.session_state.result_profile
        posts   = st.session_state.result_posts

        if profile and posts:
            st.markdown(f"""
            <div style="background:#0d1f16; border:1px solid #1a4d30; border-radius:12px;
                        padding:18px 22px; margin:14px 0;">
              <div style="font-family:'Plus Jakarta Sans',sans-serif; font-size:17px; font-weight:800;
                          color:#1DB954; letter-spacing:-0.01em;">✅ Done — @{profile['username']}</div>
              <div style="font-family:'Plus Jakarta Sans',sans-serif; font-size:13px; color:#3d7a55; margin-top:3px;">
                {len(posts)} posts fetched successfully
              </div>
            </div>
            """, unsafe_allow_html=True)

            badges = ""
            if profile["verified"]:    badges += '<span class="badge badge-verified">✓ Verified</span>'
            if profile["is_business"]: badges += '<span class="badge badge-business">Business</span>'
            if profile["category"] and profile["category"] != "N/A":
                badges += f'<span class="badge badge-category">{profile["category"]}</span>'

            st.markdown(f"""
            <div class="profile-card">
              <div class="profile-name">{profile['full_name']}</div>
              <div class="profile-username">@{profile['username']}</div>
              <div>{badges}</div>
              <div class="profile-bio">{profile['bio'] or ''}</div>
            </div>
            """, unsafe_allow_html=True)

            followers = profile["followers"]
            n         = len(posts)
            avg_er    = round(sum(p["er_percent"] for p in posts) / n, 4) if n else 0
            total_l   = sum(p["likes"]       for p in posts)
            total_c   = sum(p["comments"]    for p in posts)
            total_v   = sum(p["video_views"] for p in posts)

            st.markdown(f"""
            <div class="metric-grid">
              <div class="metric-card metric-followers">
                <div class="metric-value">{fmt_num(followers)}</div>
                <div class="metric-label">Followers</div>
              </div>
              <div class="metric-card">
                <div class="metric-value">{fmt_num(total_l)}</div>
                <div class="metric-label">Total Likes</div>
              </div>
              <div class="metric-card">
                <div class="metric-value">{fmt_num(total_c)}</div>
                <div class="metric-label">Total Comments</div>
              </div>
              <div class="metric-card">
                <div class="metric-value">{fmt_num(total_l // n if n else 0)}</div>
                <div class="metric-label">Avg Likes / Post</div>
              </div>
              <div class="metric-card">
                <div class="metric-value">{fmt_num(total_v)}</div>
                <div class="metric-label">Video Views</div>
              </div>
              <div class="metric-card metric-er">
                <div class="metric-value">{avg_er:.2f}%</div>
                <div class="metric-label">Avg ER%</div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown('<div class="section-label">Recent Posts</div>', unsafe_allow_html=True)
            type_cls = {"Image": "type-image", "Video": "type-video", "Carousel": "type-carousel"}

            rows_html = ""
            for i, p in enumerate(posts):
                tc        = type_cls.get(p["type"], "")
                shortcode = p.get("shortcode", "")
                link_text = shortcode[:14] + "↗" if shortcode else "—"
                rows_html += (
                    f'<div class="post-row post-row-profile">'
                    f'  <div class="post-cell-idx">{i+1}</div>'
                    f'  <div class="post-cell">{p["date"]}</div>'
                    f'  <div class="post-cell {tc}">{p["type"]}</div>'
                    f'  <div class="post-cell"><a class="post-link" href="{p["url"]}" target="_blank">{link_text}</a></div>'
                    f'  <div class="post-cell-num">{p["likes"]:,}</div>'
                    f'  <div class="post-cell-num">{p["comments"]:,}</div>'
                    f'  <div class="post-cell-er">{p["er_percent"]:.2f}%</div>'
                    f'</div>'
                )

            st.markdown(
                f'<div class="post-table">'
                f'  <div class="post-row post-row-profile post-row-header">'
                f'    <span>#</span><span>Date</span><span>Type</span><span>URL</span>'
                f'    <span>Likes</span><span>Comments</span><span>ER%</span>'
                f'  </div>'
                f'  {rows_html}'
                f'</div>',
                unsafe_allow_html=True
            )

            st.markdown('<div class="section-label">Export</div>', unsafe_allow_html=True)
            st.download_button(
                label="⬇  Download Excel Report",
                data=st.session_state.result_excel,
                file_name=st.session_state.result_fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.markdown("""
            <div style="text-align:center; font-family:'Plus Jakarta Sans',sans-serif;
                        font-size:11px; color:#383838; margin-top:6px;">
              Includes 3 sheets: Post Metrics · Profile Summary · Captions
            </div>
            """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()