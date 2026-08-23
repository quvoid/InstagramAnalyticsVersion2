"""
Instagram Brand Partnership Analyzer
=====================================
Scrapes creator profiles grouped by brand. Each brand gets its own
trio of sheets in one output workbook:

    Brand Overview  →  per-brand summary (one workbook-level sheet)
    MONTBLANC — Profiles Overview
    MONTBLANC — Post Metrics
    MONTBLANC — Captions
    SHEAFFER  — Profiles Overview
    ... etc.

Accounts that appear under multiple brands are scraped ONCE and reused.

Run:   python scrape_brands.py
"""

import sys, io, re, time, random, json
from datetime import datetime, timezone

sys.stdout.reconfigure(encoding="utf-8")

# ════════════════════════════════════════════════════════════════
#  BRAND → USERNAMES MAP  (edit freely — order is preserved)
# ════════════════════════════════════════════════════════════════

BRANDS = {
    "MONTBLANC": [
        "thedanielbruhl",
        "rupertfriend",
        "joeyking",
        "dreelouisehemingway",
        "esther.mcgregor",
        "houseofwaris",
        "thecooperkoch",
        "kiernanshipka",
        "sharypova_nna",
        "sammsalter",
        "waybuss",
        "thebursin",
        "jackbenedwards",
        "stellamaxwell",
        "zidane",
    ],
    "PARKER PENS": [
        "harsh_palawat",
        "ishita_chopraa",
        "sukanyaboruah99",
        "durjoydatta",
        "duologue_barundas",
    ],
    # CROSS PENS INDIA: no genuine creator accounts found — skipped
    "SHEAFFER": [
        "reenaaggarwal2",
        "rayrooqma",
        "emschillinday",
        "tinagarg.official",
        "tina_mansukhani_garg",
        "dancerukmini",
        "followyourchild",
        "madhuri.official",
        "ozgeinus",
        "mynewestaddiction",
        "manankathuria",
        "escribotuluz",
        "makingmemorieswithale",
        "arianx",
        "journal.to.go",
        "harleyestate",
    ],
    "LAMY INDIA": [
        "talesofink90",
        "rani_banerjee",
        "aalyakumarr",
        "chirpypunnu",
        "mehandiratta",
        "nikhilnaik0508",
        "reader_viddh",
        "pastelpalettebytanya",
        "justletteritout",
        "sakshi_2000.earth",
        "anmolmalik5",
    ],
    "MAKOBA INDIA": [
        "whatsupgurgaon",
        "gurugrammers_",
        "mukulgoyaldesigns",
        "sonali_singh4",
    ],
    "SUBMARINE PENS": [
        "prashijainofficial",
        "vidveda_",
        "shreyavdeshpande",
        "nits_way",
        "manasvi.daftary",
        "thesagarvadapalli",
    ],
}

# ════════════════════════════════════════════════════════════════
#  COOKIES  (refresh from Chrome → F12 → Application → Cookies)
# ════════════════════════════════════════════════════════════════

COOKIES = {
    "sessionid":  "25113411270%3AyQFaao428g6Xb9%3A0%3AAYjwS_M5UIDcS-i41tvdVFu8WKSqfzfgEhlZLGMvFg",
    "csrftoken":  "3gJbkGDZp99lA8QQ0brobyoHzOreuu8f",
    "mid":        "afyCbwALAAFRStE-k17-dfO5_jfa",
    "ds_user_id": "25113411270",

}

# ════════════════════════════════════════════════════════════════
#  CONFIG
# ════════════════════════════════════════════════════════════════

POSTS_PER_USER  = 20
OUTPUT_FILENAME = "brand_partnership_analysis.xlsx"

# ════════════════════════════════════════════════════════════════
#  DEPENDENCIES
# ════════════════════════════════════════════════════════════════

try:
    from curl_cffi import requests as cffi_requests
    USE_CURL_CFFI = True
    print("✓ curl_cffi found — Chrome TLS fingerprint active")
except ImportError:
    import requests as cffi_requests
    USE_CURL_CFFI = False
    print("⚠ curl_cffi not found — using standard requests")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IG_APP_ID = "936619743392459"
BASE_HEADERS = {
    "accept":           "*/*",
    "accept-language":  "en-US,en;q=0.9",
    "user-agent":       (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "x-ig-app-id":      IG_APP_ID,
    "x-ig-www-claim":   "0",
    "x-requested-with": "XMLHttpRequest",
    "x-csrftoken":      COOKIES.get("csrftoken", ""),
    "origin":           "https://www.instagram.com",
    "referer":          "https://www.instagram.com/",
}

# Per-brand accent colours (background fill for header rows)
BRAND_COLORS = {
    "MONTBLANC":      "1C1C1C",  # near-black (MB identity)
    "PARKER PENS":    "1A3A6B",  # deep blue
    "SHEAFFER":       "8B0000",  # dark red
    "LAMY INDIA":     "215732",  # dark green
    "MAKOBA INDIA":   "5B2D8E",  # purple
    "SUBMARINE PENS": "1A5276",  # teal-blue
}
BRAND_TEXT_COLOR = "FFFFFF"      # white text on all brand headers


# ════════════════════════════════════════════════════════════════
#  NETWORK / SCRAPE HELPERS
# ════════════════════════════════════════════════════════════════

def make_session():
    if USE_CURL_CFFI:
        s = cffi_requests.Session(impersonate="chrome120")
    else:
        import requests; s = requests.Session()
    s.headers.update(BASE_HEADERS)
    s.cookies.update(COOKIES)
    return s

def _delay(a=1.0, b=2.5):
    time.sleep(random.uniform(a, b))

def fetch_profile_info(username: str, session) -> dict:
    url  = f"https://www.instagram.com/api/v1/users/web_profile_info/?username={username}"
    hdrs = {**BASE_HEADERS, "referer": f"https://www.instagram.com/{username}/"}
    for attempt in range(1, 4):
        try:
            r = session.get(url, headers=hdrs, cookies=COOKIES, timeout=18)
            if r.status_code == 200:
                return r.json()
            elif r.status_code == 429:
                wait = 35 * attempt
                print(f"    ⏳ Rate-limited — sleeping {wait}s...")
                time.sleep(wait)
            elif r.status_code in (401, 403):
                print(f"    ❌ {r.status_code} on @{username} — cookies may be expired.")
                return {"_auth_error": r.status_code}
            elif r.status_code == 404:
                print(f"    ❌ @{username} not found (404).")
                return {}
            else:
                _delay(2, 5)
        except Exception as exc:
            print(f"    ⚠ Network error: {exc}")
            _delay(2, 4)
    return {}

def fetch_posts_v2(user_id: str, session, count=20) -> list:
    url = f"https://www.instagram.com/api/v1/feed/user/{user_id}/?count={count}"
    for attempt in range(1, 4):
        try:
            r = session.get(url, headers=BASE_HEADERS, cookies=COOKIES, timeout=18)
            if r.status_code == 200:
                return r.json().get("items", [])
            elif r.status_code == 429:
                time.sleep(35 * attempt)
            else:
                _delay(2, 5)
        except Exception:
            _delay(2, 4)
    return []

def parse_post_v2(item: dict, followers: int) -> dict:
    likes    = item.get("like_count", 0) or 0
    comments = item.get("comment_count", 0) or 0
    cap_obj  = item.get("caption")
    caption  = (cap_obj.get("text", "") if isinstance(cap_obj, dict) else "") or ""
    mt_map   = {1: "Image", 2: "Video", 8: "Carousel"}
    ts       = item.get("taken_at", 0)
    date_str = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d") if ts else "N/A"
    code     = item.get("code", item.get("shortcode", ""))
    lines    = [l.strip() for l in caption.splitlines() if l.strip()]
    return {
        "shortcode":   code,
        "url":         f"https://www.instagram.com/p/{code}/",
        "title":       (lines[0] if lines else "Instagram Post")[:100],
        "type":        mt_map.get(item.get("media_type", 1), "Image"),
        "date":        date_str,
        "likes":       likes,
        "comments":    comments,
        "video_views": item.get("view_count", 0) or item.get("play_count", 0) or 0,
        "caption":     caption,
        "er_percent":  round((likes + comments) / followers * 100, 4) if followers else 0,
    }

def scrape_one(username: str, session) -> dict | None:
    """Return a fully-populated profile dict, or None on failure."""
    raw = fetch_profile_info(username, session)
    if "_auth_error" in raw:
        return None          # stop further scraping — cookies dead
    user = raw.get("data", {}).get("user")
    if not user:
        return None

    followers = user.get("edge_followed_by", {}).get("count", 0)
    profile = {
        "username":    username,
        "full_name":   user.get("full_name", "N/A"),
        "followers":   followers,
        "following":   user.get("edge_follow", {}).get("count", 0),
        "total_posts": user.get("edge_owner_to_timeline_media", {}).get("count", 0),
        "bio":         user.get("biography", ""),
        "verified":    user.get("is_verified", False),
        "is_business": user.get("is_business_account", False),
        "user_id":     user.get("id", ""),
        "posts":       [],
        "avg_likes":   0,
        "avg_comments":0,
        "avg_er":      0,
    }

    if profile["user_id"]:
        raw_posts = fetch_posts_v2(profile["user_id"], session, count=POSTS_PER_USER)
        total_l = total_c = 0
        for item in raw_posts[:POSTS_PER_USER]:
            p = parse_post_v2(item, followers)
            total_l += p["likes"]; total_c += p["comments"]
            profile["posts"].append(p)
        n = len(profile["posts"])
        if n:
            profile["avg_likes"]    = total_l // n
            profile["avg_comments"] = total_c // n
            profile["avg_er"]       = round((total_l + total_c) / n / followers * 100, 4) if followers else 0

    return profile


# ════════════════════════════════════════════════════════════════
#  EXCEL BUILDER
# ════════════════════════════════════════════════════════════════

def build_excel(brands_scraped: dict[str, list]) -> bytes:
    """
    brands_scraped: { brand_name: [profile_dict, ...], ... }
    """
    wb = openpyxl.Workbook()

    thin = Side(style="thin",   color="CCCCCC")
    med  = Side(style="medium", color="888888")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def tbdr(brand):
        accent = BRAND_COLORS.get(brand, "444444")
        return Border(left=thin, right=thin, top=thin,
                      bottom=Side(style="medium", color=accent))

    def fill(hex_):   return PatternFill("solid", fgColor=hex_)
    def bf(sz=10, c="000000"):  return Font(name="Calibri", bold=True,  size=sz, color=c)
    def nf(sz=10, c="000000"):  return Font(name="Calibri", bold=False, size=sz, color=c)
    def lnk(sz=10):             return Font(name="Calibri", bold=False, size=sz, color="0563C1", underline="single")
    def ctr(w=False): return Alignment(horizontal="center", vertical="center", wrap_text=w)
    def lft(w=True):  return Alignment(horizontal="left",   vertical="center", wrap_text=w)
    def rgt(w=False): return Alignment(horizontal="right",  vertical="center", wrap_text=w)

    # ── helpers ──────────────────────────────────────────────
    def set_col_widths(ws, widths):
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    def header_row(ws, row_num, cols, brand=None, hdr_fill="F2F2F2", hdr_txt="000000", height=24):
        bg   = BRAND_COLORS.get(brand, hdr_fill) if brand else hdr_fill
        txt  = BRAND_TEXT_COLOR if brand else hdr_txt
        for col, val in enumerate(cols, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.font = bf(10, txt); c.fill = fill(bg); c.alignment = ctr(); c.border = bdr
        ws.row_dimensions[row_num].height = height

    # ── SHEET 0: Brand Overview (master summary) ─────────────
    ws0 = wb.active
    ws0.title = "Brand Overview"
    ws0.sheet_view.showGridLines = True

    ws0.merge_cells("A1:M1")
    ws0["A1"] = "Instagram Brand Partnership Analysis — All Brands"
    ws0["A1"].font = bf(15, "000000"); ws0["A1"].alignment = ctr(); ws0.row_dimensions[1].height = 34

    ws0.merge_cells("A2:M2")
    ts_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    total_accounts = sum(len(v) for v in brands_scraped.values())
    ws0["A2"] = (f"Generated: {ts_str}  |  Brands: {len(brands_scraped)}  |  "
                 f"Total creator accounts: {total_accounts}  |  Posts per account: {POSTS_PER_USER}")
    ws0["A2"].font = nf(9, "555555"); ws0["A2"].alignment = Alignment(horizontal="right", vertical="center")
    ws0.row_dimensions[2].height = 18

    ov_headers = ["#", "Brand", "Username", "Full Name", "Verified", "Business",
                  "Followers", "Following", "Total Posts",
                  "Avg Likes", "Avg Comments", "Avg ER%", "Cross-Brand?"]
    ov_widths  = [5, 18, 20, 26, 10, 11, 14, 12, 14, 12, 14, 12, 14]
    set_col_widths(ws0, ov_widths)
    header_row(ws0, 3, ov_headers, brand=None, hdr_fill="2C2C2C", hdr_txt="FFFFFF", height=26)
    ws0.freeze_panes = "A4"

    # Build a set of usernames that appear in >1 brand
    all_unames = [u for plist in brands_scraped.values() for p in plist for u in [p["username"]]]
    seen_count = {}
    for u in all_unames:
        seen_count[u] = seen_count.get(u, 0) + 1
    cross_brand = {u for u, c in seen_count.items() if c > 1}

    ov_row = 4
    for brand, profiles in brands_scraped.items():
        brand_color = BRAND_COLORS.get(brand, "444444")
        for p in profiles:
            xb = "Yes" if p["username"] in cross_brand else ""
            vals = [
                ov_row - 3,
                brand,
                f"@{p['username']}",
                p["full_name"],
                "✓" if p["verified"] else "",
                "Biz" if p["is_business"] else "",
                p["followers"],
                p["following"],
                p["total_posts"],
                p["avg_likes"],
                p["avg_comments"],
                p["avg_er"] / 100,
                xb,
            ]
            for col, val in enumerate(vals, 1):
                c = ws0.cell(row=ov_row, column=col, value=val); c.border = bdr
                if col == 2:      # Brand column — brand accent colour
                    c.fill = fill(brand_color); c.font = bf(10, "FFFFFF"); c.alignment = ctr()
                elif col == 3:    # Username
                    c.fill = fill("FFF8F0"); c.font = bf(10); c.alignment = lft(w=False)
                elif col == 4:    # Full Name
                    c.font = nf(10); c.alignment = lft(w=False)
                elif col == 12:   # ER%
                    c.number_format = "0.00%"; c.font = bf(10); c.alignment = ctr()
                elif col in (7,8,9,10,11) and isinstance(val, int):
                    c.number_format = "#,##0"; c.font = nf(10); c.alignment = rgt()
                elif col == 13:   # Cross-brand
                    c.font = bf(10, "C0392B") if xb else nf(10); c.alignment = ctr()
                else:
                    c.font = nf(10); c.alignment = ctr()
            ws0.row_dimensions[ov_row].height = 22
            ov_row += 1

    # ── PER-BRAND SHEETS ─────────────────────────────────────
    for brand, profiles in brands_scraped.items():
        brand_color = BRAND_COLORS.get(brand, "444444")
        safe_brand  = brand[:28]   # sheet name max 31 chars; leave room for suffix

        # ── Brand Profiles Overview ──────────────────────────
        ws_p = wb.create_sheet(f"{safe_brand} — Profiles")
        ws_p.sheet_view.showGridLines = True

        ws_p.merge_cells("A1:K1")
        ws_p["A1"] = f"{brand} — Creator Profiles Overview"
        ws_p["A1"].font = bf(13, "FFFFFF"); ws_p["A1"].fill = fill(brand_color)
        ws_p["A1"].alignment = ctr(); ws_p.row_dimensions[1].height = 32

        prof_headers = ["#", "Username", "Full Name", "Verified", "Business",
                        "Followers", "Following", "Total Posts",
                        "Avg Likes", "Avg Comments", "Avg ER%"]
        prof_widths  = [5, 20, 26, 10, 11, 14, 12, 14, 12, 14, 12]
        set_col_widths(ws_p, prof_widths)
        header_row(ws_p, 2, prof_headers, brand=brand, height=24)
        ws_p.freeze_panes = "A3"

        for i, p in enumerate(profiles, 1):
            row = i + 2
            vals = [
                i,
                f"@{p['username']}",
                p["full_name"],
                "✓" if p["verified"] else "",
                "Biz" if p["is_business"] else "",
                p["followers"],
                p["following"],
                p["total_posts"],
                p["avg_likes"],
                p["avg_comments"],
                p["avg_er"] / 100,
            ]
            for col, val in enumerate(vals, 1):
                c = ws_p.cell(row=row, column=col, value=val); c.border = bdr
                if col == 2:
                    c.fill = fill("FFF8F0"); c.font = bf(10); c.alignment = lft(w=False)
                elif col == 3:
                    c.font = nf(10); c.alignment = lft(w=False)
                elif col == 11:
                    c.number_format = "0.00%"; c.font = bf(10); c.alignment = ctr()
                elif col in (6,7,8,9,10) and isinstance(val, int):
                    c.number_format = "#,##0"; c.font = nf(10); c.alignment = rgt()
                else:
                    c.font = nf(10); c.alignment = ctr()
            ws_p.row_dimensions[row].height = 22

        # ── Brand Post Metrics ────────────────────────────────
        ws_m = wb.create_sheet(f"{safe_brand} — Posts")
        ws_m.sheet_view.showGridLines = True

        ws_m.merge_cells("A1:L1")
        ws_m["A1"] = f"{brand} — Detailed Post Metrics (Last {POSTS_PER_USER} Posts / Account)"
        ws_m["A1"].font = bf(13, "FFFFFF"); ws_m["A1"].fill = fill(brand_color)
        ws_m["A1"].alignment = ctr(); ws_m.row_dimensions[1].height = 32

        post_headers = ["#", "Username", "Creator Name", "Title", "Full Caption",
                        "Date", "Type", "Post URL", "Likes", "Comments", "Video Views", "ER%"]
        post_widths  = [5, 20, 24, 30, 52, 13, 11, 46, 12, 12, 13, 10]
        set_col_widths(ws_m, post_widths)
        header_row(ws_m, 2, post_headers, brand=brand, height=24)
        ws_m.freeze_panes = "A3"

        pidx = 1
        for p in profiles:
            for pst in p["posts"]:
                row = pidx + 2
                vals = [
                    pidx,
                    f"@{p['username']}",
                    p["full_name"],
                    pst["title"],
                    pst["caption"],
                    pst["date"],
                    pst["type"],
                    pst["url"],
                    pst["likes"],
                    pst["comments"],
                    pst["video_views"],
                    pst["er_percent"] / 100,
                ]
                for col, val in enumerate(vals, 1):
                    c = ws_m.cell(row=row, column=col, value=val); c.border = bdr
                    if col == 2:
                        c.fill = fill("FFF8F0"); c.font = bf(10); c.alignment = lft(w=False)
                    elif col in (4, 5):
                        c.font = nf(10); c.alignment = lft(w=True)
                    elif col == 8:
                        c.font = lnk(); c.alignment = lft(w=False)
                        if val: c.hyperlink = val
                    elif col == 12:
                        c.number_format = "0.00%"; c.font = bf(10); c.alignment = ctr()
                    elif col in (9, 10, 11):
                        c.number_format = "#,##0"; c.font = nf(10); c.alignment = rgt()
                    else:
                        c.font = nf(10); c.alignment = ctr()
                ws_m.row_dimensions[row].height = 26
                pidx += 1

        # ── Brand Captions ────────────────────────────────────
        ws_c = wb.create_sheet(f"{safe_brand} — Captions")
        ws_c.sheet_view.showGridLines = True

        ws_c.merge_cells("A1:G1")
        ws_c["A1"] = f"{brand} — Full Post Captions"
        ws_c["A1"].font = bf(13, "FFFFFF"); ws_c["A1"].fill = fill(brand_color)
        ws_c["A1"].alignment = ctr(); ws_c.row_dimensions[1].height = 32

        cap_headers = ["#", "Username", "Title", "Post URL", "Full Caption", "Likes", "ER%"]
        cap_widths  = [5, 20, 30, 46, 68, 12, 10]
        set_col_widths(ws_c, cap_widths)
        header_row(ws_c, 2, cap_headers, brand=brand, height=24)
        ws_c.freeze_panes = "A3"

        cidx = 1
        for p in profiles:
            for pst in p["posts"]:
                row = cidx + 2
                vals = [
                    cidx,
                    f"@{p['username']}",
                    pst["title"],
                    pst["url"],
                    pst["caption"],
                    pst["likes"],
                    pst["er_percent"] / 100,
                ]
                for col, val in enumerate(vals, 1):
                    c = ws_c.cell(row=row, column=col, value=val); c.border = bdr
                    if col == 2:
                        c.fill = fill("FFF8F0"); c.font = bf(10); c.alignment = lft(w=False)
                    elif col in (3, 5):
                        c.font = nf(10); c.alignment = lft(w=True)
                    elif col == 4:
                        c.font = lnk(); c.alignment = lft(w=False)
                        if val: c.hyperlink = val
                    elif col == 7:
                        c.number_format = "0.00%"; c.font = bf(10); c.alignment = ctr()
                    elif col == 6:
                        c.number_format = "#,##0"; c.font = nf(10); c.alignment = rgt()
                    else:
                        c.font = nf(10); c.alignment = ctr()
                ws_c.row_dimensions[row].height = 26
                cidx += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════

def main():
    # Collect all unique usernames across all brands
    all_unique = {}  # username → normalised (lowercase)
    for brand, unames in BRANDS.items():
        for u in unames:
            key = u.strip().lstrip("@").lower()
            if key not in all_unique:
                all_unique[key] = u.strip().lstrip("@")

    total_unique = len(all_unique)
    total_accounts = sum(len(v) for v in BRANDS.items())

    print(f"\n{'═'*68}")
    print("  Instagram Brand Partnership Analyzer")
    print(f"  Brands: {len(BRANDS)}  |  Unique accounts: {total_unique}  |  Posts/account: {POSTS_PER_USER}")
    print(f"{'═'*68}\n")

    session = make_session()

    # ── Phase 1: Scrape every unique account ONCE ──────────────
    cache: dict[str, dict] = {}   # normalised username → profile dict
    failed: set[str] = set()
    auth_dead = False

    uname_list = list(all_unique.items())  # [(key, original), ...]
    for idx, (key, uname) in enumerate(uname_list, 1):
        if auth_dead:
            print("🔴 Auth error detected — aborting remaining scrapes. Refresh cookies.")
            break
        print(f"[{idx}/{total_unique}] @{uname} ...", end=" ", flush=True)
        profile = scrape_one(uname, session)
        if profile is None:
            # Check if it's an auth error (fetch_profile_info returned {"_auth_error": ...})
            # We detect auth failure via fetch_profile_info returning the sentinel
            print("⚠ skipped / not found")
            failed.add(key)
        else:
            n = len(profile["posts"])
            print(f"✓  followers: {profile['followers']:,}  |  posts fetched: {n}  |  avg ER: {profile['avg_er']:.2f}%")
            cache[key] = profile
        if idx < total_unique:
            _delay(1.5, 3.0)

    print(f"\n✅ Scrape phase done. {len(cache)}/{total_unique} accounts succeeded.")
    if failed:
        print(f"⚠  Failed / skipped: {', '.join(f'@{all_unique[k]}' for k in failed)}")

    # ── Phase 2: Build per-brand result sets ───────────────────
    brands_scraped: dict[str, list] = {}
    for brand, unames in BRANDS.items():
        plist = []
        for u in unames:
            key = u.strip().lstrip("@").lower()
            if key in cache:
                plist.append(cache[key])
        brands_scraped[brand] = plist
        n_scraped = len(plist)
        n_total   = len(unames)
        print(f"  {brand}: {n_scraped}/{n_total} accounts")

    # ── Phase 3: Cache raw data ────────────────────────────────
    cache_path = OUTPUT_FILENAME.rsplit(".", 1)[0] + "_cache.json"
    try:
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump({"brands": brands_scraped}, f, ensure_ascii=False, indent=2)
        print(f"\n💾 Raw data cached → {cache_path}")
    except Exception as e:
        print(f"⚠ Cache write failed ({e}) — continuing anyway.")

    # ── Phase 4: Build Excel ───────────────────────────────────
    print("📊 Building Excel workbook ...")
    excel_bytes = build_excel(brands_scraped)

    save_path = OUTPUT_FILENAME
    try:
        with open(save_path, "wb") as f:
            f.write(excel_bytes)
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_path = OUTPUT_FILENAME.rsplit(".", 1)[0] + f"_{stamp}.xlsx"
        print(f"⚠ File locked — saving to {save_path}")
        with open(save_path, "wb") as f:
            f.write(excel_bytes)

    total_sheets = 1 + len(brands_scraped) * 3  # Brand Overview + 3 per brand
    print(f"\n{'═'*68}")
    print(f"  ✅ Done!  Saved: {save_path}")
    print(f"  Sheets ({total_sheets} total):")
    print(f"    • Brand Overview  (all brands, all accounts)")
    for brand in brands_scraped:
        safe = brand[:28]
        print(f"    • {safe} — Profiles / Posts / Captions")
    print(f"{'═'*68}\n")


if __name__ == "__main__":
    main()
