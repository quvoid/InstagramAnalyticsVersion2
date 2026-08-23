"""
Google Trends — Competitor Brand Search Interest Analysis
=========================================================
Pulls 12-month search interest for all 6 pen brands in India
and writes results into competitor_paid_media_analysis.xlsx
as a new "Google Trends" sheet.
"""

import time, sys, json
from datetime import datetime

# ── Patch: urllib3 v2 renamed method_whitelist → allowed_methods
# pytrends still uses the old name, so we shim it here.
import urllib3.util.retry as _retry_mod
_OrigRetry = _retry_mod.Retry
class _PatchedRetry(_OrigRetry):
    def __init__(self, *args, method_whitelist=None, **kwargs):
        if method_whitelist is not None and "allowed_methods" not in kwargs:
            kwargs["allowed_methods"] = method_whitelist
        super().__init__(*args, **kwargs)
_retry_mod.Retry = _PatchedRetry

import pandas as pd
from pytrends.request import TrendReq

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

WORKBOOK_FILE = "competitor_paid_media_analysis.xlsx"
SHEET_NAME    = "Google Trends"

# ── Brand search terms (what actual users type) ───────────────
BRAND_TERMS = {
    "Montblanc":      "Montblanc pen",
    "Parker Pens":    "Parker pen",
    "Sheaffer":       "Sheaffer pen",
    "Lamy":           "Lamy pen",
    "Makoba":         "Makoba pen",
    "Submarine Pens": "Submarine pen",
}

BRAND_COLORS = {
    "Montblanc":      "1C1C1C",
    "Parker Pens":    "1A3A6B",
    "Sheaffer":       "8B0000",
    "Lamy":           "1E5631",
    "Makoba":         "4A235A",
    "Submarine Pens": "154360",
}

GEO       = "IN"            # India
TIMEFRAME = "today 12-m"    # last 12 months

# ── Style helpers ─────────────────────────────────────────────
thin = Side(style="thin",   color="CCCCCC")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_):  return PatternFill("solid", fgColor=hex_)
def bf(sz=10, c="000000"): return Font(name="Calibri", bold=True,  size=sz, color=c)
def nf(sz=10, c="000000"): return Font(name="Calibri", bold=False, size=sz, color=c)
def ctr(w=False): return Alignment(horizontal="center", vertical="center", wrap_text=w)
def lft():        return Alignment(horizontal="left",   vertical="center")
def rgt():        return Alignment(horizontal="right",  vertical="center")

HDR_BG  = "1F2D3D"
HDR_FG  = "FFFFFF"


# ── 1. Pull trends data in batches of 5 ──────────────────────
print(f"\n{'═'*60}")
print("  Google Trends — Pen Brand Search Interest (India, 12 months)")
print(f"{'═'*60}\n")

pytrends = TrendReq(hl="en-US", tz=330, timeout=(15, 40), retries=2, backoff_factor=3)

terms      = list(BRAND_TERMS.values())
brand_list = list(BRAND_TERMS.keys())
all_dfs    = {}
date_index = None

# Fetch ONE brand at a time — Google throttles batch requests heavily
for brand, term in BRAND_TERMS.items():
    print(f"Fetching: {brand} ({term!r}) ...", end=" ", flush=True)
    for attempt in range(1, 4):
        try:
            pytrends.build_payload([term], cat=0, timeframe=TIMEFRAME, geo=GEO)
            df = pytrends.interest_over_time()
            if "isPartial" in df.columns:
                df = df.drop(columns=["isPartial"])
            if date_index is None and not df.empty:
                date_index = [d.strftime("%Y-%m-%d") for d in df.index]
            all_dfs[brand] = df[term].tolist() if (term in df.columns and not df.empty) else []
            print(f"✓  {len(all_dfs[brand])} weeks")
            break
        except Exception as e:
            err = str(e)[:80]
            if attempt < 3:
                wait = 20 * attempt
                print(f"⏳ 429 — waiting {wait}s (attempt {attempt}/3)...", end=" ", flush=True)
                time.sleep(wait)
            else:
                print(f"⚠ failed: {err}")
                all_dfs[brand] = []
    # Polite delay between each brand — essential to avoid throttle
    time.sleep(15)

if not date_index:
    print("❌ Could not retrieve date index from Google Trends.")
    sys.exit(1)

print(f"\n✓ Date range: {date_index[0]}  →  {date_index[-1]}")
print(f"  {len(date_index)} weekly data points\n")

# ── 2. Summary statistics ─────────────────────────────────────
summary = {}
for brand, vals in all_dfs.items():
    if vals:
        avg_v   = round(sum(vals) / len(vals), 1)
        peak_v  = max(vals)
        peak_wk = date_index[vals.index(peak_v)] if peak_v in vals else "N/A"
        last_v  = vals[-1]
        # trend: compare last 4 weeks vs. first 4 weeks
        trend = round(((sum(vals[-4:]) / 4) - (sum(vals[:4]) / 4)), 1) if len(vals) >= 8 else 0
        summary[brand] = {
            "avg": avg_v, "peak": peak_v, "peak_week": peak_wk,
            "last": last_v, "trend": trend, "vals": vals,
        }
        direction = "▲" if trend > 0 else ("▼" if trend < 0 else "→")
        print(f"  {brand:<20}  avg={avg_v:>5}  peak={peak_v:>3}  last4wk trend: {direction} {abs(trend):.1f}")
    else:
        summary[brand] = {"avg":0,"peak":0,"peak_week":"N/A","last":0,"trend":0,"vals":[]}
        print(f"  {brand:<20}  ⚠ no data")

# ── 3. Open workbook and write sheet ─────────────────────────
wb = openpyxl.load_workbook(WORKBOOK_FILE)

# Remove existing sheet if present
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]

ws = wb.create_sheet(SHEET_NAME, 5)   # insert after Meta Ad Library
ws.sheet_view.showGridLines = True

n_brands = len(brand_list)
n_weeks  = len(date_index)
TOTAL_COLS = 1 + n_brands   # Week + one col per brand

# ── Title ──────────────────────────────────────────────────────
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS + 1)
ws["A1"] = (f"Google Trends — Search Interest Index: Pen Brands in India  "
            f"| {date_index[0]} to {date_index[-1]}  |  100 = peak interest")
ws["A1"].font     = bf(13, HDR_FG)
ws["A1"].fill     = fill(HDR_BG)
ws["A1"].alignment = ctr()
ws.row_dimensions[1].height = 34

# ── Summary block (rows 3–9) ───────────────────────────────────
ws.merge_cells("A3:A3")
ws["A3"] = "BRAND SUMMARY"
ws["A3"].font = bf(10, HDR_FG); ws["A3"].fill = fill(HDR_BG)
ws["A3"].alignment = ctr(); ws.row_dimensions[3].height = 22

sum_hdrs = ["Brand", "Search Term", "Avg Interest (0-100)",
            "Peak Score", "Peak Week", "Latest Score", "12M Trend"]
sum_widths = [18, 22, 22, 14, 14, 15, 14]
for col, (h, w) in enumerate(zip(sum_hdrs, sum_widths), 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = bf(10, HDR_FG); c.fill = fill(HDR_BG)
    c.alignment = ctr(); c.border = bdr
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[4].height = 24

for i, (brand, s) in enumerate(summary.items(), 1):
    row = 4 + i
    bc  = BRAND_COLORS.get(brand, "444444")
    direction = "▲ Rising" if s["trend"] > 2 else ("▼ Falling" if s["trend"] < -2 else "→ Stable")
    trend_color = "1E5631" if s["trend"] > 2 else ("8B0000" if s["trend"] < -2 else "333333")

    vals = [
        brand,
        BRAND_TERMS.get(brand, ""),
        s["avg"],
        s["peak"],
        s["peak_week"],
        s["last"],
        direction,
    ]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val); c.border = bdr
        if col == 1:
            c.fill = fill(bc); c.font = bf(10, "FFFFFF"); c.alignment = lft()
        elif col in (3, 4, 6):
            c.font = bf(11 if col == 3 else 10); c.alignment = ctr()
            # Conditional colour: high = green, low = red
            if isinstance(val, float) and val >= 50:
                c.font = bf(11 if col == 3 else 10, "1E5631")
            elif isinstance(val, float) and val < 10:
                c.font = bf(11 if col == 3 else 10, "8B0000")
        elif col == 7:
            c.font = bf(10, trend_color); c.alignment = ctr()
        else:
            c.font = nf(10); c.alignment = ctr()
    ws.row_dimensions[row].height = 22

# ── Weekly data table (starts at row 13) ─────────────────────
DATA_START_ROW = 13

ws.merge_cells(start_row=DATA_START_ROW - 1, start_column=1,
               end_row=DATA_START_ROW - 1, end_column=TOTAL_COLS)
ws.cell(row=DATA_START_ROW - 1, column=1).value = "WEEKLY SEARCH INTEREST  (0–100 index, 100 = all-time peak in period)"
ws.cell(row=DATA_START_ROW - 1, column=1).font  = bf(10, HDR_FG)
ws.cell(row=DATA_START_ROW - 1, column=1).fill  = fill(HDR_BG)
ws.cell(row=DATA_START_ROW - 1, column=1).alignment = lft()
ws.row_dimensions[DATA_START_ROW - 1].height = 22

# Column headers: Week | Brand1 | Brand2 ...
ws.column_dimensions["A"].width = 14
ws.cell(row=DATA_START_ROW, column=1, value="Week").font = bf(10, HDR_FG)
ws.cell(row=DATA_START_ROW, column=1).fill = fill(HDR_BG)
ws.cell(row=DATA_START_ROW, column=1).alignment = ctr()
ws.cell(row=DATA_START_ROW, column=1).border = bdr

for col, brand in enumerate(brand_list, 2):
    bc = BRAND_COLORS.get(brand, "444444")
    c  = ws.cell(row=DATA_START_ROW, column=col, value=brand)
    c.font = bf(10, "FFFFFF"); c.fill = fill(bc)
    c.alignment = ctr(); c.border = bdr
    ws.column_dimensions[get_column_letter(col)].width = 16
ws.row_dimensions[DATA_START_ROW].height = 24
ws.freeze_panes = f"A{DATA_START_ROW + 1}"

# Data rows
for week_i, date_str in enumerate(date_index):
    row = DATA_START_ROW + 1 + week_i
    # Week cell
    c = ws.cell(row=row, column=1, value=date_str)
    c.font = nf(10); c.alignment = ctr(); c.border = bdr
    ws.row_dimensions[row].height = 18

    for col, brand in enumerate(brand_list, 2):
        vals_list = summary[brand]["vals"]
        val = vals_list[week_i] if week_i < len(vals_list) else None
        c   = ws.cell(row=row, column=col, value=val)
        c.border = bdr; c.alignment = ctr()
        if val is None:
            c.font = nf(10, "AAAAAA")
        elif val >= 75:
            c.font = bf(10, "1E5631"); c.fill = fill("D5F5E3")
        elif val >= 40:
            c.font = nf(10, "1A5276"); c.fill = fill("D6EAF8")
        elif val <= 5:
            c.font = nf(10, "AAAAAA")
        else:
            c.font = nf(10)

# ── Save ──────────────────────────────────────────────────────
wb.save(WORKBOOK_FILE)

print(f"\n{'═'*60}")
print(f"  ✅ 'Google Trends' sheet written to {WORKBOOK_FILE}")
print(f"  {n_brands} brands  |  {n_weeks} weekly data points each")
print(f"{'═'*60}\n")
