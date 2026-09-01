"""
Build comprehensive Master Excel for 10 Pune & Hyderabad Malls:
pune_hyderabad_malls_master_analysis.xlsx
"""

import sys, json, os, openpyxl
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

with open("pune_hyderabad_malls_1year_dataset.json", encoding="utf-8") as f:
    raw_data = json.load(f)

malls_results = raw_data["malls_results"]
creators_roster = raw_data["creators_roster"]

# Flatten all collabs
all_collabs = []
for mr in malls_results:
    for c in mr["collabs"]:
        all_collabs.append(c)

# Map creator profile info for fast lookup
creator_map = {c["raw_handle"]: c for c in creators_roster}

for c in all_collabs:
    prof = creator_map.get(c["raw_handle"], {})
    c["full_name"] = prof.get("full_name", c["raw_handle"])
    c["followers"] = prof.get("followers", 0)
    c["creator_tier"] = prof.get("tier", "🌱 Nano (<10K)")

# Sort all collabs by Tier ascending (Tier 1 -> Tier 2 -> Tier 3 -> Tier 4), then Views descending
tier_rank = {
    "Tier 1: Toggle ON + Boosted Paid Ad": 1,
    "Tier 2: Toggle ON + Organic Collab": 2,
    "Tier 3: Toggle OFF + Heavily Boosted Ad": 3,
    "Tier 4: Toggle OFF + Organic / Noise": 4
}
all_collabs.sort(key=lambda x: (tier_rank.get(x["tier"], 4), -x["views"]))

# Workbook Setup
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Styles
font_title = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
font_sub = Font(name="Calibri", size=9, italic=True, color="E0E0E0")
font_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_norm = Font(name="Calibri", size=10, bold=False, color="000000")
font_mute = Font(name="Calibri", size=9, bold=False, color="5D6D7E")
font_link = Font(name="Calibri", size=10, bold=False, color="0563C1", underline="single")

thin_line = Side(style="thin", color="D5D8DC")
border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

# Tier Banner Fills
fill_t1_banner = PatternFill("solid", fgColor="145A32") # Dark Emerald
fill_t1_row = PatternFill("solid", fgColor="D4EFDF") # Mint Green
fill_t2_banner = PatternFill("solid", fgColor="1E8449") # Forest Green
fill_t2_row = PatternFill("solid", fgColor="EAFAF1") # Soft Green
fill_t3_banner = PatternFill("solid", fgColor="B7950B") # Warm Gold
fill_t3_row = PatternFill("solid", fgColor="FEF9E7") # Light Gold
fill_t4_banner = PatternFill("solid", fgColor="566573") # Slate Gray
fill_t4_row = PatternFill("solid", fgColor="FFFFFF")

# ==============================================================================
# TAB 1: EXECUTIVE SUMMARY
# ==============================================================================
ws_sum = wb.create_sheet("Executive Summary")
ws_sum.sheet_view.showGridLines = True
ws_sum.merge_cells("A1:N1")
ws_sum["A1"] = "Executive Summary — 1-Year Creator Collaborations & 4-Tier Hierarchy across 10 Major Malls (Pune vs Hyderabad)"
ws_sum["A1"].font = font_title
ws_sum["A1"].fill = PatternFill("solid", fgColor="0B2240")
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 30

sum_headers = [
    ("#", 5), ("Mall Name", 34), ("City", 14), ("Total Collab Posts", 16),
    ("Unique Creators", 16), ("🟢 Tier 1: Toggle ON + Boosted", 22),
    ("🟢 Tier 2: Toggle ON + Organic", 22), ("🚀 Tier 3: Toggle OFF + Boosted", 22),
    ("⚪ Tier 4: Toggle OFF + Organic", 22), ("💎 High-Intent Paid (T1+T2+T3)", 22),
    ("Paid Collabs %", 15), ("Total Video Views", 18), ("Avg Views / Post", 16),
    ("Top Creator Partnerships (Sample)", 40)
]

for c_idx, (h_text, w) in enumerate(sum_headers, 1):
    c = ws_sum.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="1B2631")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_cell
    ws_sum.column_dimensions[get_column_letter(c_idx)].width = w
ws_sum.row_dimensions[2].height = 28
ws_sum.freeze_panes = "A3"

# Aggregate stats per mall
row_idx = 3
pune_totals = {"collabs": 0, "creators": set(), "t1": 0, "t2": 0, "t3": 0, "t4": 0, "views": 0}
hyd_totals = {"collabs": 0, "creators": set(), "t1": 0, "t2": 0, "t3": 0, "t4": 0, "views": 0}

for idx, mr in enumerate(malls_results, 1):
    m_name = mr["mall_name"]
    city = mr["city"]
    collabs = mr["collabs"]
    
    t1_cnt = sum(1 for c in collabs if "Tier 1" in c["tier"])
    t2_cnt = sum(1 for c in collabs if "Tier 2" in c["tier"])
    t3_cnt = sum(1 for c in collabs if "Tier 3" in c["tier"])
    t4_cnt = sum(1 for c in collabs if "Tier 4" in c["tier"])
    paid_cnt = t1_cnt + t2_cnt + t3_cnt
    tot_views = sum(c["views"] for c in collabs)
    avg_views = int(tot_views / len(collabs)) if collabs else 0
    paid_pct = (paid_cnt / len(collabs) * 100.0) if collabs else 0.0
    
    # Top creators
    top_c = ", ".join([f"{c['handle']}" for c in mr["creators"][:3]])

    # Update regional totals
    target_tot = pune_totals if city == "Pune" else hyd_totals
    target_tot["collabs"] += len(collabs)
    for c in mr["creators"]: target_tot["creators"].add(c["raw_handle"])
    target_tot["t1"] += t1_cnt; target_tot["t2"] += t2_cnt; target_tot["t3"] += t3_cnt; target_tot["t4"] += t4_cnt
    target_tot["views"] += tot_views

    vals = [
        idx, m_name, city, len(collabs), len(mr["creators"]),
        t1_cnt, t2_cnt, t3_cnt, t4_cnt, paid_cnt,
        f"{paid_pct:.1f}%", tot_views, avg_views, top_c
    ]
    
    for c_idx, v in enumerate(vals, 1):
        cell = ws_sum.cell(row=row_idx, column=c_idx, value=v)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (2, 14): cell.font = font_bold if c_idx == 2 else font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (3, 11): cell.font = font_bold; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in range(4, 11):
            cell.font = font_bold if c_idx in (4, 5, 10) else font_norm
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "#,##0"
            if c_idx == 10 and v > 0: cell.fill = PatternFill("solid", fgColor="D4EFDF")
        elif c_idx in (12, 13):
            cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "#,##0"

    ws_sum.row_dimensions[row_idx].height = 20
    row_idx += 1

# Add Pune Subtotal Row
ws_sum.merge_cells(f"A{row_idx}:C{row_idx}")
ws_sum.cell(row=row_idx, column=1, value="SUBTOTAL: PUNE MALLS (5 Malls)").font = font_bold
ws_sum.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
p_paid = pune_totals["t1"] + pune_totals["t2"] + pune_totals["t3"]
p_paid_pct = (p_paid / pune_totals["collabs"] * 100.0) if pune_totals["collabs"] else 0.0
p_avg = int(pune_totals["views"] / pune_totals["collabs"]) if pune_totals["collabs"] else 0

p_vals = [
    pune_totals["collabs"], len(pune_totals["creators"]), pune_totals["t1"], pune_totals["t2"],
    pune_totals["t3"], pune_totals["t4"], p_paid, f"{p_paid_pct:.1f}%", pune_totals["views"], p_avg, "—"
]
for c_idx, v in enumerate(p_vals, 4):
    cell = ws_sum.cell(row=row_idx, column=c_idx, value=v)
    cell.font = font_bold
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if isinstance(v, int): cell.number_format = "#,##0"

for col in range(1, 15):
    ws_sum.cell(row=row_idx, column=col).border = border_cell
    ws_sum.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="E8F8F5")
ws_sum.row_dimensions[row_idx].height = 22
row_idx += 1

# Add Hyderabad Subtotal Row
ws_sum.merge_cells(f"A{row_idx}:C{row_idx}")
ws_sum.cell(row=row_idx, column=1, value="SUBTOTAL: HYDERABAD MALLS (5 Malls)").font = font_bold
ws_sum.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
h_paid = hyd_totals["t1"] + hyd_totals["t2"] + hyd_totals["t3"]
h_paid_pct = (h_paid / hyd_totals["collabs"] * 100.0) if hyd_totals["collabs"] else 0.0
h_avg = int(hyd_totals["views"] / hyd_totals["collabs"]) if hyd_totals["collabs"] else 0

h_vals = [
    hyd_totals["collabs"], len(hyd_totals["creators"]), hyd_totals["t1"], hyd_totals["t2"],
    hyd_totals["t3"], hyd_totals["t4"], h_paid, f"{h_paid_pct:.1f}%", hyd_totals["views"], h_avg, "—"
]
for c_idx, v in enumerate(h_vals, 4):
    cell = ws_sum.cell(row=row_idx, column=c_idx, value=v)
    cell.font = font_bold
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if isinstance(v, int): cell.number_format = "#,##0"

for col in range(1, 15):
    ws_sum.cell(row=row_idx, column=col).border = border_cell
    ws_sum.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="FEF9E7")
ws_sum.row_dimensions[row_idx].height = 22
row_idx += 1

# Grand Total Row
tot_collabs = len(all_collabs)
tot_creators = len(creators_roster)
g_t1 = pune_totals["t1"] + hyd_totals["t1"]
g_t2 = pune_totals["t2"] + hyd_totals["t2"]
g_t3 = pune_totals["t3"] + hyd_totals["t3"]
g_t4 = pune_totals["t4"] + hyd_totals["t4"]
g_paid = g_t1 + g_t2 + g_t3
g_paid_pct = (g_paid / tot_collabs * 100.0) if tot_collabs else 0.0
g_views = pune_totals["views"] + hyd_totals["views"]
g_avg = int(g_views / tot_collabs) if tot_collabs else 0

ws_sum.merge_cells(f"A{row_idx}:C{row_idx}")
ws_sum.cell(row=row_idx, column=1, value="GRAND TOTAL: ALL 10 MALLS").font = font_title
ws_sum.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")

g_vals = [tot_collabs, tot_creators, g_t1, g_t2, g_t3, g_t4, g_paid, f"{g_paid_pct:.1f}%", g_views, g_avg, "—"]
for c_idx, v in enumerate(g_vals, 4):
    cell = ws_sum.cell(row=row_idx, column=c_idx, value=v)
    cell.font = font_title
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if isinstance(v, int): cell.number_format = "#,##0"

for col in range(1, 15):
    ws_sum.cell(row=row_idx, column=col).border = border_cell
    ws_sum.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="1B2631")
ws_sum.row_dimensions[row_idx].height = 24


# ==============================================================================
# TAB 2: MASTER HIERARCHY (ALL COLLABS)
# ==============================================================================
ws_mast = wb.create_sheet("Master Hierarchy (All Collabs)")
ws_mast.sheet_view.showGridLines = True
ws_mast.merge_cells("A1:O1")
ws_mast["A1"] = f"Master Collaborations Roster — All 10 Pune & Hyderabad Malls ({len(all_collabs)} Collab Posts)"
ws_mast["A1"].font = font_title
ws_mast["A1"].fill = PatternFill("solid", fgColor="0B2240")
ws_mast["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_mast.row_dimensions[1].height = 30

mast_headers = [
    ("#", 5), ("City", 12), ("Mall Name", 30), ("Creator Handle", 24),
    ("Creator Full Name", 26), ("Audience Followers", 18), ("Creator Scale Tier", 24),
    ("Date", 12), ("Video Views", 15), ("Likes", 14), ("Comments", 12),
    ("Like-to-View %", 14), ("Paid Partnership Toggle", 22), ("4-Tier Classification", 30), ("Instagram Post URL", 42)
]

for c_idx, (h_text, w) in enumerate(mast_headers, 1):
    c = ws_mast.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="1B2631")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_cell
    ws_mast.column_dimensions[get_column_letter(c_idx)].width = w
ws_mast.row_dimensions[2].height = 24
ws_mast.freeze_panes = "A3"

for idx, post in enumerate(all_collabs, 1):
    r_num = idx + 2
    r_vals = [
        idx, post["city"], post["mall_name"], post["creator_handle"], post["full_name"],
        post["followers"], post["creator_tier"], post["date"], post["views"], post["likes"],
        post["comments"], f"{post['like_to_view_pct']:.2f}%",
        "Toggle ON" if post["is_paid_toggle"] else "Toggle OFF",
        post["tier"], post["post_url"]
    ]
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws_mast.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (2, 8, 13): cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (3, 4, 5): cell.font = font_bold if c_idx == 4 else font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx in (6, 9, 10, 11): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx == 7: cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 12: cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center")
        elif c_idx == 14:
            cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
            if "Tier 1" in str(val): cell.fill = fill_t1_row
            elif "Tier 2" in str(val): cell.fill = fill_t2_row
            elif "Tier 3" in str(val): cell.fill = fill_t3_row
        elif c_idx == 15: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None

    ws_mast.row_dimensions[r_num].height = 20


# ==============================================================================
# TAB 3: UNIQUE CREATOR ROSTER
# ==============================================================================
ws_roster = wb.create_sheet("Unique Creator Roster")
ws_roster.sheet_view.showGridLines = True
ws_roster.merge_cells("A1:J1")
ws_roster["A1"] = f"Master Creator Directory across Pune & Hyderabad Malls ({len(creators_roster)} Unique Creators)"
ws_roster["A1"].font = font_title
ws_roster["A1"].fill = PatternFill("solid", fgColor="0B2240")
ws_roster["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_roster.row_dimensions[1].height = 30

roster_headers = [
    ("#", 5), ("Creator Handle", 24), ("Full Name", 26), ("Audience Followers", 18),
    ("Audience Scale Tier", 26), ("City", 14), ("Primary Mall Collaborated", 32),
    ("Total Collab Posts", 16), ("Total Video Views", 18), ("Instagram Profile URL", 42)
]

for c_idx, (h_text, w) in enumerate(roster_headers, 1):
    c = ws_roster.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="1B2631")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_cell
    ws_roster.column_dimensions[get_column_letter(c_idx)].width = w
ws_roster.row_dimensions[2].height = 24
ws_roster.freeze_panes = "A3"

for idx, c in enumerate(creators_roster, 1):
    r_num = idx + 2
    r_vals = [
        idx, c["handle"], c["full_name"], c["followers"], c["tier"], c["city"],
        c["primary_mall"], c["total_posts"], c["total_views"], c["profile_url"]
    ]
    for c_idx, val in enumerate(r_vals, 1):
        cell = ws_roster.cell(row=r_num, column=c_idx, value=val)
        cell.border = border_cell
        if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in (2, 3): cell.font = font_bold if c_idx == 2 else font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c_idx == 4: cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx in (5, 6, 7): cell.font = font_bold if c_idx in (5, 6) else font_norm; cell.alignment = Alignment(horizontal="left" if c_idx != 6 else "center", vertical="center")
        elif c_idx in (8, 9): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
        elif c_idx == 10: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None

    ws_roster.row_dimensions[r_num].height = 20


# ==============================================================================
# TABS 4 TO 13: 10 INDIVIDUAL MALL TABS
# ==============================================================================
short_tab_names = {
    "Phoenix Avenue of Stars Pune": "Phoenix Avenue Pune",
    "Seasons Mall Pune": "Seasons Mall Pune",
    "The Pavillion Pune": "The Pavillion Pune",
    "Phoenix Mall of the Millennium Wakad": "Phoenix Millennium Wakad",
    "Amanora Mall Pune": "Amanora Mall Pune",
    "Nexus Hyderabad Mall": "Nexus Hyderabad",
    "Sarath City Capital Mall Hyderabad": "Sarath City Capital",
    "Inorbit Mall Cyberabad": "Inorbit Cyberabad",
    "Lulu Mall Hyderabad": "Lulu Mall Hyderabad",
    "GVK One Mall Hyderabad": "GVK One Hyderabad"
}

for mr in malls_results:
    m_name = mr["mall_name"]
    tab_name = short_tab_names.get(m_name, m_name[:28])
    collabs = mr["collabs"]
    
    # Sort by tier, views
    collabs.sort(key=lambda x: (tier_rank.get(x["tier"], 4), -x["views"]))

    ws_m = wb.create_sheet(tab_name)
    ws_m.sheet_view.showGridLines = True
    ws_m.merge_cells("A1:L1")
    ws_m["A1"] = f"{m_name} — 1-Year Collaborations Portfolio ({len(collabs)} Collab Posts)"
    ws_m["A1"].font = font_title
    ws_m["A1"].fill = PatternFill("solid", fgColor="1B4F72" if mr["city"] == "Pune" else "78281F")
    ws_m["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_m.row_dimensions[1].height = 28

    m_headers = [
        ("#", 5), ("Creator Handle", 24), ("Date", 12), ("Views", 15), ("Likes", 14),
        ("Comments", 12), ("Like-to-View %", 14), ("Paid Partnership Toggle", 22),
        ("4-Tier Classification", 30), ("Audio Track", 28), ("Caption Snippet", 45), ("Post URL", 40)
    ]
    for c_idx, (h_text, w) in enumerate(m_headers, 1):
        c = ws_m.cell(row=2, column=c_idx, value=h_text)
        c.font = font_hdr
        c.fill = PatternFill("solid", fgColor="283747")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_cell
        ws_m.column_dimensions[get_column_letter(c_idx)].width = w
    ws_m.row_dimensions[2].height = 24
    ws_m.freeze_panes = "A3"

    for idx, post in enumerate(collabs, 1):
        r_num = idx + 2
        r_vals = [
            idx, post["creator_handle"], post["date"], post["views"], post["likes"],
            post["comments"], f"{post['like_to_view_pct']:.2f}%",
            "Toggle ON" if post["is_paid_toggle"] else "Toggle OFF",
            post["tier"], post["audio_track"], post["caption"][:120], post["post_url"]
        ]
        for c_idx, val in enumerate(r_vals, 1):
            cell = ws_m.cell(row=r_num, column=c_idx, value=val)
            cell.border = border_cell
            if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in (2, 10, 11): cell.font = font_bold if c_idx == 2 else font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx in (3, 8): cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in (4, 5, 6): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
            elif c_idx == 7: cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx == 9:
                cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
                if "Tier 1" in str(val): cell.fill = fill_t1_row
                elif "Tier 2" in str(val): cell.fill = fill_t2_row
                elif "Tier 3" in str(val): cell.fill = fill_t3_row
            elif c_idx == 12: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None

        ws_m.row_dimensions[r_num].height = 20


# ==============================================================================
# TAB 14: ALL POSTS DEEP-DIVE (3,481 POSTS)
# ==============================================================================
ws_all = wb.create_sheet("All Posts Deep-Dive (Content)")
ws_all.sheet_view.showGridLines = True
ws_all.merge_cells("A1:N1")

total_all_posts = sum(len(mr["all_posts"]) for mr in malls_results)
ws_all["A1"] = f"Complete Content & Transcript Feed Repository — All 10 Malls ({total_all_posts} Total Posts)"
ws_all["A1"].font = font_title
ws_all["A1"].fill = PatternFill("solid", fgColor="17202A")
ws_all["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_all.row_dimensions[1].height = 28

all_headers = [
    ("#", 5), ("City", 12), ("Mall Name", 30), ("Date", 12), ("Is Collab?", 14),
    ("Creator Handle", 22), ("Views", 15), ("Likes", 14), ("Comments", 12),
    ("Like-to-View %", 14), ("4-Tier Status", 28), ("Audio / Music Track", 28),
    ("Full Caption / Description Text", 60), ("Instagram Post URL", 40)
]

for c_idx, (h_text, w) in enumerate(all_headers, 1):
    c = ws_all.cell(row=2, column=c_idx, value=h_text)
    c.font = font_hdr
    c.fill = PatternFill("solid", fgColor="2C3E50")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_cell
    ws_all.column_dimensions[get_column_letter(c_idx)].width = w
ws_all.row_dimensions[2].height = 24
ws_all.freeze_panes = "A3"

post_counter = 1
for mr in malls_results:
    for post in mr["all_posts"]:
        r_num = post_counter + 2
        r_vals = [
            post_counter, post["city"], post["mall_name"], post["date"],
            "Yes (Collab)" if post["is_collab"] else "No (Mall Post)",
            post["creator_handle"], post["views"], post["likes"], post["comments"],
            f"{post['like_to_view_pct']:.2f}%", post["tier"] if post["is_collab"] else "— (Mall Owned)",
            post["audio_track"], post["caption"], post["post_url"]
        ]
        for c_idx, val in enumerate(r_vals, 1):
            cell = ws_all.cell(row=r_num, column=c_idx, value=val)
            cell.border = border_cell
            if c_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in (2, 4, 5): cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in (3, 6, 12, 13): cell.font = font_bold if c_idx == 6 and val != "— (Mall Owned)" else font_norm; cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx in (7, 8, 9): cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center"); cell.number_format = "#,##0"
            elif c_idx == 10: cell.font = font_norm; cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx == 11:
                cell.font = font_bold; cell.alignment = Alignment(horizontal="left", vertical="center")
                if "Tier 1" in str(val): cell.fill = fill_t1_row
                elif "Tier 2" in str(val): cell.fill = fill_t2_row
                elif "Tier 3" in str(val): cell.fill = fill_t3_row
            elif c_idx == 14: cell.font = font_link; cell.alignment = Alignment(horizontal="left", vertical="center"); cell.hyperlink = val if val else None

        ws_all.row_dimensions[r_num].height = 20
        post_counter += 1

output_excel = "pune_hyderabad_malls_master_analysis.xlsx"
wb.save(output_excel)
print(f"\n✓ Master Workbook saved successfully: {output_excel}")
