"""
Extract and fetch the latest post dates for all analyzed brands:
1. Official Brand Instagram Pages (latest post date, caption, URL, likes)
2. Latest Detected Partnership/Collab post date per brand from raw partnership logs
"""

import sys, json, time
from datetime import datetime, timezone
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

from scrape_profiles import make_session, fetch_profile_info, fetch_posts_v2

# Brand definitions
BRANDS = {
    # Category 1: Luxury & Premium Writing Brands
    "Montblanc":         {"handle": "montblanc", "category": "Pens & Luxury"},
    "Parker Pens India": {"handle": "parkerpenindia", "category": "Pens & Luxury"},
    "Sheaffer Pen":      {"handle": "sheafferpen", "category": "Pens & Luxury"},
    "Lamy India":        {"handle": "lamy_india", "category": "Pens & Luxury"},
    "Makoba India":      {"handle": "makobaindia", "category": "Pens & Luxury"},
    "Submarine Pens":    {"handle": "submarine_pens", "category": "Pens & Luxury"},
    
    # Category 2: Electronics Retail Chains
    "Vijay Sales":       {"handle": "vijaysales", "category": "Electronics Retail"},
    "Croma Retail":      {"handle": "croma.retail", "category": "Electronics Retail"},
    "Reliance Digital":  {"handle": "reliancedigital", "category": "Electronics Retail"},
}

session = make_session()
official_results = []

print(f"\n{'='*75}")
print("  Fetching Latest Post Dates for Official Brand Accounts")
print(f"{'='*75}\n")

for brand_name, meta in BRANDS.items():
    handle = meta["handle"]
    cat = meta["category"]
    print(f"Checking @{handle} ({brand_name}) ...", end=" ", flush=True)
    
    info = fetch_profile_info(handle, session)
    user = info.get("data", {}).get("user")
    
    if user:
        user_id = user.get("id")
        posts = fetch_posts_v2(user_id, session, count=5)
        if posts:
            p0 = posts[0]
            ts = p0.get("taken_at", 0)
            date_str = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d") if ts else "N/A"
            code = p0.get("code", p0.get("shortcode", ""))
            likes = p0.get("like_count", 0) or 0
            comments = p0.get("comment_count", 0) or 0
            cap_obj = p0.get("caption")
            cap = (cap_obj.get("text", "") if isinstance(cap_obj, dict) else "") or ""
            preview = cap.splitlines()[0][:70] if cap else "No text"
            
            official_results.append({
                "brand": brand_name,
                "category": cat,
                "handle": f"@{handle}",
                "followers": user.get("edge_followed_by", {}).get("count", 0),
                "total_posts_account": user.get("edge_owner_to_timeline_media", {}).get("count", 0),
                "last_post_date": date_str,
                "post_url": f"https://www.instagram.com/p/{code}/",
                "likes": likes,
                "comments": comments,
                "caption_preview": preview,
            })
            print(f"✓ Last post: {date_str} | Likes: {likes:,} | URL: https://www.instagram.com/p/{code}/")
        else:
            print("⚠ No posts returned")
    else:
        print("⚠ Profile not found")
    
    time.sleep(1.5)

# 2. Extract Latest Partnership Post Dates from Raw Partnership Posts sheets
collab_dates = {}

# Pens workbook
try:
    wb_pens = openpyxl.load_workbook("competitor_paid_media_analysis.xlsx", read_only=True)
    if "Raw Partnership Posts" in wb_pens.sheetnames:
        ws = wb_pens["Raw Partnership Posts"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            b_name, url, d_val = row[0], row[1], row[2]
            if b_name and d_val and d_val != "None":
                d_str = str(d_val)[:10]
                if b_name not in collab_dates or d_str > collab_dates[b_name]:
                    collab_dates[b_name] = d_str
except Exception as e:
    print(f"Note pens wb: {e}")

# Retail workbook
try:
    wb_ret = openpyxl.load_workbook("retail_brands_competitor_analysis.xlsx", read_only=True)
    if "Raw Partnership Posts" in wb_ret.sheetnames:
        ws = wb_ret["Raw Partnership Posts"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            b_name, url, d_val = row[0], row[1], row[2]
            if b_name and d_val and d_val != "None":
                d_str = str(d_val)[:10]
                if b_name not in collab_dates or d_str > collab_dates[b_name]:
                    collab_dates[b_name] = d_str
except Exception as e:
    print(f"Note retail wb: {e}")

# Output summary JSON for reference
output_data = {
    "official_brand_last_posts": official_results,
    "latest_detected_partnership_posts": collab_dates,
    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
}

with open("brand_last_post_dates.json", "w", encoding="utf-8") as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print("\n✓ Saved results to brand_last_post_dates.json")
