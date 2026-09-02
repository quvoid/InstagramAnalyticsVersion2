"""
Verify lakeshore_executive_omnichannel_master_research.xlsx
"""

import openpyxl

wb = openpyxl.load_workbook("lakeshore_executive_omnichannel_master_research.xlsx")
print("="*70)
print("WORKBOOK VERIFICATION: lakeshore_executive_omnichannel_master_research.xlsx")
print("="*70)
print(f"Total Sheets: {len(wb.sheetnames)}\n")

for name in wb.sheetnames:
    ws = wb[name]
    print(f"  • Tab: {name:<38} | Rows: {ws.max_row:<4} | Charts: {len(ws._charts)}")

print("="*70)
