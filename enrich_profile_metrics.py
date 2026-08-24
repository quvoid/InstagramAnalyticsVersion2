"""
Enrich Creator Profile Metrics with exact Collab Likes, Comments, and Engagement Rates
"""

import sys, json, openpyxl, csv
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

with open("footwear_1year_4tier_dataset.json", encoding="utf-8") as f:
    posts = json.load(f)

with open("footwear_creators_profile_metrics.json", encoding="utf-8") as f:
    profiles = json.load(f)

# Aggregate stats from posts per creator
creator_post_stats = defaultdict(lambda: {"likes": [], "comments": [], "views": [], "posts_count": 0})

for p in posts:
    rh = p["raw_handle"].lower()
    creator_post_stats[rh]["likes"].append(p["likes"])
    creator_post_stats[rh]["comments"].append(p["comments"])
    creator_post_stats[rh]["views"].append(p["views"])
    creator_post_stats[rh]["posts_count"] += 1

for p in profiles:
    rh = p["raw_handle"].lower()
    stats = creator_post_stats.get(rh)
    if stats:
        avg_l = int(sum(stats["likes"]) / len(stats["likes"])) if stats["likes"] else 0
        avg_c = int(sum(stats["comments"]) / len(stats["comments"])) if stats["comments"] else 0
        fols = p["followers"]
        avg_er = round(((avg_l + avg_c) / fols) * 100, 2) if fols > 0 else 0.0
        p["avg_likes"] = avg_l
        p["avg_comments"] = avg_c
        p["avg_er"] = avg_er
        p["collab_posts_count"] = stats["posts_count"]

with open("footwear_creators_profile_metrics.json", "w", encoding="utf-8") as f:
    json.dump(profiles, f, indent=2)

# Update CSV and Excel
wb = openpyxl.load_workbook("footwear_sneaker_brands_master_analysis.xlsx")
ws_prof = wb["Creators Profile Metrics"]

thin_line = Side(style="thin", color="D5D8DC")
border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
font_norm = Font(name="Calibri", size=10, bold=False, color="000000")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")

for idx, p in enumerate(profiles, 1):
    r_num = idx + 2
    ws_prof.cell(row=r_num, column=10, value=p["avg_likes"])
    ws_prof.cell(row=r_num, column=11, value=p["avg_comments"])
    c_er = ws_prof.cell(row=r_num, column=12, value=p["avg_er"] / 100 if p["avg_er"] else 0.0)
    c_er.number_format = "0.00%"

wb.save("footwear_sneaker_brands_master_analysis.xlsx")

with open("Footwear_Creator_Profile_Metrics.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow([
        "#", "Creator Handle", "Brands Collaborated With", "Full Name", "Verified", "Business / Professional",
        "Total Followers", "Following", "Total Posts", "Avg Likes / Post", "Avg Comments / Post", "Avg Profile ER%", "Instagram Profile URL"
    ])
    for idx, p in enumerate(profiles, 1):
        w.writerow([
            idx,
            p["handle"],
            p["brands"],
            p["full_name"],
            "Yes" if p["verified"] else "No",
            "Yes" if p["is_business"] else "No",
            p["followers"],
            p["following"],
            p["total_posts"],
            p["avg_likes"],
            p["avg_comments"],
            f"{p['avg_er']:.2f}%",
            p["profile_url"]
        ])

print("✓ Successfully enriched profile metrics across Excel and CSV!")
